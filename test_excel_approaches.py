"""
Quick test to demonstrate the difference between formula-driven and static value approaches.

Tests:
1. Formula-driven workbook - can we read formula results immediately?
2. Static value workbook - can we read values immediately?
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
import os

# Clean up any existing test files
for f in ["test_formulas.xlsx", "test_values.xlsx"]:
    if os.path.exists(f):
        os.remove(f)

print("=" * 70)
print("TEST 1: Formula-Driven Workbook (Approach 1)")
print("=" * 70)

# Create workbook with formulas
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Data"

ws1['A1'] = "Value 1"
ws1['B1'] = "Value 2"
ws1['C1'] = "Sum"

ws1['A2'] = 100
ws1['B2'] = 200
ws1['C2'] = "=A2+B2"  # Formula

ws1['A3'] = 150
ws1['B3'] = 250
ws1['C3'] = "=A3+B3"  # Formula

wb1.save("test_formulas.xlsx")
print("✅ Created test_formulas.xlsx with formulas in column C")

# Try to read it back immediately
print("\n📖 Reading back immediately with data_only=True...")
wb1_read = load_workbook("test_formulas.xlsx", data_only=True)
ws1_read = wb1_read.active

print(f"   A2 (value): {ws1_read['A2'].value}")
print(f"   B2 (value): {ws1_read['B2'].value}")
print(f"   C2 (formula result): {ws1_read['C2'].value}")
print(f"   C3 (formula result): {ws1_read['C3'].value}")

if ws1_read['C2'].value is None:
    print("   ❌ Formula cells return None - can't read computed values!")
else:
    print(f"   ✅ Formula computed to: {ws1_read['C2'].value}")

# Try without data_only
print("\n📖 Reading back with data_only=False (default)...")
wb1_read2 = load_workbook("test_formulas.xlsx")
ws1_read2 = wb1_read2.active
print(f"   C2 cell value: {ws1_read2['C2'].value}")
if ws1_read2['C2'].value and str(ws1_read2['C2'].value).startswith("="):
    print("   ⚠️  Got the formula string, not the computed value")

print("\n" + "=" * 70)
print("TEST 2: Static Value Workbook (Approach 2)")
print("=" * 70)

# Create workbook with static values using pandas
df = pd.DataFrame({
    "Value 1": [100, 150],
    "Value 2": [200, 250],
    "Sum": [300, 400]  # Pre-computed values
})

with pd.ExcelWriter("test_values.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Data", index=False)

print("✅ Created test_values.xlsx with pre-computed values")

# Try to read it back immediately
print("\n📖 Reading back immediately with data_only=True...")
wb2_read = load_workbook("test_values.xlsx", data_only=True)
ws2_read = wb2_read.active

print(f"   A2 (Value 1): {ws2_read['A2'].value}")
print(f"   B2 (Value 2): {ws2_read['B2'].value}")
print(f"   C2 (Sum): {ws2_read['C2'].value}")
print(f"   C3 (Sum): {ws2_read['C3'].value}")

if ws2_read['C2'].value is not None:
    print("   ✅ Values are immediately readable!")
else:
    print("   ❌ Unexpected None value")

# Can also read with pandas
print("\n📖 Reading back with pandas...")
df_read = pd.read_excel("test_values.xlsx", sheet_name="Data")
print(df_read)
print("   ✅ Pandas can read it directly too!")

print("\n" + "=" * 70)
print("CONCLUSION")
print("=" * 70)
print("Approach 1 (Formulas): ❌ Can't read computed values immediately")
print("                       → Need JSON replica for downstream scripts")
print("Approach 2 (Values):   ✅ Values readable immediately")
print("                       → No special handling needed")
print("=" * 70)

# Cleanup
os.remove("test_formulas.xlsx")
os.remove("test_values.xlsx")
print("\n🧹 Cleaned up test files")
