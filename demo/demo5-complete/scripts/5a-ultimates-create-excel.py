# Creates a formatted Excel workbook displaying projected ultimate indications from multiple
# reserving methods (Chain Ladder, Bornhuetter-Ferguson, etc.) alongside prior selections.
# Actuaries review this workbook and manually enter their selected ultimate values and reasoning.

"""
goal: Create Ultimates.xlsx for actuarial review and selection of ultimate losses by period.

Each sheet (one per measure) contains:
  - Period and current maturity information
  - Actual reported losses
  - Initial expected ultimate
  - Indicated ultimates from Chain Ladder, BF, and other methods
  - Prior selections and reasoning (when available)
  - Blank selection rows for actuary input

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 5a-ultimates-create-excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pathlib
import json

from modules import config
from modules.xl_styles import (
    HEADER_FILL, HEADER_FONT, SELECTION_FILL, AI_FILL, PRIOR_FILL, USER_FILL, THIN_BORDER,
)

# Paths from modules/config.py — override here if needed:
INPUT_ULTIMATES  = config.ULTIMATES + "projected-ultimates.parquet"
PRIOR_SELECTIONS_RB = config.SELECTIONS + "ultimates-prior.json"             # Optional, priority 1 — set to prior cycle's selected ultimates
PRIOR_SELECTIONS_OE = config.SELECTIONS + "ultimates-prior-oe.json"          # Optional, priority 2 — fallback prior
OUTPUT_FILE      = config.SELECTIONS + "Ultimates.xlsx"


def format_sheet(ws, measure, df_ult, df_prior):
    """
    Format a single ultimates sheet with indicated values and selection areas.
    
    Args:
        ws: openpyxl worksheet object
        measure: Measure name (e.g., 'Incurred Loss', 'Paid Loss')
        df_ult: DataFrame with projected ultimates containing columns:
                period, current_age, actual, ultimate_ie, ultimate_cl, ultimate_bf
        df_prior: DataFrame with prior selections containing columns:
                  period, measure, selection, reasoning (or None if no prior selections)
    """
    # Filter for measure
    df_m = df_ult[df_ult['measure'] == measure].copy()
    if df_m.empty:
        return
    
    # Write headers
    headers = [
        "Period", "Current Age", "Actual", 
        "Initial Expected Ultimate", "Chain Ladder Ultimate", "BF Ultimate",
        "Prior Selection", "Prior Reasoning", 
        "Rules-Based AI Selection", "Rules-Based AI Reasoning",
        "Open-Ended AI Selection", "Open-Ended AI Reasoning",
        "User Selection", "User Reasoning"
    ]
    
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['N'].width = 40
    
    # Create dict of prior
    prior_dict = {}
    if df_prior is not None and measure in df_prior['measure'].values:
        mp = df_prior[df_prior['measure'] == measure]
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {"sel": r.get('selection', r.get('selected_ultimate')), "reason": r.get('reasoning', '')}
            
    # Write rows
    for r_idx, (_, row) in enumerate(df_m.iterrows(), start=2):
        period = str(row['period'])
        
        # Period & current age
        ws.cell(row=r_idx, column=1, value=period).border = THIN_BORDER
        
        val_age = row.get('current_age')
        if pd.isna(val_age): val_age = ""
        ws.cell(row=r_idx, column=2, value=val_age).border = THIN_BORDER
        
        # Values
        for c_idx, col_name in enumerate(['actual', 'ultimate_ie', 'ultimate_cl', 'ultimate_bf'], start=3):
            val = row.get(col_name)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
            
        # Prior
        prior_sel = prior_dict.get(period, {}).get("sel", "")
        prior_reason = prior_dict.get(period, {}).get("reason", "")
        
        c = ws.cell(row=r_idx, column=7, value=prior_sel)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=8, value=prior_reason)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Rules-Based AI Selection (yellow fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=9)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=10)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Open-Ended AI Selection (purple fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=11)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=12)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # User Selection (blank - actuary input)
        c = ws.cell(row=r_idx, column=13)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=14)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)


def main():
    """Create ultimates selection Excel file from projected ultimates."""
    print("Creating Ultimates.xlsx...")
    
    try:
        df_ult = pd.read_parquet(INPUT_ULTIMATES)
        print(f"Loaded ultimates from {INPUT_ULTIMATES}")
    except Exception as e:
        print(f"Error loading ultimates: {e}")
        exit(1)
        
    df_prior = None
    # Try loading prior selections - prioritize rules-based, fall back to open-ended
    prior_loaded = False
    try:
        if pathlib.Path(PRIOR_SELECTIONS_RB).exists():
            with open(PRIOR_SELECTIONS_RB, 'r') as f:
                prior_data = json.load(f)
            df_prior = pd.DataFrame(prior_data)
            print(f"Loaded prior selections from {PRIOR_SELECTIONS_RB} (rules-based)")
            prior_loaded = True
    except Exception as e:
        print(f"  Could not load rules-based prior selections: {e}")
    
    if not prior_loaded:
        try:
            if pathlib.Path(PRIOR_SELECTIONS_OE).exists():
                with open(PRIOR_SELECTIONS_OE, 'r') as f:
                    prior_data = json.load(f)
                df_prior = pd.DataFrame(prior_data)
                print(f"Loaded prior selections from {PRIOR_SELECTIONS_OE} (open-ended)")
        except Exception as e:
            print(f"  No prior selections found: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    measures = ['Incurred Loss', 'Paid Loss', 'Reported Count', 'Closed Count']
    
    for m in measures:
        if m in df_ult['measure'].values:
            ws = wb.create_sheet(title=m)
            format_sheet(ws, m, df_ult, df_prior)
            print(f"  Created sheet for {m}")
            
    out_dir = pathlib.Path(OUTPUT_FILE).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
