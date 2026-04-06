"""
goal: Create an Excel workbook displaying indicated ultimates with spaces for final selection.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pathlib
import json

# File paths
INPUT_ULTIMATES = "../ultimates/projected-ultimates.parquet"
PRIOR_SELECTIONS = "../selections/ultimates.json"  # Optional
OUTPUT_FILE = "../selections/Ultimates.xlsx"

# Formatting Styles
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
SELECTION_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
PRIOR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_sheet(ws, measure, df_ult, df_prior):
    # Filter for measure
    df_m = df_ult[df_ult['measure'] == measure].copy()
    if df_m.empty:
        return
    
    # Write headers
    headers = [
        "Period", "Current Age", "Actual", 
        "Initial Expected Ultimate", "Chain Ladder Ultimate", "BF Ultimate",
        "Prior Selection", "Prior Reasoning", 
        "Selected Ultimate", "Reasoning"
    ]
    
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['J'].width = 40
    
    # Create dict of prior
    prior_dict = {}
    if df_prior is not None and measure in df_prior['measure'].values:
        mp = df_prior[df_prior['measure'] == measure]
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {"sel": r['selection'], "reason": r['reasoning']}
            
    # Write rows
    for r_idx, (_, row) in enumerate(df_m.iterrows(), start=2):
        period = str(row['period'])
        
        # Period & current age
        ws.cell(row=r_idx, column=1, value=period).border = THIN_BORDER
        
        val_age = row.get('current_age')
        if pd.isna(val_age): val_age = ""
        ws.cell(row=r_idx, column=2, value=val_age).border = THIN_BORDER
        
        # Values
        for c_idx, col_name in enumerate(['actual', 'expected_ultimate', 'cl_ultimate', 'bf_ultimate'], start=3):
            val = row.get(col_name)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
            
        # Prior
        prior_sel = prior_dict.get(period, {}).get("sel", "")
        prior_reason = prior_dict.get(period, {}).get("reason", "")
        
        c = ws.cell(row=r_idx, column=7, value=prior_sel)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=8, value=prior_reason)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # New selection
        c = ws.cell(row=r_idx, column=9)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=10)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)

if __name__ == "__main__":
    print("Creating Ultimates.xlsx...")
    
    try:
        df_ult = pd.read_parquet(INPUT_ULTIMATES)
    except Exception as e:
        print(f"Error loading ultimates: {e}")
        exit(1)
        
    df_prior = None
    try:
        if pathlib.Path(PRIOR_SELECTIONS).exists():
            with open(PRIOR_SELECTIONS, 'r') as f:
                prior_data = json.load(f)
            df_prior = pd.DataFrame(prior_data)
    except Exception as e:
        print(f"  No prior selections found or error: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    measures = ['Incurred Loss', 'Paid Loss', 'Reported Count', 'Closed Count']
    
    for m in measures:
        if m in df_ult['measure'].values:
            ws = wb.create_sheet(title=m)
            format_sheet(ws, m, df_ult, df_prior)
            
    out_dir = pathlib.Path(OUTPUT_FILE).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    
    wb.save(OUTPUT_FILE)
    print(f"Successfully created {OUTPUT_FILE}")
