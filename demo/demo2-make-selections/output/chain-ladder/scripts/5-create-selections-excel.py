"""
goal: Create Chain Ladder Selections.xlsx for actuarial LDF review and selection.

Each sheet (one per measure) contains from top to bottom:
  - Loss triangle
  - LDF triangle
  - Diagnostics triangles (one per diagnostic)
  - LDF averages and QA metrics
  - Blank selection rows (with prior selections shown above for reference)

usage: Run from project root:
    .venv/Scripts/Activate.ps1; python demo/demo2-make-selections/output/chain-ladder/scripts/5-create-selections-excel.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
OUTPUT_PATH = str(SCRIPT_DIR / "../data/") + "/"
METHOD_ID = "chainladder"
SELECTIONS_OUTPUT_PATH = str(SCRIPT_DIR / "../selections/") + "/"
OUTPUT_FILE_NAME = "Chain Ladder Selections.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SELECTION_FILL = PatternFill("solid", fgColor="FFF2CC")
PRIOR_FILL = PatternFill("solid", fgColor="E2EFDA")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(bold=True, size=10)
LABEL_FONT = Font(bold=True, size=9)
DATA_FONT = Font(size=9)

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, level="header"):
    if level == "header":
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    elif level == "subheader":
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
    elif level == "section":
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    elif level == "selection":
        cell.fill = SELECTION_FILL
        cell.font = LABEL_FONT
    elif level == "prior":
        cell.fill = PRIOR_FILL
        cell.font = LABEL_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def write_triangle(ws, start_row, title, row_labels, col_labels, data_dict, number_format="#,##0"):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    for row_label in row_labels:
        row_cell = ws.cell(row=start_row, column=1, value=row_label)
        row_cell.font = LABEL_FONT
        row_cell.alignment = Alignment(horizontal="left")
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((str(row_label), str(col)))
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = number_format
        start_row += 1

    return start_row + 1


def write_metrics_table(ws, start_row, title, metric_rows, col_labels, data_dict):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="Metric").border = THIN_BORDER
    ws.cell(row=start_row, column=1).font = SUBHEADER_FONT
    ws.cell(row=start_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    for metric in metric_rows:
        row_cell = ws.cell(row=start_row, column=1, value=metric)
        row_cell.font = LABEL_FONT
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((metric, str(col)))
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = "0.0000"
        start_row += 1

    return start_row + 1


def write_selections_section(ws, start_row, col_labels, prior_selections=None, measure=None):
    title_cell = ws.cell(row=start_row, column=1, value="LDF Selections")
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    if prior_selections is not None and measure is not None:
        measure_prior = prior_selections[prior_selections['measure'] == measure]
        if not measure_prior.empty:
            prior_dict = {}
            for _, row in measure_prior.iterrows():
                prior_dict[str(row['interval'])] = {
                    'selection': row['selection'],
                    'reasoning': row['reasoning']
                }

            cell = ws.cell(row=start_row, column=1, value="Prior Selection")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['selection']
                    c.number_format = "0.0000"
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="right")
            start_row += 1

            cell = ws.cell(row=start_row, column=1, value="Prior Reasoning")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['reasoning']
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="left", wrap_text=True)
            start_row += 1
            start_row += 1  # blank separator

    for label in ["Selection", "Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "selection")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = SELECTION_FILL
            c.border = THIN_BORDER
            if label == "Selection":
                c.number_format = "0.0000"
        start_row += 1

    return start_row


def build_sheet(ws, measure, df2, df3, df4, df_prior=None):
    df_m = df2[df2['measure'] == measure].copy()
    periods = df_m['period'].cat.categories.tolist()
    ages = [str(a) for a in df_m['age'].cat.categories.tolist()]
    intervals = [str(i) for i in df_m['interval'].dropna().cat.categories.tolist()]

    # Loss Triangle
    loss_dict = {(str(r['period']), str(r['age'])): r['value'] for _, r in df_m.iterrows()}
    row_ptr = write_triangle(ws, 1, f"{measure} - Loss Triangle", periods, ages, loss_dict, "#,##0")

    # LDF Triangle
    ldf_dict = {(str(r['period']), str(r['interval'])): r['ldf']
                for _, r in df_m[df_m['ldf'].notna()].iterrows()}
    row_ptr = write_triangle(ws, row_ptr, f"{measure} - LDF Triangle", periods, intervals, ldf_dict, "0.0000")

    # Diagnostics Triangles
    diagnostic_cols = [col for col in df3.columns if col not in ['period', 'age']]
    for diag_col in diagnostic_cols:
        diag_dict = {(str(r['period']), str(r['age'])): r[diag_col] for _, r in df3.iterrows()}
        if df3[diag_col].dropna().abs().max() <= 2.0 and 'rate' in diag_col.lower():
            number_format = "0.00%"
        else:
            number_format = "0.0000"
        diag_title = f"Diagnostic - {diag_col.replace('_', ' ').title()}"
        row_ptr = write_triangle(ws, row_ptr, diag_title, periods, ages, diag_dict, number_format)

    # LDF Averages & QA Metrics
    df_avg = df4[df4['measure'] == measure].copy()
    metric_cols = [col for col in df_avg.columns if col not in ['measure', 'interval']]
    metrics_dict = {}
    for _, row in df_avg.iterrows():
        for metric in metric_cols:
            metrics_dict[(metric, str(row['interval']))] = row[metric]
    row_ptr = write_metrics_table(ws, row_ptr, f"{measure} - LDF Averages & QA Metrics",
                                  metric_cols, intervals, metrics_dict)

    # Selections
    write_selections_section(ws, row_ptr, intervals, prior_selections=df_prior, measure=measure)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(ages) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B2"


def main():
    df2 = pd.read_parquet(OUTPUT_PATH + f"2_{METHOD_ID}_enhanced.parquet")
    df3 = pd.read_parquet(OUTPUT_PATH + f"3_{METHOD_ID}_diagnostics.parquet")
    df4 = pd.read_parquet(OUTPUT_PATH + f"4_{METHOD_ID}_ldf_averages.parquet")
    print(f"Loaded data: {len(df2)} enhanced rows, {len(df3)} diagnostic rows, {len(df4)} average rows")

    prior_selections_path = Path(OUTPUT_PATH) / "../prior-selections.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"Loaded {len(df_prior)} prior selections")
    else:
        print("No prior selections found (optional)")

    for col in df3.columns:
        if col not in ['period', 'age']:
            if 'rate' in col.lower() or 'ratio' in col.lower() or 'to_' in col.lower():
                if df3[col].dropna().abs().max() <= 2.0:
                    df3[col] = df3[col].astype(float)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    measures = df2['measure'].cat.categories.tolist()
    for measure in measures:
        sheet_name = measure[:31]
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, measure, df2, df3, df4, df_prior=df_prior)
        print(f"Built sheet: {sheet_name}")

    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    wb.save(output_file)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    main()
