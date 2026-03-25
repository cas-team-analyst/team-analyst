"""
goal: Update ONLY the selections section of Chain Ladder Selections.xlsx from chain-ladder.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

usage: Run from project root (close the Excel file first):
    .venv/Scripts/Activate.ps1; python demo/demo2-make-selections/output/chain-ladder/scripts/6-update-selections-excel.py
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
SELECTIONS_JSON_PATH = str(SCRIPT_DIR / "../selections/")
SELECTIONS_EXCEL_PATH = str(SCRIPT_DIR / "../selections/")
METHOD_ID = "chainladder"
SELECTIONS_FILE = SELECTIONS_JSON_PATH + f"/{METHOD_ID}.json"
EXCEL_FILE = SELECTIONS_EXCEL_PATH + "/Chain Ladder Selections.xlsx"

SELECTION_FILL = PatternFill("solid", fgColor="FFF2CC")
LABEL_FONT = Font(bold=True, size=9)
DATA_FONT = Font(size=9)
THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def find_selections_section(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "LDF Selections":
                section_row = cell.row
                header_row = section_row + 1
                # Scan for the "Selection" label row (not "Prior Selection")
                # It may be offset by prior selection rows if they exist
                for check_row in range(header_row + 1, header_row + 10):
                    label = ws.cell(row=check_row, column=1).value
                    if label == "Selection":
                        selection_row = check_row
                        reasoning_row = check_row + 1
                        return header_row, selection_row, reasoning_row
                return None, None, None
    return None, None, None


def get_interval_columns(ws, header_row):
    interval_to_col = {}
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip() and '-' in str(cell.value):
            interval_to_col[str(cell.value).strip()] = cell.column
    return interval_to_col


def update_sheet(ws, measure_selections):
    header_row, selection_row, reasoning_row = find_selections_section(ws)
    if header_row is None:
        print(f"  WARNING: Could not find 'LDF Selections' section in sheet '{ws.title}'")
        return

    interval_to_col = get_interval_columns(ws, header_row)

    for sel in measure_selections:
        interval = sel["interval"]
        if interval not in interval_to_col:
            print(f"  WARNING: Interval '{interval}' not found in sheet '{ws.title}'")
            continue

        col = interval_to_col[interval]

        sel_cell = ws.cell(row=selection_row, column=col)
        sel_cell.value = sel["selection"]
        sel_cell.fill = SELECTION_FILL
        sel_cell.font = DATA_FONT
        sel_cell.alignment = Alignment(horizontal="right")
        sel_cell.border = THIN_BORDER
        sel_cell.number_format = "0.0000"

        reason_cell = ws.cell(row=reasoning_row, column=col)
        reason_cell.value = sel["measure"] + (": " + sel["reasoning"] if sel.get("reasoning") else "")
        reason_cell.fill = SELECTION_FILL
        reason_cell.font = Font(size=8, italic=True)
        reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        reason_cell.border = THIN_BORDER

    ws.row_dimensions[reasoning_row].height = 60
    print(f"  Updated {len(measure_selections)} selections in '{ws.title}'")


def main():
    selections_file = Path(SELECTIONS_FILE)
    if not selections_file.exists():
        print(f"Selections file not found: {SELECTIONS_FILE}")
        print("Skipping update - no selections to apply.")
        return

    with open(selections_file) as f:
        selections = json.load(f)

    if not selections:
        print(f"No selections found in {SELECTIONS_FILE}")
        print("Skipping update - selections array is empty.")
        return

    print(f"Loaded {len(selections)} selections from {SELECTIONS_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    by_measure = {}
    for sel in selections:
        m = sel["measure"]
        by_measure.setdefault(m, []).append(sel)

    for sheet_name in wb.sheetnames:
        matched = None
        for measure in by_measure:
            if sheet_name == measure[:31]:
                matched = measure
                break
        if matched:
            update_sheet(wb[sheet_name], by_measure[matched])
        else:
            print(f"No selections found for sheet '{sheet_name}' - skipping")

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
