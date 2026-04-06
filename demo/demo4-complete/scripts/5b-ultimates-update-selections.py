import json
import pandas as pd
from openpyxl import load_workbook
import pathlib

# File paths
JSON_FILE = "../selections/ultimates.json"
EXCEL_FILE = "../selections/Ultimates.xlsx"

def update_excel_selections():
    print(f"Loading selections from {JSON_FILE}")
    try:
        with open(JSON_FILE, "r") as f:
            selections = json.load(f)
    except FileNotFoundError:
        print(f"Selections file not found: {JSON_FILE}")
        print("Skipping update - no selections to apply.")
        return
        
    print(f"Loaded {len(selections)} selections")
    
    # Organize selections by measure
    by_measure = {}
    for entry in selections:
        meas = entry.get('measure')
        if not meas: continue
        if meas not in by_measure:
            by_measure[meas] = {}
        # Key by period
        by_measure[meas][str(entry['period'])] = {
            'selection': entry['selection'],
            'reasoning': entry['reasoning']
        }
    
    print(f"Opening workbook {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Excel file not found: {EXCEL_FILE}")
        return
        
    updates_made = 0
    for measure, periods_data in by_measure.items():
        sheet_name = measure
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the Selection and Reasoning columns
            # Column 9 is Selection, 10 is Reasoning
            
            # Iterate through rows starting at 2
            row = 2
            while True:
                period_cell = ws.cell(row=row, column=1)
                if not period_cell.value:
                    break
                    
                period = str(period_cell.value).strip()
                if period in periods_data:
                    ws.cell(row=row, column=9).value = float(periods_data[period]['selection'])
                    ws.cell(row=row, column=10).value = str(periods_data[period]['reasoning'])
                    updates_made += 1
                row += 1
                
    wb.save(EXCEL_FILE)
    print(f"Updates complete. Wrote {updates_made} selections to excel.")

if __name__ == "__main__":
    update_excel_selections()
