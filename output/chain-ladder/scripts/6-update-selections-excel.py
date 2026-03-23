"""
goal: Update ONLY the selections section of Chain Ladder Selections.xlsx from chain-ladder.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: This script must be run from its own directory for relative paths to work correctly. Close the Excel file before running.
    cd .claude/skills/reserving-methods/assets/chain-ladder
    python 6-update-selections-excel.py
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Replace when using this file in an actual project:
SELECTIONS_JSON_PATH = "../test-output/"  # Path to selections JSON file
SELECTIONS_EXCEL_PATH = "../test-output/"  # Path to selections Excel file
METHOD_ID = "chainladder"
SELECTIONS_FILE = SELECTIONS_JSON_PATH + f"{METHOD_ID}.json"
EXCEL_FILE = SELECTIONS_EXCEL_PATH + "Chain Ladder Selections.xlsx"

SELECTION_FILL = PatternFill("solid", fgColor="FFF2CC")
LABEL_FONT = Font(bold=True, size=9)
DATA_FONT = Font(size=9)
THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def find_selections_section(ws):
    """Find the row where 'LDF Selections' section header and interval headers are."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "LDF Selections":
                section_row = cell.row
                # Header row with intervals is next row
                header_row = section_row + 1
                # Data rows follow
                selection_row = header_row + 1
                reasoning_row = header_row + 2
                return header_row, selection_row, reasoning_row
    return None, None, None


def get_interval_columns(ws, header_row):
    """Build a dict mapping interval string -> column index from the header row."""
    interval_to_col = {}
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip() and '-' in str(cell.value):
            interval_to_col[str(cell.value).strip()] = cell.column
    return interval_to_col


def update_sheet(ws, measure_selections):
    """Update the selections section of a single sheet."""
    header_row, selection_row, reasoning_row = find_selections_section(ws)
    if header_row is None:
        print(f"  WARNING: Could not find 'LDF Selections' section in sheet '{ws.title}'")
        return

    interval_to_col = get_interval_columns(ws, header_row)

    for sel in measure_selections:
        interval = sel["interval"]
        if interval not in interval_to_col:
            print(f"  WARNING: Interval '{interval}' not found in sheet '{ws.title}'")
            continue

        col = interval_to_col[interval]

        # Write selection value
        sel_cell = ws.cell(row=selection_row, column=col)
        sel_cell.value = sel["selection"]
        sel_cell.fill = SELECTION_FILL
        sel_cell.font = DATA_FONT
        sel_cell.alignment = Alignment(horizontal="right")
        sel_cell.border = THIN_BORDER
        sel_cell.number_format = "0.0000"

        # Write reasoning
        reason_cell = ws.cell(row=reasoning_row, column=col)
        reason_cell.value = sel["method"] + (": " + sel["reasoning"] if sel.get("reasoning") else "")
        reason_cell.fill = SELECTION_FILL
        reason_cell.font = Font(size=8, italic=True)
        reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        reason_cell.border = THIN_BORDER

    # Set row height for reasoning row
    ws.row_dimensions[reasoning_row].height = 60

    print(f"  Updated {len(measure_selections)} selections in '{ws.title}'")


def main():
    """Update Chain Ladder Selections Excel file with selections from JSON."""
    with open(SELECTIONS_FILE) as f:
        selections = json.load(f)

    print(f"Loaded {len(selections)} selections from {SELECTIONS_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Group selections by measure
    by_measure = {}
    for sel in selections:
        m = sel["measure"]
        by_measure.setdefault(m, []).append(sel)

    for sheet_name in wb.sheetnames:
        # Match sheet name to measure (sheet name may be truncated to 31 chars)
        matched = None
        for measure in by_measure:
            if sheet_name == measure[:31]:
                matched = measure
                break
        if matched:
            update_sheet(wb[sheet_name], by_measure[matched])
        else:
            print(f"No selections found for sheet '{sheet_name}' - skipping")

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
