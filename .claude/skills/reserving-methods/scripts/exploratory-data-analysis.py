"""
In this file: 
- print_all_sheet_previews(): Print formatted previews of all sheets in an Excel file
- get_sheet_names(): Extract all sheet names from a spreadsheet file
- read_file_preview(): Read and format a preview of file contents with row/column indicators
"""

from openpyxl import load_workbook
import pandas as pd
import os
from typing import List

def print_all_sheet_previews(file_path: str, max_rows: int = 10, max_cols: int = 10) -> None:
    """Loop over all sheets in an Excel file and print previews one by one.
    
    Args:
        file_path: Path to the Excel file
        max_rows: Maximum number of rows to include in preview (default: 10)
        max_cols: Maximum number of columns to include in preview (default: 10)
    """
    sheet_names = get_sheet_names(file_path)
    
    for sheet_name in sheet_names:
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*80}")
        try:
            preview = read_file_preview(file_path, sheet_name=sheet_name, max_rows=max_rows, max_cols=max_cols)
            print(preview)
        except Exception as e:
            print(f"Error reading sheet: {str(e)}")


def get_sheet_names(file_path: str) -> List[str]:
    """Get all sheet names from Excel file"""
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        return ['CSV']  # CSV files have only one "sheet"
    
    elif file_ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
        # Use openpyxl to get sheet names
        wb = load_workbook(file_path, read_only=True, keep_links=False)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    
    else:
        raise ValueError(f"Unsupported file type: {file_ext}")


def read_file_preview(file_path: str, sheet_name: str = None, max_rows: int = None, max_cols: int = None) -> str:
    """Read file sample and format as preview text with row/column numbers"""
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.csv':
        df_sample = pd.read_csv(file_path, header=None, nrows=max_rows, dtype=str)
        df_sample = df_sample.iloc[:, :max_cols] if len(df_sample.columns) > max_cols else df_sample
    
    elif file_ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
        # Read Excel file without treating first row as headers
        df_sample = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str, engine='openpyxl')
        # Limit rows and columns
        df_sample = df_sample.head(max_rows)
        df_sample = df_sample.iloc[:, :max_cols] if len(df_sample.columns) > max_cols else df_sample
    
    else:
        raise ValueError(f"Unsupported file type: {file_ext}")
    
    # Build preview with row numbers
    rows_with_numbers = []
    col_headers = [f'Col {i+1}' for i in range(len(df_sample.columns))]
    rows_with_numbers.append(f"Row | {' | '.join(col_headers)}")
    rows_with_numbers.append("-" * 80)
    
    for idx in range(len(df_sample)):
        row_num = idx + 1
        row_data = df_sample.iloc[idx]
        row_str = f"{row_num:3d} | " + " | ".join([str(val) if pd.notna(val) else "" for val in row_data])
        rows_with_numbers.append(row_str)
    
    return "\n".join(rows_with_numbers)
