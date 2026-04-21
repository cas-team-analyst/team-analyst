# Creates a formatted Excel workbook for tail factor review and selection. Each measure gets its
# own sheet with observed data, scenario comparisons, and selection areas for actuarial judgment.

"""
goal: Create Chain Ladder Selections - Tail.xlsx for actuarial tail factor review and selection.
      This file should be manually edited by the actuary to make tail factor selections.

Sheet layout per measure:
  - Section A: Triangle Type Banner (measure name + key stats)
  - Section B: Observed Factors Table (AY x age triangle + weighted avg + CV)
  - Section C: Scenario Comparison Table (method x starting_age scenarios with diagnostics)
  - Section D: Selection Area (prior, rule-based, AI, final selection rows)
  - Section E: Documentation Checklist (manual actuary verification checkboxes)

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 2d-tail-create-excel.py
"""

import pandas as pd
import json
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

from modules import config
from modules.xl_styles import (
    HEADER_FILL, SUBHEADER_FILL, SECTION_FILL, SELECTION_FILL, PRIOR_FILL, AI_FILL,
    HEADER_FONT, SUBHEADER_FONT, SECTION_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)

# Paths
TAIL_SCENARIOS_PATH = config.PROCESSED_DATA + "tail-scenarios.parquet"
ENHANCED_PATH = config.PROCESSED_DATA + "2_enhanced.parquet"
DIAGNOSTICS_PATH = config.PROCESSED_DATA + "3_diagnostics.parquet"
SELECTIONS_OUTPUT_PATH = config.SELECTIONS
OUTPUT_FILE_NAME = "Chain Ladder Selections - Tail.xlsx"

# Colors for scenario comparison rows
GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")   # good scenario
YELLOW_FILL = PatternFill("solid", fgColor="FFE699")  # marginal scenario
RED_FILL = PatternFill("solid", fgColor="F4B084")     # poor scenario

# Triangle type expectations (for banner notes)
TRIANGLE_TYPE_NOTES = {
    'Paid Loss': 'Paid loss tails typically longer than incurred',
    'Incurred Loss': 'Incurred loss tails typically shorter than paid',
    'Closed Count': 'Closure tails similar to paid loss',
    'Reported Count': 'Reported count tails similar to incurred loss',
}


def get_type_note(measure):
    """Return type-specific note for measure banner."""
    for key, note in TRIANGLE_TYPE_NOTES.items():
        if key.lower() in measure.lower():
            return note
    return ''


def write_section_header(ws, row, col_span, title, level="section"):
    """Write a section header spanning multiple columns."""
    cell = ws.cell(row=row, column=1, value=title)
    style_header(cell, level)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    return row + 1


def write_banner_section(ws, start_row, measure, df_enhanced, df_diagnostics):
    """Section A: Triangle Type Banner with measure name and key stats."""
    row = write_section_header(ws, start_row, 12, measure, "header")
    
    type_note = get_type_note(measure)
    if type_note:
        cell = ws.cell(row=row, column=1, value=type_note)
        cell.font = Font(italic=True, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1
    
    # Key stats row
    df_m = df_enhanced[df_enhanced['measure'] == measure].copy()
    if not df_m.empty:
        max_age = df_m['age'].max()
        n_ays = df_m['period'].nunique()
        
        # Get closure rate from diagnostics at last age if available (diagnostics are shared, no measure filter)
        closure_pct = None
        if 'claim_closure_rate' in df_diagnostics.columns:
            last_age_diag = df_diagnostics[df_diagnostics['age'] == max_age]
            if not last_age_diag.empty:
                closure_rate = last_age_diag['claim_closure_rate'].iloc[0]
                if pd.notna(closure_rate):
                    closure_pct = closure_rate * 100
        
        # Last observed avg LDF
        last_ldf = None
        if 'ldf' in df_m.columns:
            max_interval = df_m['interval'].max()
            last_ldfs = df_m[df_m['interval'] == max_interval]['ldf'].dropna()
            if not last_ldfs.empty:
                last_ldf = last_ldfs.mean()
        
        stats_labels = ['Max Observed Age:', 'Accident Years:', 'Closure % at Last Age:', 'Last Observed Avg LDF:']
        stats_values = [
            max_age,
            n_ays,
            f"{closure_pct:.1f}%" if closure_pct is not None else 'N/A',
            f"{last_ldf:.4f}" if last_ldf is not None else 'N/A'
        ]
        
        for i, (label, value) in enumerate(zip(stats_labels, stats_values)):
            col = i * 3 + 1
            label_cell = ws.cell(row=row, column=col, value=label)
            label_cell.font = LABEL_FONT
            label_cell.alignment = Alignment(horizontal="right")
            
            value_cell = ws.cell(row=row, column=col+1, value=value)
            value_cell.font = DATA_FONT
            value_cell.alignment = Alignment(horizontal="left")
        
        row += 1
    
    return row + 1


def write_observed_factors_section(ws, start_row, measure, df_enhanced):
    """Section B: Observed Factors Table (triangle + weighted avg + CV)."""
    df_m = df_enhanced[df_enhanced['measure'] == measure].copy()
    if df_m.empty:
        return start_row
    
    periods = sorted(df_m['period'].unique())
    ages = sorted([a for a in df_m['age'].unique() if pd.notna(a)])
    
    row = write_section_header(ws, start_row, len(ages) + 1, "Observed Age-to-Age Factors", "section")
    
    # Header row with ages
    ws.cell(row=row, column=1, value="Accident Year").border = THIN_BORDER
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    
    for c_idx, age in enumerate(ages, start=2):
        cell = ws.cell(row=row, column=c_idx, value=age)
        style_header(cell, "subheader")
    row += 1
    
    # Build factor lookup
    factor_dict = {}
    for _, r in df_m[df_m['ldf'].notna()].iterrows():
        key = (str(r['period']), r['age'])
        factor_dict[key] = r['ldf']
    
    # Determine candidate starting ages (highlight these columns)
    candidate_ages = set()
    # We'll mark them after we know which ages have scenarios
    
    # Period rows
    for period in periods:
        period_cell = ws.cell(row=row, column=1, value=period)
        period_cell.font = LABEL_FONT
        period_cell.alignment = Alignment(horizontal="left")
        period_cell.border = THIN_BORDER
        
        for c_idx, age in enumerate(ages, start=2):
            val = factor_dict.get((str(period), age))
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = "0.0000"
        row += 1
    
    # Weighted average row
    weighted_avg_dict = {}
    weight_dict = {}
    for _, r in df_m[df_m['ldf'].notna()].iterrows():
        age_key = r['age']
        ldf = r['ldf']
        weight = r.get('weight', 1.0)
        if age_key not in weighted_avg_dict:
            weighted_avg_dict[age_key] = 0.0
            weight_dict[age_key] = 0.0
        weighted_avg_dict[age_key] += ldf * weight
        weight_dict[age_key] += weight
    
    for age_key in weighted_avg_dict:
        if weight_dict[age_key] > 0:
            weighted_avg_dict[age_key] /= weight_dict[age_key]
    
    avg_cell = ws.cell(row=row, column=1, value="Weighted Avg")
    avg_cell.font = LABEL_FONT
    avg_cell.border = THIN_BORDER
    for c_idx, age in enumerate(ages, start=2):
        val = weighted_avg_dict.get(age)
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER
        if val is not None:
            cell.number_format = "0.0000"
    row += 1
    
    # CV row
    cv_dict = {}
    for age in ages:
        age_factors = [factor_dict.get((str(p), age)) for p in periods]
        age_factors = [f for f in age_factors if f is not None]
        if len(age_factors) > 1:
            mean_val = sum(age_factors) / len(age_factors)
            variance = sum((f - mean_val)**2 for f in age_factors) / len(age_factors)
            std_dev = variance ** 0.5
            cv_dict[age] = std_dev / mean_val if mean_val != 0 else None
    
    cv_cell = ws.cell(row=row, column=1, value="CV")
    cv_cell.font = LABEL_FONT
    cv_cell.border = THIN_BORDER
    for c_idx, age in enumerate(ages, start=2):
        val = cv_dict.get(age)
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="right")
        cell.border = THIN_BORDER
        if val is not None:
            cell.number_format = "0.0000"
        # Highlight candidate starting age columns (light yellow)
        # We'll do this after we process scenarios to know which ages are candidates
    row += 1
    
    return row + 1


def get_scenario_color(row_data):
    """Determine row color based on scenario quality diagnostics."""
    # Green if: monotone AND no slope breaks AND good R² AND no gap AND materiality ok
    # Yellow: marginal (Bondy/Modified Bondy always yellow since no R²)
    # Red: poor quality
    
    # Helper to safely convert boolean (handles pandas NA)
    def safe_bool(val):
        if pd.isna(val):
            return False
        return bool(val)
    
    is_monotone = safe_bool(row_data.get('is_monotone_from_here', False))
    slope_breaks = row_data.get('slope_sign_changes', 999)
    if pd.isna(slope_breaks):
        slope_breaks = 999
    r_squared = row_data.get('r_squared')
    gap_flag = safe_bool(row_data.get('gap_flag', True))
    materiality_ok = safe_bool(row_data.get('materiality_ok', False))
    method = row_data.get('method', '')
    
    # Bondy variants have no R² - always yellow
    if 'bondy' in method.lower():
        return YELLOW_FILL
    
    # Green: high quality scenario
    if (is_monotone and slope_breaks == 0 and 
        r_squared is not None and not pd.isna(r_squared) and r_squared > 0.85 and 
        not gap_flag and materiality_ok):
        return GREEN_FILL
    
    # Red: poor quality
    if (not is_monotone or slope_breaks > 0 or 
        (r_squared is not None and not pd.isna(r_squared) and r_squared < 0.70) or gap_flag):
        return RED_FILL
    
    # Yellow: marginal
    return YELLOW_FILL


def write_scenario_comparison_section(ws, start_row, measure, df_scenarios):
    """Section C: Scenario Comparison Table with diagnostics."""
    df_m = df_scenarios[df_scenarios['measure'] == measure].copy()
    if df_m.empty:
        return start_row
    
    col_headers = [
        'Starting Age', 'Monotone', 'CV', 'Slope Breaks', 'Method', 'Params',
        'Tail Factor', 'R2', 'LOO Std Dev', 'Gap to Last', 'Gap Flag',
        '% of CDF', 'Materiality OK', '+10% Reserve Delta', '+20% Reserve Delta'
    ]
    
    row = write_section_header(ws, start_row, len(col_headers), "Scenario Comparison", "section")
    
    # Header row
    for c_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=header)
        style_header(cell, "subheader")
    row += 1
    
    # Sort scenarios by starting_age then method
    df_m = df_m.sort_values(['starting_age', 'method'])
    
    # Data rows
    for _, scenario in df_m.iterrows():
        row_fill = get_scenario_color(scenario)
        
        # Helper to safely convert boolean columns (handles pandas NA)
        def safe_bool(val):
            if pd.isna(val):
                return False
            return bool(val)
        
        values = [
            scenario['starting_age'],
            'Y' if safe_bool(scenario.get('is_monotone_from_here', False)) else 'N',
            scenario.get('cv_at_starting_age'),
            scenario.get('slope_sign_changes', 0),
            scenario['method'],
            str(scenario.get('method_params', '')) if pd.notna(scenario.get('method_params')) else '',
            scenario['tail_factor'],
            scenario.get('r_squared'),
            scenario.get('loo_std_dev'),
            scenario.get('gap_to_last_observed'),
            'Y' if safe_bool(scenario.get('gap_flag', False)) else 'N',
            scenario.get('pct_of_cdf'),
            'Y' if safe_bool(scenario.get('materiality_ok', False)) else 'N',
            scenario.get('sensitivity_plus10_reserve_delta'),
            scenario.get('sensitivity_plus20_reserve_delta'),
        ]
        
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
            cell.border = THIN_BORDER
            cell.fill = row_fill
            
            # Number formatting
            if c_idx == 1:  # Starting Age
                cell.number_format = "0"
            elif c_idx == 3:  # CV
                if val is not None:
                    cell.number_format = "0.0000"
            elif c_idx == 7:  # Tail Factor
                if val is not None:
                    cell.number_format = "0.0000"
            elif c_idx in [8, 9, 10]:  # R2, LOO Std Dev, Gap
                if val is not None:
                    cell.number_format = "0.0000"
            elif c_idx == 12:  # % of CDF
                if val is not None:
                    cell.number_format = "0.00%"
            elif c_idx in [14, 15]:  # Reserve deltas
                if val is not None:
                    cell.number_format = "#,##0"
        
        row += 1
    
    return row + 1


def write_selection_section(ws, start_row, measure, prior_selections=None):
    """Section D: Selection Area (prior, rule-based, AI, final selection)."""
    row = write_section_header(ws, start_row, 6, "Tail Factor Selection", "section")
    
    # Column headers
    headers = ['Label', 'Cutoff Age', 'Tail Factor', 'Method', 'Reasoning', 'Additional Notes']
    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=header)
        style_header(cell, "subheader")
    row += 1
    
    # Prior selection (if available)
    if prior_selections is not None:
        prior_m = prior_selections[prior_selections['measure'] == measure]
        if not prior_m.empty:
            prior_row = prior_m.iloc[0]
            values = [
                'Prior Selection',
                prior_row.get('cutoff_age', ''),
                prior_row.get('tail_factor', ''),
                prior_row.get('method', ''),
                prior_row.get('reasoning', ''),
                ''
            ]
            for c_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c_idx, value=val)
                cell.fill = PRIOR_FILL
                cell.border = THIN_BORDER
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
                if c_idx in [2, 3] and val:  # Cutoff Age, Tail Factor
                    cell.number_format = "0.0000" if c_idx == 3 else "0"
            row += 1
            
            # Prior delta row
            cell = ws.cell(row=row, column=1, value="Prior Delta")
            cell.fill = PRIOR_FILL
            cell.border = THIN_BORDER
            cell.font = LABEL_FONT
            for c_idx in range(2, 7):
                cell = ws.cell(row=row, column=c_idx, value='')
                cell.fill = PRIOR_FILL
                cell.border = THIN_BORDER
            ws.cell(row=row, column=5, value="(current - prior)").font = Font(italic=True, size=8)
            row += 1
            
            row += 1  # Blank row
    
    # Rule-Based Selection (blue background - filled by 2e script)
    cell = ws.cell(row=row, column=1, value="Rule-Based Selection")
    style_header(cell, "ai")
    cell.fill = PatternFill("solid", fgColor="D6E4F0")  # light blue
    for c_idx in range(2, 7):
        cell = ws.cell(row=row, column=c_idx, value='')
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row += 1
    
    # AI Selection (purple background - filled by 2e script)
    cell = ws.cell(row=row, column=1, value="AI Selection")
    style_header(cell, "ai")
    for c_idx in range(2, 7):
        cell = ws.cell(row=row, column=c_idx, value='')
        cell.fill = AI_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row += 1
    
    row += 1  # Blank row
    
    # Final Selection (yellow - actuary input)
    cell = ws.cell(row=row, column=1, value="FINAL SELECTION")
    style_header(cell, "selection")
    for c_idx in range(2, 7):
        cell = ws.cell(row=row, column=c_idx, value='')
        cell.fill = SELECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
    row += 1
    
    # Additional info rows
    info_rows = [
        ('Pct of CDF', ''),
        ('Anchor Basis', '(Closure/payment data OR Materiality <0.1% of CDF)'),
        ('Sensitivity +10% Reserve Delta', ''),
        ('Sensitivity +20% Reserve Delta', ''),
    ]
    
    for label, note in info_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER
        
        for c_idx in range(2, 7):
            cell = ws.cell(row=row, column=c_idx, value='')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        if note:
            ws.cell(row=row, column=2, value=note).font = Font(italic=True, size=8)
        
        row += 1
    
    return row + 1


def write_checklist_section(ws, start_row):
    """Section E: Documentation Checklist."""
    row = write_section_header(ws, start_row, 3, "Documentation Checklist", "section")
    
    checklist_items = [
        'Alternatives considered and documented',
        'Diagnostics reviewed (R2, residuals, LOO stability)',
        'Sensitivity results documented',
        'Override justification (if selection differs from top scenario)',
        'Prior vs current delta and driver documented',
        'Peer review notes',
        'Sign-off',
    ]
    
    ws.cell(row=row, column=1, value="Item").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    
    ws.cell(row=row, column=2, value="Complete?").font = SUBHEADER_FONT
    ws.cell(row=row, column=2).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    
    ws.cell(row=row, column=3, value="Notes").font = SUBHEADER_FONT
    ws.cell(row=row, column=3).fill = SUBHEADER_FILL
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    row += 1
    
    for item in checklist_items:
        ws.cell(row=row, column=1, value=item).font = DATA_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        
        ws.cell(row=row, column=2, value='').border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        
        ws.cell(row=row, column=3, value='').border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", wrap_text=True)
        
        row += 1
    
    return row


def build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, prior_selections=None):
    """Build complete sheet for one measure."""
    row = 1
    
    # Section A: Banner
    row = write_banner_section(ws, row, measure, df_enhanced, df_diagnostics)
    
    # Section B: Observed Factors
    row = write_observed_factors_section(ws, row, measure, df_enhanced)
    
    # Section C: Scenario Comparison
    row = write_scenario_comparison_section(ws, row, measure, df_scenarios)
    
    # Section D: Selection Area
    row = write_selection_section(ws, row, measure, prior_selections)
    
    # Section E: Documentation Checklist
    write_checklist_section(ws, row)
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    
    ws.freeze_panes = "B2"


def main():
    """Create Tail Factor Selections Excel file."""
    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    if Path(output_file).exists():
        raise FileExistsError(
            f"Output file already exists: {output_file}\n"
            "Delete or rename the file before re-running to avoid overwriting manual edits."
        )
    
    if not Path(TAIL_SCENARIOS_PATH).exists():
        raise FileNotFoundError(
            f"Tail scenarios file not found: {TAIL_SCENARIOS_PATH}\n"
            "Run 2c-tail-methods-diagnostics.py first to generate tail scenarios."
        )
    
    print("Loading data...")
    df_scenarios = pd.read_parquet(TAIL_SCENARIOS_PATH)
    df_enhanced = pd.read_parquet(ENHANCED_PATH)
    df_diagnostics = pd.read_parquet(DIAGNOSTICS_PATH)
    
    print(f"  {len(df_scenarios)} tail scenarios")
    print(f"  {len(df_enhanced)} enhanced rows")
    print(f"  {len(df_diagnostics)} diagnostic rows")
    
    # Load prior selections if available
    prior_selections_path = Path(config.SELECTIONS) / "tail-factor-prior.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"  {len(df_prior)} prior tail selections loaded")
    else:
        print("  No prior tail selections found (optional)")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    measures = sorted(df_scenarios['measure'].unique())
    print(f"\nCreating sheets for measures: {measures}")
    
    for measure in measures:
        ws = wb.create_sheet(title=measure[:31])
        build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, df_prior)
        print(f"  Built sheet: {measure[:31]}")
    
    wb.save(output_file)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    main()
