# Refreshes only the selections section of the Ultimates Excel workbook based on saved selections 
# from the ultimates.json file. Allows you to update the displayed selections without rebuilding 
# the entire workbook, which is useful when you need to revert changes or apply selections from 
# another source.

"""
goal: Update ONLY the selections section of Ultimates.xlsx from ultimates.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 5b-ultimates-update-selections.py
"""

import json
import pandas as pd
from openpyxl import load_workbook
import pathlib

# Replace when using this file in an actual project:
SELECTIONS_JSON_PATH = "../selections/"  # Path to selections JSON file
SELECTIONS_EXCEL_PATH = "../selections/"  # Path to selections Excel file
JSON_FILE = SELECTIONS_JSON_PATH + "ultimates.json"
EXCEL_FILE = SELECTIONS_EXCEL_PATH + "Ultimates.xlsx"


def update_sheet_selections(ws, periods_data):
    """
    Update the selection and reasoning columns in a single sheet.
    
    Args:
        ws: openpyxl worksheet object
        periods_data: Dict mapping period -> {'selection': value, 'reasoning': text}
    
    Returns:
        Number of selections updated
    """
    updates_made = 0
    
    # Column 9 is Selection, 10 is Reasoning
    # Iterate through rows starting at 2 (row 1 is headers)
    row = 2
    while True:
        period_cell = ws.cell(row=row, column=1)
        if not period_cell.value:
            break
            
        period = str(period_cell.value).strip()
        if period in periods_data:
            ws.cell(row=row, column=9).value = float(periods_data[period]['selection'])
            ws.cell(row=row, column=10).value = str(periods_data[period]['reasoning'])
            updates_made += 1
        row += 1
        
    return updates_made


def main():
    """Update Ultimates Excel file with selections from JSON."""
    print(f"Loading selections from {JSON_FILE}")
    try:
        with open(JSON_FILE, "r") as f:
            selections = json.load(f)
    except FileNotFoundError:
        print(f"Selections file not found: {JSON_FILE}")
        print("Skipping update - no selections to apply.")
        return
    
    # Check if selections array is empty
    if not selections or len(selections) == 0:
        print(f"No selections found in {JSON_FILE}")
        print("Skipping update - selections array is empty.")
        return
        
    print(f"Loaded {len(selections)} selections")
    
    # Organize selections by measure
    by_measure = {}
    for entry in selections:
        meas = entry.get('measure')
        if not meas: 
            continue
        if meas not in by_measure:
            by_measure[meas] = {}
        # Key by period
        by_measure[meas][str(entry['period'])] = {
            'selection': entry['selection'],
            'reasoning': entry['reasoning']
        }
    
    print(f"Opening workbook {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Excel file not found: {EXCEL_FILE}")
        return
        
    total_updates = 0
    for measure, periods_data in by_measure.items():
        sheet_name = measure
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            updates = update_sheet_selections(ws, periods_data)
            total_updates += updates
            print(f"  Updated {updates} selections in '{sheet_name}'")
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found in workbook - skipping")
                
    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")
    print(f"Total updates: {total_updates}")


if __name__ == "__main__":
    main()
