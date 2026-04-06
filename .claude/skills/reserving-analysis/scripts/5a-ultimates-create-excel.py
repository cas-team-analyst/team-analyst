# Creates a formatted Excel workbook displaying projected ultimate indications from multiple
# reserving methods (Chain Ladder, Bornhuetter-Ferguson, etc.) alongside prior selections.
# Actuaries review this workbook and manually enter their selected ultimate values and reasoning.

"""
goal: Create Ultimates.xlsx for actuarial review and selection of ultimate losses by period.

Each sheet (one per measure) contains:
  - Period and current maturity information
  - Actual reported losses
  - Initial expected ultimate
  - Indicated ultimates from Chain Ladder, BF, and other methods
  - Prior selections and reasoning (when available)
  - Blank selection rows for actuary input

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 5a-ultimates-create-excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pathlib
import json

# Replace when using this file in an actual project:
ULTIMATES_INPUT_PATH = "../ultimates/"  # Path to projected ultimates
INPUT_ULTIMATES = ULTIMATES_INPUT_PATH + "projected-ultimates.parquet"
SELECTIONS_JSON_PATH = "../selections/"  # Path to read prior selections
PRIOR_SELECTIONS = SELECTIONS_JSON_PATH + "ultimates.json"  # Optional
SELECTIONS_EXCEL_PATH = "../selections/"  # Path to save Excel selection file
OUTPUT_FILE = SELECTIONS_EXCEL_PATH + "Ultimates.xlsx"

# Formatting Styles
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
SELECTION_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
PRIOR_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def format_sheet(ws, measure, df_ult, df_prior):
    """
    Format a single ultimates sheet with indicated values and selection areas.
    
    Args:
        ws: openpyxl worksheet object
        measure: Measure name (e.g., 'Incurred Loss', 'Paid Loss')
        df_ult: DataFrame with projected ultimates containing columns:
                period, current_age, actual, expected_ultimate, cl_ultimate, bf_ultimate
        df_prior: DataFrame with prior selections containing columns:
                  period, measure, selection, reasoning (or None if no prior selections)
    """
    # Filter for measure
    df_m = df_ult[df_ult['measure'] == measure].copy()
    if df_m.empty:
        return
    
    # Write headers
    headers = [
        "Period", "Current Age", "Actual", 
        "Initial Expected Ultimate", "Chain Ladder Ultimate", "BF Ultimate",
        "Prior Selection", "Prior Reasoning", 
        "Selected Ultimate", "Reasoning"
    ]
    
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['J'].width = 40
    
    # Create dict of prior
    prior_dict = {}
    if df_prior is not None and measure in df_prior['measure'].values:
        mp = df_prior[df_prior['measure'] == measure]
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {"sel": r['selection'], "reason": r['reasoning']}
            
    # Write rows
    for r_idx, (_, row) in enumerate(df_m.iterrows(), start=2):
        period = str(row['period'])
        
        # Period & current age
        ws.cell(row=r_idx, column=1, value=period).border = THIN_BORDER
        
        val_age = row.get('current_age')
        if pd.isna(val_age): val_age = ""
        ws.cell(row=r_idx, column=2, value=val_age).border = THIN_BORDER
        
        # Values
        for c_idx, col_name in enumerate(['actual', 'expected_ultimate', 'cl_ultimate', 'bf_ultimate'], start=3):
            val = row.get(col_name)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
            
        # Prior
        prior_sel = prior_dict.get(period, {}).get("sel", "")
        prior_reason = prior_dict.get(period, {}).get("reason", "")
        
        c = ws.cell(row=r_idx, column=7, value=prior_sel)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=8, value=prior_reason)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # New selection
        c = ws.cell(row=r_idx, column=9)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=10)
        c.fill = SELECTION_FILL
        c.alignment = Alignment(wrap_text=True)


def main():
    """Create ultimates selection Excel file from projected ultimates."""
    print("Creating Ultimates.xlsx...")
    
    try:
        df_ult = pd.read_parquet(INPUT_ULTIMATES)
        print(f"Loaded ultimates from {INPUT_ULTIMATES}")
    except Exception as e:
        print(f"Error loading ultimates: {e}")
        exit(1)
        
    df_prior = None
    try:
        if pathlib.Path(PRIOR_SELECTIONS).exists():
            with open(PRIOR_SELECTIONS, 'r') as f:
                prior_data = json.load(f)
            df_prior = pd.DataFrame(prior_data)
            print(f"Loaded prior selections from {PRIOR_SELECTIONS}")
    except Exception as e:
        print(f"  No prior selections found or error: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    measures = ['Incurred Loss', 'Paid Loss', 'Reported Count', 'Closed Count']
    
    for m in measures:
        if m in df_ult['measure'].values:
            ws = wb.create_sheet(title=m)
            format_sheet(ws, m, df_ult, df_prior)
            print(f"  Created sheet for {m}")
            
    out_dir = pathlib.Path(OUTPUT_FILE).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
