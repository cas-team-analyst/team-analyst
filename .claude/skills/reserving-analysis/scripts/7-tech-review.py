# Reads complete-analysis.xlsx and runs actuarial soundness checks,
# printing PASS/WARN/FAIL results to console and writing a color-coded
# Excel report.

"""
goal: Technical review of complete-analysis.xlsx produced by script 6.

Outputs:
  - Console: PASS/WARN/FAIL per check + summary counts
  - ../output/tech-review.xlsx: color-coded results (Status | Group | Check | Detail)

Handles any period format: integer years (2015), partial years (2024.5),
or date strings ("7/1/25 - 6/30/26"). Checks that require integer years
(consecutive sequence, year-range validation, maturity ordering) are
automatically skipped when periods are not integer-convertible.

run-note: Run from the scripts/ directory:
    cd scripts/
    python 7-tech-review.py
"""

import os
import pathlib

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import HEADER_FILL, HEADER_FONT, THIN_BORDER

# ── User-configurable properties ─────────────────────────────────────────────
# Paths from modules/config.py — override here if needed:
INPUT_COMPLETE_ANALYSIS = config.OUTPUT + "complete-analysis.xlsx"
OUTPUT_REVIEW           = config.OUTPUT + "tech-review.xlsx"

IBNR_NEG_TOLERANCE     = -500   # WARN if IBNR < this (small negatives OK from rounding)
SEV_OUTLIER_RATIO      = 5.0    # WARN if any severity > this × median
LOSS_RATE_MAX          = 2.0    # WARN if any loss rate exceeds this
X_TO_ULT_MATURE_TOL    = 0.02   # WARN if most-mature X-to-Ult != 1.0 by more than this (paid: 0.05)
PERIOD_MIN             = 1990   # WARN if any integer period < this
PERIOD_MAX             = 2030   # WARN if any integer period > this
ULTIMATE_OUTLIER_RATIO = 10.0   # WARN if any period ultimate > this × median

# LDF ceiling by age band: link ratios above these thresholds trigger WARN
LDF_CEILINGS = {
    (12, 24): 5.0, (24, 36): 3.0, (36, 48): 2.5, (48, 60): 2.0,
    (60, 72): 1.75, (72, 84): 1.5, (84, 96): 1.35, (96, 108): 1.20,
    (108, 120): 1.10,
}
LDF_CEILING_DEFAULT = 1.05  # ages beyond 120
# ─────────────────────────────────────────────────────────────────────────────

# ── Output styling ────────────────────────────────────────────────────────────
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
HDR_FILL  = HEADER_FILL
HDR_FONT  = HEADER_FONT
THIN      = THIN_BORDER


# ── Result collector ──────────────────────────────────────────────────────────

class Checker:
    def __init__(self):
        self.results = []

    def record(self, status, group, check, detail=""):
        self.results.append({"status": status, "group": group, "check": check, "detail": detail})
        tag = f"[{status}]".ljust(6)
        msg = f"  {tag} {check}"
        if detail:
            msg += f": {detail}"
        print(msg)

    def ok(self, group, check, detail=""):
        self.record("PASS", group, check, detail)

    def warn(self, group, check, detail=""):
        self.record("WARN", group, check, detail)

    def fail(self, group, check, detail=""):
        self.record("FAIL", group, check, detail)

    def counts(self):
        c = {"PASS": 0, "WARN": 0, "FAIL": 0}
        for r in self.results:
            c[r["status"]] += 1
        return c


# ── Period helpers ────────────────────────────────────────────────────────────

def _canon(p):
    """
    Canonical string for any period value.
    2015 / '2015' / 2015.0  -> '2015'
    '7/1/25 - 6/30/26'      -> '7/1/25 - 6/30/26'
    None / NaN              -> None
    """
    if p is None:
        return None
    try:
        if pd.isna(p):
            return None
    except (TypeError, ValueError):
        pass
    s = str(p).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s


def _period_set(df, col):
    """Set of canonical period strings from a DataFrame column, dropping nulls."""
    return {_canon(p) for p in df[col] if _canon(p) is not None}


def _as_int(p):
    """Raise ValueError if canonical period string is not an integer year."""
    return int(p)


def _all_int(periods):
    """True when every period in the set is convertible to int."""
    try:
        [_as_int(p) for p in periods]
        return True
    except (ValueError, TypeError):
        return False


def _canon_map(df, period_col, value_col):
    """Return {canon_period: float_value} from two columns, skipping nulls."""
    result = {}
    for _, row in df.iterrows():
        p = _canon(row.get(period_col))
        v = row.get(value_col)
        if p is not None and pd.notna(v):
            try:
                result[p] = float(v)
            except (ValueError, TypeError):
                pass
    return result


def _ldf_ceiling(age_from, age_to):
    """Return the WARN ceiling for a given age-to-age interval."""
    return LDF_CEILINGS.get((age_from, age_to), LDF_CEILING_DEFAULT)


# ── Sheet parsing helpers ─────────────────────────────────────────────────────

def _to_df(headers, data_rows):
    hdrs = [str(h) if h is not None else f"_col{i}" for i, h in enumerate(headers)]
    rows = [r for r in data_rows if any(v is not None for v in r)]
    return pd.DataFrame(rows, columns=hdrs)


def read_with_title(ws):
    """Row 1 = title (skip), row 2 = headers, row 3+ = data."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return pd.DataFrame()
    return _to_df(rows[1], rows[2:])


def read_no_title(ws):
    """Row 1 = headers, row 2+ = data."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return pd.DataFrame()
    return _to_df(rows[0], rows[1:])


def read_cl_sheet(ws):
    """
    Row 1 = title (skip), row 2 = (None, '12', '24', ...), row 3+ = data.
    Returns DataFrame indexed by canonical period string, columns = integer ages.

    Stops at the first blank row after data starts so that subsequent sections
    of the sheet (LDF averages, diagnostics, etc.) are not mistaken for
    triangle rows.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return pd.DataFrame()
    ages = []
    for h in rows[1][1:]:
        if h is None:
            continue
        try:
            ages.append(int(h))
        except (ValueError, TypeError):
            ages.append(h)
    records = {}
    data_started = False
    for row in rows[2:]:
        if row[0] is None:
            if data_started:
                break   # first blank row after triangle = end of triangle section
            continue
        data_started = True
        period = _canon(row[0])
        if period is None:
            continue
        records[period] = {age: (row[i + 1] if i + 1 < len(row) else None)
                           for i, age in enumerate(ages)}
    if not records:
        return pd.DataFrame()
    return pd.DataFrame.from_dict(records, orient="index")


def detect_sheets(wb):
    names = wb.sheetnames
    known_measures = {"Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"}
    tri_known = {"Incurred-to-Ult", "Paid-to-Ult", "Reported-to-Ult", "Closed-to-Ult"}
    return {
        "measure_sheets": [n for n in names if n in known_measures],
        "sel_sheets":     [n for n in names if n.startswith("Sel - ")],
        "cl_sheets":      [n for n in names if n.startswith("CL - ")],
        "tri_sheets":     [n for n in names if n in tri_known],
        "diag_sheet":     "Diagnostics" in names,
        "avg_ibnr":       "Average IBNR" in names,
        "avg_unpaid":     "Average Unpaid" in names,
    }


def _num(series):
    return pd.to_numeric(series, errors="coerce")


# ── Check groups ──────────────────────────────────────────────────────────────

def check_structure(ck, wb, info):
    g = "1. Structure"
    n = len(wb.sheetnames)
    if n >= 1:
        ck.ok(g, "File readable", f"{n} sheets")
    else:
        ck.fail(g, "File readable", "No sheets")

    if info["measure_sheets"]:
        ck.ok(g, "Measure sheets present", ", ".join(info["measure_sheets"]))
    else:
        ck.fail(g, "Measure sheets present",
                "None of: Incurred Loss, Paid Loss, Reported Count, Closed Count")

    if info["diag_sheet"]:
        ck.ok(g, "Diagnostics sheet present")
    else:
        ck.warn(g, "Diagnostics sheet present", "Not found")

    if info["tri_sheets"]:
        ck.ok(g, "X-to-Ult triangle sheets present", ", ".join(info["tri_sheets"]))
    else:
        ck.warn(g, "X-to-Ult triangle sheets present", "None found")

    names = wb.sheetnames
    dupes = sorted({nm for nm in names if names.count(nm) > 1})
    if not dupes:
        ck.ok(g, "No duplicate sheet names")
    else:
        ck.fail(g, "No duplicate sheet names", f"Duplicates: {dupes}")


def check_period_consistency(ck, info, measure_dfs, diag_df, tri_dfs, cl_dfs):
    g = "2. Period Consistency"

    if not measure_dfs:
        ck.warn(g, "Period consistency", "No measure sheets — skipping")
        return

    first_df = next(iter(measure_dfs.values()))
    canon = _period_set(first_df, "Accident Period")

    # All measure sheets same periods
    mismatches = [m for m, df in measure_dfs.items()
                  if _period_set(df, "Accident Period") != canon]
    if not mismatches:
        ck.ok(g, "All measure sheets same periods", f"{len(canon)} periods")
    else:
        ck.fail(g, "All measure sheets same periods", f"Mismatch: {mismatches}")

    # Diagnostics periods subset of measure periods
    if diag_df is not None and not diag_df.empty:
        dcol = diag_df.columns[0]
        dperiods = _period_set(diag_df, dcol)
        extra = dperiods - canon
        if not extra:
            ck.ok(g, "Diagnostics periods in measure periods")
        else:
            ck.fail(g, "Diagnostics periods in measure periods", f"Extra: {extra}")

    # Triangle periods = measure periods
    for tname, tdf in tri_dfs.items():
        if tdf.empty:
            ck.warn(g, f"Triangle '{tname}' has data", "Empty")
            continue
        pcol = tdf.columns[0]
        tperiods = _period_set(tdf, pcol)
        if tperiods == canon:
            ck.ok(g, f"'{tname}' periods match measures")
        else:
            parts = []
            if tperiods - canon:  parts.append(f"extra: {tperiods - canon}")
            if canon - tperiods:  parts.append(f"missing: {canon - tperiods}")
            ck.fail(g, f"'{tname}' periods match measures", "; ".join(parts))

    # Consecutive — integer periods only (WARN)
    if _all_int(canon):
        int_canon = sorted(_as_int(p) for p in canon)
        if len(int_canon) > 1:
            gaps = [(int_canon[i], int_canon[i + 1])
                    for i in range(len(int_canon) - 1)
                    if int_canon[i + 1] - int_canon[i] != 1]
            if not gaps:
                ck.ok(g, "Period sequence consecutive",
                      f"{int_canon[0]}-{int_canon[-1]}")
            else:
                ck.warn(g, "Period sequence consecutive", f"Gaps between: {gaps}")
    else:
        ck.ok(g, "Period sequence (non-integer — consecutive check skipped)",
              f"{len(canon)} periods")

    # Sorted ascending — integer periods only
    if _all_int(canon):
        for m, df in measure_dfs.items():
            ps = [_as_int(_canon(p)) for p in df["Accident Period"]
                  if _canon(p) is not None]
            if ps == sorted(ps):
                ck.ok(g, f"'{m}' periods sorted ascending")
            else:
                ck.fail(g, f"'{m}' periods sorted ascending", "Out of order")

    # CL sheets contain all measure periods (subset check)
    for cl_name, cl_df in cl_dfs.items():
        if cl_df.empty:
            continue
        cl_periods = {_canon(p) for p in cl_df.index if _canon(p) is not None}
        missing = canon - cl_periods
        if not missing:
            ck.ok(g, f"'{cl_name}' contains all expected periods")
        else:
            ck.fail(g, f"'{cl_name}' contains all expected periods",
                    f"Missing: {missing}")


def check_maturity(ck, measure_dfs):
    g = "3. Current Age / Maturity"

    if not measure_dfs:
        ck.warn(g, "Maturity checks", "No measure sheets — skipping")
        return

    all_ages = []
    for df in measure_dfs.values():
        if "Current Age" in df.columns:
            all_ages.extend(_num(df["Current Age"]).dropna().tolist())

    if not all_ages:
        ck.warn(g, "Current ages found", "No age data")
        return

    if all(a > 0 for a in all_ages):
        ck.ok(g, "All current ages > 0")
    else:
        ck.fail(g, "All current ages > 0",
                f"Non-positive: {[a for a in all_ages if a <= 0]}")

    non_mult = [int(a) for a in all_ages if int(a) % 12 != 0]
    if not non_mult:
        ck.ok(g, "All current ages multiples of 12")
    else:
        ck.warn(g, "All current ages multiples of 12", f"Not multiples: {non_mult}")

    max_age, min_age = max(all_ages), min(all_ages)
    if max_age <= 120:
        ck.ok(g, "Max current age <= 120", f"{int(max_age)}")
    else:
        ck.warn(g, "Max current age <= 120", f"Max = {int(max_age)}")
    if min_age >= 12:
        ck.ok(g, "Min current age >= 12", f"{int(min_age)}")
    else:
        ck.warn(g, "Min current age >= 12", f"Min = {int(min_age)}")

    # Age decreases as period increases — integer periods only
    first_df = next(iter(measure_dfs.values()))
    canon_first = _period_set(first_df, "Accident Period")
    if _all_int(canon_first) and "Current Age" in first_df.columns:
        tmp = first_df.copy()
        tmp["_p"] = [_as_int(_canon(p)) if _canon(p) else None
                     for p in tmp["Accident Period"]]
        tmp["_a"] = _num(tmp["Current Age"])
        tmp = tmp.dropna(subset=["_p", "_a"]).sort_values("_p")
        ages_seq = tmp["_a"].tolist()
        violations = sum(1 for i in range(len(ages_seq) - 1)
                         if ages_seq[i] < ages_seq[i + 1])
        if violations == 0:
            ck.ok(g, "Current age decreases as period increases (older = more mature)")
        else:
            ck.warn(g, "Current age decreases as period increases",
                    f"{violations} violation(s)")
    else:
        ck.ok(g, "Maturity ordering (non-integer periods — order check skipped)")


def check_selected_ultimates(ck, measure_dfs):
    g = "4. Selected Ultimate Integrity"

    count_measures = {"Reported Count", "Closed Count"}

    for m, df in measure_dfs.items():
        if "Selected Ultimate" not in df.columns:
            ck.fail(g, f"'{m}' has Selected Ultimate column", "Column missing")
            continue

        sel = _num(df["Selected Ultimate"])

        # No nulls
        null_count = sel.isna().sum()
        if null_count == 0:
            ck.ok(g, f"'{m}' no null Selected Ultimates")
        else:
            ck.fail(g, f"'{m}' no null Selected Ultimates", f"{null_count} null(s)")

        # All > 0
        non_pos = (sel <= 0).sum()
        if non_pos == 0:
            ck.ok(g, f"'{m}' all Selected Ultimates > 0")
        else:
            ck.fail(g, f"'{m}' all Selected Ultimates > 0", f"{non_pos} non-positive")

        # Count measures: selected should be near-integer (fractional = suspicious)
        if m in count_measures:
            fractional = sel.dropna()
            frac_count = ((fractional % 1).abs() > 0.5).sum()
            if frac_count == 0:
                ck.ok(g, f"'{m}' selections near-integer (count measure)")
            else:
                ck.warn(g, f"'{m}' selections near-integer",
                        f"{frac_count} period(s) with fractional selected count > 0.5")

        # IBNR = Selected - Actual
        if "IBNR" in df.columns and "Actual" in df.columns:
            actual = _num(df["Actual"])
            ibnr = _num(df["IBNR"])
            diff = (sel - actual - ibnr).abs()
            bad = (diff > 1.0).sum()
            if bad == 0:
                ck.ok(g, f"'{m}' IBNR = Selected - Actual (arithmetic)")
            else:
                ck.fail(g, f"'{m}' IBNR = Selected - Actual",
                        f"{bad} row(s) off by > 1.0 (max diff: {diff.max():.1f})")

        # IBNR >= 0
        if "IBNR" in df.columns:
            ibnr = _num(df["IBNR"])
            below_tol = ibnr[ibnr < IBNR_NEG_TOLERANCE]
            mild_neg  = ibnr[(ibnr < 0) & (ibnr >= IBNR_NEG_TOLERANCE)]
            if len(below_tol) == 0 and len(mild_neg) == 0:
                ck.ok(g, f"'{m}' IBNR >= 0")
            elif len(below_tol) > 0:
                bad_ps = [_canon(p) for p in df.loc[below_tol.index, "Accident Period"]]
                ck.fail(g, f"'{m}' IBNR >= 0",
                        f"Below tolerance in periods: {bad_ps}")
            else:
                mild_ps = [_canon(p) for p in df.loc[mild_neg.index, "Accident Period"]]
                ck.warn(g, f"'{m}' IBNR >= 0",
                        f"Slight negatives (within tolerance) periods: {mild_ps}")

        # CL >= Actual
        if "Chain Ladder" in df.columns and "Actual" in df.columns:
            actual = _num(df["Actual"])
            cl = _num(df["Chain Ladder"])
            bad = (cl < actual - 1.0).sum()
            if bad == 0:
                ck.ok(g, f"'{m}' Chain Ladder >= Actual")
            else:
                ck.warn(g, f"'{m}' Chain Ladder >= Actual",
                        f"{bad} period(s) with CL < Actual")

        # BF >= Actual
        if "BF" in df.columns and "Actual" in df.columns:
            bf = _num(df["BF"])
            if bf.notna().any():
                actual = _num(df["Actual"])
                bad = (bf < actual - 1.0).sum()
                if bad == 0:
                    ck.ok(g, f"'{m}' BF >= Actual")
                else:
                    ck.warn(g, f"'{m}' BF >= Actual",
                            f"{bad} period(s) with BF < Actual")

        # IBNR% pattern — integer periods only; tolerance scaled to N/5
        if "IBNR" in df.columns and sel.notna().any():
            canon_set = _period_set(df, "Accident Period")
            if _all_int(canon_set):
                tmp = df.copy()
                tmp["_p"] = [_as_int(_canon(p)) if _canon(p) else None
                             for p in tmp["Accident Period"]]
                tmp["_sel"] = _num(tmp["Selected Ultimate"])
                tmp["_ibnr"] = _num(tmp["IBNR"])
                tmp = tmp.dropna(subset=["_p", "_sel", "_ibnr"])
                tmp = tmp[tmp["_sel"] > 0].sort_values("_p")
                pcts = (tmp["_ibnr"] / tmp["_sel"]).tolist()
                if pcts:
                    if pcts[-1] == max(pcts):
                        ck.ok(g, f"'{m}' most recent period has highest IBNR%",
                              f"{pcts[-1]:.1%}")
                    else:
                        ck.warn(g, f"'{m}' most recent period has highest IBNR%",
                                f"Most recent={pcts[-1]:.1%}, max={max(pcts):.1%}")
                    violations = sum(1 for i in range(len(pcts) - 1)
                                     if pcts[i] > pcts[i + 1])
                    n_tol = max(1, len(pcts) // 5)
                    if violations == 0:
                        ck.ok(g, f"'{m}' IBNR% non-increasing as maturity increases")
                    elif violations <= n_tol:
                        ck.warn(g, f"'{m}' IBNR% non-increasing as maturity increases",
                                f"{violations} reversal(s) (within N/5={n_tol} tolerance)")
                    else:
                        ck.warn(g, f"'{m}' IBNR% non-increasing as maturity increases",
                                f"{violations} reversal(s) (exceeds N/5={n_tol} tolerance)")

        # Extreme outlier
        vals = sel.dropna()
        if len(vals) >= 2:
            med = vals.median()
            if med > 0:
                outliers = vals[vals > ULTIMATE_OUTLIER_RATIO * med]
                if outliers.empty:
                    ck.ok(g, f"'{m}' no extreme outlier ultimates")
                else:
                    ck.warn(g, f"'{m}' no extreme outlier ultimates",
                            f"{len(outliers)} period(s) > {ULTIMATE_OUTLIER_RATIO}x median"
                            f" ({med:,.0f})")

        # Year range — integer periods only
        canon_set = _period_set(df, "Accident Period")
        if _all_int(canon_set):
            out_rng = [p for p in canon_set
                       if _as_int(p) < PERIOD_MIN or _as_int(p) > PERIOD_MAX]
            if not out_rng:
                ck.ok(g, f"'{m}' periods in plausible range ({PERIOD_MIN}-{PERIOD_MAX})")
            else:
                ck.warn(g, f"'{m}' periods in plausible range",
                        f"Out of range: {sorted(out_rng)}")


def check_cross_measure(ck, measure_dfs):
    g = "5. Cross-Measure Consistency"

    def _sel_map(df):
        if df is None:
            return {}
        return _canon_map(df, "Accident Period", "Selected Ultimate")

    inc  = measure_dfs.get("Incurred Loss")
    paid = measure_dfs.get("Paid Loss")
    rep  = measure_dfs.get("Reported Count")
    cls  = measure_dfs.get("Closed Count")

    def _same_periods(df1, df2, n1, n2):
        p1 = _period_set(df1, "Accident Period")
        p2 = _period_set(df2, "Accident Period")
        if p1 == p2:
            ck.ok(g, f"{n1} and {n2} same periods")
        else:
            ck.fail(g, f"{n1} and {n2} same periods",
                    f"Diff: {p1.symmetric_difference(p2)}")

    if inc is not None and paid is not None:
        _same_periods(inc, paid, "Incurred Loss", "Paid Loss")

    if rep is not None and cls is not None:
        _same_periods(rep, cls, "Reported Count", "Closed Count")

    # Closed <= Reported (WARN not FAIL — methods may be selected independently)
    if rep is not None and cls is not None:
        rep_sel = _sel_map(rep)
        cls_sel = _sel_map(cls)
        bad = sorted(p for p in rep_sel
                     if p in cls_sel and cls_sel[p] > rep_sel[p] + 0.5)
        if not bad:
            ck.ok(g, "Closed Count selected <= Reported Count selected (all periods)")
        else:
            ck.warn(g, "Closed Count selected <= Reported Count selected",
                    f"Violations: {bad} (review if expected)")

    # Paid <= Incurred (WARN not FAIL — methods may be selected independently)
    if inc is not None and paid is not None:
        inc_sel  = _sel_map(inc)
        paid_sel = _sel_map(paid)
        bad = sorted(p for p in inc_sel
                     if p in paid_sel and paid_sel[p] > inc_sel[p] + 1.0)
        if not bad:
            ck.ok(g, "Paid Loss selected <= Incurred Loss selected (all periods)")
        else:
            ck.warn(g, "Paid Loss selected <= Incurred Loss selected",
                    f"Violations: {bad} (review if expected)")

    # Unpaid >= IBNR for Incurred Loss
    if inc is not None and "IBNR" in inc.columns and "Unpaid" in inc.columns:
        ibnr   = _num(inc["IBNR"])
        unpaid = _num(inc["Unpaid"])
        mask   = ibnr.notna() & unpaid.notna()
        bad    = (unpaid[mask] < ibnr[mask] - 1.0).sum()
        if bad == 0:
            ck.ok(g, "Incurred Loss: Unpaid >= IBNR (all periods)")
        else:
            ck.fail(g, "Incurred Loss: Unpaid >= IBNR", f"{bad} violation(s)")


def check_sel_sheets(ck, measure_dfs, sel_dfs):
    g = "6. Sel - Sheet Consistency"

    if not sel_dfs:
        ck.warn(g, "Sel - sheets present", "None found — skipping")
        return

    for sel_name, sel_df in sel_dfs.items():
        measure_name = sel_name.replace("Sel - ", "")
        meas_df = measure_dfs.get(measure_name)

        if meas_df is None:
            ck.warn(g, f"'{sel_name}' matching measure sheet",
                    f"No measure sheet '{measure_name}'")
            continue

        sel_pcol = "Period" if "Period" in sel_df.columns else sel_df.columns[0]
        sel_periods  = _period_set(sel_df, sel_pcol)
        meas_periods = _period_set(meas_df, "Accident Period")

        if sel_periods == meas_periods:
            ck.ok(g, f"'{sel_name}' same periods as '{measure_name}'")
        else:
            parts = []
            if sel_periods - meas_periods:
                parts.append(f"extra: {sel_periods - meas_periods}")
            if meas_periods - sel_periods:
                parts.append(f"missing: {meas_periods - sel_periods}")
            ck.fail(g, f"'{sel_name}' periods match '{measure_name}'",
                    "; ".join(parts))

        if "Selected Ultimate" not in sel_df.columns \
                or "Selected Ultimate" not in meas_df.columns:
            ck.warn(g, f"'{sel_name}' Selected Ultimate column",
                    "Missing in one sheet")
            continue

        sel_lookup  = _canon_map(sel_df, sel_pcol, "Selected Ultimate")
        meas_lookup = _canon_map(meas_df, "Accident Period", "Selected Ultimate")

        mismatches = [
            f"{p}: Sel={sel_lookup[p]:,.0f} vs Measure={meas_lookup[p]:,.0f}"
            for p in meas_lookup
            if p in sel_lookup
            and pd.notna(sel_lookup[p]) and pd.notna(meas_lookup[p])
            and abs(sel_lookup[p] - meas_lookup[p]) > 1.0
        ]
        if not mismatches:
            ck.ok(g, f"'{sel_name}' Selected Ultimates match measure sheet")
        else:
            ck.fail(g, f"'{sel_name}' Selected Ultimates match measure sheet",
                    f"{len(mismatches)} mismatch(es): {mismatches[:2]}")

        blanks = _num(sel_df["Selected Ultimate"]).isna().sum()
        if blanks == 0:
            ck.ok(g, f"'{sel_name}' no blank Selected Ultimates")
        else:
            ck.warn(g, f"'{sel_name}' no blank Selected Ultimates",
                    f"{blanks} blank(s)")

        if "Reasoning" in sel_df.columns:
            blanks = sel_df["Reasoning"].isna().sum()
            if blanks == 0:
                ck.ok(g, f"'{sel_name}' reasoning present for all periods")
            else:
                ck.warn(g, f"'{sel_name}' reasoning present",
                        f"{blanks} period(s) missing reasoning")


def check_xtoult_triangles(ck, tri_dfs, measure_dfs):
    g = "7. X-to-Ult Triangles"

    if not tri_dfs:
        ck.warn(g, "X-to-Ult triangles", "None found — skipping")
        return

    n_periods = None
    if measure_dfs:
        n_periods = len(_period_set(next(iter(measure_dfs.values())), "Accident Period"))

    for tname, tdf in tri_dfs.items():
        if tdf.empty:
            ck.warn(g, f"'{tname}' has data", "Empty")
            continue

        age_cols = [col for col in tdf.columns[1:] if pd.notna(col)]
        data = tdf[age_cols].apply(pd.to_numeric, errors="coerce")
        flat = data.values.flatten()
        flat_nn = flat[~pd.isnull(flat)]

        if n_periods is not None:
            if len(tdf) == n_periods:
                ck.ok(g, f"'{tname}' row count = period count", f"{len(tdf)}")
            else:
                ck.fail(g, f"'{tname}' row count = period count",
                        f"Triangle={len(tdf)}, Measures={n_periods}")

        above_1 = (flat_nn > 1.0 + 1e-6).sum()
        below_0 = (flat_nn <= 0).sum()
        if above_1 == 0 and below_0 == 0:
            ck.ok(g, f"'{tname}' all values in (0, 1]")
        else:
            parts = []
            if above_1: parts.append(f"{above_1} cell(s) > 1.0")
            if below_0: parts.append(f"{below_0} cell(s) <= 0")
            ck.warn(g, f"'{tname}' all values in (0, 1]", "; ".join(parts))

        # Wider tolerance for paid (0.05) vs incurred (0.02)
        tol = 0.05 if "Paid" in tname else X_TO_ULT_MATURE_TOL
        first_row_nn = data.iloc[0].dropna()
        if not first_row_nn.empty:
            mature_val = float(first_row_nn.iloc[-1])
            if abs(mature_val - 1.0) <= tol:
                ck.ok(g, f"'{tname}' most-mature diagonal ~= 1.0",
                      f"{mature_val:.4f}")
            else:
                ck.warn(g, f"'{tname}' most-mature diagonal ~= 1.0",
                        f"{mature_val:.4f} (diff={abs(mature_val - 1.0):.4f}, tol={tol})")

        violations = 0
        for _, row in data.iterrows():
            nn = row.dropna().values
            violations += sum(1 for i in range(len(nn) - 1)
                              if nn[i] > nn[i + 1] + 1e-6)
        if violations == 0:
            ck.ok(g, f"'{tname}' non-decreasing across ages (monotone)")
        else:
            ck.warn(g, f"'{tname}' non-decreasing across ages",
                    f"{violations} reversal(s)")

    # Incurred-to-Ult >= Paid-to-Ult at each cell
    inc_tri  = tri_dfs.get("Incurred-to-Ult")
    paid_tri = tri_dfs.get("Paid-to-Ult")
    if inc_tri is not None and paid_tri is not None \
            and not inc_tri.empty and not paid_tri.empty:
        inc_ages  = [col for col in inc_tri.columns[1:]  if pd.notna(col)]
        paid_ages = [col for col in paid_tri.columns[1:] if pd.notna(col)]
        common    = [a for a in inc_ages if a in paid_ages]
        inc_idx  = inc_tri.set_index(inc_tri.columns[0])[common].apply(
            pd.to_numeric, errors="coerce")
        paid_idx = paid_tri.set_index(paid_tri.columns[0])[common].apply(
            pd.to_numeric, errors="coerce")
        violations = 0
        for p in inc_idx.index.intersection(paid_idx.index):
            for a in common:
                iv, pv = inc_idx.loc[p, a], paid_idx.loc[p, a]
                if pd.notna(iv) and pd.notna(pv) and iv < pv - 1e-6:
                    violations += 1
        if violations == 0:
            ck.ok(g, "Incurred-to-Ult >= Paid-to-Ult at all cells")
        else:
            ck.warn(g, "Incurred-to-Ult >= Paid-to-Ult",
                    f"{violations} cell(s) where Inc < Paid")


def _check_avg_tri(ck, df, label):
    g = "8. Average IBNR / Unpaid"
    if df is None or df.empty:
        ck.warn(g, f"{label} has data", "Sheet empty or missing")
        return
    age_cols = [col for col in df.columns[1:] if pd.notna(col)]
    data = df[age_cols].apply(pd.to_numeric, errors="coerce")
    flat = data.values.flatten()
    flat_nn = flat[~pd.isnull(flat)]

    neg = (flat_nn < -1.0).sum()
    if neg == 0:
        ck.ok(g, f"{label} all values >= 0")
    else:
        ck.warn(g, f"{label} all values >= 0", f"{neg} cell(s) < 0")

    violations = 0
    for _, row in data.iterrows():
        nn = row.dropna().values
        violations += sum(1 for i in range(len(nn) - 1)
                          if nn[i] < nn[i + 1] - 1.0)
    if violations == 0:
        ck.ok(g, f"{label} non-increasing across ages (IBNR decreases with age)")
    else:
        ck.warn(g, f"{label} non-increasing across ages", f"{violations} reversal(s)")

    first_nn = data.iloc[0].dropna()
    if not first_nn.empty:
        val = float(first_nn.iloc[-1])
        if abs(val) < 1.0:
            ck.ok(g, f"{label} most-mature period final value ~= 0", f"{val:,.0f}")
        else:
            ck.warn(g, f"{label} most-mature period final value ~= 0", f"{val:,.0f}")


def check_avg_triangles(ck, avg_ibnr_df, avg_unpaid_df):
    g = "8. Average IBNR / Unpaid"
    _check_avg_tri(ck, avg_ibnr_df, "Average IBNR")
    _check_avg_tri(ck, avg_unpaid_df, "Average Unpaid")

    if avg_ibnr_df is not None and avg_unpaid_df is not None \
            and not avg_ibnr_df.empty and not avg_unpaid_df.empty:
        ibnr_ages = [col for col in avg_ibnr_df.columns[1:] if pd.notna(col)]
        unp_ages  = [col for col in avg_unpaid_df.columns[1:] if pd.notna(col)]
        common = [a for a in ibnr_ages if a in unp_ages]
        ibnr_idx = avg_ibnr_df.set_index(avg_ibnr_df.columns[0])[common].apply(
            pd.to_numeric, errors="coerce")
        unp_idx  = avg_unpaid_df.set_index(avg_unpaid_df.columns[0])[common].apply(
            pd.to_numeric, errors="coerce")
        violations = 0
        for p in ibnr_idx.index.intersection(unp_idx.index):
            for a in common:
                iv, uv = ibnr_idx.loc[p, a], unp_idx.loc[p, a]
                if pd.notna(iv) and pd.notna(uv) and uv < iv - 1.0:
                    violations += 1
        if violations == 0:
            ck.ok(g, "Average Unpaid >= Average IBNR at all matching cells")
        else:
            ck.warn(g, "Average Unpaid >= Average IBNR",
                    f"{violations} cell(s) where Unpaid < IBNR")


def check_diagnostics(ck, diag_df):
    g = "9. Diagnostics"

    if diag_df is None or diag_df.empty:
        ck.warn(g, "Diagnostics", "Sheet missing or empty — skipping")
        return

    if "Ultimate Severity" in diag_df.columns:
        sev = _num(diag_df["Ultimate Severity"]).dropna()
        if (sev > 0).all():
            ck.ok(g, "Ultimate Severity > 0 for all periods")
        else:
            ck.fail(g, "Ultimate Severity > 0",
                    f"{(sev <= 0).sum()} non-positive")
        if len(sev) >= 2:
            med = sev.median()
            outliers = sev[sev > SEV_OUTLIER_RATIO * med]
            if outliers.empty:
                ck.ok(g, "Ultimate Severity no outliers", f"Median={med:,.0f}")
            else:
                ck.warn(g, "Ultimate Severity no outliers",
                        f"{len(outliers)} period(s) > {SEV_OUTLIER_RATIO}x"
                        f" median ({med:,.0f})")
    else:
        ck.warn(g, "Ultimate Severity column", "Not found in Diagnostics")

    if "Ultimate Loss Rate" in diag_df.columns:
        lr = _num(diag_df["Ultimate Loss Rate"]).dropna()
        if lr.empty:
            ck.warn(g, "Loss Rate values present", "All null")
        else:
            bad = lr[(lr <= 0) | (lr > LOSS_RATE_MAX)]
            if bad.empty:
                ck.ok(g, f"Loss Rate in (0, {LOSS_RATE_MAX})",
                      f"Range: {lr.min():.3f}-{lr.max():.3f}")
            else:
                ck.warn(g, f"Loss Rate in (0, {LOSS_RATE_MAX})",
                        f"{len(bad)} period(s) outside range")

    if "Ultimate Frequency" in diag_df.columns:
        freq = _num(diag_df["Ultimate Frequency"]).dropna()
        if freq.empty:
            ck.warn(g, "Frequency values present", "All null")
        else:
            bad = (freq <= 0).sum()
            if bad == 0:
                ck.ok(g, "Ultimate Frequency > 0 for all periods")
            else:
                ck.fail(g, "Ultimate Frequency > 0", f"{bad} non-positive")


def check_cl_triangles(ck, cl_dfs, measure_dfs, tri_dfs):
    g = "10. CL Triangle Integrity"

    if not cl_dfs:
        ck.warn(g, "CL triangles present", "None found — skipping")
        return

    for cl_name, cl_df in cl_dfs.items():
        if cl_df.empty:
            ck.warn(g, f"'{cl_name}' has data", "Empty")
            continue

        measure_name = cl_name.replace("CL - ", "")
        meas_df = measure_dfs.get(measure_name)

        actual_lookup = _canon_map(meas_df, "Accident Period", "Actual") \
            if meas_df is not None and "Actual" in meas_df.columns else {}
        known_periods = set(actual_lookup.keys())

        valid_index = [p for p in cl_df.index if _canon(p) in known_periods]
        tri = cl_df.loc[valid_index].apply(pd.to_numeric, errors="coerce") \
            if valid_index else pd.DataFrame()

        # Latest diagonal = Actual in measure sheet
        if actual_lookup and not tri.empty:
            mismatches = []
            for period_str, row in tri.iterrows():
                nn = row.dropna()
                if nn.empty:
                    continue
                latest = nn.iloc[-1]
                expected = actual_lookup.get(_canon(period_str))
                if expected is not None and abs(latest - expected) > 1.0:
                    mismatches.append(
                        f"{period_str}: CL={latest:,.0f} vs Actual={expected:,.0f}")
            if not mismatches:
                ck.ok(g, f"'{cl_name}' diagonal = Actual in measure sheet")
            else:
                ck.fail(g, f"'{cl_name}' diagonal = Actual in measure sheet",
                        f"{len(mismatches)} mismatch(es): {mismatches[:2]}")

        # Non-decreasing across ages
        violations = 0
        for _, row in tri.iterrows():
            nn = row.dropna().values
            violations += sum(1 for i in range(len(nn) - 1)
                              if nn[i] > nn[i + 1] + 1.0)
        if violations == 0:
            ck.ok(g, f"'{cl_name}' non-decreasing across ages")
        else:
            ck.warn(g, f"'{cl_name}' non-decreasing across ages",
                    f"{violations} reversal(s)")

        # No negative values
        if not tri.empty:
            flat = tri.values.flatten()
            neg = sum(1 for v in flat if pd.notna(v) and v < 0)
            if neg == 0:
                ck.ok(g, f"'{cl_name}' no negative values")
            else:
                ck.fail(g, f"'{cl_name}' no negative values", f"{neg} negative cell(s)")

        # Age headers are positive integers
        bad_ages = [a for a in cl_df.columns
                    if not (isinstance(a, (int, float)) and a > 0)]
        if not bad_ages:
            ck.ok(g, f"'{cl_name}' age headers are positive integers")
        else:
            ck.warn(g, f"'{cl_name}' age headers", f"Non-positive: {bad_ages}")

    # CL and X-to-Ult share same final age
    if cl_dfs and tri_dfs:
        cl_df = next(iter(cl_dfs.values()))
        tdf   = next(iter(tri_dfs.values()))
        if not cl_df.empty and not tdf.empty:
            cl_final = max(
                (int(a) for a in cl_df.columns
                 if isinstance(a, (int, float)) and a > 0), default=None)
            tri_ages = [col for col in tdf.columns[1:] if pd.notna(col)]
            tri_final = max(
                (int(a) for a in tri_ages
                 if isinstance(a, (int, float)) and a > 0), default=None)
            if cl_final is not None and tri_final is not None:
                if cl_final == tri_final:
                    ck.ok(g, "CL and X-to-Ult triangles share same final age",
                          f"{cl_final}")
                else:
                    ck.warn(g, "CL and X-to-Ult same final age",
                            f"CL final={cl_final}, X-to-Ult final={tri_final}")


# ── Group 11: Development Factor Checks ──────────────────────────────────────

def check_development_factors(ck, cl_dfs, measure_dfs):
    g = "11. Development Factors"

    if not cl_dfs:
        ck.warn(g, "Development factor checks", "No CL triangles — skipping")
        return

    for cl_name, cl_df in cl_dfs.items():
        if cl_df.empty:
            continue

        measure_name = cl_name.replace("CL - ", "")
        meas_df = measure_dfs.get(measure_name)
        actual_lookup = _canon_map(meas_df, "Accident Period", "Actual") \
            if meas_df is not None and "Actual" in meas_df.columns else {}
        known_periods = set(actual_lookup.keys())

        valid_index = [p for p in cl_df.index if _canon(p) in known_periods]
        tri = cl_df.loc[valid_index].apply(pd.to_numeric, errors="coerce") \
            if valid_index else pd.DataFrame()

        if tri.empty:
            ck.warn(g, f"'{cl_name}' development factors", "No triangle data")
            continue

        age_cols = sorted(
            [c for c in tri.columns if isinstance(c, (int, float)) and c > 0],
            key=lambda x: int(x)
        )

        if len(age_cols) < 2:
            ck.warn(g, f"'{cl_name}' link ratios", "Need >= 2 age columns")
            continue

        ceiling_violations = []
        below_one = []
        avg_by_age = []   # [(age_from, avg_ratio)]

        for i in range(len(age_cols) - 1):
            a0, a1 = int(age_cols[i]), int(age_cols[i + 1])
            ratios = []
            for p in tri.index:
                v0 = tri.loc[p, age_cols[i]]
                v1 = tri.loc[p, age_cols[i + 1]]
                if pd.notna(v0) and pd.notna(v1) and v0 > 0:
                    r = float(v1) / float(v0)
                    ratios.append(r)
                    ceil = _ldf_ceiling(a0, a1)
                    if r > ceil:
                        ceiling_violations.append(
                            f"{p} {a0}→{a1}: {r:.3f} (ceil {ceil})")
                    if r < 1.0 - 1e-6:
                        below_one.append(f"{p} {a0}→{a1}: {r:.4f}")
            if ratios:
                avg_by_age.append((a0, sum(ratios) / len(ratios)))

        if not ceiling_violations:
            ck.ok(g, f"'{cl_name}' all link ratios within age-band ceilings")
        else:
            ck.warn(g, f"'{cl_name}' link ratios within ceilings",
                    f"{len(ceiling_violations)} violation(s): {ceiling_violations[:2]}")

        if not below_one:
            ck.ok(g, f"'{cl_name}' all link ratios >= 1.0")
        else:
            ck.warn(g, f"'{cl_name}' link ratios >= 1.0",
                    f"{len(below_one)} below 1.0: {below_one[:2]}")

        # Average LDFs should decrease (or stay flat) as age increases
        avg_vals = [v for _, v in avg_by_age]
        ldf_mono_violations = sum(
            1 for i in range(len(avg_vals) - 1)
            if avg_vals[i] < avg_vals[i + 1] - 0.01
        )
        if ldf_mono_violations == 0:
            ck.ok(g, f"'{cl_name}' average LDFs non-increasing with age")
        else:
            ck.warn(g, f"'{cl_name}' average LDFs non-increasing with age",
                    f"{ldf_mono_violations} reversal(s) (younger age avg < older age avg)")

    # Check selected CDF (= Selected/Actual) within observed range of X-to-Ult CDFs
    # for matching measures
    for m, df in measure_dfs.items():
        if "Selected Ultimate" not in df.columns or "Actual" not in df.columns:
            continue
        tri_name = m.split()[0] + "-to-Ult"   # "Incurred Loss" -> "Incurred-to-Ult"
        # map measure to triangle name
        tri_map = {
            "Incurred Loss": "Incurred-to-Ult",
            "Paid Loss":     "Paid-to-Ult",
            "Reported Count": "Reported-to-Ult",
            "Closed Count":   "Closed-to-Ult",
        }
        # (skipping this sub-check — CDF range from X-to-Ult would need the triangle data
        #  passed in; covered by selection reasonableness in Group 16)


# ── Group 12: Paid-to-Incurred Raw Data Checks ───────────────────────────────

def check_paid_incurred_raw(ck, measure_dfs, cl_dfs):
    g = "12. Paid-to-Incurred Raw Data"

    inc_df  = measure_dfs.get("Incurred Loss")
    paid_df = measure_dfs.get("Paid Loss")

    if inc_df is None or paid_df is None:
        ck.warn(g, "Paid and Incurred present", "One or both missing — skipping")
        return

    inc_actual  = _canon_map(inc_df,  "Accident Period", "Actual")
    paid_actual = _canon_map(paid_df, "Accident Period", "Actual")

    # Paid actual <= Incurred actual (accounting identity — FAIL)
    bad_periods = [p for p in inc_actual
                   if p in paid_actual and paid_actual[p] > inc_actual[p] + 1.0]
    if not bad_periods:
        ck.ok(g, "Paid Actual <= Incurred Actual for all periods")
    else:
        details = [f"{p}: Paid={paid_actual[p]:,.0f} > Inc={inc_actual[p]:,.0f}"
                   for p in sorted(bad_periods)]
        ck.fail(g, "Paid Actual <= Incurred Actual",
                f"{len(bad_periods)} violation(s): {details[:2]}")

    # Paid/Incurred ratio non-decreasing with maturity
    if "Current Age" in inc_df.columns and _all_int(_period_set(inc_df, "Accident Period")):
        tmp_inc = inc_df[["Accident Period", "Current Age", "Actual"]].copy()
        tmp_inc["_p"]   = [_canon(p) for p in tmp_inc["Accident Period"]]
        tmp_inc["_age"] = _num(tmp_inc["Current Age"])
        tmp_inc["_ia"]  = _num(tmp_inc["Actual"])

        tmp_paid = paid_df[["Accident Period", "Actual"]].copy()
        tmp_paid["_p"]  = [_canon(p) for p in tmp_paid["Accident Period"]]
        tmp_paid["_pa"] = _num(tmp_paid["Actual"])

        merged = tmp_inc.merge(tmp_paid[["_p", "_pa"]], on="_p", how="inner")
        merged = merged.dropna(subset=["_p", "_ia", "_pa", "_age"])
        merged = merged.sort_values("_age")
        ratios = []
        for _, row in merged.iterrows():
            if row["_ia"] > 0:
                ratios.append(row["_pa"] / row["_ia"])
        violations = sum(1 for i in range(len(ratios) - 1)
                         if ratios[i] > ratios[i + 1] + 0.02)
        if violations == 0:
            ck.ok(g, "Paid/Incurred ratio non-decreasing with maturity")
        else:
            ck.warn(g, "Paid/Incurred ratio non-decreasing with maturity",
                    f"{violations} reversal(s)")

    # Paid <= Incurred at every CL triangle cell
    cl_inc  = cl_dfs.get("CL - Incurred Loss")
    cl_paid = cl_dfs.get("CL - Paid Loss")
    if cl_inc is not None and cl_paid is not None \
            and not cl_inc.empty and not cl_paid.empty:
        known = set(inc_actual.keys())
        vi = [p for p in cl_inc.index  if _canon(p) in known]
        vp = [p for p in cl_paid.index if _canon(p) in known]
        tri_inc  = cl_inc.loc[vi].apply(pd.to_numeric, errors="coerce")  if vi else pd.DataFrame()
        tri_paid = cl_paid.loc[vp].apply(pd.to_numeric, errors="coerce") if vp else pd.DataFrame()

        if not tri_inc.empty and not tri_paid.empty:
            common_periods = tri_inc.index.intersection(tri_paid.index)
            common_ages    = [c for c in tri_inc.columns if c in tri_paid.columns]
            cell_violations = 0
            for p in common_periods:
                for a in common_ages:
                    iv = tri_inc.loc[p, a]  if a in tri_inc.columns  else None
                    pv = tri_paid.loc[p, a] if a in tri_paid.columns else None
                    if pd.notna(iv) and pd.notna(pv) and float(pv) > float(iv) + 1.0:
                        cell_violations += 1
            if cell_violations == 0:
                ck.ok(g, "Paid <= Incurred at all CL triangle cells")
            else:
                ck.warn(g, "Paid <= Incurred at CL triangle cells",
                        f"{cell_violations} cell(s) where Paid > Incurred")


# ── Group 13: Case Reserve Reasonableness ────────────────────────────────────

def check_case_reserves(ck, measure_dfs):
    g = "13. Case Reserve Reasonableness"

    inc_df  = measure_dfs.get("Incurred Loss")
    paid_df = measure_dfs.get("Paid Loss")

    if inc_df is None or paid_df is None:
        ck.warn(g, "Case reserves check",
                "Incurred Loss and/or Paid Loss missing — skipping")
        return

    inc_actual  = _canon_map(inc_df,  "Accident Period", "Actual")
    paid_actual = _canon_map(paid_df, "Accident Period", "Actual")

    common = sorted(p for p in inc_actual if p in paid_actual)
    if not common:
        ck.warn(g, "Case reserves computed", "No common periods")
        return

    reserves = {p: inc_actual[p] - paid_actual[p] for p in common}

    # Case reserves = Incurred - Paid >= 0 (FAIL)
    neg_reserves = [p for p, r in reserves.items() if r < -1.0]
    if not neg_reserves:
        ck.ok(g, "Case reserves (Incurred - Paid Actual) >= 0 for all periods",
              f"Min: {min(reserves.values()):,.0f}")
    else:
        details = [f"{p}: {reserves[p]:,.0f}" for p in sorted(neg_reserves)]
        ck.fail(g, "Case reserves >= 0",
                f"{len(neg_reserves)} negative: {details[:2]}")

    # Case reserve % of incurred non-increasing with maturity
    if _all_int(set(common)) and "Current Age" in inc_df.columns:
        tmp = inc_df[["Accident Period", "Current Age"]].copy()
        tmp["_p"]   = [_canon(p) for p in tmp["Accident Period"]]
        tmp["_age"] = _num(tmp["Current Age"])
        tmp = tmp.dropna(subset=["_p", "_age"])
        age_lkp = {r["_p"]: r["_age"] for _, r in tmp.iterrows()}

        rows = sorted(
            [(age_lkp[p], reserves[p] / inc_actual[p])
             for p in common
             if p in age_lkp and inc_actual.get(p, 0) > 0],
            key=lambda x: x[0]   # ascending age = ascending maturity
        )
        pcts = [r for _, r in rows]
        violations = sum(1 for i in range(len(pcts) - 1)
                         if pcts[i] < pcts[i + 1] - 0.02)
        if violations == 0:
            ck.ok(g, "Case reserve % of incurred non-increasing with maturity")
        else:
            ck.warn(g, "Case reserve % of incurred non-increasing with maturity",
                    f"{violations} reversal(s)")

    # IBNR decomposition identity: Unpaid - IBNR = Incurred Actual - Paid Actual
    if "IBNR" in inc_df.columns and "Unpaid" in inc_df.columns:
        inc_ibnr   = _canon_map(inc_df, "Accident Period", "IBNR")
        inc_unpaid = _canon_map(inc_df, "Accident Period", "Unpaid")
        bad = []
        for p in common:
            ibnr   = inc_ibnr.get(p)
            unpaid = inc_unpaid.get(p)
            cr     = reserves.get(p)
            if ibnr is None or unpaid is None or cr is None:
                continue
            # Unpaid = Sel_inc - Paid_actual; IBNR = Sel_inc - Inc_actual
            # Unpaid - IBNR = Inc_actual - Paid_actual = case reserves
            implied_cr = unpaid - ibnr
            if abs(implied_cr - cr) > 1.0:
                bad.append(f"{p}: Unpaid-IBNR={implied_cr:,.0f}, "
                           f"Inc-Paid={cr:,.0f}")
        if not bad:
            ck.ok(g, "Unpaid - IBNR = Incurred - Paid Actual (identity)")
        else:
            ck.fail(g, "Unpaid - IBNR = Incurred - Paid Actual",
                    f"{len(bad)} mismatch(es): {bad[:2]}")


# ── Group 14: Closure Rate Checks ────────────────────────────────────────────

def check_closure_rates(ck, measure_dfs):
    g = "14. Closure Rate Checks"

    rep_df = measure_dfs.get("Reported Count")
    cls_df = measure_dfs.get("Closed Count")

    if rep_df is None or cls_df is None:
        ck.warn(g, "Closure rate checks",
                "Reported Count and/or Closed Count missing — skipping")
        return

    rep_actual = _canon_map(rep_df, "Accident Period", "Actual")
    cls_actual = _canon_map(cls_df, "Accident Period", "Actual")

    common = sorted(p for p in rep_actual if p in cls_actual and rep_actual[p] > 0)
    if not common:
        ck.warn(g, "Closure rates computed", "No common periods with positive reported")
        return

    closure_rates = {p: cls_actual[p] / rep_actual[p] for p in common}

    # Closure rate > 1.0 = FAIL
    above_one = [p for p, cr in closure_rates.items() if cr > 1.0 + 1e-6]
    if not above_one:
        ck.ok(g, "Closure rate (Closed/Reported Actual) <= 1.0 for all periods")
    else:
        details = [f"{p}: {closure_rates[p]:.3f}" for p in sorted(above_one)]
        ck.fail(g, "Closure rate <= 1.0",
                f"{len(above_one)} period(s) > 1.0: {details[:3]}")

    # Closure rate non-decreasing with maturity (more mature = more closed)
    if _all_int(set(common)) and "Current Age" in rep_df.columns:
        tmp = rep_df[["Accident Period", "Current Age"]].copy()
        tmp["_p"]   = [_canon(p) for p in tmp["Accident Period"]]
        tmp["_age"] = _num(tmp["Current Age"])
        tmp = tmp.dropna(subset=["_p", "_age"])
        age_lkp = {r["_p"]: r["_age"] for _, r in tmp.iterrows()}

        rows = sorted(
            [(age_lkp[p], closure_rates[p])
             for p in common if p in age_lkp],
            key=lambda x: x[0]   # ascending age = ascending maturity
        )
        rates = [r for _, r in rows]
        violations = sum(1 for i in range(len(rates) - 1)
                         if rates[i] > rates[i + 1] + 0.02)
        if violations == 0:
            ck.ok(g, "Closure rate non-decreasing with maturity")
        else:
            ck.warn(g, "Closure rate non-decreasing with maturity",
                    f"{violations} reversal(s)")


# ── Group 15: Severity / Frequency Trends ────────────────────────────────────

def check_severity_frequency_trends(ck, diag_df):
    g = "15. Severity / Frequency Trends"

    if diag_df is None or diag_df.empty:
        ck.warn(g, "Trend checks", "Diagnostics sheet missing — skipping")
        return

    if "Ultimate Severity" in diag_df.columns:
        sev = _num(diag_df["Ultimate Severity"]).dropna().reset_index(drop=True)
        if len(sev) >= 2:
            pct_changes = []
            for i in range(len(sev) - 1):
                if sev[i] > 0:
                    pct_changes.append(abs((sev[i + 1] - sev[i]) / sev[i]))
            large = [(i + 1, pc) for i, pc in enumerate(pct_changes) if pc > 0.25]
            if not large:
                ck.ok(g, "YoY severity change <= 25% (all periods)")
            else:
                labels = [f"period {i}: {pc:.0%}" for i, pc in large]
                ck.warn(g, "YoY severity change > 25% detected",
                        f"{len(large)} period(s): {labels[:3]}")
        else:
            ck.warn(g, "Severity trend", "Need >= 2 periods")

    if "Ultimate Frequency" in diag_df.columns:
        freq = _num(diag_df["Ultimate Frequency"]).dropna()
        if len(freq) >= 2:
            med = freq.median()
            if med > 0:
                spikes = (freq > 2.0 * med).sum()
                if spikes == 0:
                    ck.ok(g, "Frequency no spike > 2x median", f"Median={med:.4f}")
                else:
                    ck.warn(g, "Frequency spike > 2x median detected",
                            f"{spikes} period(s) above 2x median ({med:.4f})")


# ── Group 16: Selection Reasonableness ───────────────────────────────────────

def check_selection_reasonableness(ck, measure_dfs):
    g = "16. Selection Reasonableness"

    for m, df in measure_dfs.items():
        if "Selected Ultimate" not in df.columns:
            continue

        sel = _canon_map(df, "Accident Period", "Selected Ultimate")
        if not sel:
            continue

        # Collect available method columns (non-trivially populated)
        method_cols = []
        for col in ["Chain Ladder", "BF", "Initial Expected"]:
            if col in df.columns:
                vals = _num(df[col])
                if vals.notna().any():
                    method_cols.append(col)

        # Selection within range of method indications
        if len(method_cols) >= 2:
            violations_outside = []
            for _, row in df.iterrows():
                p = _canon(row.get("Accident Period"))
                if p is None:
                    continue
                s = sel.get(p)
                if s is None or pd.isna(s):
                    continue
                method_vals = []
                for col in method_cols:
                    v = row.get(col)
                    if pd.notna(v):
                        try:
                            method_vals.append(float(v))
                        except (ValueError, TypeError):
                            pass
                if len(method_vals) < 2:
                    continue
                lo, hi = min(method_vals), max(method_vals)
                if s < lo - 1.0 or s > hi + 1.0:
                    violations_outside.append(
                        f"{p}: sel={s:,.0f} outside [{lo:,.0f}, {hi:,.0f}]")
            if not violations_outside:
                ck.ok(g, f"'{m}' selections within range of method indications")
            else:
                ck.warn(g, f"'{m}' selections within range of method indications",
                        f"{len(violations_outside)} period(s): {violations_outside[:2]}")

        # Shallow review flag: all selections == CL within tolerance
        if "Chain Ladder" in df.columns:
            cl_vals = _canon_map(df, "Accident Period", "Chain Ladder")
            matching = [
                p for p in sel
                if p in cl_vals
                and pd.notna(sel.get(p)) and pd.notna(cl_vals.get(p))
                and abs(sel[p] - cl_vals[p]) <= 1.0
            ]
            total_common = sum(1 for p in sel if p in cl_vals
                               and pd.notna(sel.get(p)) and pd.notna(cl_vals.get(p)))
            if total_common >= 3 and len(matching) == total_common:
                ck.warn(g, f"'{m}' selections not all equal to Chain Ladder",
                        "All periods selected = CL — possible shallow review")
            elif total_common >= 1:
                ck.ok(g, f"'{m}' selections show method differentiation",
                      f"{len(matching)}/{total_common} periods = CL")


# ── Excel report writer ───────────────────────────────────────────────────────

def _write_sheet(ws, rows, title):
    """Write a results table into ws. rows = list of result dicts."""
    headers = ["Status", "Group", "Check", "Detail"]
    widths  = [8, 24, 55, 60]
    ws.title = title
    for c_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    fills = {"PASS": PASS_FILL, "WARN": WARN_FILL, "FAIL": FAIL_FILL}
    for r_idx, row in enumerate(rows, start=2):
        status = row["status"]
        rfill  = fills.get(status, PASS_FILL)
        for c_idx, key in enumerate(["status", "group", "check", "detail"], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[key])
            cell.fill = rfill
            cell.border = THIN
            cell.font = Font(size=9, bold=(c_idx == 1))
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c_idx == 1 else "left",
                wrap_text=(c_idx == 4))


def write_excel_report(results, path):
    wb = Workbook()

    # Sheet 1: FAIL + WARN only
    issues = [r for r in results if r["status"] in ("FAIL", "WARN")]
    _write_sheet(wb.active, issues, "Issues")

    # Sheet 2: full list
    ws_all = wb.create_sheet("All Checks")
    _write_sheet(ws_all, results, "All Checks")

    os.makedirs(pathlib.Path(path).parent, exist_ok=True)
    wb.save(path)
    print(f"\n  Excel report -> {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"=== Tech Review: {INPUT_COMPLETE_ANALYSIS} ===\n")
    ck = Checker()

    try:
        wb = load_workbook(INPUT_COMPLETE_ANALYSIS, data_only=True)
    except Exception as e:
        ck.fail("1. Structure", "File readable", str(e))
        write_excel_report(ck.results, OUTPUT_REVIEW)
        return

    info = detect_sheets(wb)
    print(f"Sheets: {len(wb.sheetnames)} | Measures: {len(info['measure_sheets'])} | "
          f"CL: {len(info['cl_sheets'])} | Sel: {len(info['sel_sheets'])} | "
          f"Triangles: {len(info['tri_sheets'])}\n")

    measure_dfs   = {m: read_with_title(wb[m]) for m in info["measure_sheets"]}
    sel_dfs       = {s: read_no_title(wb[s])   for s in info["sel_sheets"]}
    cl_dfs        = {c: read_cl_sheet(wb[c])   for c in info["cl_sheets"]}
    tri_dfs       = {t: read_with_title(wb[t]) for t in info["tri_sheets"]}
    diag_df       = read_with_title(wb["Diagnostics"])    if info["diag_sheet"] else None
    avg_ibnr_df   = read_with_title(wb["Average IBNR"])   if info["avg_ibnr"]   else None
    avg_unpaid_df = read_with_title(wb["Average Unpaid"]) if info["avg_unpaid"] else None

    print("--- Group 1: Structure ---")
    check_structure(ck, wb, info)

    print("\n--- Group 2: Period Consistency ---")
    check_period_consistency(ck, info, measure_dfs, diag_df, tri_dfs, cl_dfs)

    print("\n--- Group 3: Current Age / Maturity ---")
    check_maturity(ck, measure_dfs)

    print("\n--- Group 4: Selected Ultimate Integrity ---")
    check_selected_ultimates(ck, measure_dfs)

    print("\n--- Group 5: Cross-Measure Consistency ---")
    check_cross_measure(ck, measure_dfs)

    print("\n--- Group 6: Sel - Sheet Consistency ---")
    check_sel_sheets(ck, measure_dfs, sel_dfs)

    print("\n--- Group 7: X-to-Ult Triangles ---")
    check_xtoult_triangles(ck, tri_dfs, measure_dfs)

    print("\n--- Group 8: Average IBNR / Unpaid ---")
    check_avg_triangles(ck, avg_ibnr_df, avg_unpaid_df)

    print("\n--- Group 9: Diagnostics ---")
    check_diagnostics(ck, diag_df)

    print("\n--- Group 10: CL Triangle Integrity ---")
    check_cl_triangles(ck, cl_dfs, measure_dfs, tri_dfs)

    print("\n--- Group 11: Development Factors ---")
    check_development_factors(ck, cl_dfs, measure_dfs)

    print("\n--- Group 12: Paid-to-Incurred Raw Data ---")
    check_paid_incurred_raw(ck, measure_dfs, cl_dfs)

    print("\n--- Group 13: Case Reserve Reasonableness ---")
    check_case_reserves(ck, measure_dfs)

    print("\n--- Group 14: Closure Rate Checks ---")
    check_closure_rates(ck, measure_dfs)

    print("\n--- Group 15: Severity / Frequency Trends ---")
    check_severity_frequency_trends(ck, diag_df)

    print("\n--- Group 16: Selection Reasonableness ---")
    check_selection_reasonableness(ck, measure_dfs)

    counts = ck.counts()
    total  = sum(counts.values())
    print(f"\n{'='*60}")
    print(f"Checks run: {total} | PASS: {counts['PASS']} | "
          f"WARN: {counts['WARN']} | FAIL: {counts['FAIL']}")
    print(f"{'='*60}")

    write_excel_report(ck.results, OUTPUT_REVIEW)


if __name__ == "__main__":
    main()
