# Refreshes only the selections section of the Excel workbook based on saved selections from a
# JSON file. Allows you to update the displayed selections without rebuilding the entire workbook,
# which is useful when you need to revert changes or apply selections from another source.

"""
goal: Update ONLY the selections section of Chain Ladder Selections.xlsx from chain-ladder.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 2b-chainladder-update-selections.py
"""

import json
import os
import openpyxl
from openpyxl.styles import Alignment, Font

from modules import config
from modules.xl_styles import SELECTION_FILL, LABEL_FONT, DATA_FONT, THIN_BORDER

# Paths from modules/config.py — override here if needed:
METHOD_ID = "chainladder"
SELECTIONS_FILE = config.SELECTIONS + f"{METHOD_ID}.json"
EXCEL_FILE      = config.SELECTIONS + "Chain Ladder Selections.xlsx"


def find_selections_section(ws):
    """Find the row where 'LDF Selections' section header and interval headers are."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "LDF Selections":
                section_row = cell.row
                header_row = section_row + 1
                # Scan for the "Selection" label row (not "Prior Selection")
                # It may be offset by prior selection rows if they exist
                for check_row in range(header_row + 1, header_row + 10):
                    label = ws.cell(row=check_row, column=1).value
                    if label == "Selection":
                        selection_row = check_row
                        reasoning_row = check_row + 1
                        return header_row, selection_row, reasoning_row
                return None, None, None
    return None, None, None


def get_interval_columns(ws, header_row):
    """Build a dict mapping interval string -> column index from the header row."""
    interval_to_col = {}
    for cell in ws[header_row]:
        if cell.value:
            interval_to_col[str(cell.value).strip()] = cell.column
    return interval_to_col


def has_existing_selections(ws):
    """Return True if any Selection row cells in this sheet already have values."""
    _, selection_row, _ = find_selections_section(ws)
    if selection_row is None:
        return False
    for cell in ws[selection_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def update_sheet(ws, measure_selections):
    """Update the selections section of a single sheet."""
    header_row, selection_row, reasoning_row = find_selections_section(ws)
    if header_row is None:
        print(f"  WARNING: Could not find 'LDF Selections' section in sheet '{ws.title}'")
        return

    interval_to_col = get_interval_columns(ws, header_row)

    for sel in measure_selections:
        interval = sel["interval"]
        if interval not in interval_to_col:
            print(f"  WARNING: Interval '{interval}' not found in sheet '{ws.title}'")
            continue

        col = interval_to_col[interval]

        # Write selection value
        sel_cell = ws.cell(row=selection_row, column=col)
        sel_cell.value = sel["selection"]
        sel_cell.fill = SELECTION_FILL
        sel_cell.font = DATA_FONT
        sel_cell.alignment = Alignment(horizontal="right")
        sel_cell.border = THIN_BORDER
        sel_cell.number_format = "0.0000"

        # Write reasoning
        reason_cell = ws.cell(row=reasoning_row, column=col)
        reason_cell.value = sel["measure"] + (": " + sel["reasoning"] if sel.get("reasoning") else "")
        reason_cell.fill = SELECTION_FILL
        reason_cell.font = Font(size=8, italic=True)
        reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        reason_cell.border = THIN_BORDER

    # Set row height for reasoning row
    ws.row_dimensions[reasoning_row].height = 60

    print(f"  Updated {len(measure_selections)} selections in '{ws.title}'")


def main():
    """Update Chain Ladder Selections Excel file with selections from JSON."""
    # Check if JSON file exists
    if not os.path.exists(SELECTIONS_FILE):
        print(f"Selections file not found: {SELECTIONS_FILE}")
        print("Skipping update - no selections to apply.")
        return
    
    # Load selections
    with open(SELECTIONS_FILE) as f:
        selections = json.load(f)
    
    # Check if selections array is empty
    if not selections or len(selections) == 0:
        print(f"No selections found in {SELECTIONS_FILE}")
        print("Skipping update - selections array is empty.")
        return

    print(f"Loaded {len(selections)} selections from {SELECTIONS_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Abort if any sheet already has manual selections to avoid overwriting actuary work
    sheets_with_selections = [s for s in wb.sheetnames if has_existing_selections(wb[s])]
    if sheets_with_selections:
        raise ValueError(
            f"Existing selections found in sheet(s): {sheets_with_selections}\n"
            "Clear selections manually before re-running to avoid overwriting actuary edits."
        )

    # Group selections by measure
    by_measure = {}
    for sel in selections:
        m = sel["measure"]
        by_measure.setdefault(m, []).append(sel)

    for sheet_name in wb.sheetnames:
        # Match sheet name to measure (sheet name may be truncated to 31 chars)
        matched = None
        for measure in by_measure:
            if sheet_name == measure[:31]:
                matched = measure
                break
        if matched:
            update_sheet(wb[sheet_name], by_measure[matched])
        else:
            print(f"No selections found for sheet '{sheet_name}' - skipping")

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
