# Shared Excel formatting constants and helpers for reserving analysis workbooks.

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")   # dark blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")   # medium blue
SECTION_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue
SELECTION_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow
PRIOR_FILL     = PatternFill("solid", fgColor="E2EFDA")   # light green
AI_FILL        = PatternFill("solid", fgColor="EDE7F6")   # light purple

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT   = Font(bold=True, size=10)
LABEL_FONT     = Font(bold=True, size=9)
DATA_FONT      = Font(size=9)

THIN        = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_STYLE_MAP = {
    "header":    (HEADER_FILL,    HEADER_FONT),
    "subheader": (SUBHEADER_FILL, SUBHEADER_FONT),
    "section":   (SECTION_FILL,   SECTION_FONT),
    "selection": (SELECTION_FILL, LABEL_FONT),
    "prior":     (PRIOR_FILL,     LABEL_FONT),
    "ai":        (AI_FILL,        LABEL_FONT),
}


def style_header(cell, level="header"):
    """Apply standard fill + font + centered alignment + border to a cell."""
    fill, font = _STYLE_MAP.get(level, (SUBHEADER_FILL, SUBHEADER_FONT))
    cell.fill      = fill
    cell.font      = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = THIN_BORDER
