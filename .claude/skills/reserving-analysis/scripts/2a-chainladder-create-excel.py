# Creates a formatted Excel workbook that displays all the claims data, development factors,
# diagnostics, and calculated averages in an organized layout. Actuaries review this workbook and
# manually enter their selected factors based on the information presented.

"""
goal: Create Chain Ladder Selections.xlsx for actuarial LDF review and selection. This file should be manually edited by the actuary to make selections.

Sheet layout:
  - One main sheet per measure: loss triangle, LDF triangle, averages (no CV/slopes), selections
  - One sheet per diagnostic triangle: "Diag - {Name}" (shared across measures)
  - One CV & slopes sheet per measure: "{Measure} - CV & Slopes"

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 2a-chainladder-create-excel.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pathlib import Path

from modules import config
from modules.xl_styles import (
    HEADER_FILL, SUBHEADER_FILL, SECTION_FILL, SELECTION_FILL, PRIOR_FILL, AI_FILL, USER_FILL,
    HEADER_FONT, SUBHEADER_FONT, SECTION_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)

# Paths from modules/config.py — override here if needed:
OUTPUT_PATH = config.PROCESSED_DATA
SELECTIONS_OUTPUT_PATH = config.SELECTIONS
METHOD_ID = "chainladder"
OUTPUT_FILE_NAME = "Chain Ladder Selections - LDFs.xlsx"

# Short labels for diagnostic sheet names (max ~14 chars to fit in 31-char sheet name)
DIAG_SHEET_LABELS = {
    'reported_claims': 'Rptd Claims',
    'incurred_severity': 'Inc Severity',
    'paid_severity': 'Paid Severity',
    'paid_to_incurred': 'Paid to Incurred',
    'open_counts': 'Open Counts',
    'average_case_reserve': 'Avg Case Rsrv',
    'claim_closure_rate': 'Clm Closure Rt',
    'incremental_incurred_severity': 'Incr Inc Svty',
    'incremental_paid_severity': 'Incr Paid Svty',
    'incremental_closure_rate': 'Incr Closure Rt',
}

# Number formats for each diagnostic column
DIAG_NUMBER_FORMATS = {
    'reported_claims': '#,##0',
    'incurred_severity': '#,##0',
    'paid_severity': '#,##0',
    'paid_to_incurred': '0.00%',
    'open_counts': '#,##0',
    'average_case_reserve': '#,##0',
    'claim_closure_rate': '0.00%',
    'incremental_incurred_severity': '#,##0',
    'incremental_paid_severity': '#,##0',
    'incremental_closure_rate': '0.00%',
}


def diag_sheet_name(diag_col):
    label = DIAG_SHEET_LABELS.get(diag_col, diag_col.replace('_', ' ').title()[:14])
    return f"Diag - {label}"[:31]


def cv_slopes_sheet_name(measure):
    return f"{measure} - CV & Slopes"[:31]


def write_triangle(ws, start_row, title, row_labels, col_labels, data_dict, number_format="#,##0"):
    """Write a titled triangle section. data_dict: {(row_label, col_label): value}"""
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    for row_label in row_labels:
        row_cell = ws.cell(row=start_row, column=1, value=row_label)
        row_cell.font = LABEL_FONT
        row_cell.alignment = Alignment(horizontal="left")
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((str(row_label), str(col)))
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = number_format
        start_row += 1

    return start_row + 1


def write_metrics_table(ws, start_row, title, metric_rows, col_labels, data_dict):
    """Write a metrics table. data_dict: {(metric_name, col_label): value}"""
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="Metric").border = THIN_BORDER
    ws.cell(row=start_row, column=1).font = SUBHEADER_FONT
    ws.cell(row=start_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    for metric in metric_rows:
        row_cell = ws.cell(row=start_row, column=1, value=metric)
        row_cell.font = LABEL_FONT
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((metric, str(col)))
            cell = ws.cell(row=start_row, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if val is not None:
                cell.number_format = "0.0000"
        start_row += 1

    return start_row + 1


def write_selections_section(ws, start_row, col_labels, prior_selections=None, measure=None):
    """Write selection section with optional prior selections followed by blank rows for new selections."""
    title_cell = ws.cell(row=start_row, column=1, value="LDF Selections")
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    if prior_selections is not None and measure is not None:
        measure_prior = prior_selections[prior_selections['measure'] == measure]
        if not measure_prior.empty:
            prior_dict = {}
            for _, row in measure_prior.iterrows():
                prior_dict[str(row['interval'])] = {
                    'selection': row['selection'],
                    'reasoning': row['reasoning']
                }

            cell = ws.cell(row=start_row, column=1, value="Prior Selection")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['selection']
                    c.number_format = "0.0000"
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="right")
            start_row += 1

            cell = ws.cell(row=start_row, column=1, value="Prior Reasoning")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['reasoning']
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="left", wrap_text=True)
            start_row += 1

            start_row += 1

    for label in ["AI Selection", "AI Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "ai")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = AI_FILL
            c.border = THIN_BORDER
            if label == "AI Selection":
                c.number_format = "0.0000"
            if label == "AI Reasoning":
                c.alignment = Alignment(horizontal="left", wrap_text=True)
                c.font = DATA_FONT
        start_row += 1

    start_row += 1

    for label in ["Selection", "Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "selection")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = SELECTION_FILL
            c.border = THIN_BORDER
            if label == "Selection":
                c.number_format = "0.0000"
        start_row += 1

    start_row += 1

    for label in ["User Selection", "User Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "user")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = USER_FILL
            c.border = THIN_BORDER
            if label == "User Selection":
                c.number_format = "0.0000"
            if label == "User Reasoning":
                c.alignment = Alignment(horizontal="left", wrap_text=True)
                c.font = DATA_FONT
        start_row += 1

    return start_row


_AVG_PERIOD_ORDER = ['all', '3yr', '5yr', '10yr']
_AVG_TYPE_ORDER = ['simple', 'weighted', 'exclude_high_low']


def _sort_avg_cols(cols):
    """Sort average columns: group by period (all/3yr/5yr/10yr), within each: simple > weighted > exclude_high_low."""
    def key(col):
        normalized = col.replace('avg_', '')
        for p_idx, p in enumerate(_AVG_PERIOD_ORDER):
            for t_idx, t in enumerate(_AVG_TYPE_ORDER):
                if normalized == f"{t}_{p}":
                    return (p_idx, t_idx)
        return (99, 99)
    return sorted(cols, key=key)


def _avg_display_name(col):
    """Strip 'avg_' prefix from exclude_high_low variants for cleaner display."""
    return col.replace('avg_exclude_high_low', 'exclude_high_low')


def build_main_sheet(ws, measure, df2, df4, df_prior=None):
    """Build main measure sheet: loss triangle, age-to-age factors, averages (no CV/slopes), selections."""
    df_m = df2[df2['measure'] == measure].copy()
    periods = df_m['period'].cat.categories.tolist()
    ages = [str(a) for a in df_m['age'].cat.categories.tolist()]
    intervals = [str(i) for i in df_m['interval'].dropna().cat.categories.tolist()]

    # Loss Triangle — title is just the measure name
    loss_dict = {}
    for _, row in df_m.iterrows():
        loss_dict[(str(row['period']), str(row['age']))] = row['value']
    row_ptr = write_triangle(ws, 1, measure, periods, ages, loss_dict, "#,##0")

    # Age-to-Age Factors
    ldf_dict = {}
    for _, row in df_m[df_m['ldf'].notna()].iterrows():
        ldf_dict[(str(row['period']), str(row['interval']))] = row['ldf']
    row_ptr = write_triangle(ws, row_ptr, "Age-to-Age Factors", periods, intervals, ldf_dict, "0.0000")

    # Averages (exclude CV and slope columns; reorder and rename for display)
    df_avg = df4[df4['measure'] == measure].copy()
    raw_avg_cols = [col for col in df_avg.columns
                    if col not in ['measure', 'interval']
                    and not col.startswith('cv_')
                    and not col.startswith('slope_')]
    avg_cols = _sort_avg_cols(raw_avg_cols)
    display_cols = [_avg_display_name(c) for c in avg_cols]

    avg_dict = {}
    for _, row in df_avg.iterrows():
        for raw, display in zip(avg_cols, display_cols):
            avg_dict[(display, str(row['interval']))] = row[raw]
    intervals_with_tail = intervals + ["Tail"]
    row_ptr = write_metrics_table(ws, row_ptr, "Averages", display_cols, intervals_with_tail, avg_dict)

    # Selections
    write_selections_section(ws, row_ptr, intervals_with_tail, prior_selections=df_prior, measure=measure)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(ages) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B2"


def build_diagnostic_sheet(ws, diag_col, df2, df3):
    """Build a single diagnostic triangle sheet (shared across all measures)."""
    first_measure = df2['measure'].cat.categories[0]
    df_m = df2[df2['measure'] == first_measure].copy()
    periods = df_m['period'].cat.categories.tolist()
    ages = [str(a) for a in df_m['age'].cat.categories.tolist()]

    number_format = DIAG_NUMBER_FORMATS.get(diag_col, "0.0000")

    diag_dict = {}
    for _, row in df3.iterrows():
        diag_dict[(str(row['period']), str(row['age']))] = row[diag_col]

    title = diag_col.replace('_', ' ').title()
    write_triangle(ws, 1, title, periods, ages, diag_dict, number_format)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(ages) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B2"


def build_cv_slopes_sheet(ws, measure, df2, df4):
    """Build CV & slopes metrics sheet for one measure."""
    df_m = df2[df2['measure'] == measure].copy()
    intervals = [str(i) for i in df_m['interval'].dropna().cat.categories.tolist()]
    intervals_with_tail = intervals + ["Tail"]

    df_avg = df4[df4['measure'] == measure].copy()
    cv_slope_cols = [col for col in df_avg.columns
                     if col not in ['measure', 'interval']
                     and (col.startswith('cv_') or col.startswith('slope_'))]

    if not cv_slope_cols:
        return

    cv_dict = {}
    for _, row in df_avg.iterrows():
        for metric in cv_slope_cols:
            cv_dict[(metric, str(row['interval']))] = row[metric]

    write_metrics_table(ws, 1, f"{measure} - CV & Slopes", cv_slope_cols, intervals_with_tail, cv_dict)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(intervals_with_tail) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12
    ws.freeze_panes = "B2"


def main():
    """Create Chain Ladder Selections Excel file with all measures."""
    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    if Path(output_file).exists():
        raise FileExistsError(
            f"Output file already exists: {output_file}\n"
            "Delete or rename the file before re-running to avoid overwriting manual edits."
        )

    df2 = pd.read_parquet(OUTPUT_PATH + "2_enhanced.parquet")
    df3 = pd.read_parquet(OUTPUT_PATH + "3_diagnostics.parquet")
    df4 = pd.read_parquet(OUTPUT_PATH + "4_ldf_averages.parquet")

    print(f"Loaded data: {len(df2)} enhanced rows, {len(df3)} diagnostic rows, {len(df4)} average rows")

    prior_selections_path = Path(OUTPUT_PATH) / "../prior-selections.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"Loaded {len(df_prior)} prior selections")
    else:
        print("No prior selections found (optional)")

    for col in df3.columns:
        if col not in ['period', 'age']:
            if ('rate' in col.lower() or 'ratio' in col.lower() or 'to_' in col.lower()):
                if df3[col].dropna().abs().max() <= 2.0:
                    df3[col] = df3[col].astype(float)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    measures = df2['measure'].cat.categories.tolist()
    diagnostic_cols = [col for col in df3.columns if col not in ['period', 'age']]

    # Main sheet per measure
    for measure in measures:
        ws = wb.create_sheet(title=measure[:31])
        build_main_sheet(ws, measure, df2, df4, df_prior=df_prior)
        print(f"Built main sheet: {measure[:31]}")

    # One sheet per diagnostic (shared across measures)
    for diag_col in diagnostic_cols:
        sheet_title = diag_sheet_name(diag_col)
        ws = wb.create_sheet(title=sheet_title)
        build_diagnostic_sheet(ws, diag_col, df2, df3)
        print(f"Built diagnostic sheet: {sheet_title}")

    # CV & slopes sheet per measure
    for measure in measures:
        sheet_title = cv_slopes_sheet_name(measure)
        ws = wb.create_sheet(title=sheet_title)
        build_cv_slopes_sheet(ws, measure, df2, df4)
        print(f"Built CV & Slopes sheet: {sheet_title}")

    wb.save(output_file)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    main()
