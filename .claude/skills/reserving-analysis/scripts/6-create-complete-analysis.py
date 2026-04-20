# Reads projected ultimates, actuary selections, and triangle data to produce a
# final Excel workbook summarizing the complete reserving analysis.

"""
goal: Create complete-analysis.xlsx (and three supporting workbooks) from the outputs
of the prior numbered scripts.

Outputs:
  - selected-ultimates.xlsx     Loss + Counts sheets with selected ultimates per period
  - post-method-series.xlsx     Ultimate Severity, Loss Rate, Frequency diagnostics
  - post-method-triangles.xlsx  X-to-Ultimate triangles, Average IBNR, Average Unpaid
  - complete-analysis.xlsx      Master workbook combining all prior Excel outputs + above

Gracefully handles optional data:
  - ultimate_ie (IE method, script 3): omits IE columns if not present
  - ultimate_bf (BF method, script 4): omits BF columns if not present
  - Exposure in triangles: omits Loss Rate and Frequency if not present
  - ultimates.json selections: falls back to bf > cl > ie if file is absent or a row is missing

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 6-create-complete-analysis.py
"""

import copy
import json
import os
import pathlib
import re

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import (
    HEADER_FILL, SUBHEADER_FILL, SECTION_FILL, SELECTION_FILL,
    HEADER_FONT, SUBHEADER_FONT, SECTION_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)

# ── User-configurable properties ─────────────────────────────────────────────
# Paths from modules/config.py — override here if needed:

INPUT_ULTIMATES       = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_SELECTIONS_JSON = config.SELECTIONS + "ultimates.json"
INPUT_TRIANGLES       = config.PROCESSED_DATA + "1_triangles.parquet"
OUTPUT_PATH           = config.OUTPUT

# Excel files from prior scripts to fold into the complete analysis workbook.
# Each entry: (path_to_file, sheet_name_prefix_or_None).
# Files that do not exist are silently skipped.
ANALYSIS_SOURCE_FILES = [
    (config.SELECTIONS + "Chain Ladder Selections.xlsx", "CL - "),
    (config.SELECTIONS + "Ultimates.xlsx",               "Sel - "),
]

# Maps each measure to its "unpaid" proxy measure used to compute the Unpaid column.
# Unpaid = Selected Ultimate − latest actual of the proxy measure for that period.
UNPAID_PROXY = {
    "Incurred Loss":  "Paid Loss",
    "Paid Loss":      "Paid Loss",
    "Reported Count": "Closed Count",
    "Closed Count":   "Closed Count",
}

# ─────────────────────────────────────────────────────────────────────────────
# Derived output paths — do not modify.
OUTPUT_SELECTED_ULTIMATES = OUTPUT_PATH + "selected-ultimates.xlsx"
OUTPUT_POST_SERIES        = OUTPUT_PATH + "post-method-series.xlsx"
OUTPUT_POST_TRIANGLES     = OUTPUT_PATH + "post-method-triangles.xlsx"
OUTPUT_COMPLETE_ANALYSIS  = OUTPUT_PATH + "complete-analysis.xlsx"

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# Sheet-ref patterns for rewriting formulas after the prefix rename.
# Quoted names may contain anything except ' and !; unquoted names are alphanumerics + underscore.
_QUOTED_SHEET_REF = re.compile(r"'([^'!]+)'!")
_UNQUOTED_SHEET_REF = re.compile(r"(?<![A-Za-z0-9_'\"\]])([A-Za-z_][A-Za-z0-9_]*)!")


def _quote_sheet_name(name):
    """Quote a sheet name for use in a formula, unless it's a simple identifier."""
    if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        return name
    return f"'{name}'"


def rewrite_formula_sheet_refs(formula, rename_map):
    """
    Rewrite sheet references in a formula string per rename_map {old_name: new_name}.
    Preserves non-formula values unchanged. Handles both quoted and unquoted refs.

    Applied during the sheet-copy loop so that formulas like ='Closed Count'!E56 in
    a source workbook continue to resolve after that sheet is renamed to 'CL - Closed Count'
    in the master workbook.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _sub(m):
        old = m.group(1)
        new = rename_map.get(old, old)
        return f"{_quote_sheet_name(new)}!"

    result = _QUOTED_SHEET_REF.sub(_sub, formula)
    result = _UNQUOTED_SHEET_REF.sub(_sub, result)
    return result


def _style_cell(cell, level="subheader"):
    style_header(cell, level)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _try_int(val):
    """Convert value to integer where possible, else return original value."""
    try:
        num = float(val)
        if num == int(num):
            return int(num)
        return num
    except (ValueError, TypeError):
        return val


def _col_has_data(df, col):
    """True if column exists in df and contains at least one non-NaN value."""
    return col in df.columns and df[col].notna().any()


def _write_header_row(ws, headers, row=1, level="subheader", col_width=22):
    """Write a styled header row at the given row number."""
    for c_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=text)
        _style_cell(cell, level)
        ws.column_dimensions[get_column_letter(c_idx)].width = col_width


def _write_data_cell(cell, value, num_fmt=None, is_numeric=False):
    """Write a data cell with consistent font, border, alignment."""
    # Ensure numeric values are stored as numbers, not text
    if value is not None and is_numeric:
        try:
            cell.value = float(value) if not isinstance(value, (int, float)) else value
        except (ValueError, TypeError):
            cell.value = value
    else:
        cell.value = value
    
    cell.font   = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                               vertical="center")
    if num_fmt and cell.value is not None and isinstance(cell.value, (int, float)):
        cell.number_format = num_fmt


def _safe(val):
    """Return None for NaN so openpyxl writes a blank cell."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


# ── Data loading ──────────────────────────────────────────────────────────────

def load_combined(ultimates_path, selections_json_path):
    """
    Load projected ultimates and merge in actuary selections from JSON.

    Returns:
        combined  (DataFrame): one row per (period, measure) with columns
                               period, measure, current_age, actual,
                               ultimate_cl, [ultimate_ie], [ultimate_bf],
                               selected_ultimate, selected_ibnr, selected_unpaid
        has_ie    (bool): True when ultimate_ie data is present
        has_bf    (bool): True when ultimate_bf data is present
    """
    df = pd.read_parquet(ultimates_path)
    df["period"]  = df["period"].astype(str)
    df["measure"] = df["measure"].astype(str)

    has_ie = _col_has_data(df, "ultimate_ie")
    has_bf = _col_has_data(df, "ultimate_bf")

    # Load actuary selections; missing file or missing row both gracefully fall back.
    sel_lookup = {}
    sel_path = pathlib.Path(selections_json_path)
    if sel_path.exists():
        with open(sel_path, "r") as f:
            entries = json.load(f)
        for entry in entries:
            key = (str(entry["measure"]), str(entry["period"]))
            sel_lookup[key] = float(entry["selection"])
    else:
        print(f"  Note: {selections_json_path} not found — using method fallback for selected ultimate.")

    # selected_ultimate: JSON entry → bf → cl → ie (first non-NaN available)
    def _pick_selected(row):
        key = (row["measure"], row["period"])
        if key in sel_lookup:
            return sel_lookup[key]
        for col in ("ultimate_bf", "ultimate_cl", "ultimate_ie"):
            if col in df.columns:
                v = row.get(col, np.nan)
                if pd.notna(v):
                    return v
        return np.nan

    df["selected_ultimate"] = df.apply(_pick_selected, axis=1)
    df["selected_ibnr"]     = df["selected_ultimate"] - df["actual"]

    # Unpaid = selected ultimate − actual of the proxy measure for that period
    actual_lookup = df.set_index(["period", "measure"])["actual"].to_dict()

    def _pick_unpaid(row):
        proxy = UNPAID_PROXY.get(row["measure"])
        if proxy is None:
            return np.nan
        proxy_actual = actual_lookup.get((row["period"], proxy), np.nan)
        if pd.isna(proxy_actual) or pd.isna(row["selected_ultimate"]):
            return np.nan
        return row["selected_ultimate"] - proxy_actual

    df["selected_unpaid"] = df.apply(_pick_unpaid, axis=1)

    return df, has_ie, has_bf


def get_exposure(triangles_df):
    """
    Extract latest exposure value per period from the triangles.

    Returns dict {str(period): float} or {} when no Exposure rows exist.
    """
    exp = triangles_df[triangles_df["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return {}
    exp["period"]  = exp["period"].astype(str)
    exp["age_int"] = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    latest = exp.sort_values("age_int").groupby("period").last()
    return latest["value"].to_dict()


# ── Excel writers ─────────────────────────────────────────────────────────────

def write_selected_ultimates(combined, has_ie, has_bf, path):
    """
    Write selected-ultimates.xlsx with one sheet per measure found in data.

    Sheet order follows MEASURE_ORDER when measures are present; unknown measures
    appended at end. Columns included/excluded dynamically based on which methods ran.
    """
    MEASURE_ORDER = ["Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"]

    present = combined["measure"].unique().tolist()
    measures = [m for m in MEASURE_ORDER if m in present]
    measures += [m for m in present if m not in MEASURE_ORDER]  # any extras

    def _get(df_idx, period, col):
        if period not in df_idx.index or col not in df_idx.columns:
            return None
        return _safe(df_idx.loc[period, col])

    wb = Workbook()
    wb.remove(wb.active)

    for measure in measures:
        df_m = combined[combined["measure"] == measure].set_index("period")
        if df_m.empty:
            continue

        headers = ["Accident Period", "Current Age", "Actual", "Chain Ladder"]
        if has_ie:
            headers.append("Initial Expected")
        if has_bf:
            headers.append("BF")
        headers += ["Selected Ultimate", "IBNR", "Unpaid"]

        ws = wb.create_sheet(title=measure[:31])

        # Section title spanning all columns
        title_cell = ws.cell(row=1, column=1, value=measure)
        _style_cell(title_cell, "section")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # Column headers
        _write_header_row(ws, headers, row=2, level="subheader", col_width=20)
        ws.column_dimensions["A"].width = 18
        ws.freeze_panes = "A3"

        periods = sorted(
            df_m.index,
            key=lambda x: (int(x) if str(x).isdigit() else x)
        )
        for row_idx, period in enumerate(periods, start=3):
            vals = [
                _try_int(period),
                _get(df_m, period, "current_age"),
                _get(df_m, period, "actual"),
                _get(df_m, period, "ultimate_cl"),
            ]
            if has_ie:
                vals.append(_get(df_m, period, "ultimate_ie"))
            if has_bf:
                vals.append(_get(df_m, period, "ultimate_bf"))
            vals += [
                _get(df_m, period, "selected_ultimate"),
                _get(df_m, period, "selected_ibnr"),
                _get(df_m, period, "selected_unpaid"),
            ]

            for c_idx, val in enumerate(vals, start=1):
                is_numeric = c_idx >= 1  # All columns are numeric (period is year)
                num_fmt = _NUM_FMT if c_idx > 2 else None  # Currency formatting for cols 3+
                if c_idx == 2:  # Age column gets 0 decimals but no comma
                    num_fmt = "0"
                _write_data_cell(ws.cell(row=row_idx, column=c_idx), val, num_fmt, is_numeric)

    os.makedirs(pathlib.Path(path).parent, exist_ok=True)
    wb.save(path)
    print(f"  Saved -> {path}")


def write_post_method_series(combined, exp_map, path):
    """
    Write post-method-series.xlsx with Ultimate Severity and, when exposure
    is available, Ultimate Loss Rate and Frequency.
    """
    has_exp = len(exp_map) > 0

    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    headers = ["Accident Period", "Ultimate Severity"]
    if has_exp:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostics"

    title_cell = ws.cell(row=1, column=1, value="Post-Method Series Diagnostics")
    _style_cell(title_cell, "section")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    _write_header_row(ws, headers, row=2, level="subheader", col_width=22)
    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "A3"

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else x)
    )
    for row_idx, period in enumerate(periods, start=3):
        ult_loss   = inc.loc[period, "selected_ultimate"] if period in inc.index else np.nan
        ult_counts = rep.loc[period, "selected_ultimate"] if period in rep.index else np.nan
        exp_val    = exp_map.get(period, np.nan)

        sev  = (ult_loss / ult_counts
                if pd.notna(ult_loss) and pd.notna(ult_counts) and ult_counts != 0
                else None)
        lr   = (ult_loss / exp_val
                if has_exp and pd.notna(ult_loss) and pd.notna(exp_val) and exp_val != 0
                else None)
        freq = (ult_counts / exp_val
                if has_exp and pd.notna(ult_counts) and pd.notna(exp_val) and exp_val != 0
                else None)

        row_vals = [_try_int(period), sev]
        if has_exp:
            row_vals += [lr, freq]

        for c_idx, val in enumerate(row_vals, start=1):
            is_numeric = True  # All columns are numeric
            num_fmt = _DEC_FMT if c_idx > 1 else None
            _write_data_cell(ws.cell(row=row_idx, column=c_idx), val, num_fmt, is_numeric)

    os.makedirs(pathlib.Path(path).parent, exist_ok=True)
    wb.save(path)
    print(f"  Saved -> {path}")


def _write_triangle_sheet(ws, label, data, num_fmt):
    """Write a pivoted triangle DataFrame to a worksheet with 2a-consistent styling."""
    ages = sorted(data.columns)

    # Section title spanning all columns
    title_cell = ws.cell(row=1, column=1, value=label)
    _style_cell(title_cell, "section")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ages) + 1)

    # Column headers (age labels)
    ws.column_dimensions["A"].width = 18
    period_hdr = ws.cell(row=2, column=1, value="Period")
    _style_cell(period_hdr, "subheader")
    for c_idx, age in enumerate(ages, start=2):
        cell = ws.cell(row=2, column=c_idx, value=int(age) if pd.notna(age) else age)
        _style_cell(cell, "subheader")
        ws.column_dimensions[cell.column_letter].width = 14

    ws.freeze_panes = "B3"

    # Data rows
    for r_idx, (period_int, row) in enumerate(data.iterrows(), start=3):
        period_cell = ws.cell(row=r_idx, column=1,
                              value=int(period_int) if pd.notna(period_int) else period_int)
        period_cell.font   = LABEL_FONT
        period_cell.border = THIN_BORDER
        period_cell.alignment = Alignment(horizontal="right")
        if pd.notna(period_int):
            period_cell.number_format = "0"  # Integer format for period
        for c_idx, age in enumerate(ages, start=2):
            val = row.get(age, np.nan)
            _write_data_cell(ws.cell(row=r_idx, column=c_idx),
                             None if pd.isna(val) else val, num_fmt, is_numeric=True)


def write_post_method_triangles(triangles_df, combined, path):
    """
    Write post-method-triangles.xlsx with:
      - X-to-Ultimate development triangles (one sheet per measure present)
      - Average IBNR:   Incurred Ultimate − Incurred at each age
      - Average Unpaid: Incurred Ultimate − Paid at each age

    Sheets are omitted when the underlying data is missing.
    """
    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    df = triangles_df.copy()
    df["period"]     = df["period"].astype(str)
    df["measure"]    = df["measure"].astype(str)
    df["age_int"]    = pd.to_numeric(df["age"].astype(str), errors="coerce")
    df["period_int"] = pd.to_numeric(df["period"], errors="coerce")

    def _pivot(measure):
        sub = df[df["measure"] == measure]
        if sub.empty:
            return None
        return sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first"
        ).astype(float)

    def _period_str(period_int):
        return str(int(period_int)) if pd.notna(period_int) else str(period_int)

    wb = Workbook()
    wb.remove(wb.active)

    # X-to-Ultimate ratio triangles
    for meas, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        pivot = _pivot(meas)
        if pivot is None:
            continue
        result = pivot.copy()
        for period_int in result.index:
            sel = sel_lookup.get((_period_str(period_int), meas), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period_int] = result.loc[period_int] / sel
        ws = wb.create_sheet(title=label)
        _write_triangle_sheet(ws, label, result, "0.0000")

    # Average IBNR = Incurred Ultimate − Incurred at each historical age
    pivot_inc = _pivot("Incurred Loss")
    if pivot_inc is not None:
        avg_ibnr = pivot_inc.copy()
        for period_int in avg_ibnr.index:
            sel = sel_lookup.get((_period_str(period_int), "Incurred Loss"), np.nan)
            if pd.notna(sel):
                avg_ibnr.loc[period_int] = sel - avg_ibnr.loc[period_int]
        ws = wb.create_sheet(title="Average IBNR")
        _write_triangle_sheet(ws, "Average IBNR", avg_ibnr, _NUM_FMT)

    # Average Unpaid = Incurred Ultimate − Paid at each historical age
    pivot_paid = _pivot("Paid Loss")
    if pivot_paid is not None and pivot_inc is not None:
        avg_unpaid = pivot_paid.copy()
        for period_int in avg_unpaid.index:
            sel = sel_lookup.get((_period_str(period_int), "Incurred Loss"), np.nan)
            if pd.notna(sel):
                avg_unpaid.loc[period_int] = sel - avg_unpaid.loc[period_int]
        ws = wb.create_sheet(title="Average Unpaid")
        _write_triangle_sheet(ws, "Average Unpaid", avg_unpaid, _NUM_FMT)

    os.makedirs(pathlib.Path(path).parent, exist_ok=True)
    wb.save(path)
    print(f"  Saved -> {path}")


def write_notes_sheet(ws):
    """
    Write a Notes sheet with workbook overview and table of contents.
    """
    from datetime import datetime
    
    row = 1
    
    # Main title
    title_cell = ws.cell(row=row, column=1, value="Reserve Analysis Workbook")
    _style_cell(title_cell, "header")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 24
    row += 1
    
    # Metadata section
    meta_cell = ws.cell(row=row, column=1, value="Workbook Information")
    _style_cell(meta_cell, "section")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    
    # Creation date
    date_label = ws.cell(row=row, column=1, value="Created:")
    date_label.font = LABEL_FONT
    date_label.border = THIN_BORDER
    date_label.alignment = Alignment(horizontal="left", vertical="center")
    
    date_value = ws.cell(row=row, column=2, value=datetime.now().strftime("%B %d, %Y %I:%M %p"))
    date_value.font = DATA_FONT
    date_value.border = THIN_BORDER
    date_value.alignment = Alignment(horizontal="left", vertical="center")
    row += 1
    
    # Description
    desc_label = ws.cell(row=row, column=1, value="Description:")
    desc_label.font = LABEL_FONT
    desc_label.border = THIN_BORDER
    desc_label.alignment = Alignment(horizontal="left", vertical="center")
    
    desc_value = ws.cell(row=row, column=2, value="Complete actuarial reserve analysis combining selections and projections")
    desc_value.font = DATA_FONT
    desc_value.border = THIN_BORDER
    desc_value.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 2
    
    # Table of contents section
    toc_cell = ws.cell(row=row, column=1, value="Table of Contents")
    _style_cell(toc_cell, "section")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    
    # Headers for TOC
    toc_name_header = ws.cell(row=row, column=1, value="Sheet Name")
    toc_name_header.font = SUBHEADER_FONT
    toc_name_header.fill = SUBHEADER_FILL
    toc_name_header.border = THIN_BORDER
    toc_name_header.alignment = Alignment(horizontal="center", vertical="center")
    
    toc_desc_header = ws.cell(row=row, column=2, value="Description")
    toc_desc_header.font = SUBHEADER_FONT
    toc_desc_header.fill = SUBHEADER_FILL
    toc_desc_header.border = THIN_BORDER
    toc_desc_header.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    # Return the row number where sheet list should start
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.freeze_panes = "A8"
    
    return row


def write_full_analysis(output_path, source_files, internal_files):
    """
    Combine all Excel files into a single master complete-analysis.xlsx.

    source_files  : list of (file_path, sheet_prefix_or_None) from prior scripts
    internal_files: list of file_paths generated by this script (no prefix)

    Files that do not exist on disk are silently skipped.
    """
    master = Workbook()
    master.remove(master.active)
    
    # Create Notes sheet first
    notes_ws = master.create_sheet(title="Notes", index=0)
    toc_start_row = write_notes_sheet(notes_ws)

    all_files = list(source_files) + [(f, None) for f in internal_files]
    
    # Track all sheet names and their descriptions for TOC
    sheet_descriptions = []

    for file_path, prefix in all_files:
        if not os.path.exists(file_path):
            print(f"  Skipping (not found): {file_path}")
            continue
        wb = load_workbook(file_path, data_only=False)  # Keep formatting

        # Build per-workbook rename map so formulas like ='Closed Count'!E56 can be
        # rewritten to ='CL - Closed Count'!E56 during the copy.
        rename_map = {
            s: (f"{prefix}{s}" if prefix else s)[:31] for s in wb.sheetnames
        }

        for sname in wb.sheetnames:
            new_name = rename_map[sname]
            ws_src = wb[sname]
            ws_dst = master.create_sheet(title=new_name)

            # Copy cell values, formats, and styles
            for row in ws_src.iter_rows():
                for cell in row:
                    dst_cell = ws_dst[cell.coordinate]
                    dst_cell.value = rewrite_formula_sheet_refs(cell.value, rename_map)
                    if cell.has_style:
                        dst_cell.font = copy.copy(cell.font)
                        dst_cell.border = copy.copy(cell.border)
                        dst_cell.fill = copy.copy(cell.fill)
                        dst_cell.number_format = cell.number_format
                        dst_cell.protection = copy.copy(cell.protection)
                        dst_cell.alignment = copy.copy(cell.alignment)
            
            # Copy column widths
            for col_letter in ws_src.column_dimensions:
                if col_letter in ws_src.column_dimensions:
                    ws_dst.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width
            
            # Copy row heights
            for row_num in ws_src.row_dimensions:
                if row_num in ws_src.row_dimensions:
                    ws_dst.row_dimensions[row_num].height = ws_src.row_dimensions[row_num].height
            
            # Copy merged cells
            for merged_cell_range in ws_src.merged_cells.ranges:
                ws_dst.merge_cells(str(merged_cell_range))
            
            # Copy freeze panes
            if ws_src.freeze_panes:
                ws_dst.freeze_panes = ws_src.freeze_panes
            
            # Determine description based on sheet name patterns
            desc = _get_sheet_description(new_name, prefix)
            sheet_descriptions.append((new_name, desc))
            
        print(f"  Added sheets from {file_path}")
    
    # Write TOC entries to Notes sheet
    for idx, (sheet_name, desc) in enumerate(sheet_descriptions, start=toc_start_row):
        name_cell = notes_ws.cell(row=idx, column=1, value=sheet_name)
        name_cell.font = DATA_FONT
        name_cell.border = THIN_BORDER
        name_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        desc_cell = notes_ws.cell(row=idx, column=2, value=desc)
        desc_cell.font = DATA_FONT
        desc_cell.border = THIN_BORDER
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


def _get_sheet_description(sheet_name, prefix):
    """Generate a description for a sheet based on its name."""
    # Remove prefix for pattern matching
    base_name = sheet_name
    if prefix:
        base_name = sheet_name[len(prefix):]
    
    # Common patterns
    descriptions = {
        # Chain Ladder sheets
        "Development Age-to-Age": "Historical age-to-age development factors",
        "Simple Averages": "Simple averages of age-to-age factors",
        "Vol-Weighted Avgs": "Volume-weighted averages of age-to-age factors",
        "Medians": "Median age-to-age factors",
        "Selections": "Selected loss development factors",
        
        # Ultimates selection sheets
        "Summary": "Ultimate loss selections summary",
        
        # Measure-specific sheets
        "Incurred Loss": "Incurred loss projections and selections",
        "Paid Loss": "Paid loss projections and selections",
        "Reported Count": "Reported claim count projections",
        "Closed Count": "Closed claim count projections",
        
        # Diagnostic sheets
        "Diagnostics": "Post-method diagnostic calculations",
        "Incurred-to-Ult": "Incurred to ultimate development ratios",
        "Paid-to-Ult": "Paid to ultimate development ratios",
        "Reported-to-Ult": "Reported to ultimate development ratios",
        "Closed-to-Ult": "Closed to ultimate development ratios",
        "Average IBNR": "Average IBNR by development age",
        "Average Unpaid": "Average unpaid by development age",
    }
    
    # Try exact match first
    if base_name in descriptions:
        return descriptions[base_name]
    
    # Try partial matches
    for key, desc in descriptions.items():
        if key in base_name or base_name in key:
            return desc
    
    # Default
    return "Analysis results"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)

    print("Loading data...")
    combined, has_ie, has_bf = load_combined(INPUT_ULTIMATES, INPUT_SELECTIONS_JSON)
    print(f"  has_ie={has_ie}, has_bf={has_bf}, rows={len(combined)}")

    triangles_df = pd.read_parquet(INPUT_TRIANGLES)
    exp_map      = get_exposure(triangles_df)
    print(f"  has_exposure={len(exp_map) > 0} ({len(exp_map)} periods)")

    print("\nWriting selected ultimates...")
    write_selected_ultimates(combined, has_ie, has_bf, OUTPUT_SELECTED_ULTIMATES)

    print("\nWriting post-method series diagnostics...")
    write_post_method_series(combined, exp_map, OUTPUT_POST_SERIES)

    print("\nWriting post-method triangle diagnostics...")
    write_post_method_triangles(triangles_df, combined, OUTPUT_POST_TRIANGLES)

    print("\nWriting complete analysis workbook...")
    internal = [OUTPUT_SELECTED_ULTIMATES, OUTPUT_POST_SERIES, OUTPUT_POST_TRIANGLES]
    write_full_analysis(OUTPUT_COMPLETE_ANALYSIS, ANALYSIS_SOURCE_FILES, internal)

    # Console IBNR summary
    print("\n=== IBNR Summary ===")
    pd.set_option("display.float_format", lambda x: f"{x:,.0f}")
    for m in ["Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"]:
        sub = combined[combined["measure"] == m]
        if sub.empty or sub["selected_ultimate"].isna().all():
            continue
        parts = [
            f"Actual={sub['actual'].sum():,.0f}",
            f"CL={sub['ultimate_cl'].sum():,.0f}",
        ]
        if has_ie:
            parts.append(f"IE={sub['ultimate_ie'].sum():,.0f}")
        if has_bf:
            parts.append(f"BF={sub['ultimate_bf'].sum():,.0f}")
        parts += [
            f"Selected={sub['selected_ultimate'].sum():,.0f}",
            f"IBNR={sub['selected_ibnr'].sum():,.0f}",
        ]
        print(f"  {m}: " + "  ".join(parts))


if __name__ == "__main__":
    print("=== Step 6: Creating complete analysis workbook ===")
    main()
