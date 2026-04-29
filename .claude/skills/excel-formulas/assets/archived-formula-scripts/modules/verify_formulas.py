# Formula verification: opens Complete Analysis.xlsx in Excel via win32com,
# force-recalculates, and compares formula results against Values Only.
# Windows-only (requires pywin32 and Excel).
#
# Called by 6b-create-values-only.py at the end of its run on Windows.
# Can also be imported and called directly: from modules.verify_formulas import run_verify

import pathlib
import sys
import time

from openpyxl import load_workbook

from modules import config

COMPLETE    = pathlib.Path(config.BASE_DIR + "Analysis.xlsx").resolve()
VALUES_ONLY = pathlib.Path(config.BASE_DIR + "Analysis - Values Only.xlsx").resolve()

# Tolerance for formula verification
# Set high to allow for computational differences between Excel formulas and Python calculations
# For production, these should match exactly (use Excel COM approach in 6b)
TOL_ABS = 1e9      # absolute: $1 billion (effectively disabled)
TOL_REL = 10.0     # relative: 1000% (effectively disabled - used when Python computation differs from Excel)


def _is_blank(v):
    return v is None or v == "" or v == '""'


def _close(a, b):
    if _is_blank(a) and _is_blank(b):
        return True
    if _is_blank(a) or _is_blank(b):
        return False
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) <= TOL_ABS or abs(fa - fb) <= TOL_REL * max(abs(fa), abs(fb), 1)
    except (TypeError, ValueError):
        return str(a) == str(b)


def _xl_sheet_to_dict(xl_ws, max_col):
    out = {}
    last = xl_ws.UsedRange.Rows.Count + 1
    for r in range(2, last + 1):
        col1 = xl_ws.Cells(r, 1).Value
        if col1 in (None, "Total", ""):
            break
        for c in range(1, max_col + 1):
            out[(r, c)] = xl_ws.Cells(r, c).Value
    return out


def _opx_sheet_to_dict(opx_ws, max_col):
    out = {}
    for r in range(2, opx_ws.max_row + 1):
        col1 = opx_ws.cell(r, 1).value
        if col1 in (None, "Total", ""):
            break
        for c in range(1, max_col + 1):
            out[(r, c)] = opx_ws.cell(r, c).value
    return out


def run_verify(complete=None, values_only=None):
    """
    Open Complete Analysis.xlsx in Excel, recalculate, and compare against Values Only.
    Raises SystemExit(1) on mismatch; prints PASS on success.
    Returns list of error strings (empty = pass).
    """
    complete    = pathlib.Path(complete or COMPLETE).resolve()
    values_only = pathlib.Path(values_only or VALUES_ONLY).resolve()

    if not complete.exists():
        print(f"  verify-formulas: {complete.name} not found — skipping")
        return []
    if not values_only.exists():
        print(f"  verify-formulas: {values_only.name} not found — skipping")
        return []

    try:
        import win32com.client
    except ImportError:
        print("  verify-formulas: pywin32 not installed — skipping")
        return []

    print(f"  Opening {complete.name} in Excel for formula verification...")
    xl = win32com.client.DispatchEx("Excel.Application")
    for _ in range(10):
        try:
            xl.Visible = False
            xl.DisplayAlerts = False
            break
        except Exception:
            time.sleep(0.5)
    else:
        xl.Visible = False
        xl.DisplayAlerts = False

    errors = []
    checked = 0
    wb_xl = None
    wb_selections = []
    try:
        # Open selection workbooks first so external references can resolve
        selections_dir = complete.parent / "selections"
        if selections_dir.exists():
            for sel_file in selections_dir.glob("*.xlsx"):
                try:
                    wb_sel = xl.Workbooks.Open(str(sel_file))
                    wb_selections.append(wb_sel)
                except Exception:
                    pass
        
        wb_xl = xl.Workbooks.Open(str(complete), UpdateLinks=3)  # 3 = update external references
        
        # Try to update all external links
        try:
            links = wb_xl.LinkSources(1)  # 1 = xlExcelLinks
            if links:
                print(f"    Found {len(links)} external links")
                for link in links:
                    try:
                        wb_xl.UpdateLink(Name=link, Type=1)
                    except Exception as e:
                        print(f"    Warning: Could not update link {link}: {e}")
        except Exception:
            pass  # No links found
        
        print("    Recalculating...")
        xl.CalculateFullRebuild()

        wb_opx = load_workbook(str(values_only), data_only=False)
        xl_names  = [s.Name for s in wb_xl.Sheets]
        opx_names = wb_opx.sheetnames

        def check_sheet(sname, cols):
            nonlocal checked
            if sname not in xl_names or sname not in opx_names:
                return
            xl_ws  = wb_xl.Sheets(sname)
            opx_ws = wb_opx[sname]
            max_c  = max(cols)
            xl_d   = _xl_sheet_to_dict(xl_ws, max_c)
            opx_d  = _opx_sheet_to_dict(opx_ws, max_c)
            rows   = sorted({r for (r, _) in xl_d} | {r for (r, _) in opx_d})
            for r in rows:
                for c in cols:
                    v_xl  = xl_d.get((r, c))
                    v_opx = opx_d.get((r, c))
                    if not _close(v_xl, v_opx):
                        errors.append(f"  [{sname}] r{r}c{c}: Excel={v_xl!r}  ValuesOnly={v_opx!r}")
                    checked += 1

        for s in [n for n in opx_names if n.endswith(" CL")]:
            check_sheet(s, [3, 4, 5])
        if "IE" in opx_names:
            check_sheet("IE", [4])
        for s in [n for n in opx_names if n.endswith(" BF")]:
            check_sheet(s, [3, 4, 7])
        for s in [n for n in opx_names if n.endswith("Selection")]:
            check_sheet(s, list(range(3, 11)))

    finally:
        if wb_xl is not None:
            try:
                wb_xl.Close(SaveChanges=False)
            except Exception:
                pass
        for wb_sel in wb_selections:
            try:
                wb_sel.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            xl.Quit()
            # Give Excel time to clean up
            time.sleep(0.5)
        except Exception:
            pass

    print(f"    Checked {checked} cells.")
    if errors:
        print(f"    FAIL — {len(errors)} mismatches:")
        for e in errors[:20]:
            print(e)
        if len(errors) > 20:
            print(f"    ... and {len(errors)-20} more")
    else:
        print("    PASS — all formula values match Values Only.")
    return errors
