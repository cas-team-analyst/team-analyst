# Reads projected ultimates, actuary selections, and triangle data to produce
# Analysis.xlsx — formulas with cached values using xlsxwriter for instant display.
# No need for separate values-only version - cached formulas work in Excel AND Python.
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6a-create-complete-analysis.py

import copy
import os
import pathlib

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import DATA_FONT, THIN_BORDER, create_xlsxwriter_formats
from modules.xl_selections import (
    SKIP_ROW_LABELS as _SKIP_ROW_LABELS,
    SELECTION_LABELS as _SELECTION_LABELS,
    find_selected_values as _find_selected_values,
    find_selected_reasoning as _find_selected_reasoning,
    copy_ws_filtered as _copy_ws_filtered,
)
from modules.xl_utils import (
    _safe, _to_py, _period_int, _copy_ws,
    measure_short_name,
    ultimates_sheet_for_measure, ultimates_col_header,
)
from modules.xl_writers import _data_cell, _write_title_and_headers, _write_headers
from modules.xl_values import UNPAID_PROXY, _has_method
from modules.xl_notes import _sheet_desc, write_notes_sheet
from modules.analysis_loaders import (
    MEASURE_TO_CATEGORY,
    load_selections,
    load_selection_reasoning,
    load_combined,
    get_exposure,
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def col_letter(col_idx):
    """Convert 0-based column index to Excel column letter (A, B, C, etc.)"""
    result = ''
    while col_idx >= 0:
        result = chr(col_idx % 26 + ord('A')) + result
        col_idx = col_idx // 26 - 1
    return result

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
OUTPUT_PATH            = config.OUTPUT

OUTPUT_COMPLETE = config.BASE_DIR + "Analysis.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS  + "Chain Ladder Selections - LDFs.xlsx"
INPUT_TAIL_EXCEL    = config.SELECTIONS  + "Chain Ladder Selections - Tail.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# Excel external reference paths - using absolute paths for reliability
# (openpyxl doesn't set proper workbook relationships for relative paths to work)
import pathlib
_BASE_ABS = str(pathlib.Path(OUTPUT_COMPLETE).parent.resolve())
_ULT_WB  = _BASE_ABS + "\\selections\\[Ultimates.xlsx]"
_CL_LDF_WB = _BASE_ABS + "\\selections\\[Chain Ladder Selections - LDFs.xlsx]"
_CL_TAIL_WB = _BASE_ABS + "\\selections\\[Chain Ladder Selections - Tail.xlsx]"

def load_tail_selections(tail_excel_path):
    """
    Load tail factor selections from Chain Ladder Selections - Tail.xlsx.
    Priority: User Selection (if populated) > Rules-Based AI Selection.
    Returns {measure: (cutoff_age: int, tail_factor: float)}.  Returns {} if file absent.
    """
    path = pathlib.Path(tail_excel_path)
    if not path.exists():
        print(f"  Note: {tail_excel_path} not found -- no tail selections loaded")
        return {}

    wb = load_workbook(tail_excel_path, data_only=True)
    tail_map = {}

    for measure in wb.sheetnames:
        ws = wb[measure]
        in_tail = False
        cutoff_col = tail_col = reason_col = None
        user_entry = rb_entry = None

        for row in ws.iter_rows():
            col1 = row[0].value if row else None
            if col1 == "Tail Factor Selection":
                in_tail = True
                continue
            if not in_tail:
                continue
            if col1 == "Label":
                for cell in row:
                    if cell.value == "Cutoff Age":
                        cutoff_col = cell.column
                    elif cell.value == "Tail Factor":
                        tail_col = cell.column
                    elif cell.value == "Reasoning":
                        reason_col = cell.column
                continue
            if cutoff_col is None or tail_col is None:
                continue
            if col1 in ("User Selection", "Rules-Based AI Selection"):
                cv = row[cutoff_col - 1].value
                tv = row[tail_col - 1].value
                rv = row[reason_col - 1].value if reason_col else None
                if cv is not None and tv is not None:
                    try:
                        entry = (int(float(str(cv))), float(str(tv)), rv)
                        if col1 == "User Selection":
                            user_entry = entry
                        elif rb_entry is None:
                            rb_entry = entry
                    except (ValueError, TypeError):
                        pass

        chosen = user_entry if user_entry is not None else rb_entry
        if chosen:
            tail_map[measure] = chosen

    wb.close()
    summary = {m: (v[0], v[1]) for m, v in tail_map.items()}
    print(f"  Loaded tail selections: {summary}")
    return tail_map


def load_exposure_row_map(cl_ldf_excel_path):
    """
    Load exposure row map from Chain Ladder Selections - LDFs.xlsx.
    Returns {period: row_number} for building formula references.
    Returns {} if file absent or Exposure sheet not found.
    """
    path = pathlib.Path(cl_ldf_excel_path)
    if not path.exists():
        print(f"  Note: {cl_ldf_excel_path} not found -- no exposure row map loaded")
        return {}
    
    wb = load_workbook(cl_ldf_excel_path, data_only=True)
    if "Exposure" not in wb.sheetnames:
        wb.close()
        print("  Note: Exposure sheet not found in Chain Ladder LDFs workbook")
        return {}
    
    ws = wb["Exposure"]
    row_map = {}
    
    # Find Period column (should be in row 2)
    period_col = None
    for cell in ws[2]:
        if cell.value == "Period":
            period_col = cell.column
            break
    
    if period_col is None:
        wb.close()
        print("  Note: Period column not found in Exposure sheet")
        return {}
    
    # Read period values starting from row 3 (data starts after title and headers)
    for row_idx in range(3, ws.max_row + 1):
        period_val = ws.cell(row=row_idx, column=period_col).value
        if period_val is not None:
            row_map[str(period_val)] = row_idx
    
    wb.close()
    print(f"  Loaded exposure row map for {len(row_map)} periods")
    return row_map


# ── Generated sheet writers ───────────────────────────────────────────────────

def _get_cdf_cell_ref(ws_triangle, target_age):
    # Search row 1 or 2 for the age
    for row_idx in [1, 2]:
        for cell in ws_triangle[row_idx]:
            try:
                if cell.value is not None and int(float(cell.value)) == int(float(target_age)):
                    # Found the column! Now find CDF row (usually 'CDF' in column A)
                    for r in range(1, 100):
                        if str(ws_triangle.cell(row=r, column=1).value).strip() == "CDF":
                            return f"'{ws_triangle.title}'!{get_column_letter(cell.column)}{r}"
            except (ValueError, TypeError):
                pass
    return None

def _formula_cell(ws, r, c, formula, num_fmt, cached_value=None, fmt_obj=None):
    """Write formula with cached value. r and c are 1-based (Excel coords)."""
    if fmt_obj is None:
        fmt_obj = ws.book.add_format({'num_format': num_fmt, 'align': 'right'})
    # CRITICAL: Never pass None to write_formula - corrupts Excel XML!
    # Omit cached_value parameter when None, per excel-formulas skill
    if cached_value is not None:
        ws.write_formula(r-1, c-1, formula, fmt_obj, cached_value)
    else:
        ws.write_formula(r-1, c-1, formula, fmt_obj)


def _data_cell_xlw(ws, r, c, value, num_fmt=None, fmt_obj=None):
    """Write data cell for xlsxwriter. r and c are 1-based (Excel coords)."""
    # Treat string 'None' as None (can happen with missing data)
    if value == 'None' or value == 'none':
        value = None
    
    if value is None or (isinstance(value, float) and np.isnan(value)):
        if fmt_obj:
            ws.write_blank(r-1, c-1, None, fmt_obj)
        return
    
    # Detect if value is numeric (handles both actual numbers and numeric strings)
    is_numeric = isinstance(value, (int, float, np.integer, np.floating))
    
    # Try converting string to number to avoid "number stored as text" warnings
    if not is_numeric and isinstance(value, str):
        try:
            value = float(value)
            is_numeric = True
        except (ValueError, TypeError):
            pass
    
    if fmt_obj is None:
        if is_numeric:
            fmt_obj = ws.book.add_format({'num_format': num_fmt or '#,##0', 'align': 'right'})
        else:
            fmt_obj = ws.book.add_format({'align': 'left'})
    
    if is_numeric:
        ws.write_number(r-1, c-1, float(value), fmt_obj)
    else:
        ws.write(r-1, c-1, str(value), fmt_obj)


def _write_headers_xlw(ws, headers, fmt_dict, col_width=18):
    """Write header row for xlsxwriter. Returns 2 (first data row, 1-based)."""
    for c, text in enumerate(headers, 1):
        ws.write(0, c-1, text, fmt_dict.get('subheader'))
        ws.set_column(c-1, c-1, col_width)
    ws.freeze_panes(1, 0)  # Freeze row 1
    return 2


def write_notes_sheet_xlw(ws, sheet_list, fmt_dict):
    """Write Notes sheet with metadata header and table of contents (xlsxwriter version)."""
    from datetime import datetime
    
    r = 0  # 0-based row for xlsxwriter
    
    # Title
    ws.merge_range(r, 0, r, 1, "Reserve Analysis - Complete Analysis", fmt_dict.get('header'))
    ws.set_row(r, 24)
    r += 1
    
    # Section header
    ws.merge_range(r, 0, r, 1, "Workbook Information", fmt_dict.get('section') or fmt_dict.get('subheader'))
    r += 1
    
    # Metadata
    label_fmt = fmt_dict.get('label')
    for label, value in [
        ("Created:",     datetime.now().strftime("%B %d, %Y %I:%M %p")),
        ("Description:", "Complete actuarial reserve analysis combining selections, ultimates, and diagnostics"),
    ]:
        ws.write(r, 0, label, label_fmt)
        ws.write(r, 1, value, label_fmt)
        r += 1
    
    r += 1  # Blank row
    
    # Table of Contents header
    ws.merge_range(r, 0, r, 1, "Table of Contents", fmt_dict.get('section') or fmt_dict.get('subheader'))
    r += 1
    
    # Column headers
    ws.write(r, 0, "Name", fmt_dict.get('subheader'))
    ws.write(r, 1, "Description", fmt_dict.get('subheader'))
    ws.set_column(0, 0, 28)
    ws.set_column(1, 1, 80)
    r += 1
    
    # Sheet list
    data_fmt = fmt_dict.get('label')
    for name, purpose, _cols in sheet_list:
        ws.write(r, 0, name, data_fmt)
        ws.write(r, 1, purpose, data_fmt)
        r += 1
    
    ws.freeze_panes(7, 0)  # Freeze at row 8 (1-based)


def _ult_ref(ult_col_map, sheet, col_header, row):
    col = ult_col_map.get((sheet, col_header))
    return f"='{_ULT_WB}{sheet}'!{col}{row}" if col else '""'


def _exp_ref(exp_row_map, period):
    """Build formula reference to exposure value in Chain Ladder Selections - LDFs.xlsx."""
    row = exp_row_map.get(str(period))
    if row:
        return f"='{_CL_LDF_WB}Exposure'!B{row}"
    return '""'


def _build_cl_ldfs_col_row_maps(wb_cl, measure_sheet_name):
    """
    Build column and row maps for referencing Chain Ladder Selections - LDFs.xlsx.
    Returns (col_map, user_row, rb_row, user_reason_row, rb_reason_row) where:
      col_map: {column_header: column_letter}
      user_row: row number of "User Selection"
      rb_row: row number of "Rules-Based AI Selection"
      user_reason_row: row number of "User Reasoning"
      rb_reason_row: row number of "Rules-Based AI Reasoning"
    """
    if measure_sheet_name not in wb_cl.sheetnames:
        return {}, None, None, None, None
    
    ws = wb_cl[measure_sheet_name]
    col_map = {}
    user_row = None
    rb_row = None
    user_reason_row = None
    rb_reason_row = None
    
    # Find column headers and selection rows
    for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
        col1 = row_cells[0].value if row_cells else None
        
        # Track header rows (usually None in col A)
        if col1 is None and row_idx <= 10:  # Headers in first few rows
            for cell in row_cells[1:]:
                if cell.value is not None:
                    col_map[str(cell.value)] = get_column_letter(cell.column)
        
        # Track selection rows
        if col1 == "User Selection":
            user_row = row_idx
        elif col1 == "Rules-Based AI Selection":
            rb_row = row_idx
        elif col1 == "User Reasoning":
            user_reason_row = row_idx
        elif col1 == "Rules-Based AI Reasoning":
            rb_reason_row = row_idx
    
    return col_map, user_row, rb_row, user_reason_row, rb_reason_row


def _copy_ws_with_ldfs_formulas(ws_src, ws_dst, measure_short_name, cl_ldfs_path):
    """
    Copy worksheet writing formulas that link to Chain Ladder Selections - LDFs.xlsx.
    Selection rows use IF formula: User Selection if not blank, else Rules-Based AI.
    Data rows reference the source workbook.
    """
    # Load source workbook to build maps
    if not pathlib.Path(cl_ldfs_path).exists():
        # Fallback to value copy if source doesn't exist
        _copy_ws_filtered(ws_src, ws_dst, None, None)
        return
    
    wb_cl = load_workbook(cl_ldfs_path, data_only=False)
    col_map, user_row, rb_row, user_reason_row, rb_reason_row = _build_cl_ldfs_col_row_maps(wb_cl, measure_short_name)
    wb_cl.close()
    
    if not col_map or user_row is None or rb_row is None:
        # Fallback if structure not found
        sel_vals = _find_selected_values(ws_src)
        sel_reason = _find_selected_reasoning(ws_src)
        _copy_ws_filtered(ws_src, ws_dst, sel_vals, sel_reason)
        return
    
    selection_done = False
    dst_row = 1
    # Excel external ref format: '[Workbook.xlsx]SheetName'!CellRef
    ext_ref = f"'{_CL_LDF_WB}{measure_short_name}'"
    
    src_rows = list(ws_src.iter_rows())
    skip_next = False
    
    for row_idx, src_cells in enumerate(src_rows):
        col1 = src_cells[0].value if src_cells else None
        
        # Skip if marked by previous iteration
        if skip_next:
            skip_next = False
            continue
        
        # Skip AI option/reasoning rows
        if col1 in _SKIP_ROW_LABELS:
            continue
        
        # Check if this is a section title row (Age-to-Age Factors, Averages, LDF Selections)
        is_section_title = (col1 in ("Age-to-Age Factors", "Averages", "LDF Selections") and 
                           all(c.value in (None, "") for c in src_cells[1:]))
        
        # If section title and next row has headers, merge them
        if is_section_title and row_idx + 1 < len(src_rows):
            next_row = src_rows[row_idx + 1]
            next_col1 = next_row[0].value if next_row else None
            has_headers = next_col1 in (None, "") and any(c.value not in (None, "") for c in next_row[1:])
            
            if has_headers:
                # Write title in col A and headers in cols 2+ on same row
                for col_idx, src_cell in enumerate(src_cells):
                    dst_cell = ws_dst.cell(dst_row, col_idx + 1)
                    if col_idx == 0:
                        dst_cell.value = col1
                        if src_cell.has_style:
                            dst_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=False)
                            dst_cell.border = copy.copy(src_cell.border)
                            dst_cell.fill = copy.copy(src_cell.fill)
                            dst_cell.alignment = copy.copy(src_cell.alignment)
                    elif col_idx < len(next_row):
                        # Take header from next row
                        next_cell = next_row[col_idx]
                        dst_cell.value = next_cell.value
                        if next_cell.has_style:
                            dst_cell.font = copy.copy(next_cell.font)
                            dst_cell.border = copy.copy(next_cell.border)
                            dst_cell.fill = copy.copy(next_cell.fill)
                            dst_cell.number_format = next_cell.number_format
                            dst_cell.alignment = copy.copy(next_cell.alignment)
                dst_row += 1
                skip_next = True
                continue
        
        # Handle selection rows
        if col1 in _SELECTION_LABELS:
            if selection_done:
                continue
            selection_done = True
            
            # Write "Selected" row with IF formulas
            for src_cell in src_cells:
                dst_cell = ws_dst.cell(dst_row, src_cell.column)
                if src_cell.column == 1:
                    dst_cell.value = "Selected"
                else:
                    # Build IF formula: if User not blank, use User, else use RB-AI
                    col_letter = get_column_letter(src_cell.column)
                    formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_row},{ext_ref}!{col_letter}{rb_row})'
                    dst_cell.value = formula
                
                if src_cell.has_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_row += 1
            
            # Write "Selected Reasoning" row with IF formula
            if user_reason_row and rb_reason_row:
                for src_cell in src_cells:
                    dst_cell = ws_dst.cell(dst_row, src_cell.column)
                    if src_cell.column == 1:
                        dst_cell.value = "Selected Reasoning"
                    else:
                        col_letter = get_column_letter(src_cell.column)
                        formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_reason_row},{ext_ref}!{col_letter}{rb_reason_row})'
                        dst_cell.value = formula
                    
                    if src_cell.has_style:
                        dst_cell.font = copy.copy(src_cell.font)
                        dst_cell.border = copy.copy(src_cell.border)
                        dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.number_format = ""
                    dst_cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
                dst_row += 1
            continue
        
        # Copy all other rows with formulas linking to source
        for src_cell in src_cells:
            dst_cell = ws_dst.cell(dst_row, src_cell.column)
            
            # For data cells (not labels), write formula; for labels, copy value
            if src_cell.column == 1 or src_cell.value is None or isinstance(src_cell.value, str):
                dst_cell.value = src_cell.value
            else:
                # Reference the source workbook
                col_letter = get_column_letter(src_cell.column)
                row_num = src_cells[0].row
                dst_cell.value = f"={ext_ref}!{col_letter}{row_num}"
            
            if src_cell.has_style:
                # Remove bold from column A cells
                if src_cell.column == 1:
                    dst_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=False)
                else:
                    dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy.copy(src_cell.alignment)
        
        dst_row += 1


def _add_cdf_formulas_to_triangle(ws):
    """
    Add CDF row formulas below Selected Reasoning in the LDF Selections section.
    CDF formulas multiply Selected LDF by the next CDF to the right (right-to-left cumulative product).
    Rightmost column references the Selected tail factor.
    """
    selected_row = None
    selected_reasoning_row = None
    cdf_row_num = None
    last_data_col = None
    
    # Find the Selected row and determine where CDF row should be
    for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
        col1 = row_cells[0].value if row_cells else None
        
        if col1 == "Selected":
            selected_row = row_idx
            # Find the last column with data
            for cell in reversed(list(row_cells)):
                if cell.value not in (None, ""):
                    last_data_col = cell.column
                    break
        elif col1 == "Selected Reasoning":
            selected_reasoning_row = row_idx
        elif col1 == "CDF":
            cdf_row_num = row_idx
            break
    
    if not selected_row or last_data_col is None:
        return  # No Selected row found
    
    # If CDF row doesn't exist, create it after Selected Reasoning (or Selected if no reasoning)
    if cdf_row_num is None:
        cdf_row_num = (selected_reasoning_row or selected_row) + 1
        # Insert a new row by shifting cells down if needed
        # Actually, if the row exists but is empty, we can just use it
    
    # Set CDF label
    ws.cell(cdf_row_num, 1).value = "CDF"
    
    # Copy style from Selected row (but remove bold)
    ref_cell = ws.cell(selected_row, 1)
    if ref_cell.has_style:
        lbl = ws.cell(cdf_row_num, 1)
        lbl.font = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=False)
        lbl.border = copy.copy(ref_cell.border)
        lbl.fill = copy.copy(ref_cell.fill)
        lbl.alignment = copy.copy(ref_cell.alignment)
    
    # Write CDF formulas from right to left
    # Rightmost column: reference Selected tail factor
    tail_col_letter = get_column_letter(last_data_col)
    tail_cell = ws.cell(cdf_row_num, last_data_col)
    tail_cell.value = f'={tail_col_letter}{selected_row}'
    tail_cell.number_format = _DEC_FMT
    
    # Other columns: =Selected_LDF * Next_CDF
    for col in range(last_data_col - 1, 1, -1):
        col_letter = get_column_letter(col)
        next_col_letter = get_column_letter(col + 1)
        
        cell = ws.cell(cdf_row_num, col)
        cell.value = f'=IFERROR({col_letter}{selected_row}*{next_col_letter}{cdf_row_num},"")'
        cell.number_format = _DEC_FMT
        if ref_cell.has_style:
            cell.font = copy.copy(ref_cell.font)
            cell.border = copy.copy(ref_cell.border)


def _tri_col_map_from_df(triangles_df, measure, max_age=None):
    """Build {age_str: col_letter} for the triangle sheet from parquet data.
    Triangle sheet col A = period, so ages start at col B (index 2).
    Ages sorted ascending match the order written by _copy_ws_filtered.
    max_age caps the map to the cutoff age — add_tail_to_triangle_ws deletes
    data columns beyond the cutoff, so referencing those cols returns empty.
    """
    tri_m = triangles_df[triangles_df["measure"].astype(str) == measure]
    if tri_m.empty:
        return {}
    ages = sorted(tri_m["age"].dropna().unique(), key=lambda a: float(str(a)))
    if max_age is not None:
        ages = [a for a in ages if float(str(a)) <= max_age]
    return {str(int(float(str(a)))): get_column_letter(i + 2) for i, a in enumerate(ages)}


def write_method_cl(gen_wb, combined, measure, ult_col_map, fmt_dict, tri_col_map=None):
    """Write Chain Ladder method sheet with xlsxwriter (formulas + cached values)."""
    short_name = measure_short_name(measure)
    ws = gen_wb.add_worksheet(f"{short_name} CL"[:31])
    headers = ["Accident Period", "Current Age", short_name, "CDF", "Ultimate", "IBNR", "Unpaid"]
    _write_headers_xlw(ws, headers, fmt_dict)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()
    ult_sheet  = ultimates_sheet_for_measure(measure)
    actual_hdr = ultimates_col_header(measure, "actual")

    # Direct cell refs into the triangle sheet (row 1=Period header, row 2+=data).
    # CL sheet row r maps to triangle row r — both sorted ascending period_int from row 2.
    tri_col_map = tri_col_map or {}
    # Tail column is the last column in the triangle (contains ultimate values for mature periods)
    tail_col_letter = col_letter(len(tri_col_map)) if tri_col_map else None

    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')
    fmt_period = fmt_dict.get('data_period')

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)
        
        # Always reference the triangle sheet for consistency
        actual_val = row["actual"]
        # Sanitize actual_val - could be NaN from dataframe
        actual_val = actual_val if pd.notna(actual_val) else None
        
        if tail_col_letter:
            # Use LOOKUP to find last non-empty value in the triangle row
            # LOOKUP(2, 1/(range<>""), range) finds the last non-empty cell in a row
            lookup_formula = f"=LOOKUP(2,1/('{short_name}'!B{r}:{tail_col_letter}{r}<>\"\"),'{short_name}'!B{r}:{tail_col_letter}{r})"
            _formula_cell(ws, r, 3, lookup_formula, _NUM_FMT, cached_value=actual_val, fmt_obj=fmt_num)
        else:
            # No triangle data available -> use static value
            _data_cell_xlw(ws, r, 3, actual_val, fmt_obj=fmt_num)
        
        cdf_val = row.get("cdf")
        # Sanitize cdf_val - could be NaN from dataframe
        cdf_val = cdf_val if pd.notna(cdf_val) else None
        
        # Use INDEX/MATCH to pull CDF from triangle sheet based on current age
        # Triangle sheet has CDF row at row 72 (after Selected at 68, Selected Reasoning at 69, blank rows at 70-71)
        # INDEX('Triangle'!$B$72:$Y$72, MATCH(CurrentAge, 'Triangle'!$B$1:$Y$1, 0))
        if tail_col_letter:
            cdf_formula = f"=IFERROR(INDEX('{short_name}'!$B$72:${tail_col_letter}$72,MATCH(B{r},'{short_name}'!$B$1:${tail_col_letter}$1,0)),1.000)"
            _formula_cell(ws, r, 4, cdf_formula, _DEC_FMT, cached_value=cdf_val, fmt_obj=fmt_dec)
        else:
            _data_cell_xlw(ws, r, 4, cdf_val, fmt_obj=fmt_dec)
        
        # Ultimate = Actual * CDF (handle 0, None, NaN properly)
        if pd.notna(actual_val) and pd.notna(cdf_val):
            ultimate_val = actual_val * cdf_val
        else:
            ultimate_val = None
        _formula_cell(ws, r, 5, f'=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),C{r}*D{r},"")', _NUM_FMT, cached_value=ultimate_val, fmt_obj=fmt_num)
        
        # IBNR = Ultimate - Actual (handle 0, None, NaN properly)
        if pd.notna(ultimate_val) and pd.notna(actual_val):
            ibnr_val = ultimate_val - actual_val
        else:
            ibnr_val = None
        _formula_cell(ws, r, 6, f"=E{r}-C{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)
        
        # Unpaid: use proxy if available, otherwise same as IBNR
        if proxy_exists:
            # Can't easily cache cross-sheet values here - use ibnr_val as approximation (already checked for NaN above)
            _formula_cell(ws, r, 7, f"=E{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)
        else:
            _formula_cell(ws, r, 7, f"=E{r}-C{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
    
    # Total formulas - cache with sum of actual values (pandas sum can return NaN)
    total_actual = sub["actual"].sum()
    total_actual = total_actual if pd.notna(total_actual) else 0
    total_ultimate = sub.apply(lambda row: row["actual"] * row.get("cdf", 0) if pd.notna(row["actual"]) and pd.notna(row.get("cdf")) else 0, axis=1).sum()
    total_ultimate = total_ultimate if pd.notna(total_ultimate) else 0
    total_ibnr = total_ultimate - total_actual
    
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT, cached_value=total_actual, fmt_obj=fmt_num)
    _formula_cell(ws, t, 5, f"=SUM(E2:E{t-2})", _NUM_FMT, cached_value=total_ultimate, fmt_obj=fmt_num)
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT, cached_value=total_ibnr, fmt_obj=fmt_num)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT, cached_value=total_ibnr, fmt_obj=fmt_num)

def write_method_bf(gen_wb, combined, measure, ult_col_map, fmt_dict, tri_col_maps=None):
    """Write Bornhuetter-Ferguson method sheet with xlsxwriter (formulas + cached values)."""
    tri_col_maps = tri_col_maps or {}
    short_name = measure_short_name(measure)
    ws = gen_wb.add_worksheet(f"{short_name} BF"[:31])
    
    # bf canonical structure: 
    # Accident Period, Current Age, Initial Expected, CDF, % Unreported, Unreported, Actual, Ultimate, IBNR, Unpaid
    headers = ["Accident Period", "Current Age", "Initial Expected", "CDF", "% Unreported", "Unreported", short_name, "Ultimate", "IBNR", "Unpaid"]
    _write_headers_xlw(ws, headers, fmt_dict)
    
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")

    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()
    ult_sheet  = ultimates_sheet_for_measure(measure)
    actual_hdr = ultimates_col_header(measure, "actual")
    ie_hdr     = ultimates_col_header(measure, "ie")
    has_ie_method = _has_method(combined, measure, "ultimate_ie")

    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')
    fmt_pct = fmt_dict.get('data_pct')
    fmt_period = fmt_dict.get('data_period')

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)
        
        # Initial Expected (IE)
        ie_val = row.get("ultimate_ie")
        # Sanitize ie_val - could be NaN from dataframe
        ie_val = ie_val if pd.notna(ie_val) else None
        ie_formula = f"='IE'!D{r}" if has_ie_method else _ult_ref(ult_col_map, ult_sheet, ie_hdr, r)
        _formula_cell(ws, r, 3, ie_formula, _NUM_FMT, cached_value=ie_val, fmt_obj=fmt_num)
        
        # CDF from CL sheet
        cdf_val = row.get("cdf")
        # Sanitize cdf_val - could be NaN from dataframe
        cdf_val = cdf_val if pd.notna(cdf_val) else None
        _formula_cell(ws, r, 4, f"='{short_name} CL'!D{r}", _DEC_FMT, cached_value=cdf_val, fmt_obj=fmt_dec)
        
        # % Unreported = 1 - (1/CDF) (handle 0, None, NaN properly)
        if pd.notna(cdf_val) and cdf_val != 0:
            pct_unrep = 1 - (1 / cdf_val)
        else:
            pct_unrep = None
        _formula_cell(ws, r, 5, f"=1-(1/D{r})", "0.0%", cached_value=pct_unrep, fmt_obj=fmt_pct)
        
        # Unreported = IE * % Unreported (handle 0, None, NaN properly)
        if pd.notna(ie_val) and pd.notna(pct_unrep):
            unrep_val = ie_val * pct_unrep
        else:
            unrep_val = None
        _formula_cell(ws, r, 6, f"=C{r}*E{r}", _NUM_FMT, cached_value=unrep_val, fmt_obj=fmt_num)
        
        # Actual - use LOOKUP to find last value in triangle row
        actual_val = row["actual"]
        # Sanitize actual_val - could be NaN from dataframe
        actual_val = actual_val if pd.notna(actual_val) else None
        tri_col_map = tri_col_maps.get(measure, {})
        tail_col_letter = col_letter(len(tri_col_map)) if tri_col_map else None
        if tail_col_letter:
            lookup_formula = f"=LOOKUP(2,1/('{short_name}'!B{r}:{tail_col_letter}{r}<>\"\"),'{short_name}'!B{r}:{tail_col_letter}{r})"
            _formula_cell(ws, r, 7, lookup_formula, _NUM_FMT, cached_value=actual_val, fmt_obj=fmt_num)
        else:
            _data_cell_xlw(ws, r, 7, actual_val, fmt_obj=fmt_num)
        
        # Ultimate = Unreported + Actual (handle 0, None, NaN properly)
        if pd.notna(unrep_val) and pd.notna(actual_val):
            ultimate_bf = unrep_val + actual_val
        else:
            ultimate_bf = None
        _formula_cell(ws, r, 8, f"=F{r}+G{r}", _NUM_FMT, cached_value=ultimate_bf, fmt_obj=fmt_num)
        
        # IBNR = Ultimate - Actual (handle 0, None, NaN properly)
        if pd.notna(ultimate_bf) and pd.notna(actual_val):
            ibnr_val = ultimate_bf - actual_val
        else:
            ibnr_val = None
        _formula_cell(ws, r, 9, f"=H{r}-G{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)
        
        # Unpaid
        if proxy_exists:
            _formula_cell(ws, r, 10, f"=H{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)
        else:
            _formula_cell(ws, r, 10, f"=H{r}-G{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
    
    # Calculate totals for caching (pandas sum can return NaN)
    total_ie = sub["ultimate_ie"].sum() if "ultimate_ie" in sub.columns else 0
    total_ie = total_ie if pd.notna(total_ie) else 0
    
    # Calculate unreported with proper NaN handling
    def calc_unrep(row):
        ie = row.get("ultimate_ie")
        cdf = row.get("cdf")
        if pd.notna(ie) and pd.notna(cdf) and cdf != 0:
            return ie * (1 - (1 / cdf))
        return 0
    
    total_unrep = sub.apply(calc_unrep, axis=1).sum()
    total_unrep = total_unrep if pd.notna(total_unrep) else 0
    
    total_actual = sub["actual"].sum()
    total_actual = total_actual if pd.notna(total_actual) else 0
    
    total_ultimate_bf = total_unrep + total_actual
    total_ibnr = total_ultimate_bf - total_actual
    total_pct_unrep = (total_unrep / total_ie) if total_ie else 0
    
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT, cached_value=total_ie, fmt_obj=fmt_num)
    # Col 5: % Unreported total = weighted avg = SUM(F)/SUM(C)
    _formula_cell(ws, t, 5, f"=SUM(F2:F{t-2})/SUM(C2:C{t-2})", "0.0%", cached_value=total_pct_unrep, fmt_obj=fmt_pct)
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT, cached_value=total_unrep, fmt_obj=fmt_num)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT, cached_value=total_actual, fmt_obj=fmt_num)
    _formula_cell(ws, t, 8, f"=SUM(H2:H{t-2})", _NUM_FMT, cached_value=total_ultimate_bf, fmt_obj=fmt_num)
    _formula_cell(ws, t, 9, f"=SUM(I2:I{t-2})", _NUM_FMT, cached_value=total_ibnr, fmt_obj=fmt_num)
    _formula_cell(ws, t, 10, f"=SUM(J2:J{t-2})", _NUM_FMT, cached_value=total_ibnr, fmt_obj=fmt_num)


def write_method_ie(gen_wb, combined, measure, exp_row_map, ult_col_map, fmt_dict):
    """Write Initial Expected method sheet with xlsxwriter (formulas + cached values)."""
    short_name = measure_short_name(measure)
    ws = gen_wb.add_worksheet("IE")
    headers = ["Accident Period", "Current Age", "Exposure", "IE Ultimate", "Selected Loss Rate"]
    _write_headers_xlw(ws, headers, fmt_dict)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    ult_sheet = ultimates_sheet_for_measure(measure)
    ie_hdr    = ultimates_col_header(measure, "ie")

    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')
    fmt_period = fmt_dict.get('data_period')

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)

        # Exposure - link to Chain Ladder Selections - LDFs.xlsx Exposure sheet
        # We don't have exposure values in combined, so can't cache easily
        exp_formula = _exp_ref(exp_row_map, row["period"])
        _formula_cell(ws, r, 3, exp_formula, _NUM_FMT, fmt_obj=fmt_num)

        # IE Ultimate
        ie_val = row.get("ultimate_ie")
        # Sanitize ie_val - could be NaN from dataframe
        ie_val = ie_val if pd.notna(ie_val) else None
        ie_ref = _ult_ref(ult_col_map, ult_sheet, ie_hdr, r)
        if ie_ref == '""':
            _data_cell_xlw(ws, r, 4, ie_val, fmt_obj=fmt_num)
        else:
            _formula_cell(ws, r, 4, ie_ref, _NUM_FMT, cached_value=ie_val, fmt_obj=fmt_num)
        
        # Selected Loss Rate = IE Ultimate / Exposure
        # Can't cache without exposure value
        _formula_cell(ws, r, 5, f"=D{r}/C{r}", _DEC_FMT, fmt_obj=fmt_dec)


def write_selection_grouped(gen_wb, combined, measures_group, title, ult_col_map, fmt_dict, tri_col_maps=None):
    """Write selection summary sheet with xlsxwriter (formulas + cached values)."""
    tri_col_maps = tri_col_maps or {}
    ws = gen_wb.add_worksheet(title)
    
    # columns e.g.: Accident Period, Current Age, Incurred, Paid, Incurred CL, Paid CL, Initial Expected, Incurred BF, Paid BF, Selected Ultimate, IBNR, Unpaid
    # Column ordering convention: CL, IE, BF
    
    # Find active measures in this group
    active_m = [
        m for m in measures_group
        if m in combined["measure"].unique()
        and combined[combined["measure"] == m]["actual"].notna().any()
    ]
    if not active_m:
        return
        
    main_m = active_m[0]
    sub = combined[combined["measure"] == main_m].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    short_names = [measure_short_name(m) for m in active_m]
    bf_m  = [m for m in active_m if _has_method(combined, m, "ultimate_bf")]
    has_group_ie = any(_has_method(combined, m, "ultimate_ie") for m in active_m)

    ult_sheet  = ultimates_sheet_for_measure(active_m[0])
    ext_ref    = f"'{_ULT_WB}{ult_sheet}'"
    user_col   = ult_col_map.get((ult_sheet, "User Selection"), "")
    rb_col     = ult_col_map.get((ult_sheet, "Rules-Based AI Selection"), "")
    user_r_col = ult_col_map.get((ult_sheet, "User Reasoning"), "")
    rb_r_col   = ult_col_map.get((ult_sheet, "Rules-Based AI Reasoning"), "")
    ie_measure = next((m for m in active_m if _has_method(combined, m, "ultimate_ie")), active_m[0])
    ie_short   = measure_short_name(ie_measure)

    # Build headers in CL, IE, BF order
    headers = ["Accident Period", "Current Age"]
    for s in short_names:
        headers.append(s)
    for s in short_names:
        headers.append(f"{s} CL")
    if has_group_ie:
        headers.append("Initial Expected")
    for m in bf_m:
        headers.append(f"{measure_short_name(m)} BF")
    headers.extend(["Selected Ultimate", "IBNR", "Unpaid", "Selected Reasoning"])

    _write_headers_xlw(ws, headers, fmt_dict)
    ws.set_column(len(headers)-1, len(headers)-1, 40)  # Reasoning column width

    fmt_num = fmt_dict.get('data_num')
    fmt_text = fmt_dict.get('label')
    fmt_period = fmt_dict.get('data_period')

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)

        col_idx = 3
        # Actuals - use LOOKUP to find last value in triangle row for each measure
        for m in active_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            actual_val = m_row["actual"].iloc[0] if len(m_row) > 0 else None
            # Sanitize - .iloc[0] can return NaN
            actual_val = actual_val if pd.notna(actual_val) else None
            s = measure_short_name(m)
            tri_col_map = tri_col_maps.get(m, {})
            tail_col_letter = col_letter(len(tri_col_map)) if tri_col_map else None
            if tail_col_letter:
                lookup_formula = f"=LOOKUP(2,1/('{s}'!B{r}:{tail_col_letter}{r}<>\"\"),'{s}'!B{r}:{tail_col_letter}{r})"
                _formula_cell(ws, r, col_idx, lookup_formula, _NUM_FMT, cached_value=actual_val, fmt_obj=fmt_num)
            else:
                _data_cell_xlw(ws, r, col_idx, actual_val, fmt_obj=fmt_num)
            col_idx += 1
        
        # CL ultimates
        for m in active_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            cl_ult = m_row["ultimate_cl"].iloc[0] if len(m_row) > 0 and "ultimate_cl" in m_row.columns else None
            # Sanitize - .iloc[0] can return NaN
            cl_ult = cl_ult if pd.notna(cl_ult) else None
            s = measure_short_name(m)
            _formula_cell(ws, r, col_idx, f"='{s} CL'!E{r}", _NUM_FMT, cached_value=cl_ult, fmt_obj=fmt_num)
            col_idx += 1
        
        # IE ultimate (if available)
        if has_group_ie:
            ie_val = row.get("ultimate_ie")
            # Sanitize - could be NaN
            ie_val = ie_val if pd.notna(ie_val) else None
            _formula_cell(ws, r, col_idx, f"='IE'!D{r}", _NUM_FMT, cached_value=ie_val, fmt_obj=fmt_num)
            col_idx += 1
        
        # BF ultimates
        for m in bf_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            bf_ult = m_row["ultimate_bf"].iloc[0] if len(m_row) > 0 and "ultimate_bf" in m_row.columns else None
            # Sanitize - .iloc[0] can return NaN
            bf_ult = bf_ult if pd.notna(bf_ult) else None
            _formula_cell(ws, r, col_idx, f"='{measure_short_name(m)} BF'!H{r}", _NUM_FMT, cached_value=bf_ult, fmt_obj=fmt_num)
            col_idx += 1
        
        # Selected Ultimate → links Ultimates.xlsx (User Selection, falls back to RB-AI)
        sel_formula = (
            f'=IF({ext_ref}!{user_col}{r}<>"",{ext_ref}!{user_col}{r},{ext_ref}!{rb_col}{r})'
            if user_col and rb_col else '""'
        )
        # Cache with selected_ultimate from combined
        sel_val = row.get("selected_ultimate")
        # Sanitize - could be NaN
        sel_val = sel_val if pd.notna(sel_val) else None
        _formula_cell(ws, r, col_idx, sel_formula, _NUM_FMT, cached_value=sel_val, fmt_obj=fmt_num)
        selected_col_idx = col_idx
        col_idx += 1
        
        # IBNR = Selected - First Actual
        first_actual = combined[(combined["measure"] == active_m[0]) & (combined["period"] == row["period"])]["actual"].iloc[0] if len(active_m) > 0 else 0
        # Sanitize - .iloc[0] can return NaN
        first_actual = first_actual if pd.notna(first_actual) else 0
        ibnr_val = (sel_val - first_actual) if pd.notna(sel_val) else None
        _formula_cell(ws, r, col_idx, f"={col_letter(selected_col_idx-1)}{r}-C{r}", _NUM_FMT, cached_value=ibnr_val, fmt_obj=fmt_num)
        col_idx += 1
        
        # Unpaid = Selected - Second Actual (if two measures) or First Actual
        unpaid_actual = "D" if len(active_m) > 1 else "C"
        second_actual = combined[(combined["measure"] == active_m[1]) & (combined["period"] == row["period"])]["actual"].iloc[0] if len(active_m) > 1 else first_actual
        # Sanitize - .iloc[0] can return NaN
        second_actual = second_actual if pd.notna(second_actual) else 0
        unpaid_val = (sel_val - second_actual) if pd.notna(sel_val) else None
        _formula_cell(ws, r, col_idx, f"={col_letter(selected_col_idx-1)}{r}-{unpaid_actual}{r}", _NUM_FMT, cached_value=unpaid_val, fmt_obj=fmt_num)
        col_idx += 1
        
        # Selected Reasoning → links Ultimates.xlsx (User Reasoning, falls back to RB-AI)
        reason_formula = (
            f'=IF({ext_ref}!{user_col}{r}<>"",{ext_ref}!{user_r_col}{r},{ext_ref}!{rb_r_col}{r})'
            if user_col and user_r_col and rb_r_col else '""'
        )
        reason_val = row.get("selected_reasoning", "")
        _formula_cell(ws, r, col_idx, reason_formula, "@", cached_value=reason_val, fmt_obj=fmt_text)

    # Totals row — SUM every numeric column, skip Period and Age
    n_cols = 2 + len(active_m) * 2 + len(bf_m) + (1 if has_group_ie else 0) + 3
    t = 3 + len(sub)
    _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
    
    for c in range(3, n_cols + 1):
        col_lt = col_letter(c - 1)  # Convert to 0-based then to letter
        # Calculate total for caching
        # For now, use a simple sum - could be more sophisticated
        _formula_cell(ws, t, c, f"=SUM({col_lt}2:{col_lt}{t-2})", _NUM_FMT, fmt_obj=fmt_num)


def _build_tail_workbook_maps(tail_wb_path, measure_sheet_name):
    """
    Build row/column maps for Tail Factor Selection in Chain Ladder Selections - Tail.xlsx.
    Returns (cutoff_col, tail_col, reason_col, user_row, rb_row) where columns/rows are
    dynamically discovered from the workbook structure.
    """
    if not pathlib.Path(tail_wb_path).exists():
        return None, None, None, None, None
    
    try:
        wb = load_workbook(tail_wb_path, data_only=False)
        if measure_sheet_name not in wb.sheetnames:
            wb.close()
            return None, None, None, None, None
        
        ws = wb[measure_sheet_name]
        in_tail_section = False
        cutoff_col = tail_col = reason_col = None
        user_row = rb_row = None
        
        for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
            col1 = row_cells[0].value if row_cells else None
            
            if col1 == "Tail Factor Selection":
                in_tail_section = True
                continue
            
            if not in_tail_section:
                continue
            
            # Find column headers
            if col1 == "Label":
                for cell in row_cells:
                    if cell.value == "Cutoff Age":
                        cutoff_col = get_column_letter(cell.column)
                    elif cell.value == "Tail Factor":
                        tail_col = get_column_letter(cell.column)
                    elif cell.value == "Reasoning":
                        reason_col = get_column_letter(cell.column)
                continue
            
            # Find selection rows
            if col1 == "User Selection":
                user_row = row_idx
            elif col1 == "Rules-Based AI Selection":
                rb_row = row_idx
        
        wb.close()
        return cutoff_col, tail_col, reason_col, user_row, rb_row
    except Exception:
        return None, None, None, None, None


def add_tail_to_triangle_ws(ws, cutoff_age, tail_factor, reasoning=None,
                            df2=None, measure=None):
    """
    After _copy_ws_filtered:
      1. Delete all interval columns whose ending age > cutoff_age.
      2. Append a {cutoff_age}-Ult column header to ATA / Averages / LDF sections.
      3. Fill the ATA data rows at the tail column with the observed cumulative
         factor for each period that has development beyond the cutoff age
         (product of per-period ATAs from cutoff_age onward, sourced from df2).
      4. Write a CDF row in LDF Selections with right-to-left cumulative factors.
      5. Write the tail reasoning in the Selected Reasoning row at the tail column.
    """
    tail_label = f"{cutoff_age}-Ult"

    # ── Single pass: locate section header rows and ATA data rows ────────────
    in_ata = in_avg = in_ldf = False
    in_ata_data = False
    ata_hdr_row = avg_hdr_row = ldf_hdr_row = None
    ata_data_rows = {}          # {period_str: row_num} for ATA section
    avg_data_row_nums = []      # row nums of averages data rows (to clear at tail_col)
    ldf_option_row_nums = []    # row nums of LDF section rows that aren't Selected/Reasoning
    last_keep_col = None
    selected_row_num = None
    last_sel_row = None
    sel_vals_by_col = {}

    for row_cells in ws.iter_rows():
        row_num = row_cells[0].row
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Age-to-Age Factors":
            in_ata = True; in_avg = in_ldf = False; in_ata_data = False; continue
        if col1 == "Averages":
            in_avg = True; in_ata = in_ldf = False; in_ata_data = False; continue
        if col1 == "LDF Selections":
            in_ldf = True; in_ata = in_avg = False; in_ata_data = False; continue

        if in_ata:
            if ata_hdr_row is None:
                if col1 is None:
                    ata_hdr_row = row_num
                    in_ata_data = True
            elif in_ata_data:
                if col1 is None:
                    in_ata_data = False
                else:
                    try:
                        ata_data_rows[str(int(float(str(col1))))] = row_num
                    except (ValueError, TypeError):
                        pass
            continue

        if in_avg:
            if avg_hdr_row is None and col1 in ("Metric", None):
                avg_hdr_row = row_num
            elif avg_hdr_row is not None and col1 not in (None, ""):
                avg_data_row_nums.append(row_num)
            continue

        if in_ldf:
            if ldf_hdr_row is None:
                if col1 is None:
                    ldf_hdr_row = row_num
                    for cell in row_cells:
                        v = cell.value
                        if isinstance(v, str) and '-' in v:
                            parts = v.split('-')
                            if len(parts) == 2:
                                try:
                                    if int(parts[1]) <= cutoff_age:
                                        last_keep_col = cell.column
                                except ValueError:
                                    pass
            elif col1 == "Selected":
                selected_row_num = row_num
                last_sel_row = row_num
                sel_vals_by_col = {
                    c.column: float(c.value)
                    for c in row_cells[1:]
                    if c.value is not None
                    and isinstance(c.value, (int, float))
                    and (last_keep_col is None or c.column <= last_keep_col)
                }
            elif col1 == "Selected Reasoning":
                last_sel_row = row_num
            elif col1 is not None and col1 != "":
                ldf_option_row_nums.append(row_num)
            continue

    if last_keep_col is None:
        return

    # ── Column layout ─────────────────────────────────────────────────────────
    # Each ATA column K stores formula (K+1)/K, so the "191-203" interval at
    # last_keep_col references last_keep_col+1 as its numerator (age-203 data).
    # We must KEEP last_keep_col+1 or that formula breaks.
    # tail_col = last_keep_col+1 (the age-203 data column, relabelled to tail).
    # Delete everything from last_keep_col+2 onward.
    tail_col   = last_keep_col + 1
    del_start  = last_keep_col + 2
    del_count  = ws.max_column - last_keep_col - 1
    if del_count > 0:
        ws.delete_cols(del_start, del_count)

    ws.column_dimensions[get_column_letter(tail_col)].width = 14

    # ── Overwrite stale content at tail_col ───────────────────────────────────
    # ATA data rows: formula was (col tail_col+1)/col tail_col → now broken;
    # clear so Excel doesn't show a shifted wrong reference.
    for row_num in ata_data_rows.values():
        ws.cell(row_num, tail_col).value = None

    # Averages data rows: were averages of "203-215" ATAs — blank for tail col.
    for row_num in avg_data_row_nums:
        ws.cell(row_num, tail_col).value = None

    # LDF option rows (Vol Wtd, Simple, etc.): stale "203-215" LDFs — blank.
    for row_num in ldf_option_row_nums:
        ws.cell(row_num, tail_col).value = None

    # ── Replace section headers at tail_col with tail label ──────────────────
    for hdr_row in (ata_hdr_row, avg_hdr_row, ldf_hdr_row):
        if hdr_row is None:
            continue
        ref = ws.cell(hdr_row, last_keep_col)
        new = ws.cell(hdr_row, tail_col)
        new.value = tail_label
        if ref.has_style:
            new.font      = copy.copy(ref.font)
            new.border    = copy.copy(ref.border)
            new.fill      = copy.copy(ref.fill)
            new.number_format = ref.number_format
            new.alignment = copy.copy(ref.alignment)

    if selected_row_num is None:
        return

    # Set Selected row at tail_col to formula referencing Tail workbook
    # IF(User Selection not blank, User Selection, Rules-Based AI Selection)
    c = ws.cell(selected_row_num, tail_col)
    if measure:
        # Dynamically discover Tail workbook structure
        tail_path = config.SELECTIONS + "/Chain Ladder Selections - Tail.xlsx"
        t_cutoff_col, t_tail_col, t_reason_col, t_user_row, t_rb_row = _build_tail_workbook_maps(tail_path, measure)
        
        if t_tail_col and t_user_row and t_rb_row:
            tail_ref = f"'{_CL_TAIL_WB}{measure}'"
            c.value = f'=IF({tail_ref}!{t_tail_col}{t_user_row}<>"",{tail_ref}!{t_tail_col}{t_user_row},{tail_ref}!{t_tail_col}{t_rb_row})'
        else:
            # Fallback to static value if structure not found
            c.value = tail_factor
    else:
        c.value = tail_factor
    c.number_format = _DEC_FMT

    # ── Compute cumulative CDFs right-to-left, seeded by tail_factor ─────────
    cdf_by_col = {tail_col: tail_factor}
    for col in range(last_keep_col, 1, -1):
        ldf = sel_vals_by_col.get(col)
        if ldf is not None and ldf > 0:
            cdf_by_col[col] = ldf * cdf_by_col.get(col + 1, tail_factor)

    # ── Write CDF row after Selected Reasoning ────────────────────────────────
    cdf_row = (last_sel_row or selected_row_num) + 1
    ref_cell = ws.cell(selected_row_num, 1)
    lbl = ws.cell(cdf_row, 1)
    lbl.value = "CDF"
    if ref_cell.has_style:
        lbl.font      = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=False)
        lbl.border    = copy.copy(ref_cell.border)
        lbl.fill      = copy.copy(ref_cell.fill)
        lbl.alignment = copy.copy(ref_cell.alignment)

    for col in range(2, tail_col + 1):
        if cdf_by_col.get(col) is None:
            continue
        c = ws.cell(cdf_row, col)
        if col == tail_col:
            # Reference the Selected row's tail formula
            c.value = f'={get_column_letter(col)}{selected_row_num}'
        else:
            sel_ref  = f"{get_column_letter(col)}{selected_row_num}"
            next_ref = f"{get_column_letter(col + 1)}{cdf_row}"
            c.value  = f'=IFERROR({sel_ref}*{next_ref},"")'
        c.number_format = _DEC_FMT
        c.font          = DATA_FONT
        c.border        = THIN_BORDER

    # ── Write tail reasoning in Selected Reasoning row at tail column ─────────
    if last_sel_row and last_sel_row > selected_row_num:
        rc = ws.cell(last_sel_row, tail_col)
        if measure:
            # Dynamically discover Tail workbook structure
            tail_path = config.SELECTIONS + "/Chain Ladder Selections - Tail.xlsx"
            t_cutoff_col, t_tail_col, t_reason_col, t_user_row, t_rb_row = _build_tail_workbook_maps(tail_path, measure)
            
            if t_tail_col and t_reason_col and t_user_row and t_rb_row:
                tail_ref = f"'{_CL_TAIL_WB}{measure}'"
                # Use tail column to check if User Selection is populated, then pull from reasoning column
                rc.value = f'=IF({tail_ref}!{t_tail_col}{t_user_row}<>"",{tail_ref}!{t_reason_col}{t_user_row},{tail_ref}!{t_reason_col}{t_rb_row})'
            else:
                # Fallback to static value if structure not found
                rc.value = reasoning
        else:
            rc.value = reasoning
        ref_r = ws.cell(last_sel_row, last_keep_col)
        if ref_r.has_style:
            rc.font   = copy.copy(ref_r.font)
            rc.border = copy.copy(ref_r.border)
            rc.fill   = copy.copy(ref_r.fill)
        rc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")


def create_triangle_sheets_xlw(gen_wb, measures, fmt_dict):
    """
    Create triangle sheets with xlsxwriter by reading from Chain Ladder Selections - LDFs.xlsx.
    Writes formulas with cached values for immediate display.
    """
    if not pathlib.Path(INPUT_CL_EXCEL).exists():
        return
    
    # Load source workbooks (formulas and values)
    wb_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
    wb_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)
    
    for measure in measures:
        if measure not in wb_form.sheetnames:
            continue
        
        short_name = measure_short_name(measure)[:31]
        ws_xlw = gen_wb.add_worksheet(short_name)
        ws_src_form = wb_form[measure]
        ws_src_vals = wb_vals[measure]
        
        # Build column/row maps for selection logic
        col_map, user_row, rb_row, user_reason_row, rb_reason_row = _build_cl_ldfs_col_row_maps(wb_form, measure)
        ext_ref = f"'{_CL_LDF_WB}{measure}'"
        
        selection_done = False
        dst_row = 0
        skip_next = False
        selected_row_num = None  # Track the "Selected" row for CDF calculations
        src_rows = list(ws_src_form.iter_rows())
        
        for row_idx, src_cells in enumerate(src_rows):
            if skip_next:
                skip_next = False
                continue
            
            col1 = src_cells[0].value if src_cells else None
            
            # Skip AI option rows
            if col1 in _SKIP_ROW_LABELS:
                continue
            
            # Handle section titles (merge with next row headers if applicable)
            is_section_title = (col1 in ("Age-to-Age Factors", "Averages", "LDF Selections") and 
                               all(c.value in (None, "") for c in src_cells[1:]))
            
            if is_section_title and row_idx + 1 < len(src_rows):
                next_row = src_rows[row_idx + 1]
                next_col1 = next_row[0].value if next_row else None
                has_headers = next_col1 in (None, "") and any(c.value not in (None, "") for c in next_row[1:])
                
                if has_headers:
                    # Write title + headers on same row
                    for col_idx, src_cell in enumerate(src_cells):
                        if col_idx == 0:
                            ws_xlw.write(dst_row, col_idx, col1, fmt_dict.get('subheader'))
                        elif col_idx < len(next_row):
                            next_cell = next_row[col_idx]
                            ws_xlw.write(dst_row, col_idx, next_cell.value, fmt_dict.get('subheader'))
                    dst_row += 1
                    skip_next = True
                    continue
            
            # Handle selection rows
            if col1 in _SELECTION_LABELS:
                if selection_done:
                    continue
                selection_done = True
                selected_row_num = dst_row  # Track for CDF row
                
                # Write "Selected" row with IF formulas
                for src_cell in src_cells:
                    col_idx = src_cell.column - 1  # 0-based
                    if col_idx == 0:
                        ws_xlw.write(dst_row, col_idx, "Selected", fmt_dict.get('label'))
                    else:
                        # IF formula: User not blank → User, else RB-AI
                        col_letter = get_column_letter(src_cell.column)
                        formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_row},{ext_ref}!{col_letter}{rb_row})'
                        # Get cached value from values workbook
                        val_cell = ws_src_vals.cell(src_cell.row, src_cell.column)
                        cached_val = val_cell.value if val_cell.value not in (None, "") else None
                        if cached_val is not None:
                            ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('data_ldf'), cached_val)
                        else:
                            ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('data_ldf'))
                dst_row += 1
                
                # Write "Selected Reasoning" row with IF formula
                if user_reason_row and rb_reason_row:
                    for src_cell in src_cells:
                        col_idx = src_cell.column - 1
                        if col_idx == 0:
                            ws_xlw.write(dst_row, col_idx, "Selected Reasoning", fmt_dict.get('label'))
                        else:
                            col_letter = get_column_letter(src_cell.column)
                            formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_reason_row},{ext_ref}!{col_letter}{rb_reason_row})'
                            val_cell = ws_src_vals.cell(src_cell.row, src_cell.column)
                            cached_val = val_cell.value if val_cell.value not in (None, "") else None
                            if cached_val is not None:
                                ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('label'), cached_val)
                            else:
                                ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('label'))
                    dst_row += 1
                continue
            
            # Copy all other rows (data rows, labels, headers)
            for src_cell in src_cells:
                col_idx = src_cell.column - 1
                val_cell = ws_src_vals.cell(src_cell.row, src_cell.column)
                
                # Labels (column A) or text: copy value directly
                if col_idx == 0 or src_cell.value is None or isinstance(src_cell.value, str):
                    fmt = fmt_dict.get('label') if col_idx == 0 else None
                    ws_xlw.write(dst_row, col_idx, src_cell.value, fmt)
                else:
                    # Data cells: write formula with cached value
                    col_letter = get_column_letter(src_cell.column)
                    row_num = src_cell.row
                    formula = f"={ext_ref}!{col_letter}{row_num}"
                    cached_val = val_cell.value if val_cell.value not in (None, "") else None
                    if cached_val is not None:
                        ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('data_ldf'), cached_val)
                    else:
                        ws_xlw.write_formula(dst_row, col_idx, formula, fmt_dict.get('data_ldf'))
            
            dst_row += 1
        
        # Add CDF row after all other rows
        # CDF row calculates cumulative development factors working backwards from tail
        cdf_row = dst_row
        ws_xlw.write(cdf_row, 0, "CDF", fmt_dict.get('label'))
        
        # Find the last data column (last age in triangle)
        last_col = 0
        for c in range(1, 50):  # Check up to column 50
            header = ws_src_form.cell(1, c + 1).value  # +1 because openpyxl is 1-based
            if header is None or header == "":
                last_col = c - 1
                break
            last_col = c
        
        if last_col > 0:
            # Get tail reference for this measure
            tail_ref = f"'{_CL_TAIL_WB}{measure}'"
            
            # For the rightmost column (tail), reference the tail workbook
            # The tail workbook has the selected tail factor that should be used
            # We'll use column L (12th column) as a default, but ideally should discover dynamically
            # For now, use 1.000 as default tail if no tail workbook
            tail_col_letter = get_column_letter(last_col + 1)  # openpyxl is 1-based
            tail_formula = f'=IFERROR({tail_ref}!$X$2,1.000)'  # X2 is typically where tail factor is
            ws_xlw.write_formula(cdf_row, last_col, tail_formula, fmt_dict.get('data_ldf'), 1.000)
            
            if selected_row_num is not None:
                # Work backwards from last_col-1 to column 1 (0-based indexing in xlsxwriter)
                for col_idx in range(last_col - 1, 0, -1):
                    ldf_ref = f"{get_column_letter(col_idx + 1)}{selected_row_num + 1}"  # Reference Selected row LDF
                    next_cdf_ref = f"{get_column_letter(col_idx + 2)}{cdf_row + 1}"  # Reference next CDF
                    cdf_formula = f"=IFERROR({ldf_ref}*{next_cdf_ref},1.000)"
                    
                    # Get cached value from calculations (approximate)
                    # We'll use 1.000 as default since we don't have actual values yet
                    ws_xlw.write_formula(cdf_row, col_idx, cdf_formula, fmt_dict.get('data_ldf'))
        
        # Set column widths (standard for triangles)
        ws_xlw.set_column(0, 0, 25)  # Label column
        ws_xlw.set_column(1, 20, 12)  # Data columns
    
    wb_form.close()
    wb_vals.close()


def create_triangle_sheets(gen_wb, measures):
    # Copy the triangle sheets from the CL selections workbook with formulas
    tail_sel = load_tail_selections(INPUT_TAIL_EXCEL)
    df2_enh = (pd.read_parquet(INPUT_CL_ENHANCED)
               if pathlib.Path(INPUT_CL_ENHANCED).exists()
               else pd.DataFrame())
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        wb_cl_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
        for measure in measures:
            if measure in wb_cl_vals.sheetnames:
                short_name = measure_short_name(measure)[:31]
                ws = gen_wb.create_sheet(title=short_name)
                # Use formula-based copy to link to Chain Ladder Selections - LDFs.xlsx
                _copy_ws_with_ldfs_formulas(wb_cl_form[measure], ws, measure, INPUT_CL_EXCEL)
                # Add CDF formulas below Selected row
                _add_cdf_formulas_to_triangle(ws)
                if measure in tail_sel:
                    cutoff_age, tf, reasoning = tail_sel[measure]
                    add_tail_to_triangle_ws(
                        ws, cutoff_age, tf, reasoning,
                        df2=df2_enh, measure=measure,
                    )



def write_exposure_sheet(gen_wb, triangles_path, fmt_dict):
    """
    Write Exposure sheet from triangles parquet with xlsxwriter.
    Single age per period -> two-column table (Period | Exposure).
    Multiple ages        -> full triangle (Period | age1 | age2 | ...).
    """
    if not pathlib.Path(triangles_path).exists():
        return
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return

    exp["period_int"] = exp["period"].apply(lambda x: _period_int(str(x)))
    exp["age_num"]    = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    exp = exp.sort_values(["period_int", "age_num"])

    ages = sorted(exp["age_num"].dropna().unique())
    ws   = gen_wb.add_worksheet("Exposure")

    fmt_num = fmt_dict.get('data_num')
    fmt_label = fmt_dict.get('label')

    if len(ages) <= 1:
        headers  = ["Period", "Exposure"]
        # Write title row
        ws.write(0, 0, "Exposure", fmt_dict.get('header'))
        ws.merge_range(0, 0, 0, len(headers)-1, "Exposure", fmt_dict.get('header'))
        # Write headers
        for c, hdr in enumerate(headers):
            ws.write(1, c, hdr, fmt_dict.get('subheader'))
            ws.set_column(c, c, 18)
        ws.freeze_panes(2, 0)
        
        data_row = 2
        for r, (_, row) in enumerate(exp.iterrows(), start=data_row):
            _data_cell_xlw(ws, r+1, 1, row["period_int"], fmt_obj=fmt_num)  # r+1 for 1-based
            _data_cell_xlw(ws, r+1, 2, _safe(row["value"]), fmt_obj=fmt_num)
    else:
        pivot = (
            exp.pivot_table(
                index="period_int", columns="age_num",
                values="value", aggfunc="first", observed=True,
            )
            .sort_index()
        )
        pivot.columns = [int(c) for c in pivot.columns]
        headers  = ["Period"] + [str(c) for c in pivot.columns]
        
        # Write title row
        ws.write(0, 0, "Exposure", fmt_dict.get('header'))
        ws.merge_range(0, 0, 0, len(headers)-1, "Exposure", fmt_dict.get('header'))
        # Write headers
        for c, hdr in enumerate(headers):
            ws.write(1, c, hdr, fmt_dict.get('subheader'))
            ws.set_column(c, c, 14)
        ws.freeze_panes(2, 0)
        
        data_row = 2
        for r, (idx, row) in enumerate(pivot.iterrows(), start=data_row):
            _data_cell_xlw(ws, r+1, 1, idx, fmt_obj=fmt_num)  # r+1 for 1-based
            for c, val in enumerate(row, start=2):
                _data_cell_xlw(ws, r+1, c, _safe(val), fmt_obj=fmt_num)


def write_diagnostics_sheet(ws, combined, exp_map, fmt_dict):
    """
    Write the 'Diagnostics' sheet with xlsxwriter: Ultimate Severity, Loss Rate, Frequency.
    Loss Rate and Frequency columns are omitted when no exposure data is present.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    if inc.empty:
        return

    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    # Write title row
    ws.write(0, 0, "Post-Method Series Diagnostics", fmt_dict.get('header'))
    ws.merge_range(0, 0, 0, len(headers)-1, "Post-Method Series Diagnostics", fmt_dict.get('header'))
    # Write headers
    for c, hdr in enumerate(headers):
        ws.write(1, c, hdr, fmt_dict.get('subheader'))
        ws.set_column(c, c, 18)
    ws.freeze_panes(2, 0)
    
    data_row = 2
    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else str(x)),
    )
    for r, p in enumerate(periods, start=data_row):
        ult_loss   = inc.loc[p, "selected_ultimate"] if p in inc.index else np.nan
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        def _div(a, b):
            if pd.notna(a) and pd.notna(b) and b != 0:
                return a / b
            return None

        _data_cell_xlw(ws, r+1, 1, _period_int(p), fmt_obj=fmt_num)  # r+1 for 1-based
        _data_cell_xlw(ws, r+1, 2, _div(ult_loss, ult_counts), fmt_obj=fmt_dec)
        if has_exposure:
            _data_cell_xlw(ws, r+1, 3, _div(ult_loss, exp), fmt_obj=fmt_dec)
            _data_cell_xlw(ws, r+1, 4, _div(ult_counts, exp), fmt_obj=fmt_dec)


def build_post_method_triangle_data(triangles_df, combined):
    """
    Build post-method triangle DataFrames as a list of (sheet_name, index_name, df).
    Each df has period ints as index and age ints as columns.
    """
    if triangles_df.empty:
        return []

    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    def _to_int_series(s):
        try:
            return s.apply(lambda x: int(float(str(x))))
        except (ValueError, TypeError):
            return s.astype(str)

    def x_to_ult(measure, label):
        sub = triangles_df[triangles_df["measure"] == measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        return label, result

    sheets = []
    for measure, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        result = x_to_ult(measure, label)
        if result is not None:
            sheets.append(result)

    def avg_triangle(sub_measure, sheet_name, diff_fn):
        sub = triangles_df[triangles_df["measure"] == sub_measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), sub_measure), np.nan)
            if pd.notna(sel):
                result.loc[period] = diff_fn(sel, result.loc[period])
        return sheet_name, result

    inc_sub  = triangles_df[triangles_df["measure"] == "Incurred Loss"]
    paid_sub = triangles_df[triangles_df["measure"] == "Paid Loss"]

    if not inc_sub.empty:
        r = avg_triangle("Incurred Loss", "Average IBNR", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    if not paid_sub.empty:
        r = avg_triangle("Paid Loss", "Average Unpaid", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    return sheets


def write_triangle_sheet(ws, sheet_name, df):
    """
    Write a triangle DataFrame to a worksheet.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    headers = ["Period"] + [str(c) for c in df.columns]
    data_row = _write_title_and_headers(ws, sheet_name, headers, col_width=14)

    for r, (idx, row) in enumerate(df.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), idx)
        for c, val in enumerate(row, start=2):
            _data_cell(ws.cell(r, c), _safe(val), _DEC_FMT)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def assemble_workbook(output_path, generated_wb):
    """
    Build master workbook from the pre-built generated sheets.
    All content is already computed values — no formula copying needed.
    """
    master = Workbook()
    master.remove(master.active)
    sheet_list = []

    for sname in generated_wb.sheetnames:
        ws_dst = master.create_sheet(title=sname)
        _copy_ws(generated_wb[sname], ws_dst)
        sheet_list.append((sname, *_sheet_desc(sname)))

    notes_ws = master.create_sheet(title="Notes", index=0)
    write_notes_sheet(notes_ws, sheet_list)

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    sel_lookup = load_selections(INPUT_SELECTIONS_EXCEL)

    ult_wb = load_workbook(INPUT_SELECTIONS_EXCEL, data_only=True)
    ult_col_map = {}
    for _sname in ult_wb.sheetnames:
        for _cell in ult_wb[_sname][1]:
            if _cell.value:
                ult_col_map[(_sname, str(_cell.value).strip())] = col_letter(_cell.column - 1)  # Convert to xlsxwriter 0-based
    ult_wb.close()

    combined, available_methods = load_combined(INPUT_ULTIMATES, sel_lookup)
    reason_lookup = load_selection_reasoning(INPUT_SELECTIONS_EXCEL)
    combined["selected_reasoning"] = combined.apply(
        lambda row: reason_lookup.get(
            (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]),
            ""
        ) or "",
        axis=1
    )
    exp_map = get_exposure(INPUT_TRIANGLES)
    has_exposure = bool(exp_map)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]
    
    # Load exposure row map for linking to Chain Ladder LDFs workbook
    exp_row_map = load_exposure_row_map(INPUT_CL_EXCEL)
    
    # We also need triangles_df for diagnostics
    triangles_df = pd.read_parquet(INPUT_TRIANGLES)
    # Load tail selections early to know which ages survive add_tail_to_triangle_ws.
    # _tri_col_map_from_df caps at cutoff so CL actual refs don't point to deleted columns.
    tail_cutoff = {}
    if pathlib.Path(INPUT_TAIL_EXCEL).exists():
        _ts = load_tail_selections(INPUT_TAIL_EXCEL)
        tail_cutoff = {m: info[0] for m, info in _ts.items()}

    print(f"Measures: {measures}")
    print(f"Methods: {available_methods}")
    
    # Create xlsxwriter workbook with cached formula support
    wb = xlsxwriter.Workbook(OUTPUT_COMPLETE, {'use_future_functions': True})
    fmt = create_xlsxwriter_formats(wb)
    
    loss_m  = [m for m in measures if "Loss"  in m]
    count_m = [m for m in measures if "Count" in m]
    other_m = [m for m in measures if m not in loss_m and m not in count_m]
    has_ie  = any(col == "ultimate_ie" for col, _ in available_methods)

    print("Building Loss sheets...")
    if loss_m:
        # Build tri_col_maps dictionary for all loss measures
        loss_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in loss_m}
        
        write_selection_grouped(wb, combined, ["Incurred Loss", "Paid Loss"], "Loss Selection", ult_col_map, fmt, loss_tri_maps)
        for m in loss_m:
            write_method_cl(wb, combined, m, ult_col_map, fmt, loss_tri_maps[m])
        for m in loss_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, loss_tri_maps)
        # Create single IE sheet for Incurred Loss only
        if "Incurred Loss" in loss_m and _has_method(combined, "Incurred Loss", "ultimate_ie"):
            write_method_ie(wb, combined, "Incurred Loss", exp_row_map, ult_col_map, fmt)

    print("Building Counts sheets...")
    if count_m:
        # Build tri_col_maps dictionary for all count measures
        count_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in count_m}
        
        write_selection_grouped(wb, combined, ["Reported Count", "Closed Count"], "Count Selection", ult_col_map, fmt, count_tri_maps)
        for m in count_m:
            write_method_cl(wb, combined, m, ult_col_map, fmt, count_tri_maps[m])
        for m in count_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, count_tri_maps)

    if other_m:
        print("Building other method sheets...")
        # Build tri_col_maps dictionary for other measures
        other_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in other_m}
        
        for m in other_m:
            write_method_cl(wb, combined, m, ult_col_map, fmt, other_tri_maps[m])
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, other_tri_maps)
    
    print("Copying CL LDF triangle sheets...")
    create_triangle_sheets_xlw(wb, measures, fmt)
    
    if has_exposure:
        write_exposure_sheet(wb, INPUT_TRIANGLES, fmt)
    
    # TODO: These copying functions still use openpyxl - convert in separate step
    # print("Copying remaining Diag and CV & Slopes sheets...")
    # if pathlib.Path(INPUT_CL_EXCEL).exists():
    #     wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
    #     for sname in wb_cl_form.sheetnames:
    #         if sname in ("Diagnostics", "CV & Slopes"):
    #             ws = wb.add_worksheet(sname[:31])
    #             _copy_ws_filtered(wb_cl_form[sname], ws, {}, {})
    #     wb_cl_form.close()

    # TODO: Triangle sheet writing - convert in separate step
    # print("Building post-method triangle sheets...")
    # for sheet_name, df in build_post_method_triangle_data(triangles_df, combined):
    #     ws_t = wb.add_worksheet(sheet_name[:31])
    #     write_triangle_sheet(ws_t, sheet_name, df)

    print("Building diagnostics sheet...")
    ws_diag = wb.add_worksheet("Summary Diagnostics")
    write_diagnostics_sheet(ws_diag, combined, exp_map, fmt)

    # Build Notes sheet with table of contents
    print("Adding Notes sheet...")
    # Get list of all sheets created (xlsxwriter worksheets)
    sheet_list = [(ws.name, *_sheet_desc(ws.name)) for ws in wb.worksheets()]
    ws_notes = wb.add_worksheet("Notes")
    write_notes_sheet_xlw(ws_notes, sheet_list, fmt)

    print("\nSaving Analysis.xlsx with xlsxwriter (formulas + cached values)...")
    wb.close()
    print(f"  Saved -> {OUTPUT_COMPLETE}")
    print("Done. Formulas display immediately — no need for separate values-only workbook.")

if __name__ == "__main__":
    main()

