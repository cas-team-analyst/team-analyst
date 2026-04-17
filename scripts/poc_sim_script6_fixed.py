"""
POC: Simulate 6-create-complete-analysis.py's sheet-copy loop with the
cross-sheet-ref rewriting patch applied.

Proves: after the CL-prefix rename, cross-sheet formulas like
='Closed Count'!E56 get rewritten to ='CL - Closed Count'!E56, preserving
the live recalc chain through the aggregated workbook.

Output: demo/demo4-complete/poc-output/sim-script6-master-fixed.xlsx
"""

import copy
import pathlib
import re
from openpyxl import Workbook, load_workbook

DEMO_ROOT = pathlib.Path(__file__).parent.parent / "demo" / "demo4-complete"
TEMPLATE_PATH = DEMO_ROOT / "poc-output" / "templates" / "cl-selections-template.xlsx"
OUTPUT_PATH = DEMO_ROOT / "poc-output" / "sim-script6-master-fixed.xlsx"

PREFIX = "CL - "

# Regex for sheet refs in formulas:
#   'Sheet Name'!   — quoted (required when name has space/hyphen/etc)
#   SheetName!      — unquoted (alphanumeric + underscore only)
# Quoted names may contain escaped quotes (''), but we ignore that edge case here.
_QUOTED_SHEET_REF = re.compile(r"'([^'!]+)'!")
_UNQUOTED_SHEET_REF = re.compile(r"(?<![A-Za-z0-9_'\"\]])([A-Za-z_][A-Za-z0-9_]*)!")


def rewrite_formula_sheet_refs(formula, rename_map):
    """
    Rewrite sheet refs in a formula string per rename_map {old_name: new_name}.
    Quoted refs become 'new name'!; unquoted refs become 'new name'! if new name
    contains special chars, else new_name! directly.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _quote_if_needed(name):
        # Excel requires quoting when sheet name contains anything other than
        # letters, digits, or underscores.
        if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
            return name
        return f"'{name}'"

    def _quoted_sub(m):
        old = m.group(1)
        new = rename_map.get(old, old)
        return f"{_quote_if_needed(new)}!"

    def _unquoted_sub(m):
        old = m.group(1)
        new = rename_map.get(old, old)
        return f"{_quote_if_needed(new)}!"

    result = _QUOTED_SHEET_REF.sub(_quoted_sub, formula)
    result = _UNQUOTED_SHEET_REF.sub(_unquoted_sub, result)
    return result


def copy_with_rewrites(src_wb, master_wb, prefix):
    # Build rename map for this source workbook
    rename_map = {}
    for sname in src_wb.sheetnames:
        new_name = f"{prefix}{sname}"[:31]
        rename_map[sname] = new_name

    for sname, new_name in rename_map.items():
        ws_src = src_wb[sname]
        ws_dst = master_wb.create_sheet(title=new_name)
        for row in ws_src.iter_rows():
            for cell in row:
                dst = ws_dst[cell.coordinate]
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    val = rewrite_formula_sheet_refs(val, rename_map)
                dst.value = val
                if cell.has_style:
                    dst.font = copy.copy(cell.font)
                    dst.number_format = cell.number_format
                    dst.border = copy.copy(cell.border)
                    dst.fill = copy.copy(cell.fill)
                    dst.alignment = copy.copy(cell.alignment)
        for merged_range in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged_range))
        if ws_src.freeze_panes:
            ws_dst.freeze_panes = ws_src.freeze_panes


def main():
    src = load_workbook(TEMPLATE_PATH, data_only=False)
    master = Workbook()
    master.remove(master.active)

    copy_with_rewrites(src, master, PREFIX)

    master.save(OUTPUT_PATH)
    print(f"Saved -> {OUTPUT_PATH}")

    # Spot check: show a few rewritten formulas from Selected Ults sheets
    print("\nSpot-check of rewritten cross-sheet formulas:")
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for sname in wb.sheetnames:
        if "Selected Ults" in sname:
            ws = wb[sname]
            sample = ws["C3"].value
            print(f"  {sname!r} C3: {sample}")


if __name__ == "__main__":
    main()
