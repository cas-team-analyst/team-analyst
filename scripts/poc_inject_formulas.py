"""
POC: One-time script to inject Excel formulas into the Closed Count sheet
of cl-selections-template.xlsx.

Produces a template where:
  - Triangle (rows 3-12) stays hard-coded input data
  - Age-to-age factors (rows 16-24) become =next_age/current_age formulas
  - Averages (rows 29-40) become AVERAGE/SUMPRODUCT/TRIMMEAN formulas
  - Selection row 47 stays hard-coded (actuary input)
  - NEW rows 50-52: Cumulative Development Factors (CDFs) as backward product of selections
  - NEW rows 54-65: Projected Ultimates per period = latest-diagonal * CDF-at-age

Run from the scripts/ directory. Edits the template file in place.
"""

import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

TEMPLATE_PATH = pathlib.Path(__file__).parent.parent / "demo" / "demo4-complete" / "poc-output" / "templates" / "cl-selections-template.xlsx"

SHEET = "Closed Count"

# Triangle rows: 3..12 (one per period 2015..2024)
TRI_ROW_START, TRI_ROW_END = 3, 12
# Age-to-age rows: 16..24 (9 periods have factors; 2024 at row 25 has none)
ATA_ROW_START, ATA_ROW_END = 16, 24
# Interval cols in age-to-age: B..J (9 intervals)
ATA_COL_START, ATA_COL_END = 2, 10  # B..J
# Triangle cols: B..K (B=age 12, K=age 120)
TRI_COL_START, TRI_COL_END = 2, 11  # B..K

# Averages rows
ROW_SIMPLE_ALL = 29
ROW_WEIGHTED_ALL = 30
ROW_EXCL_ALL = 31
ROW_SIMPLE_3YR = 32
ROW_WEIGHTED_3YR = 33
ROW_EXCL_3YR = 34
ROW_SIMPLE_5YR = 35
ROW_WEIGHTED_5YR = 36
ROW_EXCL_5YR = 37
ROW_SIMPLE_10YR = 38
ROW_WEIGHTED_10YR = 39
ROW_EXCL_10YR = 40

# Selection row (actuary input)
ROW_SELECTION = 47

# NEW sections
ROW_CDF_TITLE = 50
ROW_CDF_AGES = 51
ROW_CDF_VALUES = 52
ROW_ULT_TITLE = 54
ROW_ULT_HEADER = 55
ROW_ULT_START = 56  # one row per period

# Window row spans (age-to-age range B..J)
ATA_ALL_RANGE = (16, 24)
ATA_3YR_RANGE = (22, 24)
ATA_5YR_RANGE = (20, 24)
ATA_10YR_RANGE = (16, 24)  # Only 9 periods available; effectively = all

# Triangle weight ranges (one row less than age-to-age window — weights are prior-age triangle values)
TRI_ALL_RANGE = (3, 11)   # 2015-2023
TRI_3YR_RANGE = (9, 11)   # 2021-2023
TRI_5YR_RANGE = (7, 11)   # 2019-2023
TRI_10YR_RANGE = (3, 11)  # same as all (only 9 periods)


def col_letter(idx):
    """1-based column index → letter."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def atoa_formula(r, c):
    """Age-to-age: factor(period r, interval c) = triangle[period, c+1] / triangle[period, c]"""
    tri_row = r - 13  # row 16 → 3, row 24 → 11
    num_col = col_letter(c + 1)
    den_col = col_letter(c)
    return f"={num_col}{tri_row}/{den_col}{tri_row}"


def simple_avg_formula(col, row_range):
    a, b = row_range
    return f"=IFERROR(AVERAGE({col}{a}:{col}{b}),AVERAGE({col}16:{col}24))"


def weighted_avg_formula(col, ata_range, tri_range):
    ata_a, ata_b = ata_range
    tri_a, tri_b = tri_range
    ata_ref = f"{col}{ata_a}:{col}{ata_b}"
    tri_ref = f"{col}{tri_a}:{col}{tri_b}"
    all_ata = f"{col}16:{col}24"
    all_tri = f"{col}3:{col}11"
    primary = f"SUMPRODUCT({ata_ref},{tri_ref})/SUMPRODUCT(({ata_ref}<>\"\")*{tri_ref})"
    fallback = f"SUMPRODUCT({all_ata},{all_tri})/SUMPRODUCT(({all_ata}<>\"\")*{all_tri})"
    return f"=IFERROR({primary},{fallback})"


def excl_hl_formula(col, row_range):
    a, b = row_range
    rng = f"{col}{a}:{col}{b}"
    all_rng = f"{col}16:{col}24"
    primary = f"TRIMMEAN({rng},2/COUNT({rng}))"
    fallback = f"AVERAGE({all_rng})"
    return f"=IFERROR({primary},{fallback})"


def inject():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET]

    # ── Age-to-age factors (rows 16..24, cols B..J) ──
    # Only write formula where triangle has both numerator and denominator.
    # Rule: row r (age-to-age) corresponds to triangle row r-13. For 2015 (ata row 16 → tri row 3)
    # all factors exist. For 2016 (ata row 17 → tri row 4), factors exist for cols B..I (missing K).
    # Diagonal pattern: tri row r has values in cols B through K-(r-3), i.e. up to col_idx = 11 - (tri_row - 3) = 14 - tri_row.
    # Age-to-age factor for col c exists when both triangle col c AND col c+1 have values.
    # Triangle col c exists when c <= 14 - tri_row, i.e. c <= 14 - (r-13) = 27 - r.
    # So both c and c+1 exist when c+1 <= 27 - r, i.e. c <= 26 - r.
    for r in range(ATA_ROW_START, ATA_ROW_END + 1):
        max_c = 26 - r
        for c in range(ATA_COL_START, ATA_COL_END + 1):
            if c <= max_c:
                ws.cell(row=r, column=c).value = atoa_formula(r, c)
            else:
                ws.cell(row=r, column=c).value = None

    # ── Averages rows (29..40) ──
    for c in range(ATA_COL_START, ATA_COL_END + 1):
        col = col_letter(c)
        ws.cell(row=ROW_SIMPLE_ALL, column=c).value = simple_avg_formula(col, ATA_ALL_RANGE)
        ws.cell(row=ROW_WEIGHTED_ALL, column=c).value = weighted_avg_formula(col, ATA_ALL_RANGE, TRI_ALL_RANGE)
        ws.cell(row=ROW_EXCL_ALL, column=c).value = excl_hl_formula(col, ATA_ALL_RANGE)
        ws.cell(row=ROW_SIMPLE_3YR, column=c).value = simple_avg_formula(col, ATA_3YR_RANGE)
        ws.cell(row=ROW_WEIGHTED_3YR, column=c).value = weighted_avg_formula(col, ATA_3YR_RANGE, TRI_3YR_RANGE)
        ws.cell(row=ROW_EXCL_3YR, column=c).value = excl_hl_formula(col, ATA_3YR_RANGE)
        ws.cell(row=ROW_SIMPLE_5YR, column=c).value = simple_avg_formula(col, ATA_5YR_RANGE)
        ws.cell(row=ROW_WEIGHTED_5YR, column=c).value = weighted_avg_formula(col, ATA_5YR_RANGE, TRI_5YR_RANGE)
        ws.cell(row=ROW_EXCL_5YR, column=c).value = excl_hl_formula(col, ATA_5YR_RANGE)
        ws.cell(row=ROW_SIMPLE_10YR, column=c).value = simple_avg_formula(col, ATA_10YR_RANGE)
        ws.cell(row=ROW_WEIGHTED_10YR, column=c).value = weighted_avg_formula(col, ATA_10YR_RANGE, TRI_10YR_RANGE)
        ws.cell(row=ROW_EXCL_10YR, column=c).value = excl_hl_formula(col, ATA_10YR_RANGE)

    # ── NEW: CDF section (rows 50..52) ──
    # Row 50: title
    title_cell = ws.cell(row=ROW_CDF_TITLE, column=1, value="Cumulative Development Factors")
    title_cell.font = Font(bold=True, size=10)
    ws.merge_cells(start_row=ROW_CDF_TITLE, start_column=1, end_row=ROW_CDF_TITLE, end_column=11)

    # Row 51: age headers (matching triangle age columns)
    ws.cell(row=ROW_CDF_AGES, column=1, value="Age")
    ws.cell(row=ROW_CDF_AGES, column=1).font = Font(bold=True, size=9)
    for c in range(TRI_COL_START, TRI_COL_END + 1):
        # Copy age value from triangle header row 2
        ws.cell(row=ROW_CDF_AGES, column=c, value=f"=${col_letter(c)}$2")
        ws.cell(row=ROW_CDF_AGES, column=c).font = Font(bold=True, size=9)

    # Row 52: CDF values. Start at col K (age 120) = tail (K47), work backwards.
    ws.cell(row=ROW_CDF_VALUES, column=1, value="CDF")
    ws.cell(row=ROW_CDF_VALUES, column=1).font = Font(bold=True, size=9)
    # K52 = K47 (tail factor)
    ws.cell(row=ROW_CDF_VALUES, column=11, value=f"=K{ROW_SELECTION}")
    ws.cell(row=ROW_CDF_VALUES, column=11).number_format = "0.0000"
    # J52 through B52: selected LDF × CDF to the right
    for c in range(10, 1, -1):  # J=10 down to B=2
        col = col_letter(c)
        next_col = col_letter(c + 1)
        ws.cell(
            row=ROW_CDF_VALUES,
            column=c,
            value=f"={col}{ROW_SELECTION}*{next_col}{ROW_CDF_VALUES}"
        )
        ws.cell(row=ROW_CDF_VALUES, column=c).number_format = "0.0000"

    # ── NEW: Projected Ultimates section (rows 54..65) ──
    ws.cell(row=ROW_ULT_TITLE, column=1, value="Projected Ultimates")
    ws.cell(row=ROW_ULT_TITLE, column=1).font = Font(bold=True, size=10)
    ws.merge_cells(start_row=ROW_ULT_TITLE, start_column=1, end_row=ROW_ULT_TITLE, end_column=6)

    # Header row
    headers = ["Period", "Latest Age", "Latest Value", "CDF at Age", "Projected Ultimate", "IBNR"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=ROW_ULT_HEADER, column=i, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")

    # One row per period (triangle rows 3..12)
    for i, tri_row in enumerate(range(TRI_ROW_START, TRI_ROW_END + 1)):
        r = ROW_ULT_START + i
        ws.cell(row=r, column=1, value=f"=A{tri_row}")  # Period label
        # Latest Age: last non-empty in triangle header row 2 (matching last non-empty in the period's triangle row)
        ws.cell(row=r, column=2, value=f"=LOOKUP(2,1/(B{tri_row}:K{tri_row}<>\"\"),$B$2:$K$2)")
        # Latest Value: last non-empty value in period row
        ws.cell(row=r, column=3, value=f"=LOOKUP(2,1/(B{tri_row}:K{tri_row}<>\"\"),B{tri_row}:K{tri_row})")
        ws.cell(row=r, column=3).number_format = "#,##0"
        # CDF at Age: HLOOKUP age against CDF row
        ws.cell(row=r, column=4, value=f"=HLOOKUP(B{r},$B${ROW_CDF_AGES}:$K${ROW_CDF_VALUES},2,FALSE)")
        ws.cell(row=r, column=4).number_format = "0.0000"
        # Projected Ultimate = Latest Value × CDF
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        ws.cell(row=r, column=5).number_format = "#,##0"
        # IBNR = Ultimate - Latest Value
        ws.cell(row=r, column=6, value=f"=E{r}-C{r}")
        ws.cell(row=r, column=6).number_format = "#,##0"

    wb.save(TEMPLATE_PATH)
    print(f"Saved formula-injected template -> {TEMPLATE_PATH}")


if __name__ == "__main__":
    inject()
