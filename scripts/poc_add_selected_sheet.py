"""
POC addendum: Adds one 'Selected Ults - <measure>' sheet per measure to the
template workbook. Each sheet pulls from its measure's Projected Ultimates
region via intra-workbook cross-sheet formulas.

Proves: LDF edit on any measure sheet -> CDF recalc -> Projected Ultimate recalc
-> cross-sheet IBNR recalc on that measure's Selected Ults sheet.
"""

import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

TEMPLATE_PATH = (
    pathlib.Path(__file__).parent.parent
    / "demo" / "demo4-complete" / "poc-output" / "templates"
    / "cl-selections-template.xlsx"
)

MEASURE_SHEETS = ["Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"]

# Projected Ultimates region in the CL measure sheets (from poc_inject_formulas.py)
SRC_ULT_START = 56
SRC_ULT_END = 65
SRC_PERIOD_COL = "A"
SRC_ACTUAL_COL = "C"     # Latest Value
SRC_ULT_COL = "E"        # Projected Ultimate

# Sheet name prefix — keep under 31 chars even after script 6's "CL - " prefix
# "CL - Selected Ults - Incurred Loss" = 34 chars, truncates to "CL - Selected Ults - Incurred L"
# That's acceptable; Excel still treats as valid sheet.
NEW_SHEET_FMT = "Sel Ults - {measure}"  # Kept short so "CL - Sel Ults - X" stays under Excel's 31-char sheet name limit after script 6's prefix.


def build_sheet(wb, measure):
    new_name = NEW_SHEET_FMT.format(measure=measure)[:31]
    if new_name in wb.sheetnames:
        del wb[new_name]
    ws = wb.create_sheet(new_name)

    title = ws.cell(row=1, column=1, value=f"Selected Ultimates — {measure}")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["Accident Period", "Actual (Latest)", "Chain Ladder Ult", "Selected Ultimate", "IBNR"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([18, 16, 18, 20, 14], start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    ws.freeze_panes = "A3"

    # Data rows — cross-sheet references to the measure's Projected Ultimates block
    for i, src_row in enumerate(range(SRC_ULT_START, SRC_ULT_END + 1)):
        r = 3 + i
        ws.cell(row=r, column=1, value=f"='{measure}'!{SRC_PERIOD_COL}{src_row}")
        cell = ws.cell(row=r, column=2, value=f"='{measure}'!{SRC_ACTUAL_COL}{src_row}")
        cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=3, value=f"='{measure}'!{SRC_ULT_COL}{src_row}")
        cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=4, value=f"=C{r}")  # default to CL; actuary can override
        cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=5, value=f"=D{r}-B{r}")
        cell.number_format = "#,##0"

    # Totals
    total_row = 3 + (SRC_ULT_END - SRC_ULT_START + 1)
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for col_letter, col_idx in [("B", 2), ("C", 3), ("D", 4), ("E", 5)]:
        cell = ws.cell(
            row=total_row, column=col_idx,
            value=f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
        )
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)

    print(f"  Built -> {new_name!r}")


def main():
    wb = load_workbook(TEMPLATE_PATH)

    # Remove obsolete sheet names from earlier POC iterations if present
    for old_name in ("Selected Ultimates - POC",
                     "Selected Ults - Incurred Loss",
                     "Selected Ults - Paid Loss",
                     "Selected Ults - Reported Count",
                     "Selected Ults - Closed Count"):
        if old_name in wb.sheetnames:
            del wb[old_name]
            print(f"  Removed obsolete sheet {old_name!r}")

    for measure in MEASURE_SHEETS:
        if measure not in wb.sheetnames:
            print(f"  Skipping {measure!r} (source sheet not found)")
            continue
        build_sheet(wb, measure)

    wb.save(TEMPLATE_PATH)
    print(f"\nSaved -> {TEMPLATE_PATH}")


if __name__ == "__main__":
    main()
