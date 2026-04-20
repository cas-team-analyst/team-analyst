"""
Smoke test for the patched 6-create-complete-analysis.py: run its
write_full_analysis() against the POC formula template and verify the resulting
master workbook preserves formulas and has all cross-sheet refs resolving.
"""

import importlib.util
import pathlib
import re
import sys
from openpyxl import load_workbook

DEMO_ROOT = pathlib.Path(__file__).parent.parent / "demo" / "demo4-complete"
TEMPLATE = DEMO_ROOT / "poc-output" / "templates" / "cl-selections-template.xlsx"
SCRIPT6 = DEMO_ROOT / "scripts" / "6-create-complete-analysis.py"
OUTPUT = DEMO_ROOT / "poc-output" / "complete-analysis-poc.xlsx"

# Import script 6 as a module
spec = importlib.util.spec_from_file_location("script6", SCRIPT6)
script6 = importlib.util.module_from_spec(spec)
sys.path.insert(0, str(SCRIPT6.parent))
spec.loader.exec_module(script6)

# Call write_full_analysis with our formula template as the single source.
# Use no internal files (we'd need projected-ultimates/triangles/etc to build those).
sources = [(str(TEMPLATE), "CL - ")]
internal = []

script6.write_full_analysis(str(OUTPUT), sources, internal)

# Verify: count cross-sheet refs and ensure all resolve
wb = load_workbook(OUTPUT, data_only=False)
sheet_set = set(wb.sheetnames)
good = 0
bad = []
for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                for m in re.finditer(r"'([^'!]+)'!", v):
                    ref = m.group(1)
                    if ref in sheet_set:
                        good += 1
                    else:
                        bad.append((sname, cell.coordinate, ref, v))

print(f"\nCross-sheet ref check:")
print(f"  Good refs: {good}")
print(f"  Broken refs: {len(bad)}")
if bad:
    for b in bad[:5]:
        print(f"    BROKEN: {b}")

# Also sanity-check a few specific formulas survived intact
print("\nSpot-check formulas in the patched output:")
for sname in ["CL - Closed Count", "CL - Sel Ults - Incurred Loss"]:
    if sname in wb.sheetnames:
        ws = wb[sname]
        for coord in ["B16", "B29", "K52", "C56", "E56"]:
            try:
                v = ws[coord].value
                if v is not None:
                    print(f"  {sname!r} {coord}: {v}")
            except Exception:
                pass

print(f"\nOutput at: {OUTPUT}")
