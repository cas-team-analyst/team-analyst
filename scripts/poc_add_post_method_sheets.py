"""
POC: Add formula-driven post-method summary sheets to the template.

Replaces what 6-create-complete-analysis.py's write_post_method_series and
write_post_method_triangles currently emit as hard-coded values.

Sheets added:
  - 'Ult Severity'    : per-period Incurred Ult / Reported Ult severity
  - 'X-to-Ult <measure>' (x4): triangle_cell / selected_ultimate ratios
  - 'Avg IBNR'        : Incurred Ult - Incurred at each age
  - 'Avg Unpaid'      : Incurred Ult - Paid at each age

All cells are formulas; every ref is intra-workbook so the patched script 6
will rewrite them correctly when sheets are prefixed with 'CL - '.

Exposure-dependent diagnostics (Loss Rate, Frequency) not included — they
require exposure data not present in the CL workbook. Flagged as TODO.
"""

import pathlib
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = (
    pathlib.Path(__file__).parent.parent
    / "demo" / "demo4-complete" / "poc-output" / "templates"
    / "cl-selections-template.xlsx"
)

MEASURES = ["Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"]

# Triangle layout (shared across measure sheets)
TRI_ROW_START, TRI_ROW_END = 3, 12
TRI_COL_START, TRI_COL_END = 2, 11  # B..K
AGE_HEADER_ROW = 2
PERIOD_COL = 1  # col A

# Sel Ults sheet layout (from poc_add_selected_sheet.py)
SEL_SHEET_FMT = "Sel Ults - {measure}"
SEL_DATA_START_ROW = 3       # first period row in Sel Ults
SEL_SELECTED_ULT_COL = "D"   # Selected Ultimate column

# Sheet names (kept short so "CL - X" stays under Excel's 31-char cap)
ULT_SEVERITY_SHEET = "Ult Severity"
XTO_ULT_FMT = "X-to-Ult {short}"    # "X-to-Ult Incurred" etc.
AVG_IBNR_SHEET = "Avg IBNR"
AVG_UNPAID_SHEET = "Avg Unpaid"

# Short names for X-to-Ult sheets
MEASURE_SHORT = {
    "Incurred Loss": "Incurred",
    "Paid Loss": "Paid",
    "Reported Count": "Reported",
    "Closed Count": "Closed",
}


def col_letter(idx):
    return get_column_letter(idx)


def sel_ref(measure, sel_row):
    """Formula ref to a measure's Selected Ultimate cell on its Sel Ults sheet."""
    return f"'{SEL_SHEET_FMT.format(measure=measure)}'!{SEL_SELECTED_ULT_COL}{sel_row}"


def measure_cell_ref(measure, row, col):
    """Formula ref to a cell on a measure sheet (triangle region)."""
    return f"'{measure}'!{col_letter(col)}{row}"


def _delete_if_exists(wb, names):
    for n in names:
        if n in wb.sheetnames:
            del wb[n]


def build_ult_severity(wb):
    if ULT_SEVERITY_SHEET in wb.sheetnames:
        del wb[ULT_SEVERITY_SHEET]
    ws = wb.create_sheet(ULT_SEVERITY_SHEET)

    title = ws.cell(row=1, column=1, value="Ultimate Severity — Incurred / Reported")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ["Period", "Incurred Ultimate", "Reported Ultimate", "Ultimate Severity"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([14, 20, 20, 20], start=1):
        ws.column_dimensions[col_letter(i)].width = w
    ws.freeze_panes = "A3"

    # 10 periods, rows 3..12
    for i in range(10):
        r = 3 + i
        sel_row = SEL_DATA_START_ROW + i
        ws.cell(row=r, column=1, value=f"='Sel Ults - Incurred Loss'!A{sel_row}")
        cell = ws.cell(row=r, column=2, value=f"={sel_ref('Incurred Loss', sel_row)}")
        cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=3, value=f"={sel_ref('Reported Count', sel_row)}")
        cell.number_format = "#,##0"
        cell = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r},\"\")")
        cell.number_format = "#,##0.00"


def build_x_to_ult(wb):
    """One X-to-Ult ratio triangle per measure: triangle_cell / selected_ult_for_period."""
    for measure in MEASURES:
        short = MEASURE_SHORT[measure]
        sheet_name = XTO_ULT_FMT.format(short=short)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        title = ws.cell(row=1, column=1, value=f"{short} to Ultimate — ratio triangle")
        title.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TRI_COL_END)

        # Age header row (pull from measure sheet's row 2)
        ws.cell(row=2, column=1, value="Period").font = Font(bold=True, size=10)
        for c in range(TRI_COL_START, TRI_COL_END + 1):
            cell = ws.cell(row=2, column=c, value=f"='{measure}'!{col_letter(c)}$2")
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 14
        for c in range(TRI_COL_START, TRI_COL_END + 1):
            ws.column_dimensions[col_letter(c)].width = 12
        ws.freeze_panes = "B3"

        # Rows: for each period row in the measure sheet
        for i, tri_row in enumerate(range(TRI_ROW_START, TRI_ROW_END + 1)):
            r = 3 + i
            sel_row = SEL_DATA_START_ROW + i
            ws.cell(row=r, column=1, value=f"='{measure}'!A{tri_row}")
            sel_ult = sel_ref(measure, sel_row)
            for c in range(TRI_COL_START, TRI_COL_END + 1):
                tri_ref = measure_cell_ref(measure, tri_row, c)
                # IFERROR guards against blank triangle cells (which would produce 0/sel=0 anyway).
                cell = ws.cell(
                    row=r, column=c,
                    value=f"=IFERROR(IF({tri_ref}=\"\",\"\",{tri_ref}/{sel_ult}),\"\")"
                )
                cell.number_format = "0.0000"


def build_avg_triangle(wb, sheet_name, title, proxy_measure):
    """
    Avg IBNR = Incurred Ult - Incurred at age (proxy_measure = 'Incurred Loss')
    Avg Unpaid = Incurred Ult - Paid at age    (proxy_measure = 'Paid Loss')
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TRI_COL_END)

    ws.cell(row=2, column=1, value="Period").font = Font(bold=True, size=10)
    for c in range(TRI_COL_START, TRI_COL_END + 1):
        cell = ws.cell(row=2, column=c, value=f"='Incurred Loss'!{col_letter(c)}$2")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    for c in range(TRI_COL_START, TRI_COL_END + 1):
        ws.column_dimensions[col_letter(c)].width = 12
    ws.freeze_panes = "B3"

    for i, tri_row in enumerate(range(TRI_ROW_START, TRI_ROW_END + 1)):
        r = 3 + i
        sel_row = SEL_DATA_START_ROW + i
        ws.cell(row=r, column=1, value=f"='Incurred Loss'!A{tri_row}")
        inc_ult = sel_ref("Incurred Loss", sel_row)
        for c in range(TRI_COL_START, TRI_COL_END + 1):
            proxy_ref = measure_cell_ref(proxy_measure, tri_row, c)
            cell = ws.cell(
                row=r, column=c,
                value=f"=IFERROR(IF({proxy_ref}=\"\",\"\",{inc_ult}-{proxy_ref}),\"\")"
            )
            cell.number_format = "#,##0"


def main():
    wb = load_workbook(TEMPLATE_PATH)

    # Clean up any older iteration sheet names
    _delete_if_exists(wb, [
        "Ult Severity",
        "X-to-Ult Incurred", "X-to-Ult Paid", "X-to-Ult Reported", "X-to-Ult Closed",
        "Avg IBNR", "Avg Unpaid",
    ])

    build_ult_severity(wb)
    print(f"  Built -> {ULT_SEVERITY_SHEET!r}")

    build_x_to_ult(wb)
    for measure in MEASURES:
        print(f"  Built -> 'X-to-Ult {MEASURE_SHORT[measure]}'")

    build_avg_triangle(wb, AVG_IBNR_SHEET, "Average IBNR (Incurred Ult - Incurred)", "Incurred Loss")
    print(f"  Built -> {AVG_IBNR_SHEET!r}")

    build_avg_triangle(wb, AVG_UNPAID_SHEET, "Average Unpaid (Incurred Ult - Paid)", "Paid Loss")
    print(f"  Built -> {AVG_UNPAID_SHEET!r}")

    wb.save(TEMPLATE_PATH)
    print(f"\nSaved -> {TEMPLATE_PATH}")


if __name__ == "__main__":
    main()
