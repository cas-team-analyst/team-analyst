# Creates a formatted Excel workbook displaying projected ultimate indications from multiple
# reserving methods (Chain Ladder, Bornhuetter-Ferguson, etc.) alongside prior selections.
# Actuaries review this workbook and manually enter their selected ultimate values and reasoning.

"""
goal: Create Ultimates.xlsx for actuarial review and selection of ultimate losses by period.

Each sheet (one per measure) contains:
  - Period and current maturity information
  - Actual reported losses
  - Initial expected ultimate
  - Indicated ultimates from Chain Ladder, BF, and other methods
  - Prior selections and reasoning (when available)
  - Blank selection rows for actuary input

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 5a-ultimates-create-excel.py
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pathlib

from modules import config
from modules.xl_styles import (
    HEADER_FILL, HEADER_FONT, SELECTION_FILL, AI_FILL, PRIOR_FILL, USER_FILL, THIN_BORDER,
)
from modules.markdown_utils import df_to_markdown

# Paths from modules/config.py — override here if needed:
INPUT_ULTIMATES  = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES = config.PROCESSED_DATA + "triangles.parquet"
PRIOR_SELECTIONS_RB = config.SELECTIONS + "ultimates-prior.json"             # Optional, priority 1 — set to prior cycle's selected ultimates
PRIOR_SELECTIONS_OE = config.SELECTIONS + "ultimates-prior-oe.json"          # Optional, priority 2 — fallback prior
OUTPUT_FILE      = config.SELECTIONS + "Ultimates.xlsx"
SELECTIONS_OUTPUT_PATH = config.SELECTIONS


def export_md_data(df_ult, exp_md):
    import pathlib
    # Subagents should use these markdown files as canonical context.
    # Workbook formulas may not be evaluated in headless execution, so markdown
    # context avoids dependence on stale Excel caches.
    
    # Export Loss category (Incurred + Paid)
    loss_measures = ['Incurred Loss', 'Paid Loss']
    loss_data = []
    for measure in loss_measures:
        df_m = df_ult[df_ult['measure'] == measure].copy()
        if not df_m.empty:
            loss_data.append((measure, df_m))
    
    if loss_data:
        md_path = pathlib.Path(SELECTIONS_OUTPUT_PATH) / "ultimates-context-loss.md"
        md_content = "# Ultimates Context: Loss\n\n"
        md_content += "## Table of Contents\n"
        md_content += "- [Exposure](#exposure)\n"
        md_content += "- [Projected Ultimates](#projected-ultimates)\n\n"
        md_content += "## Exposure\n" + exp_md + "\n"
        md_content += "## Projected Ultimates\n\n"
        
        for measure, df_m in loss_data:
            df_m = df_m.drop(columns=['measure'], errors='ignore')
            md_content += f"### {measure}\n\n"
            md_content += df_to_markdown(df_m, index=False) + "\n\n"
        
        with open(md_path, 'w') as f:
            f.write(md_content)
        print(f"  Exported MD: {md_path}")
    
    # Export Count category (Reported + Closed)
    count_measures = ['Reported Count', 'Closed Count']
    count_data = []
    for measure in count_measures:
        df_m = df_ult[df_ult['measure'] == measure].copy()
        if not df_m.empty:
            count_data.append((measure, df_m))
    
    if count_data:
        md_path = pathlib.Path(SELECTIONS_OUTPUT_PATH) / "ultimates-context-count.md"
        md_content = "# Ultimates Context: Count\n\n"
        md_content += "## Table of Contents\n"
        md_content += "- [Exposure](#exposure)\n"
        md_content += "- [Projected Ultimates](#projected-ultimates)\n\n"
        md_content += "## Exposure\n" + exp_md + "\n"
        md_content += "## Projected Ultimates\n\n"
        
        for measure, df_m in count_data:
            df_m = df_m.drop(columns=['measure'], errors='ignore')
            md_content += f"### {measure}\n\n"
            md_content += df_to_markdown(df_m, index=False) + "\n\n"
        
        with open(md_path, 'w') as f:
            f.write(md_content)
        print(f"  Exported MD: {md_path}")



def format_loss_sheet(ws, df_ult, df_prior):
    """
    Format the Loss sheet with Incurred and Paid columns combined.
    
    Args:
        ws: openpyxl worksheet object
        df_ult: DataFrame with projected ultimates for all measures
        df_prior: DataFrame with prior selections (or None if no prior selections)
    """
    # Get data for both loss measures
    df_incurred = df_ult[df_ult['measure'] == 'Incurred Loss'].copy() if 'Incurred Loss' in df_ult['measure'].values else pd.DataFrame()
    df_paid = df_ult[df_ult['measure'] == 'Paid Loss'].copy() if 'Paid Loss' in df_ult['measure'].values else pd.DataFrame()
    
    if df_incurred.empty and df_paid.empty:
        return
    
    # Merge on period to get one row per period
    if not df_incurred.empty and not df_paid.empty:
        df_combined = df_incurred[['period', 'current_age', 'actual']].copy()
        df_combined.rename(columns={'actual': 'incurred_actual'}, inplace=True)
        df_paid_subset = df_paid[['period', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_paid_subset.rename(columns={'actual': 'paid_actual', 'ultimate_cl': 'paid_cl', 'ultimate_bf': 'paid_bf'}, inplace=True)
        df_combined = df_combined.merge(df_paid_subset, on='period', how='outer')
        df_combined.rename(columns={'incurred_actual': 'incurred', 'paid_actual': 'paid'}, inplace=True)
        
        # Add incurred ultimates
        df_combined = df_combined.merge(
            df_incurred[['period', 'ultimate_cl', 'ultimate_bf']].rename(columns={'ultimate_cl': 'incurred_cl', 'ultimate_bf': 'incurred_bf'}),
            on='period', how='left'
        )
    elif not df_incurred.empty:
        df_combined = df_incurred[['period', 'current_age', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_combined.rename(columns={'actual': 'incurred', 'ultimate_cl': 'incurred_cl', 'ultimate_bf': 'incurred_bf'}, inplace=True)
        df_combined['paid'] = None
        df_combined['paid_cl'] = None
        df_combined['paid_bf'] = None
    else:  # only paid
        df_combined = df_paid[['period', 'current_age', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_combined.rename(columns={'actual': 'paid', 'ultimate_cl': 'paid_cl', 'ultimate_bf': 'paid_bf'}, inplace=True)
        df_combined['incurred'] = None
        df_combined['incurred_cl'] = None
        df_combined['incurred_bf'] = None
    
    # Write headers
    headers = [
        "Accident Period", "Current Age", "Incurred", "Paid",
        "Incurred CL", "Paid CL", "Incurred BF", "Paid BF",
        "Prior Selection", "Prior Reasoning",
        "Rules-Based AI Selection", "Rules-Based AI Reasoning",
        "Open-Ended AI Selection", "Open-Ended AI Reasoning",
        "User Selection", "User Reasoning"
    ]
    
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['P'].width = 40
    
    # Create dict of prior
    prior_dict = {}
    if df_prior is not None and 'Loss' in df_prior.get('category', df_prior.get('measure', pd.Series())).values:
        mp = df_prior[df_prior.get('category', df_prior.get('measure')) == 'Loss']
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {"sel": r.get('selection', r.get('selected_ultimate')), "reason": r.get('reasoning', '')}
    
    # Write rows
    for r_idx, (_, row) in enumerate(df_combined.iterrows(), start=2):
        period = str(row['period'])
        
        # Period & current age
        ws.cell(row=r_idx, column=1, value=period).border = THIN_BORDER
        
        val_age = row.get('current_age')
        if pd.isna(val_age): val_age = ""
        ws.cell(row=r_idx, column=2, value=val_age).border = THIN_BORDER
        
        # Values
        for c_idx, col_name in enumerate(['incurred', 'paid', 'incurred_cl', 'paid_cl', 'incurred_bf', 'paid_bf'], start=3):
            val = row.get(col_name)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
        
        # Prior
        prior_sel = prior_dict.get(period, {}).get("sel", "")
        prior_reason = prior_dict.get(period, {}).get("reason", "")
        
        c = ws.cell(row=r_idx, column=9, value=prior_sel)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=10, value=prior_reason)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Rules-Based AI Selection (yellow fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=11)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=12)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Open-Ended AI Selection (purple fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=13)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=14)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # User Selection (blank - actuary input)
        c = ws.cell(row=r_idx, column=15)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=16)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)


def format_count_sheet(ws, df_ult, df_prior):
    """
    Format the Count sheet with Reported and Closed columns combined.
    
    Args:
        ws: openpyxl worksheet object
        df_ult: DataFrame with projected ultimates for all measures
        df_prior: DataFrame with prior selections (or None if no prior selections)
    """
    # Get data for both count measures
    df_reported = df_ult[df_ult['measure'] == 'Reported Count'].copy() if 'Reported Count' in df_ult['measure'].values else pd.DataFrame()
    df_closed = df_ult[df_ult['measure'] == 'Closed Count'].copy() if 'Closed Count' in df_ult['measure'].values else pd.DataFrame()
    
    if df_reported.empty and df_closed.empty:
        return
    
    # Merge on period to get one row per period
    if not df_reported.empty and not df_closed.empty:
        df_combined = df_reported[['period', 'current_age', 'actual']].copy()
        df_combined.rename(columns={'actual': 'reported_actual'}, inplace=True)
        df_closed_subset = df_closed[['period', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_closed_subset.rename(columns={'actual': 'closed_actual', 'ultimate_cl': 'closed_cl', 'ultimate_bf': 'closed_bf'}, inplace=True)
        df_combined = df_combined.merge(df_closed_subset, on='period', how='outer')
        df_combined.rename(columns={'reported_actual': 'reported', 'closed_actual': 'closed'}, inplace=True)
        
        # Add reported ultimates
        df_combined = df_combined.merge(
            df_reported[['period', 'ultimate_cl', 'ultimate_bf']].rename(columns={'ultimate_cl': 'reported_cl', 'ultimate_bf': 'reported_bf'}),
            on='period', how='left'
        )
    elif not df_reported.empty:
        df_combined = df_reported[['period', 'current_age', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_combined.rename(columns={'actual': 'reported', 'ultimate_cl': 'reported_cl', 'ultimate_bf': 'reported_bf'}, inplace=True)
        df_combined['closed'] = None
        df_combined['closed_cl'] = None
        df_combined['closed_bf'] = None
    else:  # only closed
        df_combined = df_closed[['period', 'current_age', 'actual', 'ultimate_cl', 'ultimate_bf']].copy()
        df_combined.rename(columns={'actual': 'closed', 'ultimate_cl': 'closed_cl', 'ultimate_bf': 'closed_bf'}, inplace=True)
        df_combined['reported'] = None
        df_combined['reported_cl'] = None
        df_combined['reported_bf'] = None
    
    # Write headers
    headers = [
        "Accident Period", "Current Age", "Reported", "Closed",
        "Reported CL", "Closed CL", "Reported BF", "Closed BF",
        "Prior Selection", "Prior Reasoning",
        "Rules-Based AI Selection", "Rules-Based AI Reasoning",
        "Open-Ended AI Selection", "Open-Ended AI Reasoning",
        "User Selection", "User Reasoning"
    ]
    
    ws.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['P'].width = 40
    
    # Create dict of prior
    prior_dict = {}
    if df_prior is not None and 'Count' in df_prior.get('category', df_prior.get('measure', pd.Series())).values:
        mp = df_prior[df_prior.get('category', df_prior.get('measure')) == 'Count']
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {"sel": r.get('selection', r.get('selected_ultimate')), "reason": r.get('reasoning', '')}
    
    # Write rows
    for r_idx, (_, row) in enumerate(df_combined.iterrows(), start=2):
        period = str(row['period'])
        
        # Period & current age
        ws.cell(row=r_idx, column=1, value=period).border = THIN_BORDER
        
        val_age = row.get('current_age')
        if pd.isna(val_age): val_age = ""
        ws.cell(row=r_idx, column=2, value=val_age).border = THIN_BORDER
        
        # Values
        for c_idx, col_name in enumerate(['reported', 'closed', 'reported_cl', 'closed_cl', 'reported_bf', 'closed_bf'], start=3):
            val = row.get(col_name)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
        
        # Prior
        prior_sel = prior_dict.get(period, {}).get("sel", "")
        prior_reason = prior_dict.get(period, {}).get("reason", "")
        
        c = ws.cell(row=r_idx, column=9, value=prior_sel)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=10, value=prior_reason)
        c.fill = PRIOR_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Rules-Based AI Selection (yellow fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=11)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=12)
        c.fill = SELECTION_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # Open-Ended AI Selection (purple fill - will be populated by 5b script)
        c = ws.cell(row=r_idx, column=13)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=14)
        c.fill = AI_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)
        
        # User Selection (blank - actuary input)
        c = ws.cell(row=r_idx, column=15)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.number_format = "#,##0"
        
        c = ws.cell(row=r_idx, column=16)
        c.fill = USER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(wrap_text=True)




def main():
    """Create ultimates selection Excel file from projected ultimates."""
    print("Creating Ultimates.xlsx...")
    
    try:
        df_ult = pd.read_parquet(INPUT_ULTIMATES)
        print(f"Loaded ultimates from {INPUT_ULTIMATES}")
    except Exception as e:
        print(f"Error loading ultimates: {e}")
        exit(1)

    # Guard: warn if CL ultimates are missing for any loss measure — selector will fall
    # back to IE for every AY if this context is used before 2f has run successfully.
    for _m in ['Incurred Loss', 'Paid Loss']:
        if _m not in df_ult['measure'].values:
            continue
        _df_m = df_ult[df_ult['measure'] == _m]
        if 'ultimate_cl' not in _df_m.columns or _df_m['ultimate_cl'].isna().all():
            print(f"\nWARNING: ultimate_cl is all NaN for '{_m}'. "
                  f"Run 2f-chainladder-ultimates.py before this script — "
                  f"otherwise the ultimates selector will treat CL as unavailable for every AY.\n")

    df_prior = None
    # Try loading prior selections - prioritize rules-based, fall back to open-ended
    prior_loaded = False
    try:
        if pathlib.Path(PRIOR_SELECTIONS_RB).exists():
            with open(PRIOR_SELECTIONS_RB, 'r') as f:
                prior_data = json.load(f)
            df_prior = pd.DataFrame(prior_data)
            print(f"Loaded prior selections from {PRIOR_SELECTIONS_RB} (rules-based)")
            prior_loaded = True
    except Exception as e:
        print(f"  Could not load rules-based prior selections: {e}")
    
    if not prior_loaded:
        try:
            if pathlib.Path(PRIOR_SELECTIONS_OE).exists():
                with open(PRIOR_SELECTIONS_OE, 'r') as f:
                    prior_data = json.load(f)
                df_prior = pd.DataFrame(prior_data)
                print(f"Loaded prior selections from {PRIOR_SELECTIONS_OE} (open-ended)")
        except Exception as e:
            print(f"  No prior selections found: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Loss sheet (Incurred + Paid)
    has_loss = any(m in df_ult['measure'].values for m in ['Incurred Loss', 'Paid Loss'])
    if has_loss:
        ws = wb.create_sheet(title='Loss')
        format_loss_sheet(ws, df_ult, df_prior)
        print(f"  Created sheet for Loss (Incurred + Paid)")
    
    # Create Count sheet (Reported + Closed)
    has_count = any(m in df_ult['measure'].values for m in ['Reported Count', 'Closed Count'])
    if has_count:
        ws = wb.create_sheet(title='Count')
        format_count_sheet(ws, df_ult, df_prior)
        print(f"  Created sheet for Count (Reported + Closed)")
    
    out_dir = pathlib.Path(OUTPUT_FILE).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    
    exp_md = "No Exposure data\n"
    try:
        tri_df = pd.read_parquet(INPUT_TRIANGLES)
        exp_sub = tri_df[(tri_df['measure'] == 'Exposure') & tri_df['value'].notna()]
        if not exp_sub.empty:
            exp_piv = exp_sub.pivot(index='period', columns='age', values='value')
            exp_md = df_to_markdown(exp_piv, index=True)
    except Exception:
        pass
        
    export_md_data(df_ult, exp_md)


if __name__ == "__main__":
    main()
