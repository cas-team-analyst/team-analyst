from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

from modules.xl_styles import DATA_FONT, THIN_BORDER, style_header
from modules.xl_utils import _safe, _to_py


def _data_cell(cell, value, num_fmt=None):
    """Write a styled data cell with border and alignment."""
    value = _to_py(_safe(value))
    cell.value     = value
    cell.font      = DATA_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="right" if isinstance(value, (int, float)) else "left",
        vertical="center",
    )
    if num_fmt and value is not None and isinstance(value, (int, float)):
        cell.number_format = num_fmt


def _write_title_and_headers(ws, title, headers, col_width=18):
    """
    Write a title row (row 1) and a styled header row (row 2).
    Returns 3 — the first data row.
    This format matches script 7's read_with_title() convention.
    """
    title_cell = ws.cell(1, 1, title)
    style_header(title_cell, "header")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 20

    for c, text in enumerate(headers, 1):
        cell = ws.cell(2, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A3"
    return 3


def _write_headers(ws, headers, col_width=18):
    """
    Write a single styled header row (row 1), no title row.
    Returns 2 — the first data row.
    Matches script 7's read_no_title() convention ("Sel - " sheets).
    """
    for c, text in enumerate(headers, 1):
        cell = ws.cell(1, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A2"
    return 2


# ==============================================================================
# xlsxwriter functions (for formula-driven Excel workbooks)
# ==============================================================================

def col_letter(col_idx):
    """Convert 0-based column index to Excel column letter (A, B, C, etc.)."""
    result = ''
    while col_idx >= 0:
        result = chr(col_idx % 26 + ord('A')) + result
        col_idx = col_idx // 26 - 1
    return result


def write_triangle_xlsxwriter(ws, start_row, row_labels, col_labels, data_dict, fmt, number_format="#,##0"):
    """
    Write a triangle to xlsxwriter worksheet starting at start_row (0-based).
    
    Args:
        ws: xlsxwriter worksheet
        start_row: Starting row (0-based)
        row_labels: List of row labels (e.g., periods/accident years)
        col_labels: List of column labels (e.g., ages/development periods)
        data_dict: Dict mapping (str(row_label), str(col_label)) -> value
        fmt: Format dict from create_xlsxwriter_formats()
        number_format: Excel number format string for data cells
    
    Returns:
        (next_row, data_start_row, data_end_row) tuple
    """
    # Create format with specific number format
    data_fmt = fmt['wb'].add_format({
        'align': 'right',
        'valign': 'vcenter',
        'num_format': number_format
    })
    
    # Write header row
    ws.write(start_row, 0, "Period", fmt['subheader_left'])
    for c_idx, col in enumerate(col_labels):
        ws.write(start_row, c_idx + 1, col, fmt['subheader_right'])
    
    # Write data rows
    data_start_row = start_row + 1
    for r_idx, row_label in enumerate(row_labels):
        row = data_start_row + r_idx
        # Write row label as number if numeric (try converting strings)
        if isinstance(row_label, (int, float, np.integer, np.floating)):
            ws.write_number(row, 0, float(row_label), fmt['label'])
        else:
            try:
                ws.write_number(row, 0, float(row_label), fmt['label'])
            except (ValueError, TypeError):
                ws.write(row, 0, str(row_label), fmt['label'])
        for c_idx, col in enumerate(col_labels):
            val = data_dict.get((str(row_label), str(col)))
            if val is not None:
                ws.write(row, c_idx + 1, val, data_fmt)
    
    data_end_row = data_start_row + len(row_labels) - 1
    next_row = start_row + len(row_labels) + 2  # +1 for header, +1 for blank row after
    
    return next_row, data_start_row, data_end_row


def write_ultimates_sheet_xlw(wb, ws, df_ult, df_prior, category, measures):
    """
    Write ultimates selection sheet for xlsxwriter with proper formatting.
    
    Creates a sheet with actuals, method projections (CL, IE, BF), and selection columns.
    KEY FIX: Only ONE Initial Expected column per category, not separate IE per measure.
    
    Args:
        wb: xlsxwriter workbook object (for creating formats)
        ws: xlsxwriter worksheet object
        df_ult: DataFrame with projected ultimates (must have 'measure' column)
        df_prior: DataFrame with prior selections (or None)
        category: 'Losses' or 'Counts'
        measures: List of measure names (e.g., ['Incurred Loss', 'Paid Loss'])
    
    Example for Losses:
        Accident Period, Current Age, Incurred, Paid, 
        Incurred CL, Paid CL, Initial Expected, Incurred BF, Paid BF,
        [Prior Selection, Prior Reasoning],
        Rules-Based AI Selection, Rules-Based AI Reasoning,
        Open-Ended AI Selection, Open-Ended AI Reasoning,
        User Selection, User Reasoning
    """
    from modules.xl_styles import create_xlsxwriter_formats
    
    # Get formats
    fmt = create_xlsxwriter_formats(wb)
    
    # Extract data for each measure
    measure_dfs = {}
    for measure in measures:
        df_m = df_ult[df_ult['measure'] == measure].copy() if measure in df_ult['measure'].values else pd.DataFrame()
        if not df_m.empty:
            measure_dfs[measure] = df_m
    
    if not measure_dfs:
        return
    
    # Determine short names for measures
    measure_short = {}
    for m in measure_dfs.keys():
        if 'Incurred' in m:
            measure_short[m] = 'Incurred'
        elif 'Paid' in m:
            measure_short[m] = 'Paid'
        elif 'Reported' in m:
            measure_short[m] = 'Reported'
        elif 'Closed' in m:
            measure_short[m] = 'Closed'
        else:
            measure_short[m] = m
    
    # For Counts: check if we have actual closed count data
    has_closed_data = False
    if category == 'Counts':
        for m in measures:
            if 'Closed' in m and m in measure_dfs:
                has_closed_data = measure_dfs[m]['actual'].notna().any()
                if not has_closed_data:
                    # Remove from measure_dfs if no actual data
                    measure_dfs.pop(m, None)
    
    # Merge dataframes on period to get one row per period
    df_combined = None
    for m, df_m in measure_dfs.items():
        short = measure_short[m]
        # Build column list dynamically - only include columns that exist
        base_cols = ['period', 'current_age', 'actual', 'ultimate_cl']
        optional_cols = ['ultimate_ie', 'ultimate_bf']
        cols_to_extract = base_cols + [c for c in optional_cols if c in df_m.columns]
        
        df_subset = df_m[cols_to_extract].copy()
        rename_dict = {
            'actual': f'{short.lower()}_actual',
            'ultimate_cl': f'{short.lower()}_cl',
        }
        if 'ultimate_ie' in df_subset.columns:
            rename_dict['ultimate_ie'] = f'{short.lower()}_ie_temp'  # Temp name - will consolidate
        if 'ultimate_bf' in df_subset.columns:
            rename_dict['ultimate_bf'] = f'{short.lower()}_bf'
        
        df_subset = df_subset.rename(columns=rename_dict)
        
        if df_combined is None:
            df_combined = df_subset
        else:
            # Build merge column list dynamically
            merge_cols = ['period', f'{short.lower()}_actual', f'{short.lower()}_cl']
            if f'{short.lower()}_ie_temp' in df_subset.columns:
                merge_cols.append(f'{short.lower()}_ie_temp')
            if f'{short.lower()}_bf' in df_subset.columns:
                merge_cols.append(f'{short.lower()}_bf')
            df_combined = df_combined.merge(df_subset[merge_cols], on='period', how='outer')
    
    # Consolidate IE columns - take first non-null IE value across all measures
    # (IE is typically the same for all measures in a category)
    ie_cols = [c for c in df_combined.columns if c.endswith('_ie_temp')]
    if ie_cols:
        df_combined['ie'] = df_combined[ie_cols].bfill(axis=1).iloc[:, 0]
        df_combined = df_combined.drop(columns=ie_cols)
    else:
        df_combined['ie'] = None
    
    # Check if we have prior data for this category
    has_prior_data = df_prior is not None and category in df_prior.get('category', df_prior.get('measure', pd.Series())).values
    
    # Check if we have IE or BF data
    has_ie_data = 'ie' in df_combined.columns and df_combined['ie'].notna().any()
    bf_cols = [c for c in df_combined.columns if c.endswith('_bf')]
    has_bf_data = any(df_combined[c].notna().any() for c in bf_cols) if bf_cols else False
    
    # Build headers
    headers = ["Accident Period", "Current Age"]
    
    # Add actual columns
    for m in measure_dfs.keys():
        headers.append(measure_short[m])
    
    # Add CL columns
    for m in measure_dfs.keys():
        headers.append(f"{measure_short[m]} CL")
    
    # Add ONE Initial Expected column (only if IE data exists)
    if has_ie_data:
        headers.append("Initial Expected")
    
    # Add BF columns (only if BF data exists)
    if has_bf_data:
        for m in measure_dfs.keys():
            headers.append(f"{measure_short[m]} BF")
    
    # Add Prior columns (conditional)
    if has_prior_data:
        headers.extend(["Prior Selection", "Prior Reasoning"])
    
    # Add selection columns
    headers.extend([
        "Rules-Based AI Selection", "Rules-Based AI Reasoning",
        "Open-Ended AI Selection", "Open-Ended AI Reasoning",
        "User Selection", "User Reasoning"
    ])
    
    # Define which columns are left-aligned (text/reasoning columns)
    text_columns = {
        "Accident Period",
        "Prior Reasoning",
        "Rules-Based AI Reasoning",
        "Open-Ended AI Reasoning",
        "User Reasoning"
    }
    
    # Write headers with appropriate alignment
    for c_idx, header in enumerate(headers):
        if header in text_columns:
            ws.write(0, c_idx, header, fmt['subheader_left'])
        else:
            ws.write(0, c_idx, header, fmt['subheader_right'])
        ws.set_column(c_idx, c_idx, 18)
    
    # Build column map
    col_map = {header: idx for idx, header in enumerate(headers)}
    
    # Set wider widths for selection and reasoning columns
    if has_prior_data:
        ws.set_column(col_map["Prior Selection"], col_map["Prior Selection"], 22)
        ws.set_column(col_map["Prior Reasoning"], col_map["Prior Reasoning"], 30)
    ws.set_column(col_map["Rules-Based AI Selection"], col_map["Rules-Based AI Selection"], 22)
    ws.set_column(col_map["Rules-Based AI Reasoning"], col_map["Rules-Based AI Reasoning"], 30)
    ws.set_column(col_map["Open-Ended AI Selection"], col_map["Open-Ended AI Selection"], 22)
    ws.set_column(col_map["Open-Ended AI Reasoning"], col_map["Open-Ended AI Reasoning"], 30)
    ws.set_column(col_map["User Selection"], col_map["User Selection"], 22)
    ws.set_column(col_map["User Reasoning"], col_map["User Reasoning"], 40)
    
    # Create dict of prior selections
    prior_dict = {}
    if has_prior_data:
        mp = df_prior[df_prior.get('category', df_prior.get('measure')) == category]
        for _, r in mp.iterrows():
            prior_dict[str(r['period'])] = {
                "sel": r.get('selection', r.get('selected_ultimate')),
                "reason": r.get('reasoning', '')
            }
    
    # Write data rows
    for r_idx, (_, row) in enumerate(df_combined.iterrows(), start=1):
        period = row['period']
        
        # Period - write as number if numeric (try converting strings)
        if isinstance(period, (int, float, np.integer, np.floating)):
            ws.write_number(r_idx, 0, float(period), fmt['label'])
        else:
            try:
                ws.write_number(r_idx, 0, float(period), fmt['label'])
            except (ValueError, TypeError):
                ws.write(r_idx, 0, str(period), fmt['label'])
        
        # Current Age - write as number if numeric (try converting strings)
        val_age = row.get('current_age')
        if pd.isna(val_age):
            ws.write_blank(r_idx, 1, None, fmt['label'])
        elif isinstance(val_age, (int, float, np.integer, np.floating)):
            ws.write_number(r_idx, 1, float(val_age), fmt['label'])
        else:
            try:
                ws.write_number(r_idx, 1, float(val_age), fmt['label'])
            except (ValueError, TypeError):
                ws.write(r_idx, 1, str(val_age), fmt['label'])
        
        # Write actuals
        col_idx = 2
        for m in measure_dfs.keys():
            short = measure_short[m].lower()
            val = row.get(f'{short}_actual')
            if pd.isna(val):
                val = ""
            ws.write(r_idx, col_idx, val, fmt['data_num'])
            col_idx += 1
        
        # Write CL columns
        for m in measure_dfs.keys():
            short = measure_short[m].lower()
            val = row.get(f'{short}_cl')
            if pd.isna(val):
                val = ""
            ws.write(r_idx, col_idx, val, fmt['data_num'])
            col_idx += 1
        
        # Write ONE IE column (only if IE data exists)
        if has_ie_data:
            val = row.get('ie')
            if pd.isna(val):
                val = ""
            ws.write(r_idx, col_idx, val, fmt['data_num'])
            col_idx += 1
        
        # Write BF columns (only if BF data exists)
        if has_bf_data:
            for m in measure_dfs.keys():
                short = measure_short[m].lower()
                val = row.get(f'{short}_bf')
                if pd.isna(val):
                    val = ""
                ws.write(r_idx, col_idx, val, fmt['data_num'])
                col_idx += 1
        
        # Prior (only if prior data available)
        if has_prior_data:
            prior_sel = prior_dict.get(str(period), {}).get("sel", "")
            prior_reason = prior_dict.get(str(period), {}).get("reason", "")
            ws.write(r_idx, col_map["Prior Selection"], prior_sel, fmt['prior'])
            ws.write(r_idx, col_map["Prior Reasoning"], prior_reason, fmt['prior'])
        
        # Rules-Based AI Selection (yellow fill - populated by 5b script)
        ws.write(r_idx, col_map["Rules-Based AI Selection"], "", fmt['selection'])
        ws.write(r_idx, col_map["Rules-Based AI Reasoning"], "", fmt['selection'])
        
        # Open-Ended AI Selection (purple fill - populated by 5b script)
        ws.write(r_idx, col_map["Open-Ended AI Selection"], "", fmt['ai'])
        ws.write(r_idx, col_map["Open-Ended AI Reasoning"], "", fmt['ai'])
        
        # User Selection (blank - actuary input)
        ws.write(r_idx, col_map["User Selection"], "", fmt['user'])
        ws.write(r_idx, col_map["User Reasoning"], "", fmt['user'])
