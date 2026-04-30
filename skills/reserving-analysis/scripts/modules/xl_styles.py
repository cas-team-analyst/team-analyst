# Shared Excel formatting constants and helpers for reserving analysis workbooks.

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Color palette - shared across openpyxl and xlsxwriter
COLORS = {
    'header_blue': '1F4E79',
    'subheader_blue': '4472C4',
    'section_blue': 'D9E1F2',
    'selection_green': 'E2EFDA',
    'prior_yellow': 'FFF2CC',
    'ai_blue': 'DDEBF7',
    'user_orange': 'F4B084',
    'data_gray': 'D9D9D9',
    'label_gray': 'F2F2F2',
}

# openpyxl styles
HEADER_FILL    = PatternFill("solid", fgColor=COLORS['header_blue'])
SUBHEADER_FILL = PatternFill("solid", fgColor=COLORS['subheader_blue'])
SECTION_FILL   = PatternFill("solid", fgColor=COLORS['section_blue'])
SELECTION_FILL = PatternFill("solid", fgColor=COLORS['selection_green'])
PRIOR_FILL     = PatternFill("solid", fgColor=COLORS['prior_yellow'])
AI_FILL        = PatternFill("solid", fgColor=COLORS['ai_blue'])
USER_FILL      = PatternFill("solid", fgColor=COLORS['user_orange'])

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT   = Font(bold=True, size=10)
LABEL_FONT     = Font(bold=True, size=9)
DATA_FONT      = Font(size=9)

THIN        = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_STYLE_MAP = {
    "header":    (HEADER_FILL,    HEADER_FONT),
    "subheader": (SUBHEADER_FILL, SUBHEADER_FONT),
    "section":   (SECTION_FILL,   SECTION_FONT),
    "selection": (SELECTION_FILL, LABEL_FONT),
    "prior":     (PRIOR_FILL,     LABEL_FONT),
    "ai":        (AI_FILL,        LABEL_FONT),
    "user":      (USER_FILL,      LABEL_FONT),
}


def style_header(cell, level="header"):
    """Apply standard fill + font + centered alignment + border to a cell."""
    fill, font = _STYLE_MAP.get(level, (SUBHEADER_FILL, SUBHEADER_FONT))
    cell.fill      = fill
    cell.font      = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = THIN_BORDER


def create_xlsxwriter_formats(workbook):
    """
    Create a consistent set of xlsxwriter Format objects for reserving workbooks.
    
    Args:
        workbook: xlsxwriter.Workbook instance
    
    Returns:
        dict: Format objects keyed by name
    """
    formats = {}
    
    # Header format
    formats['header'] = workbook.add_format({
        'bold': True,
        'bg_color': '#' + COLORS['header_blue'],
        'font_color': 'FFFFFF',
        'align': 'center',
        'valign': 'vcenter'
    })
    
    # Base formats
    formats['subheader'] = workbook.add_format({
        'bold': True,
        'bg_color': '#' + COLORS['data_gray'],
        'align': 'center',
        'valign': 'vcenter'
    })
    
    formats['subheader_right'] = workbook.add_format({
        'bold': True,
        'bg_color': '#' + COLORS['data_gray'],
        'align': 'right',
        'valign': 'vcenter'
    })
    
    formats['subheader_left'] = workbook.add_format({
        'bold': True,
        'bg_color': '#' + COLORS['data_gray'],
        'align': 'left',
        'valign': 'vcenter'
    })
    
    formats['label'] = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter'
    })
    
    formats['data'] = workbook.add_format({
        'align': 'right',
        'valign': 'vcenter'
    })
    
    # Data with number formats
    formats['data_num'] = workbook.add_format({
        'align': 'right',
        'valign': 'vcenter',
        'num_format': '#,##0'
    })
    
    formats['data_ldf'] = workbook.add_format({
        'align': 'right',
        'valign': 'vcenter',
        'num_format': '0.0000'
    })
    
    formats['data_pct'] = workbook.add_format({
        'align': 'right',
        'valign': 'vcenter',
        'num_format': '0.00%'
    })
    
    # Period format - numeric but no comma formatting (General)
    formats['data_period'] = workbook.add_format({
        'align': 'right',
        'valign': 'vcenter',
        'num_format': '0'
    })
    
    # Section header (grey background)
    formats['section'] = workbook.add_format({
        'bg_color': '#' + COLORS['data_gray'],
        'align': 'left',
        'valign': 'vcenter',
        'bold': True
    })
    
    # Label row (lighter grey background)
    formats['label_row'] = workbook.add_format({
        'bg_color': '#' + COLORS['label_gray'],
        'align': 'left',
        'valign': 'vcenter',
        'bold': True
    })
    
    # Diagnostic section header (grey background)
    formats['diagnostic_section'] = workbook.add_format({
        'bg_color': '#' + COLORS['data_gray'],
        'align': 'left',
        'valign': 'vcenter',
        'bold': True
    })
    
    # Prior selections (yellow)
    formats['prior'] = workbook.add_format({
        'bg_color': '#' + COLORS['prior_yellow'],
        'align': 'right',
        'num_format': '#,##0'
    })
    
    formats['prior_data'] = workbook.add_format({
        'bg_color': '#' + COLORS['prior_yellow'],
        'align': 'right',
        'num_format': '0.0000'
    })
    
    formats['prior_text'] = workbook.add_format({
        'bg_color': '#' + COLORS['prior_yellow'],
        'align': 'left',
        'text_wrap': True
    })
    
    formats['prior_num'] = workbook.add_format({
        'bg_color': '#' + COLORS['prior_yellow'],
        'align': 'right',
        'num_format': '0'
    })
    
    formats['prior_note'] = workbook.add_format({
        'bg_color': '#' + COLORS['prior_yellow'],
        'italic': True,
        'font_size': 8,
        'align': 'left'
    })
    
    # Current selections (green)
    formats['selection'] = workbook.add_format({
        'bg_color': '#' + COLORS['selection_green'],
        'align': 'right',
        'num_format': '#,##0'
    })
    
    formats['selection_data'] = workbook.add_format({
        'bg_color': '#' + COLORS['selection_green'],
        'align': 'right',
        'num_format': '0.0000'
    })
    
    formats['selection_text'] = workbook.add_format({
        'bg_color': '#' + COLORS['selection_green'],
        'align': 'left',
        'text_wrap': True
    })
    
    # AI selections (blue)
    formats['ai'] = workbook.add_format({
        'bg_color': '#' + COLORS['ai_blue'],
        'align': 'right',
        'num_format': '#,##0'
    })
    
    formats['ai_data'] = workbook.add_format({
        'bg_color': '#' + COLORS['ai_blue'],
        'align': 'right',
        'num_format': '0.0000'
    })
    
    formats['ai_text'] = workbook.add_format({
        'bg_color': '#' + COLORS['ai_blue'],
        'align': 'left',
        'text_wrap': True
    })
    
    # User selections (orange)
    formats['user'] = workbook.add_format({
        'bg_color': '#' + COLORS['user_orange'],
        'align': 'right',
        'num_format': '#,##0'
    })
    
    formats['user_data'] = workbook.add_format({
        'bg_color': '#' + COLORS['user_orange'],
        'align': 'right',
        'num_format': '0.0000'
    })
    
    formats['user_text'] = workbook.add_format({
        'bg_color': '#' + COLORS['user_orange'],
        'align': 'left',
        'text_wrap': True
    })
    
    # Header format (for measure banner)
    formats['header'] = workbook.add_format({
        'bold': True,
        'bg_color': '#' + COLORS['header_blue'],
        'font_color': 'FFFFFF',
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 11
    })
    
    return formats
