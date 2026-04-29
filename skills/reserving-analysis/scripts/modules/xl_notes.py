from datetime import datetime

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules.xl_styles import SUBHEADER_FONT, SUBHEADER_FILL, LABEL_FONT, DATA_FONT, THIN_BORDER, style_header

# Each entry: (purpose, key columns string). Matched by exact name, then prefix, then suffix.

_SHEET_DESCS_EXACT = {
    "Loss Selection": (
        "Actuary-selected ultimates for loss measures comparing Chain Ladder, Initial Expected, and Bornhuetter-Ferguson methods.",
        "Period · Age · Incurred · Paid · Incurred CL · Paid CL · Incurred BF · Paid BF · Initial Expected · Selected Ultimate · IBNR · Unpaid",
    ),
    "Counts Selection": (
        "Actuary-selected ultimates for count measures comparing Chain Ladder, Initial Expected, and Bornhuetter-Ferguson methods.",
        "Period · Age · Reported · Closed · Reported CL · Closed CL · Reported BF · Closed BF · Initial Expected · Selected Ultimate · IBNR · Unpaid",
    ),
    "Summary Diagnostics": (
        "Post-selection diagnostic metrics including ultimate severity, loss rate, and frequency.",
        "Period · Ultimate Severity · Ultimate Loss Rate · Ultimate Frequency",
    ),
    "Diagnostics": (
        "Post-selection diagnostic metrics including ultimate severity, loss rate, and frequency.",
        "Period · Ultimate Severity · Ultimate Loss Rate · Ultimate Frequency",
    ),
    "Incurred-to-Ult": (
        "Incurred loss as percent of selected ultimate by development age.",
        "Period rows · Development age columns",
    ),
    "Paid-to-Ult": (
        "Paid loss as percent of selected ultimate by development age.",
        "Period rows · Development age columns",
    ),
    "Reported-to-Ult": (
        "Reported count as percent of selected ultimate by development age.",
        "Period rows · Development age columns",
    ),
    "Closed-to-Ult": (
        "Closed count as percent of selected ultimate by development age.",
        "Period rows · Development age columns",
    ),
    "Average IBNR": (
        "Selected ultimate minus incurred at each development age showing average reserve need.",
        "Period rows · Development age columns",
    ),
    "Average Unpaid": (
        "Selected ultimate minus paid at each development age showing average unpaid reserve.",
        "Period rows · Development age columns",
    ),
}

_SHEET_DESCS_SUFFIX = {
    " CL": (
        "Chain Ladder method results with ultimate, IBNR, and unpaid reserves.",
        "Period · Age · Actual · CDF · Ultimate · IBNR · Unpaid",
    ),
    " BF": (
        "Bornhuetter-Ferguson method blending Initial Expected with emerged experience.",
        "Period · Age · Initial Expected · CDF · % Unreported/Unpaid · Unreported/Unpaid · Actual · Ultimate · IBNR · Unpaid",
    ),
    " IE": (
        "Initial Expected method using exposure and selected loss rate.",
        "Period · Age · Exposure · Selected Loss Rate · IE Ultimate",
    ),
    " - CV & Slopes": (
        "Coefficient of variation and regression slopes for age-to-age factors.",
        "Interval columns · CV row · Slope row",
    ),
}

_SHEET_DESCS_PREFIX = {
    "Diag - ": (
        "Diagnostic data for development pattern review.",
        "Period · Development age · Age-to-age factor",
    ),
}

_TRIANGLE_SHEET_NAMES = {"Incurred", "Paid", "Reported", "Closed", "Exposure"}


def _sheet_desc(name):
    if name in _SHEET_DESCS_EXACT:
        return _SHEET_DESCS_EXACT[name]
    for prefix, desc in _SHEET_DESCS_PREFIX.items():
        if name.startswith(prefix):
            return desc
    for suffix, desc in _SHEET_DESCS_SUFFIX.items():
        if name.endswith(suffix):
            return desc
    if name in _TRIANGLE_SHEET_NAMES:
        return (
            f"{name} development triangle with age-to-age factors, averages, and selected LDFs.",
            "Period rows · Age columns (triangle values) · ATA factor rows · Average rows · LDF Selection row",
        )
    return ("Analysis results.", "")


def write_notes_sheet(ws, sheet_list):
    """Write Notes sheet with metadata header and table of contents."""
    r = 1

    title_cell = ws.cell(r, 1, "Reserve Analysis - Complete Analysis")
    style_header(title_cell, "header")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 24
    r += 1

    meta = ws.cell(r, 1, "Workbook Information")
    style_header(meta, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for label, value in [
        ("Created:",     datetime.now().strftime("%B %d, %Y %I:%M %p")),
        ("Description:", "Complete actuarial reserve analysis combining selections, ultimates, and diagnostics"),
    ]:
        lbl = ws.cell(r, 1, label)
        lbl.font = LABEL_FONT; lbl.border = THIN_BORDER
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        val = ws.cell(r, 2, value)
        val.font = DATA_FONT; val.border = THIN_BORDER
        val.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        r += 1

    r += 1
    toc_hdr = ws.cell(r, 1, "Table of Contents")
    style_header(toc_hdr, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for col_num, text, width in [
        (1, "Name", 28),
        (2, "Description", 80),
    ]:
        cell = ws.cell(r, col_num, text)
        cell.font = SUBHEADER_FONT; cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = width
    r += 1

    for name, purpose, _cols in sheet_list:
        nc = ws.cell(r, 1, name)
        nc.font = DATA_FONT; nc.border = THIN_BORDER
        nc.alignment = Alignment(horizontal="left", vertical="center")
        pc = ws.cell(r, 2, purpose)
        pc.font = DATA_FONT; pc.border = THIN_BORDER
        pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    ws.freeze_panes = "A8"
