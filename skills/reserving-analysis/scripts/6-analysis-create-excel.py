# Reads projected ultimates, actuary selections, and triangle data to produce
# Analysis.xlsx — hard-coded values from source data using xlsxwriter for instant display.
# No formulas - all values computed in Python and written directly.
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6a-create-complete-analysis.py

import os
import pathlib

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import create_xlsxwriter_formats
from modules.xl_selections import (
    SKIP_ROW_LABELS as _SKIP_ROW_LABELS,
    SELECTION_LABELS as _SELECTION_LABELS,
    find_selected_values as _find_selected_values,
    find_selected_reasoning as _find_selected_reasoning,
)
from modules.xl_utils import (
    _safe, _to_py, _period_int, _copy_ws,
    measure_short_name,
    ultimates_sheet_for_measure, ultimates_col_header,
)
from modules.xl_writers import _data_cell, _write_title_and_headers, _write_headers
from modules.xl_values import UNPAID_PROXY, _has_method
from modules.xl_notes import _sheet_desc, write_notes_sheet
from modules.analysis_loaders import (
    MEASURE_TO_CATEGORY,
    load_selections,
    load_selection_reasoning,
    load_combined,
    get_exposure,
)
from modules.average_names import pretty_average_name
from modules.diagnostics_sheet import build_combined_diagnostics_sheet
from modules.calculations import (
    sanitize_value,
    calc_cl_ultimate,
    calc_bf_pct_unreported,
    calc_bf_unreported,
    calc_bf_ultimate,
    calc_ibnr,
    calc_ie_loss_rate,
    calc_total_ibnr,
    safe_divide,
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def col_letter(col_idx):
    """Convert 0-based column index to Excel column letter (A, B, C, etc.)"""
    result = ''
    while col_idx >= 0:
        result = chr(col_idx % 26 + ord('A')) + result
        col_idx = col_idx // 26 - 1
    return result

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
OUTPUT_PATH            = config.OUTPUT

OUTPUT_COMPLETE = config.BASE_DIR + "Analysis.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS  + "Chain Ladder Selections - LDFs.xlsx"
INPUT_TAIL_EXCEL    = config.SELECTIONS  + "Chain Ladder Selections - Tail.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"
INPUT_DIAGNOSTICS   = config.PROCESSED_DATA + "3_diagnostics.parquet"
INPUT_LDF_CDF_DETAIL = config.PROCESSED_DATA + "ldf-cdf-detail.parquet"  # LDF/CDF data with source tracking from 2f

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# No longer using external workbook formula references - all values hard-coded from source data

def load_ldf_cdf_detail(detail_path):
    """
    Load LDF/CDF detail data from 2f-chainladder-ultimates.py output.
    Returns {measure: DataFrame} with columns: age, ldf, cdf, source.
    Returns {} if file absent.
    """
    path = pathlib.Path(detail_path)
    if not path.exists():
        print(f"  Note: {detail_path} not found -- no LDF/CDF detail loaded")
        return {}
    
    df = pd.read_parquet(detail_path)
    detail_map = {}
    
    for measure in df['measure'].unique():
        measure_df = df[df['measure'] == measure].copy()
        detail_map[measure] = measure_df[['age', 'ldf', 'cdf', 'source']].reset_index(drop=True)
    
    print(f"  Loaded LDF/CDF detail for {len(detail_map)} measures")
    return detail_map


def load_tail_selections(tail_excel_path):
    """
    Load tail curve selections (method and reasoning) from Tail Excel file.
    Priority: User Selection → Rules-Based AI Selection → Open-Ended AI Selection.
    Returns {measure: {'method': str, 'reasoning': str, 'params': str or None}}.
    Returns {} if file absent.
    """
    path = pathlib.Path(tail_excel_path)
    if not path.exists():
        print(f"  Note: {tail_excel_path} not found -- no tail selections loaded")
        return {}
    
    wb_vals = load_workbook(tail_excel_path, data_only=True)
    tail_selections = {}
    
    for sheet_name in wb_vals.sheetnames:
        ws = wb_vals[sheet_name]
        
        # Find the Tail Curve Selection section
        in_tail_section = False
        method_col = reason_col = params_col = None
        user_row = rb_row = oe_row = None
        
        for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
            col1 = row_cells[0].value if row_cells else None
            
            # Find section header
            if col1 == "Tail Curve Selection":
                in_tail_section = True
                continue
            
            if not in_tail_section:
                continue
            
            # Find column headers (Label, Method, Reasoning, etc.)
            if col1 == "Label":
                for cell in row_cells:
                    if cell.value == "Method":
                        method_col = cell.column
                    elif cell.value == "Reasoning":
                        reason_col = cell.column
                    elif cell.value == "Parameters":
                        params_col = cell.column
                continue
            
            # Find selection rows
            if col1 == "User Selection":
                user_row = row_idx
            elif col1 == "Rules-Based AI Selection":
                rb_row = row_idx
            elif col1 == "Open-Ended AI Selection":
                oe_row = row_idx
        
        # Extract selection data with priority: User → RB → OE
        selected_method = selected_reasoning = selected_params = None
        
        for row_num in [user_row, rb_row, oe_row]:
            if row_num is None:
                continue
            
            method_val = ws.cell(row_num, method_col).value if method_col else None
            if method_val and str(method_val).strip():
                selected_method = str(method_val).strip()
                if reason_col:
                    selected_reasoning = ws.cell(row_num, reason_col).value
                    if selected_reasoning:
                        selected_reasoning = str(selected_reasoning).strip()
                if params_col:
                    selected_params = ws.cell(row_num, params_col).value
                    if selected_params:
                        selected_params = str(selected_params).strip()
                break
        
        if selected_method:
            tail_selections[sheet_name] = {
                'method': selected_method,
                'reasoning': selected_reasoning or '',
                'params': selected_params
            }
    
    wb_vals.close()
    
    if tail_selections:
        print(f"  Loaded tail selections for {len(tail_selections)} measures")
    
    return tail_selections


def load_exposure_row_map(cl_ldf_excel_path):
    """
    Load exposure row map from Chain Ladder Selections - LDFs.xlsx.
    Returns {period: row_number} for building formula references.
    Returns {} if file absent or Exposure sheet not found.
    """
    path = pathlib.Path(cl_ldf_excel_path)
    if not path.exists():
        print(f"  Note: {cl_ldf_excel_path} not found -- no exposure row map loaded")
        return {}
    
    wb = load_workbook(cl_ldf_excel_path, data_only=True)
    if "Exposure" not in wb.sheetnames:
        wb.close()
        print("  Note: Exposure sheet not found in Chain Ladder LDFs workbook")
        return {}
    
    ws = wb["Exposure"]
    row_map = {}
    
    # Find Period column (should be in row 2)
    period_col = None
    for cell in ws[2]:
        if cell.value == "Period":
            period_col = cell.column
            break
    
    if period_col is None:
        wb.close()
        print("  Note: Period column not found in Exposure sheet")
        return {}
    
    # Read period values starting from row 3 (data starts after title and headers)
    for row_idx in range(3, ws.max_row + 1):
        period_val = ws.cell(row=row_idx, column=period_col).value
        if period_val is not None:
            row_map[str(period_val)] = row_idx
    
    wb.close()
    print(f"  Loaded exposure row map for {len(row_map)} periods")
    return row_map


# ── Generated sheet writers ───────────────────────────────────────────────────

def _data_cell_xlw(ws, r, c, value, num_fmt=None, fmt_obj=None):
    """Write data cell for xlsxwriter. r and c are 1-based (Excel coords)."""
    # Treat string 'None' as None (can happen with missing data)
    if value == 'None' or value == 'none':
        value = None
    
    if value is None or (isinstance(value, float) and np.isnan(value)):
        if fmt_obj:
            ws.write_blank(r-1, c-1, None, fmt_obj)
        return
    
    # Detect if value is numeric (handles both actual numbers and numeric strings)
    is_numeric = isinstance(value, (int, float, np.integer, np.floating))
    
    # Try converting string to number to avoid "number stored as text" warnings
    if not is_numeric and isinstance(value, str):
        try:
            value = float(value)
            is_numeric = True
        except (ValueError, TypeError):
            pass
    
    if fmt_obj is None:
        if is_numeric:
            fmt_obj = ws.book.add_format({'num_format': num_fmt or '#,##0', 'align': 'right'})
        else:
            fmt_obj = ws.book.add_format({'align': 'left'})
    
    if is_numeric:
        ws.write_number(r-1, c-1, float(value), fmt_obj)
    else:
        ws.write(r-1, c-1, str(value), fmt_obj)


def _write_headers_xlw(ws, headers, fmt_dict, col_width=18):
    """Write header row for xlsxwriter. Returns 2 (first data row, 1-based)."""
    for c, text in enumerate(headers, 1):
        ws.write(0, c-1, text, fmt_dict.get('subheader'))
        ws.set_column(c-1, c-1, col_width)
    ws.freeze_panes(1, 0)  # Freeze row 1
    return 2


def write_notes_sheet_xlw(ws, sheet_list, fmt_dict):
    """Write Notes sheet with table of contents only."""
    r = 0  # 0-based row for xlsxwriter
    
    # Column headers
    ws.write(r, 0, "Sheet", fmt_dict.get('subheader_left'))
    ws.write(r, 1, "Description", fmt_dict.get('subheader_left'))
    ws.set_column(0, 0, 28)
    ws.set_column(1, 1, 80)
    r += 1
    
    # Sheet list
    data_fmt = fmt_dict.get('label')
    for name, purpose, _cols in sheet_list:
        ws.write(r, 0, name, data_fmt)
        ws.write(r, 1, purpose, data_fmt)
        r += 1
    
    ws.freeze_panes(1, 0)  # Freeze header row


def _ult_ref(ult_col_map, sheet, col_header, row, ult_wb=None):
    """Return value from Ultimates workbook (not a formula)."""
    if ult_wb is None:
        return None
    col_idx = ult_col_map.get((sheet, col_header))
    if not col_idx:
        return None
    try:
        ws = ult_wb[sheet]
        from openpyxl.utils import column_index_from_string
        col_num = column_index_from_string(col_idx)
        val = ws.cell(row=row, column=col_num).value
        return val
    except:
        return None


def _exp_ref(exp_row_map, period, cl_ldf_wb=None):
    """Return exposure value from Chain Ladder Selections workbook (not a formula)."""
    if cl_ldf_wb is None:
        return None
    row = exp_row_map.get(str(period))
    if not row:
        return None
    try:
        ws = cl_ldf_wb['Exposure']
        val = ws.cell(row=row, column=2).value  # Column B = 2
        return val
    except:
        return None


def _build_cl_ldfs_col_row_maps(wb_cl, measure_sheet_name):
    """
    Build column and row maps for referencing Chain Ladder Selections - LDFs.xlsx.
    Returns (col_map, user_row, rb_row, user_reason_row, rb_reason_row) where:
      col_map: {column_header: column_letter}
      user_row: row number of "User Selection"
      rb_row: row number of "Rules-Based AI Selection"
      user_reason_row: row number of "User Reasoning"
      rb_reason_row: row number of "Rules-Based AI Reasoning"
    """
    if measure_sheet_name not in wb_cl.sheetnames:
        return {}, None, None, None, None
    
    ws = wb_cl[measure_sheet_name]
    col_map = {}
    user_row = None
    rb_row = None
    user_reason_row = None
    rb_reason_row = None
    
    # Find column headers and selection rows
    for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
        col1 = row_cells[0].value if row_cells else None
        
        # Track header rows (usually None in col A)
        if col1 is None and row_idx <= 10:  # Headers in first few rows
            for cell in row_cells[1:]:
                if cell.value is not None:
                    col_map[str(cell.value)] = get_column_letter(cell.column)
        
        # Track selection rows
        if col1 == "User Selection":
            user_row = row_idx
        elif col1 == "Rules-Based AI Selection":
            rb_row = row_idx
        elif col1 == "User Reasoning":
            user_reason_row = row_idx
        elif col1 == "Rules-Based AI Reasoning":
            rb_reason_row = row_idx
    
    return col_map, user_row, rb_row, user_reason_row, rb_reason_row



def _tri_col_map_from_df(triangles_df, measure, max_age=None):
    """Build {age_str: col_letter} for the triangle sheet from parquet data.
    Triangle sheet col A = period, so ages start at col B (index 2).
    Ages sorted ascending match the order written by _copy_ws_filtered.
    max_age caps the map to the cutoff age — add_tail_to_triangle_ws deletes
    data columns beyond the cutoff, so referencing those cols returns empty.
    """
    tri_m = triangles_df[triangles_df["measure"].astype(str) == measure]
    if tri_m.empty:
        return {}
    ages = sorted(tri_m["age"].dropna().unique(), key=lambda a: float(str(a)))
    if max_age is not None:
        ages = [a for a in ages if float(str(a)) <= max_age]
    return {str(int(float(str(a)))): get_column_letter(i + 2) for i, a in enumerate(ages)}


# ── Generic Method Writer (Phase 2) ─────────────────────────────────────────

def _get_bf_labels(measure):
    """Get BF column labels based on measure type."""
    is_paid = measure == "Paid Loss"
    is_closed = measure == "Closed Count"
    is_count = "Count" in measure
    
    if is_paid:
        return "% Unpaid", "Unpaid", measure_short_name(measure)
    elif is_closed:
        return "% Unpaid", "Unpaid Counts", "Closed Counts"
    elif is_count:  # Reported Count
        return "% Unreported", "Unreported Counts", "Reported Counts"
    else:  # Incurred Loss and others
        return "% Unreported", "$ Unreported", measure_short_name(measure)


def write_method_sheet(gen_wb, method_type, combined, measure, fmt_dict, **kwargs):
    """
    Generic method sheet writer - config-driven columns.
    
    Args:
        gen_wb: xlsxwriter workbook
        method_type: 'CL' or 'BF'
        combined: DataFrame with all method results
        measure: Measure name
        fmt_dict: Format dictionary
        **kwargs: Additional context (ult_col_map, tri_col_map, etc.)
    """
    short_name = measure_short_name(measure)
    
    # Prepare data
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')
    fmt_pct = fmt_dict.get('data_pct')
    fmt_period = fmt_dict.get('data_period')
    
    if method_type == 'CL':
        # Chain Ladder configuration
        sheet_name = f"{short_name} CL"[:31]
        ws = gen_wb.add_worksheet(sheet_name)
        headers = ["Accident Period", "Current Age", short_name, "CDF", "Ultimate", "IBNR", "Unpaid"]
        _write_headers_xlw(ws, headers, fmt_dict)
        
        # Write data rows
        for r, (_, row) in enumerate(sub.iterrows(), start=2):
            actual_val = sanitize_value(row["actual"])
            cdf_val = sanitize_value(row.get("cdf"))
            ultimate_val = calc_cl_ultimate(actual_val, cdf_val)
            ibnr_val = calc_ibnr(ultimate_val, actual_val, measure, combined, row["period"])
            
            _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
            _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 3, actual_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 4, cdf_val, fmt_obj=fmt_dec)
            _data_cell_xlw(ws, r, 5, ultimate_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 6, ibnr_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 7, ibnr_val, fmt_obj=fmt_num)  # Unpaid = IBNR
        
        # Write totals
        t = 3 + len(sub)
        _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
        
        total_actual = sub["actual"].sum() if pd.notna(sub["actual"].sum()) else 0
        total_ultimate = sub.apply(lambda row: calc_cl_ultimate(row["actual"], row.get("cdf")) or 0, axis=1).sum()
        total_ultimate = total_ultimate if pd.notna(total_ultimate) else 0
        total_ibnr = calc_total_ibnr(total_ultimate, total_actual, measure, combined)
        
        _data_cell_xlw(ws, t, 3, total_actual, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 5, total_ultimate, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 6, total_ibnr, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 7, total_ibnr, fmt_obj=fmt_num)
    
    elif method_type == 'BF':
        # Bornhuetter-Ferguson configuration
        sheet_name = f"{short_name} BF"[:31]
        ws = gen_wb.add_worksheet(sheet_name)
        
        pct_label, unrep_label, actual_label = _get_bf_labels(measure)
        headers = ["Accident Period", "Current Age", "Initial Expected", "CDF", 
                   pct_label, unrep_label, actual_label, "Ultimate", "IBNR", "Unpaid"]
        _write_headers_xlw(ws, headers, fmt_dict)
        
        # Write data rows
        for r, (_, row) in enumerate(sub.iterrows(), start=2):
            ie_val = sanitize_value(row.get("ultimate_ie"))
            cdf_val = sanitize_value(row.get("cdf"))
            pct_unrep = calc_bf_pct_unreported(cdf_val)
            unrep_val = calc_bf_unreported(ie_val, pct_unrep)
            actual_val = sanitize_value(row["actual"])
            ultimate_bf = calc_bf_ultimate(unrep_val, actual_val)
            ibnr_val = calc_ibnr(ultimate_bf, actual_val, measure, combined, row["period"])
            
            _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
            _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 3, ie_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 4, cdf_val, fmt_obj=fmt_dec)
            _data_cell_xlw(ws, r, 5, pct_unrep, fmt_obj=fmt_pct)
            _data_cell_xlw(ws, r, 6, unrep_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 7, actual_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 8, ultimate_bf, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 9, ibnr_val, fmt_obj=fmt_num)
            _data_cell_xlw(ws, r, 10, ibnr_val, fmt_obj=fmt_num)  # Unpaid = IBNR
        
        # Write totals
        t = 3 + len(sub)
        _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
        
        total_ie = sub["ultimate_ie"].sum() if "ultimate_ie" in sub.columns else 0
        total_ie = total_ie if pd.notna(total_ie) else 0
        total_unrep = sub.apply(
            lambda row: calc_bf_unreported(row.get("ultimate_ie"), calc_bf_pct_unreported(row.get("cdf"))) or 0,
            axis=1
        ).sum()
        total_unrep = total_unrep if pd.notna(total_unrep) else 0
        total_actual = sub["actual"].sum() if pd.notna(sub["actual"].sum()) else 0
        total_ultimate_bf = total_unrep + total_actual
        total_ibnr = calc_total_ibnr(total_ultimate_bf, total_actual, measure, combined)
        total_pct_unrep = safe_divide(total_unrep, total_ie) or 0
        
        _data_cell_xlw(ws, t, 3, total_ie, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 5, total_pct_unrep, fmt_obj=fmt_pct)
        _data_cell_xlw(ws, t, 6, total_unrep, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 7, total_actual, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 8, total_ultimate_bf, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 9, total_ibnr, fmt_obj=fmt_num)
        _data_cell_xlw(ws, t, 10, total_ibnr, fmt_obj=fmt_num)
    
    else:
        raise ValueError(f"Unknown method_type: {method_type}. Expected 'CL' or 'BF'.")


# ── Convenience Method Writers ──────────────────────────────────────────────

def write_method_cl(gen_wb, combined, measure, ult_col_map, fmt_dict, tri_col_map=None):
    """Write Chain Ladder method sheet (legacy wrapper - calls generic writer)."""
    write_method_sheet(gen_wb, 'CL', combined, measure, fmt_dict, 
                      ult_col_map=ult_col_map, tri_col_map=tri_col_map)

def write_method_bf(gen_wb, combined, measure, ult_col_map, fmt_dict, tri_col_maps=None, ult_wb=None):
    """Write Bornhuetter-Ferguson method sheet (legacy wrapper - calls generic writer)."""
    write_method_sheet(gen_wb, 'BF', combined, measure, fmt_dict,
                      ult_col_map=ult_col_map, tri_col_maps=tri_col_maps, ult_wb=ult_wb)


def write_method_ie(gen_wb, combined, measure, exp_row_map, ult_col_map, fmt_dict, ult_wb=None, cl_ldf_wb=None):
    """Write Initial Expected method sheet with xlsxwriter (values from source data)."""
    short_name = measure_short_name(measure)
    ws = gen_wb.add_worksheet("IE")
    headers = ["Accident Period", "Current Age", "Exposure", "IE Ultimate", "Selected Loss Rate"]
    _write_headers_xlw(ws, headers, fmt_dict)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    ult_sheet = ultimates_sheet_for_measure(measure)
    ie_hdr    = ultimates_col_header(measure, "ie")

    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')
    fmt_period = fmt_dict.get('data_period')

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)

        # Exposure - read value from Chain Ladder Selections workbook
        exp_val = _exp_ref(exp_row_map, row["period"], cl_ldf_wb)
        _data_cell_xlw(ws, r, 3, exp_val, fmt_obj=fmt_num)

        # IE Ultimate - sanitize using centralized function
        ie_val = sanitize_value(row.get("ultimate_ie"))
        _data_cell_xlw(ws, r, 4, ie_val, fmt_obj=fmt_num)
        
        # Selected Loss Rate using centralized function
        loss_rate = calc_ie_loss_rate(ie_val, exp_val)
        _data_cell_xlw(ws, r, 5, loss_rate, fmt_obj=fmt_dec)


def write_selection_grouped(gen_wb, combined, measures_group, title, ult_col_map, fmt_dict, tri_col_maps=None, exp_map=None):
    """Write selection summary sheet with xlsxwriter (formulas + cached values)."""
    tri_col_maps = tri_col_maps or {}
    exp_map = exp_map or {}
    ws = gen_wb.add_worksheet(title)
    
    # columns e.g.: Accident Period, Current Age, Incurred, Paid, Incurred CL, Paid CL, Initial Expected, Incurred BF, Paid BF, Selected Ultimate, IBNR, Unpaid
    # Column ordering convention: CL, IE, BF
    
    # Find active measures in this group
    active_m = [
        m for m in measures_group
        if m in combined["measure"].unique()
        and combined[combined["measure"] == m]["actual"].notna().any()
    ]
    if not active_m:
        return
        
    main_m = active_m[0]
    sub = combined[combined["measure"] == main_m].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    short_names = [measure_short_name(m) for m in active_m]
    bf_m  = [m for m in active_m if _has_method(combined, m, "ultimate_bf")]
    has_group_ie = any(_has_method(combined, m, "ultimate_ie") for m in active_m)

    ie_measure = next((m for m in active_m if _has_method(combined, m, "ultimate_ie")), active_m[0])
    ie_short   = measure_short_name(ie_measure)
    
    # Check if proxy measure exists with actual data for Unpaid column
    # Unpaid = Selected Ultimate - Proxy Measure Actual
    # For Incurred Loss: proxy = Paid Loss (always exists)
    # For Reported Count: proxy = Closed Count (may not exist or may have no data)
    # If proxy has no data, Unpaid column is omitted (IBNR sufficient)
    proxy_measure = UNPAID_PROXY.get(main_m)
    has_proxy = (proxy_measure and 
                 proxy_measure in combined["measure"].unique() and
                 combined[combined["measure"] == proxy_measure]["actual"].notna().any())

    # Build headers in CL, IE, BF order
    headers = ["Accident Period", "Current Age"]
    for s in short_names:
        headers.append(s)
    for s in short_names:
        headers.append(f"{s} CL")
    if has_group_ie:
        headers.append("Initial Expected")
    for m in bf_m:
        headers.append(f"{measure_short_name(m)} BF")
    headers.append("Selected Ultimate")
    headers.append("IBNR")
    if has_proxy:
        headers.append("Unpaid")
    headers.append("Selected Reasoning")
    
    # Add diagnostic columns based on sheet type
    is_count_selection = "Count" in title
    is_loss_selection = "Loss" in title
    has_exposure = bool(exp_map)
    
    if is_loss_selection and has_exposure:
        headers.append("Ultimate Loss Rate")
    if is_count_selection:
        headers.append("Ultimate Severity")
        if has_exposure:
            headers.append("Ultimate Frequency")

    _write_headers_xlw(ws, headers, fmt_dict)
    ws.set_column(len(headers)-2, len(headers)-2, 40)  # Reasoning column width (now second-to-last or further back)

    fmt_num = fmt_dict.get('data_num')
    fmt_text = fmt_dict.get('label')
    fmt_period = fmt_dict.get('data_period')
    fmt_dec = fmt_dict.get('data_ldf')  # 4 decimals for IBNR

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell_xlw(ws, r, 1, row["period_int"], fmt_obj=fmt_period)
        _data_cell_xlw(ws, r, 2, row["current_age"], fmt_obj=fmt_num)

        col_idx = 3
        # Actuals - use LOOKUP to find last value in triangle row for each measure
        for m in active_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            actual_val = m_row["actual"].iloc[0] if len(m_row) > 0 else None
            # Sanitize - .iloc[0] can return NaN
            actual_val = actual_val if pd.notna(actual_val) else None
            _data_cell_xlw(ws, r, col_idx, actual_val, fmt_obj=fmt_num)
            col_idx += 1
        
        # CL ultimates
        for m in active_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            cl_ult = m_row["ultimate_cl"].iloc[0] if len(m_row) > 0 and "ultimate_cl" in m_row.columns else None
            # Sanitize - .iloc[0] can return NaN
            cl_ult = cl_ult if pd.notna(cl_ult) else None
            _data_cell_xlw(ws, r, col_idx, cl_ult, fmt_obj=fmt_num)
            col_idx += 1
        
        # IE ultimate (if available)
        if has_group_ie:
            ie_val = row.get("ultimate_ie")
            # Sanitize - could be NaN
            ie_val = ie_val if pd.notna(ie_val) else None
            _data_cell_xlw(ws, r, col_idx, ie_val, fmt_obj=fmt_num)
            col_idx += 1
        
        # BF ultimates
        for m in bf_m:
            m_row = combined[(combined["measure"] == m) & (combined["period"] == row["period"])]
            bf_ult = m_row["ultimate_bf"].iloc[0] if len(m_row) > 0 and "ultimate_bf" in m_row.columns else None
            # Sanitize - .iloc[0] can return NaN
            bf_ult = bf_ult if pd.notna(bf_ult) else None
            _data_cell_xlw(ws, r, col_idx, bf_ult, fmt_obj=fmt_num)
            col_idx += 1
        
        # Selected Ultimate → use value from combined data
        # Cache with selected_ultimate from combined
        sel_val = row.get("selected_ultimate")
        # Sanitize - could be NaN
        sel_val = sel_val if pd.notna(sel_val) else None
        _data_cell_xlw(ws, r, col_idx, sel_val, fmt_obj=fmt_num)
        selected_col_idx = col_idx
        col_idx += 1
        
        # IBNR = Selected - Main Measure Actual
        main_actual = combined[(combined["measure"] == main_m) & (combined["period"] == row["period"])]["actual"].iloc[0] if len(combined[(combined["measure"] == main_m) & (combined["period"] == row["period"])]) > 0 else None
        main_actual = main_actual if pd.notna(main_actual) else None
        ibnr_val = (sel_val - main_actual) if pd.notna(sel_val) and pd.notna(main_actual) else None
        _data_cell_xlw(ws, r, col_idx, ibnr_val, fmt_obj=fmt_num)
        col_idx += 1
        
        # Unpaid = Selected - Proxy Measure Actual (only if proxy has data)
        # For counts: if no Closed Count data provided, Unpaid col omitted
        # For losses: Paid Loss always exists, so Unpaid col always included
        if has_proxy:
            proxy_actual = combined[(combined["measure"] == proxy_measure) & (combined["period"] == row["period"])]["actual"].iloc[0] if len(combined[(combined["measure"] == proxy_measure) & (combined["period"] == row["period"])]) > 0 else None
            proxy_actual = proxy_actual if pd.notna(proxy_actual) else None
            unpaid_val = (sel_val - proxy_actual) if pd.notna(sel_val) and pd.notna(proxy_actual) else None
            _data_cell_xlw(ws, r, col_idx, unpaid_val, fmt_obj=fmt_num)
            col_idx += 1
        
        # Selected Reasoning → use value from combined data
        reason_val = row.get("selected_reasoning", "")
        _data_cell_xlw(ws, r, col_idx, reason_val, fmt_obj=fmt_text)
        col_idx += 1
        
        # Diagnostic columns
        period_str = str(row["period"])
        exposure = exp_map.get(period_str)
        
        if is_loss_selection and has_exposure:
            # Ultimate Loss Rate = Selected Ultimate / Exposure
            loss_rate = (sel_val / exposure) if pd.notna(sel_val) and pd.notna(exposure) and exposure != 0 else None
            _data_cell_xlw(ws, r, col_idx, loss_rate, fmt_obj=fmt_dec)
            col_idx += 1
        
        if is_count_selection:
            # Ultimate Severity = Incurred Ultimate / Reported Ultimate
            # Need to get Incurred Loss selected ultimate for this period
            inc_loss_row = combined[(combined["measure"] == "Incurred Loss") & (combined["period"] == row["period"])]
            inc_ult = inc_loss_row["selected_ultimate"].iloc[0] if len(inc_loss_row) > 0 else None
            inc_ult = inc_ult if pd.notna(inc_ult) else None
            
            severity = (inc_ult / sel_val) if pd.notna(inc_ult) and pd.notna(sel_val) and sel_val != 0 else None
            _data_cell_xlw(ws, r, col_idx, severity, fmt_obj=fmt_dec)
            col_idx += 1
            
            if has_exposure:
                # Ultimate Frequency = Selected Ultimate / Exposure
                frequency = (sel_val / exposure) if pd.notna(sel_val) and pd.notna(exposure) and exposure != 0 else None
                _data_cell_xlw(ws, r, col_idx, frequency, fmt_obj=fmt_dec)
                col_idx += 1

    # Totals row — calculate totals from dataframe
    n_cols = 2 + len(active_m) * 2 + len(bf_m) + (1 if has_group_ie else 0) + 3
    t = 3 + len(sub)
    _data_cell_xlw(ws, t, 1, "Total", fmt_obj=fmt_dict.get('label'))
    
    # Start at column 3 (actuals for each measure)
    col_idx = 3
    for m in active_m:
        m_data = combined[combined["measure"] == m]
        total_actual = m_data["actual"].sum() if len(m_data) > 0 else 0
        _data_cell_xlw(ws, t, col_idx, total_actual if pd.notna(total_actual) else None, fmt_obj=fmt_num)
        col_idx += 1
    
    # CL ultimates for each measure
    for m in active_m:
        m_data = combined[combined["measure"] == m]
        total_cl = m_data["ultimate_cl"].sum() if len(m_data) > 0 and "ultimate_cl" in m_data.columns else 0
        _data_cell_xlw(ws, t, col_idx, total_cl if pd.notna(total_cl) else None, fmt_obj=fmt_num)
        col_idx += 1
    
    # IE ultimate (if available)
    if has_group_ie:
        total_ie = sub["ultimate_ie"].sum() if "ultimate_ie" in sub.columns else 0
        _data_cell_xlw(ws, t, col_idx, total_ie if pd.notna(total_ie) else None, fmt_obj=fmt_num)
        col_idx += 1
    
    # BF ultimates for each BF measure
    for m in bf_m:
        m_data = combined[combined["measure"] == m]
        total_bf = m_data["ultimate_bf"].sum() if len(m_data) > 0 and "ultimate_bf" in m_data.columns else 0
        _data_cell_xlw(ws, t, col_idx, total_bf if pd.notna(total_bf) else None, fmt_obj=fmt_num)
        col_idx += 1
    
    # Selected ultimate
    total_sel = sub["selected_ultimate"].sum() if "selected_ultimate" in sub.columns else 0
    _data_cell_xlw(ws, t, col_idx, total_sel if pd.notna(total_sel) else None, fmt_obj=fmt_num)
    col_idx += 1
    
    # IBNR total (Selected - Main Measure Actual)
    main_m_data = combined[combined["measure"] == main_m]
    total_main_actual = main_m_data["actual"].sum() if len(main_m_data) > 0 else 0
    total_main_actual = total_main_actual if pd.notna(total_main_actual) else 0
    total_ibnr = (total_sel - total_main_actual) if pd.notna(total_sel) else None
    _data_cell_xlw(ws, t, col_idx, total_ibnr, fmt_obj=fmt_num)
    col_idx += 1
    
    # Unpaid total (only if proxy measure has data)
    # Flexibility: Count sheets may have no Closed Count data → no Unpaid col
    if has_proxy:
        proxy_m_data = combined[combined["measure"] == proxy_measure]
        total_proxy_actual = proxy_m_data["actual"].sum() if len(proxy_m_data) > 0 else 0
        total_proxy_actual = total_proxy_actual if pd.notna(total_proxy_actual) else 0
        total_unpaid = (total_sel - total_proxy_actual) if pd.notna(total_sel) else None
        _data_cell_xlw(ws, t, col_idx, total_unpaid, fmt_obj=fmt_num)
        col_idx += 1


# ── Triangle Section Writers (Phase 3) ───────────────────────────────────────

def _write_triangle_section_header(ws_xlw, dst_row, title, age_headers, fmt_dict):
    """Write a section header (e.g., 'Age-to-Age Factors') with age columns."""
    ws_xlw.write(dst_row, 0, title, fmt_dict.get('subheader'))
    for col_idx, age_val in enumerate(age_headers, start=1):
        if age_val not in (None, ""):
            ws_xlw.write(dst_row, col_idx, age_val, fmt_dict.get('subheader'))
    return dst_row + 1


def _copy_triangle_row(ws_xlw, dst_row, src_row_data, fmt_dict, is_triangle_data=False, is_averages=False):
    """Copy a single row from source data to xlsxwriter worksheet with appropriate formatting."""
    col1_label = src_row_data.get('label')
    
    # Write label column (column A)
    if is_averages and col1_label:
        label_text = pretty_average_name(str(col1_label))
    else:
        label_text = col1_label
    ws_xlw.write(dst_row, 0, label_text, fmt_dict.get('label'))
    
    # Write data columns
    for col_idx, value in src_row_data.get('values', {}).items():
        if value is None or value == "":
            fmt = fmt_dict.get('data_num') if is_triangle_data else fmt_dict.get('data_ldf')
            ws_xlw.write_blank(dst_row, col_idx, None, fmt)
        elif isinstance(value, (int, float)):
            data_fmt = fmt_dict.get('data_num') if is_triangle_data else fmt_dict.get('data_ldf')
            ws_xlw.write_number(dst_row, col_idx, float(value), data_fmt)
        else:
            try:
                num_val = float(value)
                data_fmt = fmt_dict.get('data_num') if is_triangle_data else fmt_dict.get('data_ldf')
                ws_xlw.write_number(dst_row, col_idx, num_val, data_fmt)
            except (ValueError, TypeError):
                ws_xlw.write(dst_row, col_idx, value, None)
    
    return dst_row + 1


def _write_ldf_cdf_rows(ws_xlw, dst_row, ldf_cdf_detail, age_headers, last_col, gen_wb, fmt_dict):
    """Write LDF and CDF rows with source marking (empirical vs fitted)."""
    if ldf_cdf_detail is None or ldf_cdf_detail.empty:
        return dst_row
    
    # Add LDF row
    ldf_row = dst_row
    ws_xlw.write(ldf_row, 0, "LDF", fmt_dict.get('label'))
    
    # Create fitted format for highlighting fitted LDFs
    fitted_fmt = gen_wb.add_format({
        'num_format': '#,##0.0000',
        'align': 'right',
        'italic': True,
        'bg_color': '#FFF2CC'  # Light yellow background
    })
    
    # Write LDF values
    for col_idx, age_val in enumerate(age_headers, start=1):
        if col_idx > last_col or age_val is None:
            continue
        try:
            age_int = int(float(str(age_val)))
            age_row = ldf_cdf_detail[ldf_cdf_detail['age'].astype(int) == age_int]
            
            if not age_row.empty:
                ldf_val = age_row.iloc[0]['ldf']
                source = age_row.iloc[0]['source']
                
                if source == 'fitted' and pd.notna(ldf_val):
                    ws_xlw.write_number(ldf_row, col_idx, float(ldf_val), fitted_fmt)
                elif pd.notna(ldf_val):
                    ws_xlw.write_number(ldf_row, col_idx, float(ldf_val), fmt_dict.get('data_ldf'))
                else:
                    ws_xlw.write_blank(ldf_row, col_idx, None, fmt_dict.get('data_ldf'))
            else:
                ws_xlw.write_blank(ldf_row, col_idx, None, fmt_dict.get('data_ldf'))
        except (ValueError, TypeError):
            ws_xlw.write_blank(ldf_row, col_idx, None, fmt_dict.get('data_ldf'))
    
    dst_row += 1
    
    # Add CDF row
    cdf_row = dst_row
    ws_xlw.write(cdf_row, 0, "CDF", fmt_dict.get('label'))
    
    # Write CDF values
    for col_idx, age_val in enumerate(age_headers, start=1):
        if col_idx > last_col or age_val is None:
            continue
        try:
            age_int = int(float(str(age_val)))
            age_row = ldf_cdf_detail[ldf_cdf_detail['age'].astype(int) == age_int]
            
            if not age_row.empty:
                cdf_val = age_row.iloc[0]['cdf']
                if pd.notna(cdf_val):
                    ws_xlw.write_number(cdf_row, col_idx, float(cdf_val), fmt_dict.get('data_ldf'))
                else:
                    ws_xlw.write_blank(cdf_row, col_idx, None, fmt_dict.get('data_ldf'))
            else:
                ws_xlw.write_blank(cdf_row, col_idx, None, fmt_dict.get('data_ldf'))
        except (ValueError, TypeError):
            ws_xlw.write_blank(cdf_row, col_idx, None, fmt_dict.get('data_ldf'))
    
    return dst_row + 1


def _write_tail_info(ws_xlw, dst_row, tail_selections, last_col, gen_wb, fmt_dict):
    """Write tail method information rows."""
    if not tail_selections:
        return dst_row
    
    # Blank row before tail info
    dst_row += 1
    
    # Tail Method row
    ws_xlw.write(dst_row, 0, "Tail Method", fmt_dict.get('label'))
    tail_method = tail_selections.get('method', '')
    if tail_method:
        ws_xlw.write(dst_row, 1, tail_method, fmt_dict.get('label'))
    dst_row += 1
    
    # Tail Parameters row (if applicable)
    tail_params = tail_selections.get('params')
    if tail_params and str(tail_params).strip():
        ws_xlw.write(dst_row, 0, "Tail Parameters", fmt_dict.get('label'))
        ws_xlw.write(dst_row, 1, str(tail_params), fmt_dict.get('label'))
        dst_row += 1
    
    # Tail Reasoning row
    tail_reasoning = tail_selections.get('reasoning', '')
    if tail_reasoning:
        ws_xlw.write(dst_row, 0, "Tail Reasoning", fmt_dict.get('label'))
        reasoning_fmt = gen_wb.add_format({
            'text_wrap': True,
            'align': 'left',
            'valign': 'top'
        })
        ws_xlw.write(dst_row, 1, tail_reasoning, reasoning_fmt)
        ws_xlw.set_row(dst_row, max(15 * (len(tail_reasoning) // 80 + 1), 30))
        if last_col > 1:
            ws_xlw.merge_range(dst_row, 1, dst_row, min(last_col, 5), tail_reasoning, reasoning_fmt)
        dst_row += 1
    
    return dst_row


def create_triangle_sheets_xlw(gen_wb, measures, fmt_dict, ldf_cdf_detail=None, combined_df=None, tail_selections=None):
    """
    Create triangle sheets with xlsxwriter - simplified with helper functions.
    
    Reads from Chain Ladder Selections - LDFs.xlsx and writes:
    1. Triangle data section (Period header + data rows)
    2. Age-to-Age Factors section
    3. Averages section
    4. LDF Selections section (Selected + Selected Reasoning rows)
    5. LDF and CDF rows (with source marking)
    6. Tail information (if applicable)
    
    Args:
        gen_wb: xlsxwriter workbook
        measures: list of measure names
        fmt_dict: formatting dictionary
        ldf_cdf_detail: dict mapping measure to DataFrame with columns: age, ldf, cdf, source
        combined_df: DataFrame with CDF values (not used - CDFs come from detail)
        tail_selections: dict mapping measure to {'method': str, 'reasoning': str, 'params': str or None}
    """
    ldf_cdf_detail = ldf_cdf_detail or {}
    tail_selections = tail_selections or {}
    
    if not pathlib.Path(INPUT_CL_EXCEL).exists():
        return
    
    # Load source workbooks
    wb_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
    wb_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)
    
    for measure in measures:
        if measure not in wb_form.sheetnames:
            continue
        
        short_name = measure_short_name(measure)[:31]
        ws_xlw = gen_wb.add_worksheet(short_name)
        ws_src_form = wb_form[measure]
        ws_src_vals = wb_vals[measure]
        
        # State tracking
        dst_row = 0
        current_section = None  # 'triangle', 'ata', 'averages', 'selections'
        age_headers = []
        selected_row_written = False
        
        # Build column/row maps for reading selections
        col_map, user_row, rb_row, user_reason_row, rb_reason_row = _build_cl_ldfs_col_row_maps(wb_form, measure)
        
        # Process source rows
        for row_idx, src_cells in enumerate(ws_src_form.iter_rows()):
            col1 = src_cells[0].value if src_cells else None
            
            # Skip AI option rows
            if col1 in _SKIP_ROW_LABELS:
                continue
            
            # Detect section transitions
            if col1 == "Period":
                if current_section is None:
                    # First Period = Triangle section
                    current_section = 'triangle'
                    # Extract age headers
                    age_headers = [cell.value for cell in src_cells[1:] if cell.value not in (None, "")]
                elif current_section == 'triangle':
                    # Second Period after blank = Age-to-Age Factors section
                    dst_row = _write_triangle_section_header(ws_xlw, dst_row + 1, "Age-to-Age Factors", age_headers, fmt_dict)
                    current_section = 'ata'
                    continue
            elif col1 == "Metric":
                # Averages section
                dst_row = _write_triangle_section_header(ws_xlw, dst_row + 1, "Averages", age_headers, fmt_dict)
                current_section = 'averages'
                continue
            elif col1 in _SELECTION_LABELS and not selected_row_written:
                # LDF Selections section
                dst_row += 1  # Blank row before section
                ws_xlw.write(dst_row, 0, "LDF Selections", fmt_dict.get('subheader'))
                for col_idx, age_val in enumerate(age_headers, start=1):
                    if age_val not in (None, ""):
                        ws_xlw.write(dst_row, col_idx, age_val, fmt_dict.get('subheader'))
                dst_row += 1
                current_section = 'selections'
                
                # Write "Selected" row (User selection > RB-AI selection)
                ws_xlw.write(dst_row, 0, "Selected", fmt_dict.get('label'))
                for col_idx, src_cell in enumerate(src_cells[1:], start=1):
                    val_cell = ws_src_vals.cell(src_cell.row, src_cell.column)
                    cached_val = val_cell.value if val_cell.value not in (None, "") else None
                    if cached_val is not None:
                        ws_xlw.write_number(dst_row, col_idx, float(cached_val), fmt_dict.get('data_ldf'))
                    else:
                        ws_xlw.write_blank(dst_row, col_idx, None, fmt_dict.get('data_ldf'))
                dst_row += 1
                
                # Write "Selected Reasoning" row
                if user_reason_row and rb_reason_row:
                    ws_xlw.write(dst_row, 0, "Selected Reasoning", fmt_dict.get('label'))
                    for col_idx, src_cell in enumerate(src_cells[1:], start=1):
                        user_val = ws_src_vals.cell(user_reason_row, src_cell.column).value
                        user_reason = user_val if user_val not in (None, "") else None
                        
                        if user_reason:
                            ws_xlw.write(dst_row, col_idx, str(user_reason), fmt_dict.get('label'))
                        else:
                            rb_val = ws_src_vals.cell(rb_reason_row, src_cell.column).value
                            rb_reason = rb_val if rb_val not in (None, "") else None
                            if rb_reason:
                                ws_xlw.write(dst_row, col_idx, str(rb_reason), fmt_dict.get('label'))
                            else:
                                ws_xlw.write_blank(dst_row, col_idx, None, fmt_dict.get('label'))
                    dst_row += 1
                
                selected_row_written = True
                continue
            
            # Skip if we've written selections (don't copy other selection rows)
            if selected_row_written and col1 in _SELECTION_LABELS:
                continue
            
            # Skip blank rows between sections
            if col1 in (None, ""):
                continue
            
            # Copy data row
            is_triangle_row = False
            if col1 and isinstance(col1, (int, float)):
                try:
                    year = int(col1)
                    if 1990 <= year <= 2100:
                        is_triangle_row = True
                except (ValueError, TypeError):
                    pass
            
            # Build row data structure
            row_data = {'label': col1, 'values': {}}
            for col_idx, src_cell in enumerate(src_cells[1:], start=1):
                val_cell = ws_src_vals.cell(src_cell.row, src_cell.column)
                row_data['values'][col_idx] = val_cell.value
            
            # Write row with appropriate formatting
            is_averages = (current_section == 'averages')
            dst_row = _copy_triangle_row(ws_xlw, dst_row, row_data, fmt_dict, is_triangle_row, is_averages)
        
        # Find last data column
        last_col = len(age_headers)
        
        # Write LDF and CDF rows
        dst_row += 1  # Blank row before LDF/CDF
        measure_detail = ldf_cdf_detail.get(measure)
        if measure_detail is not None:
            dst_row = _write_ldf_cdf_rows(ws_xlw, dst_row, measure_detail, age_headers, last_col, gen_wb, fmt_dict)
        
        # Write tail information
        tail_info = tail_selections.get(measure)
        if tail_info:
            dst_row = _write_tail_info(ws_xlw, dst_row, tail_info, last_col, gen_wb, fmt_dict)
        
        # Set column widths
        ws_xlw.set_column(0, 0, 25)  # Label column
        ws_xlw.set_column(1, 20, 12)  # Data columns
    
    wb_form.close()
    wb_vals.close()


def write_exposure_sheet(gen_wb, triangles_path, fmt_dict):
    """
    Write Exposure sheet from triangles parquet with xlsxwriter.
    Single age per period -> two-column table (Period | Exposure).
    Multiple ages        -> full triangle (Period | age1 | age2 | ...).
    """
    if not pathlib.Path(triangles_path).exists():
        return
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return

    exp["period_int"] = exp["period"].apply(lambda x: _period_int(str(x)))
    exp["age_num"]    = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    exp = exp.sort_values(["period_int", "age_num"])

    ages = sorted(exp["age_num"].dropna().unique())
    ws   = gen_wb.add_worksheet("Exposure")

    fmt_num = fmt_dict.get('data_num')
    fmt_label = fmt_dict.get('label')

    if len(ages) <= 1:
        headers  = ["Period", "Exposure"]
        # Write headers
        for c, hdr in enumerate(headers):
            ws.write(0, c, hdr, fmt_dict.get('subheader'))
            ws.set_column(c, c, 18)
        ws.freeze_panes(1, 0)
        
        data_row = 1
        for r, (_, row) in enumerate(exp.iterrows(), start=data_row):
            _data_cell_xlw(ws, r+1, 1, row["period_int"], fmt_obj=fmt_num)  # r+1 for 1-based
            _data_cell_xlw(ws, r+1, 2, _safe(row["value"]), fmt_obj=fmt_num)
    else:
        pivot = (
            exp.pivot_table(
                index="period_int", columns="age_num",
                values="value", aggfunc="first", observed=True,
            )
            .sort_index()
        )
        pivot.columns = [int(c) for c in pivot.columns]
        headers  = ["Period"] + [str(c) for c in pivot.columns]
        
        # Write headers
        for c, hdr in enumerate(headers):
            ws.write(0, c, hdr, fmt_dict.get('subheader'))
            ws.set_column(c, c, 14)
        ws.freeze_panes(1, 0)
        
        data_row = 1
        for r, (idx, row) in enumerate(pivot.iterrows(), start=data_row):
            _data_cell_xlw(ws, r+1, 1, idx, fmt_obj=fmt_num)  # r+1 for 1-based
            for c, val in enumerate(row, start=2):
                _data_cell_xlw(ws, r+1, c, _safe(val), fmt_obj=fmt_num)


def write_post_method_diagnostics(ws, combined, triangles_df, exp_map, fmt):
    """
    Write post-method diagnostics sheet with:
    1. Series diagnostics table (Ultimate Severity, Loss Rate, Frequency)
    2. Triangle diagnostics (X-to-Ultimate ratios and Average IBNR/Unpaid)
    """
    # Selected ultimates lookup
    sel_lookup = {}
    for _, row in combined.iterrows():
        sel_lookup[(str(row['period']), row['measure'])] = row['selected_ultimate']
    
    # ── Part 1: Series Diagnostics Table ──────────────────────────────────
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")
    
    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]
    
    # Write headers
    row_ptr = 0
    for c, hdr in enumerate(headers):
        ws.write(row_ptr, c, hdr, fmt['subheader'])
    row_ptr += 1
    
    # Set column widths
    ws.set_column(0, 0, 18)  # Accident Period
    ws.set_column(1, 1, 20)  # Ultimate Severity
    if has_exposure:
        ws.set_column(2, 2, 22)  # Ultimate Loss Rate
        ws.set_column(3, 3, 22)  # Ultimate Frequency
    
    # Write data
    periods = sorted(inc.index, key=lambda x: (int(x) if str(x).isdigit() else str(x)))
    for p in periods:
        if p not in inc.index:
            continue
        ult_loss   = inc.loc[p, "selected_ultimate"]
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)
        
        def _div(a, b):
            return a / b if pd.notna(a) and pd.notna(b) and b != 0 else None
        
        ws.write(row_ptr, 0, _period_int(p), fmt['label'])
        ws.write(row_ptr, 1, _div(ult_loss, ult_counts) or '', fmt.get('data_ldf', fmt['data_num']))
        if has_exposure:
            ws.write(row_ptr, 2, _div(ult_loss, exp) or '', fmt.get('data_ldf', fmt['data_num']))
            ws.write(row_ptr, 3, _div(ult_counts, exp) or '', fmt.get('data_ldf', fmt['data_num']))
        row_ptr += 1
    
    # Add spacing
    row_ptr += 2
    
    # ── Part 2: Triangle Diagnostics ──────────────────────────────────────
    df = triangles_df.copy()
    
    # Get periods and ages for structure
    first_measure_data = df[df['measure'] != 'Exposure']
    if not first_measure_data.empty:
        first_measure = first_measure_data['measure'].cat.categories[0]
        df_m = df[df['measure'] == first_measure].copy()
        periods = df_m['period'].cat.categories.tolist()
        ages = [str(a) for a in df_m['age'].cat.categories.tolist()]
    else:
        return
    
    # Define triangle diagnostics to write
    triangle_diagnostics = [
        ("Incurred Loss", "INCURRED-TO-ULTIMATE"),
        ("Paid Loss", "PAID-TO-ULTIMATE"),
        ("Reported Count", "REPORTED-TO-ULTIMATE"),
        ("Closed Count", "CLOSED-TO-ULTIMATE"),
    ]
    
    # X-to-Ultimate triangles
    for measure, label in triangle_diagnostics:
        measure_data = df[df['measure'] == measure]
        if measure_data.empty:
            continue
        
        # Write section title
        ws.merge_range(row_ptr, 0, row_ptr, len(ages), label, fmt['diagnostic_section'])
        row_ptr += 1
        
        # Write header row
        ws.write(row_ptr, 0, "Period", fmt['subheader_left'])
        for c_idx, age in enumerate(ages):
            ws.write(row_ptr, c_idx + 1, int(age) if age.isdigit() else age, fmt['subheader_right'])
        row_ptr += 1
        
        # Build data dict
        data_dict = {}
        for _, row in measure_data.iterrows():
            data_dict[(str(row['period']), str(row['age']))] = row['value']
        
        # Write data rows (X-to-Ultimate ratios)
        for period in periods:
            sel_ult = sel_lookup.get((str(period), measure))
            ws.write(row_ptr, 0, _period_int(period), fmt['label'])
            
            for c_idx, age in enumerate(ages):
                val = data_dict.get((str(period), age))
                if val is not None and pd.notna(val) and pd.notna(sel_ult) and sel_ult != 0:
                    ratio = val / sel_ult
                    ws.write(row_ptr, c_idx + 1, ratio, fmt.get('data_ldf', fmt['data_num']))
            row_ptr += 1
        
        row_ptr += 1  # Spacing between sections
    
    # Average IBNR triangle (if Incurred Loss and Reported Count exist)
    if "Incurred Loss" in df['measure'].unique() and "Reported Count" in df['measure'].unique():
        inc_data = df[df['measure'] == "Incurred Loss"]
        rep_data = df[df['measure'] == "Reported Count"]
        closed_data = df[df['measure'] == "Closed Count"] if "Closed Count" in df['measure'].unique() else None
        
        # Write section title
        ws.merge_range(row_ptr, 0, row_ptr, len(ages), "AVERAGE IBNR", fmt['diagnostic_section'])
        row_ptr += 1
        
        # Write header row
        ws.write(row_ptr, 0, "Period", fmt['subheader_left'])
        for c_idx, age in enumerate(ages):
            ws.write(row_ptr, c_idx + 1, int(age) if age.isdigit() else age, fmt['subheader_right'])
        row_ptr += 1
        
        # Build data dicts
        inc_dict = {}
        for _, row in inc_data.iterrows():
            inc_dict[(str(row['period']), str(row['age']))] = row['value']
        
        rep_dict = {}
        for _, row in rep_data.iterrows():
            rep_dict[(str(row['period']), str(row['age']))] = row['value']
        
        closed_dict = {}
        if closed_data is not None:
            for _, row in closed_data.iterrows():
                closed_dict[(str(row['period']), str(row['age']))] = row['value']
        
        # Write data rows (Average IBNR = (Ultimate Loss - Incurred Loss) / (Ultimate Counts - Closed Counts))
        for period in periods:
            sel_ult_loss = sel_lookup.get((str(period), "Incurred Loss"))
            sel_ult_counts = sel_lookup.get((str(period), "Reported Count"))
            ws.write(row_ptr, 0, _period_int(period), fmt['label'])
            
            for c_idx, age in enumerate(ages):
                inc_val = inc_dict.get((str(period), age))
                rep_val = rep_dict.get((str(period), age))
                closed_val = closed_dict.get((str(period), age)) if closed_dict else None
                
                # Numerator: Ultimate Loss - Incurred Loss
                numerator = (sel_ult_loss - inc_val) if (pd.notna(sel_ult_loss) and inc_val is not None and pd.notna(inc_val)) else None
                
                # Denominator: Ultimate Counts - Closed Counts (or Reported if Closed unavailable)
                if closed_val is not None and pd.notna(closed_val) and pd.notna(sel_ult_counts):
                    denominator = sel_ult_counts - closed_val
                elif rep_val is not None and pd.notna(rep_val) and pd.notna(sel_ult_counts):
                    denominator = sel_ult_counts - rep_val  # Open counts approximation
                else:
                    denominator = None
                
                if numerator is not None and denominator is not None and denominator != 0:
                    avg_ibnr = numerator / denominator
                    ws.write(row_ptr, c_idx + 1, avg_ibnr, fmt['data_num'])
            row_ptr += 1
        
        row_ptr += 1  # Spacing
    
    # Average Unpaid triangle (if Paid Loss exists)
    if "Paid Loss" in df['measure'].unique() and "Incurred Loss" in df['measure'].unique() and "Reported Count" in df['measure'].unique():
        paid_data = df[df['measure'] == "Paid Loss"]
        rep_data = df[df['measure'] == "Reported Count"]
        closed_data = df[df['measure'] == "Closed Count"] if "Closed Count" in df['measure'].unique() else None
        
        # Write section title
        ws.merge_range(row_ptr, 0, row_ptr, len(ages), "AVERAGE UNPAID", fmt['diagnostic_section'])
        row_ptr += 1
        
        # Write header row
        ws.write(row_ptr, 0, "Period", fmt['subheader_left'])
        for c_idx, age in enumerate(ages):
            ws.write(row_ptr, c_idx + 1, int(age) if age.isdigit() else age, fmt['subheader_right'])
        row_ptr += 1
        
        # Build data dicts
        paid_dict = {}
        for _, row in paid_data.iterrows():
            paid_dict[(str(row['period']), str(row['age']))] = row['value']
        
        rep_dict = {}
        for _, row in rep_data.iterrows():
            rep_dict[(str(row['period']), str(row['age']))] = row['value']
        
        closed_dict = {}
        if closed_data is not None:
            for _, row in closed_data.iterrows():
                closed_dict[(str(row['period']), str(row['age']))] = row['value']
        
        # Write data rows (Average Unpaid = (Ultimate Loss - Paid Loss) / (Ultimate Counts - Closed Counts))
        for period in periods:
            sel_ult_loss = sel_lookup.get((str(period), "Incurred Loss"))
            sel_ult_counts = sel_lookup.get((str(period), "Reported Count"))
            ws.write(row_ptr, 0, _period_int(period), fmt['label'])
            
            for c_idx, age in enumerate(ages):
                paid_val = paid_dict.get((str(period), age))
                rep_val = rep_dict.get((str(period), age))
                closed_val = closed_dict.get((str(period), age)) if closed_dict else None
                
                # Numerator: Ultimate Loss - Paid Loss
                numerator = (sel_ult_loss - paid_val) if (pd.notna(sel_ult_loss) and paid_val is not None and pd.notna(paid_val)) else None
                
                # Denominator: Ultimate Counts - Closed Counts (or Reported if Closed unavailable)
                if closed_val is not None and pd.notna(closed_val) and pd.notna(sel_ult_counts):
                    denominator = sel_ult_counts - closed_val
                elif rep_val is not None and pd.notna(rep_val) and pd.notna(sel_ult_counts):
                    denominator = sel_ult_counts - rep_val  # Open counts approximation
                else:
                    denominator = None
                
                if numerator is not None and denominator is not None and denominator != 0:
                    avg_unpaid = numerator / denominator
                    ws.write(row_ptr, c_idx + 1, avg_unpaid, fmt['data_num'])
            row_ptr += 1
    
    # Set column widths
    ws.set_column(0, 0, 22)
    for c_idx in range(1, len(ages) + 1):
        ws.set_column(c_idx, c_idx, 12)


def write_diagnostics_sheet(ws, combined, exp_map, fmt_dict):
    """
    Write the 'Diagnostics' sheet with xlsxwriter: Ultimate Severity, Loss Rate, Frequency.
    Loss Rate and Frequency columns are omitted when no exposure data is present.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    if inc.empty:
        return

    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    # Write headers
    for c, hdr in enumerate(headers):
        ws.write(0, c, hdr, fmt_dict.get('subheader'))
        ws.set_column(c, c, 18)
    ws.freeze_panes(1, 0)
    
    data_row = 1
    fmt_num = fmt_dict.get('data_num')
    fmt_dec = fmt_dict.get('data_ldf')

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else str(x)),
    )
    for r, p in enumerate(periods, start=data_row):
        ult_loss   = inc.loc[p, "selected_ultimate"] if p in inc.index else np.nan
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        def _div(a, b):
            if pd.notna(a) and pd.notna(b) and b != 0:
                return a / b
            return None

        _data_cell_xlw(ws, r+1, 1, _period_int(p), fmt_obj=fmt_num)  # r+1 for 1-based
        _data_cell_xlw(ws, r+1, 2, _div(ult_loss, ult_counts), fmt_obj=fmt_dec)
        if has_exposure:
            _data_cell_xlw(ws, r+1, 3, _div(ult_loss, exp), fmt_obj=fmt_dec)
            _data_cell_xlw(ws, r+1, 4, _div(ult_counts, exp), fmt_obj=fmt_dec)


def build_post_method_triangle_data(triangles_df, combined):
    """
    Build post-method triangle DataFrames as a list of (sheet_name, index_name, df).
    Each df has period ints as index and age ints as columns.
    """
    if triangles_df.empty:
        return []

    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    def _to_int_series(s):
        try:
            return s.apply(lambda x: int(float(str(x))))
        except (ValueError, TypeError):
            return s.astype(str)

    def x_to_ult(measure, label):
        sub = triangles_df[triangles_df["measure"] == measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        return label, result

    sheets = []
    for measure, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        result = x_to_ult(measure, label)
        if result is not None:
            sheets.append(result)

    def avg_triangle(sub_measure, sheet_name, diff_fn):
        sub = triangles_df[triangles_df["measure"] == sub_measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), sub_measure), np.nan)
            if pd.notna(sel):
                result.loc[period] = diff_fn(sel, result.loc[period])
        return sheet_name, result

    inc_sub  = triangles_df[triangles_df["measure"] == "Incurred Loss"]
    paid_sub = triangles_df[triangles_df["measure"] == "Paid Loss"]

    if not inc_sub.empty:
        r = avg_triangle("Incurred Loss", "Average IBNR", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    if not paid_sub.empty:
        r = avg_triangle("Paid Loss", "Average Unpaid", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    return sheets


def write_triangle_sheet(ws, sheet_name, df):
    """
    Write a triangle DataFrame to a worksheet.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    headers = ["Period"] + [str(c) for c in df.columns]
    data_row = _write_title_and_headers(ws, sheet_name, headers, col_width=14)

    for r, (idx, row) in enumerate(df.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), idx)
        for c, val in enumerate(row, start=2):
            _data_cell(ws.cell(r, c), _safe(val), _DEC_FMT)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def assemble_workbook(output_path, generated_wb):
    """
    Build master workbook from the pre-built generated sheets.
    All content is already computed values — no formula copying needed.
    """
    master = Workbook()
    master.remove(master.active)
    sheet_list = []

    for sname in generated_wb.sheetnames:
        ws_dst = master.create_sheet(title=sname)
        _copy_ws(generated_wb[sname], ws_dst)
        sheet_list.append((sname, *_sheet_desc(sname)))

    notes_ws = master.create_sheet(title="Notes", index=0)
    write_notes_sheet(notes_ws, sheet_list)

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    sel_lookup = load_selections(INPUT_SELECTIONS_EXCEL)

    ult_wb = load_workbook(INPUT_SELECTIONS_EXCEL, data_only=True)
    ult_col_map = {}
    for _sname in ult_wb.sheetnames:
        for _cell in ult_wb[_sname][1]:
            if _cell.value:
                ult_col_map[(_sname, str(_cell.value).strip())] = col_letter(_cell.column - 1)  # Convert to xlsxwriter 0-based
    # Keep ult_wb open for direct value lookups in write functions
    
    # Load CL LDF workbook for exposure and LDF lookups
    cl_ldf_wb = load_workbook(INPUT_CL_EXCEL, data_only=True) if pathlib.Path(INPUT_CL_EXCEL).exists() else None

    combined, available_methods = load_combined(INPUT_ULTIMATES, sel_lookup)
    reason_lookup = load_selection_reasoning(INPUT_SELECTIONS_EXCEL)
    combined["selected_reasoning"] = combined.apply(
        lambda row: reason_lookup.get(
            (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]),
            ""
        ) or "",
        axis=1
    )
    exp_map = get_exposure(INPUT_TRIANGLES)
    has_exposure = bool(exp_map)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]
    
    # Load exposure row map for linking to Chain Ladder LDFs workbook
    exp_row_map = load_exposure_row_map(INPUT_CL_EXCEL)
    
    # We also need triangles_df for diagnostics
    triangles_df = pd.read_parquet(INPUT_TRIANGLES)
    
    # Load LDF/CDF detail from 2f output (includes empirical + fitted LDFs with source tracking)
    ldf_cdf_detail = load_ldf_cdf_detail(INPUT_LDF_CDF_DETAIL)
    
    # Load tail selections (method and reasoning) from Tail Excel
    tail_selections = load_tail_selections(INPUT_TAIL_EXCEL)
    
    # Build cutoff map from LDF/CDF detail (last empirical age)
    tail_cutoff = {}
    for measure, detail_df in ldf_cdf_detail.items():
        empirical_rows = detail_df[detail_df['source'] == 'empirical']
        if not empirical_rows.empty:
            # Find last empirical age (cutoff is where fitted LDFs start)
            last_empirical_age = empirical_rows['age'].astype(int).max()
            tail_cutoff[measure] = last_empirical_age

    print(f"Measures: {measures}")
    print(f"Methods: {available_methods}")
    
    # Create xlsxwriter workbook with cached formula support
    wb = xlsxwriter.Workbook(OUTPUT_COMPLETE, {'use_future_functions': True})
    fmt = create_xlsxwriter_formats(wb)
    
    loss_m  = [m for m in measures if "Loss"  in m]
    count_m = [m for m in measures if "Count" in m]
    other_m = [m for m in measures if m not in loss_m and m not in count_m]
    has_ie  = any(col == "ultimate_ie" for col, _ in available_methods)

    # Create Notes sheet first (will write content at the end)
    ws_notes = wb.add_worksheet("Notes")
    
    print("Building Loss sheets...")
    if loss_m:
        # Build tri_col_maps dictionary for all loss measures
        loss_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in loss_m}
        
        write_selection_grouped(wb, combined, ["Incurred Loss", "Paid Loss"], "Loss Selection", ult_col_map, fmt, loss_tri_maps, exp_map)
        for m in loss_m:
            write_method_cl(wb, combined, m, ult_col_map, fmt, loss_tri_maps[m])
        for m in loss_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, loss_tri_maps, ult_wb)
        # IE sheet removed - no longer generated

    print("Building Counts sheets...")
    if count_m:
        # Build tri_col_maps dictionary for all count measures
        count_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in count_m}
        
        write_selection_grouped(wb, combined, ["Reported Count", "Closed Count"], "Count Selection", ult_col_map, fmt, count_tri_maps, exp_map)
        for m in count_m:
            # Skip Closed CL if no data or all NA
            if m == "Closed Count":
                m_data = combined[combined["measure"] == m]
                if m_data.empty or m_data["actual"].isna().all():
                    print(f"  Skipping {m} CL sheet (no data)")
                    continue
            write_method_cl(wb, combined, m, ult_col_map, fmt, count_tri_maps[m])
        for m in count_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, count_tri_maps, ult_wb)

    if other_m:
        print("Building other method sheets...")
        # Build tri_col_maps dictionary for other measures
        other_tri_maps = {m: _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)) for m in other_m}
        
        for m in other_m:
            write_method_cl(wb, combined, m, ult_col_map, fmt, other_tri_maps[m])
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(wb, combined, m, ult_col_map, fmt, other_tri_maps, ult_wb)
    
    print("Copying CL LDF triangle sheets...")
    create_triangle_sheets_xlw(wb, measures, fmt, ldf_cdf_detail, combined, tail_selections)
    
    if has_exposure:
        write_exposure_sheet(wb, INPUT_TRIANGLES, fmt)
    
    # Add Pre-Method Diagnostics sheet
    if pathlib.Path(INPUT_CL_ENHANCED).exists() and pathlib.Path(INPUT_DIAGNOSTICS).exists():
        print("Building Pre-Method Diagnostics sheet...")
        df2 = pd.read_parquet(INPUT_CL_ENHANCED)
        df3 = pd.read_parquet(INPUT_DIAGNOSTICS)
        diagnostic_cols = [col for col in df3.columns if col not in ['period', 'age', 'measure']]
        if diagnostic_cols:
            # Add 'wb' to fmt dict for diagnostics module
            fmt['wb'] = wb
            ws_diag = wb.add_worksheet("Pre-Method Diagnostics")
            build_combined_diagnostics_sheet(ws_diag, diagnostic_cols, df2, df3, fmt)
    
    # Add Post-Method Diagnostics sheet
    print("Building Post-Method Diagnostics sheet...")
    ws_post = wb.add_worksheet("Post-Method Diagnostics")
    write_post_method_diagnostics(ws_post, combined, triangles_df, exp_map, fmt)

    # TODO: Triangle sheet writing - convert in separate step
    # print("Building post-method triangle sheets...")
    # for sheet_name, df in build_post_method_triangle_data(triangles_df, combined):
    #     ws_t = wb.add_worksheet(sheet_name[:31])
    #     write_triangle_sheet(ws_t, sheet_name, df)

    # Build Notes sheet with table of contents
    print("Adding Notes sheet...")
    # Get list of all sheets created (xlsxwriter worksheets), excluding Notes itself
    sheet_list = [(ws.name, *_sheet_desc(ws.name)) for ws in wb.worksheets() if ws.name != "Notes"]
    write_notes_sheet_xlw(ws_notes, sheet_list, fmt)

    print("\nSaving Analysis.xlsx with xlsxwriter (values from source data)...")
    wb.close()
    
    # Close workbooks used for value lookups
    ult_wb.close()
    if cl_ldf_wb:
        cl_ldf_wb.close()
    
    print(f"  Saved -> {OUTPUT_COMPLETE}")
    print("Done. Formulas display immediately — no need for separate values-only workbook.")

if __name__ == "__main__":
    main()

