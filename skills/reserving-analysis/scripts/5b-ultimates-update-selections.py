# Refreshes only the selections section of the Ultimates Excel workbook based on saved selections 
# from the ultimates-ai-rules-based.json and ultimates-ai-open-ended.json files. Allows you to 
# update the displayed selections without rebuilding the entire workbook, which is useful when 
# you need to revert changes or apply selections from another source.

"""
goal: Update ONLY the selections section of Ultimates.xlsx from ultimates-ai-rules-based.json 
      and ultimates-ai-open-ended.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 5b-ultimates-update-selections.py
"""

import json
import pandas as pd
from openpyxl import load_workbook

from modules import config

# Paths from modules/config.py — override here if needed:
RULES_BASED_JSON  = config.SELECTIONS + "ultimates-ai-rules-based.json"
OPEN_ENDED_JSON   = config.SELECTIONS + "ultimates-ai-open-ended.json"
EXCEL_FILE = config.SELECTIONS + "Ultimates.xlsx"


def update_sheet_selections(ws, periods_data, selection_type="rules-based"):
    """
    Update the selection and reasoning columns in a single sheet.
    
    Args:
        ws: openpyxl worksheet object
        periods_data: Dict mapping period -> {'selection': value, 'reasoning': text}
        selection_type: "rules-based" or "open-ended" - determines which columns to update
    
    Returns:
        Number of selections updated
    """
    updates_made = 0
    
    # Rules-based: Column 9 is Selection, 10 is Reasoning
    # Open-ended: Column 11 is Selection, 12 is Reasoning
    if selection_type == "rules-based":
        sel_col = 9
        reason_col = 10
    else:  # open-ended
        sel_col = 11
        reason_col = 12
    
    # Iterate through rows starting at 2 (row 1 is headers)
    row = 2
    while True:
        period_cell = ws.cell(row=row, column=1)
        if not period_cell.value:
            break
            
        period = str(period_cell.value).strip()
        if period in periods_data:
            ws.cell(row=row, column=sel_col).value = float(periods_data[period]['selection'])
            ws.cell(row=row, column=reason_col).value = str(periods_data[period]['reasoning'])
            updates_made += 1
        row += 1
        
    return updates_made


def main():
    """Update Ultimates Excel file with selections from JSON files."""
    # Load rules-based selections
    print(f"Loading rules-based selections from {RULES_BASED_JSON}")
    try:
        with open(RULES_BASED_JSON, "r") as f:
            rules_based_selections = json.load(f)
    except FileNotFoundError:
        print(f"Rules-based selections file not found: {RULES_BASED_JSON}")
        print("Skipping rules-based update - no selections to apply.")
        rules_based_selections = []
    
    # Load open-ended selections
    print(f"Loading open-ended selections from {OPEN_ENDED_JSON}")
    try:
        with open(OPEN_ENDED_JSON, "r") as f:
            open_ended_selections = json.load(f)
    except FileNotFoundError:
        print(f"Open-ended selections file not found: {OPEN_ENDED_JSON} (optional)")
        open_ended_selections = []
    
    # Check if both are empty
    if (not rules_based_selections or len(rules_based_selections) == 0) and \
       (not open_ended_selections or len(open_ended_selections) == 0):
        print("No selections found in either file")
        print("Skipping update - selections arrays are empty.")
        return
        
    print(f"Loaded {len(rules_based_selections)} rules-based selections")
    print(f"Loaded {len(open_ended_selections)} open-ended selections")
    
    # Organize selections by measure
    def organize_by_measure(selections):
        by_measure = {}
        for entry in selections:
            meas = entry.get('measure')
            if not meas: 
                continue
            if meas not in by_measure:
                by_measure[meas] = {}
            # Key by period
            by_measure[meas][str(entry['period'])] = {
                'selection': entry.get('selection', entry.get('selected_ultimate')),  # support both key names
                'reasoning': entry['reasoning']
            }
        return by_measure
    
    rules_based_by_measure = organize_by_measure(rules_based_selections)
    open_ended_by_measure = organize_by_measure(open_ended_selections)
    
    print(f"Opening workbook {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Excel file not found: {EXCEL_FILE}")
        return
        
    total_updates_rb = 0
    total_updates_oe = 0
    
    # Update rules-based selections
    for measure, periods_data in rules_based_by_measure.items():
        sheet_name = measure
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            updates = update_sheet_selections(ws, periods_data, "rules-based")
            total_updates_rb += updates
            print(f"  Updated {updates} rules-based selections in '{sheet_name}'")
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found in workbook - skipping")
    
    # Update open-ended selections
    for measure, periods_data in open_ended_by_measure.items():
        sheet_name = measure
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            updates = update_sheet_selections(ws, periods_data, "open-ended")
            total_updates_oe += updates
            print(f"  Updated {updates} open-ended selections in '{sheet_name}'")
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found in workbook - skipping")
                
    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")
    print(f"Total rules-based updates: {total_updates_rb}")
    print(f"Total open-ended updates: {total_updates_oe}")


if __name__ == "__main__":
    main()
