# Reads projected ultimates, actuary selections, and triangle data to produce
# Analysis.xlsx — formulas intact, links to Ultimates.xlsx and LDFs workbooks.
# Open in Excel to evaluate cross-workbook references.
#
# For the plain-numbers version (used by script 7+) run 6b-create-values-only.py next.
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6a-create-complete-analysis.py

import copy
import os
import pathlib

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import DATA_FONT, THIN_BORDER
from modules.xl_selections import (
    SKIP_ROW_LABELS as _SKIP_ROW_LABELS,
    SELECTION_LABELS as _SELECTION_LABELS,
    find_selected_values as _find_selected_values,
    find_selected_reasoning as _find_selected_reasoning,
    copy_ws_filtered as _copy_ws_filtered,
)
from modules.xl_utils import (
    _safe, _to_py, _period_int, _copy_ws,
    measure_short_name,
    ultimates_sheet_for_measure, ultimates_col_header,
)
from modules.xl_writers import _data_cell, _write_title_and_headers, _write_headers
from modules.xl_values import UNPAID_PROXY, _has_method
from modules.xl_notes import _sheet_desc, write_notes_sheet
from modules.analysis_loaders import (
    MEASURE_TO_CATEGORY,
    load_selections,
    load_selection_reasoning,
    load_combined,
    get_exposure,
)

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
OUTPUT_PATH            = config.OUTPUT

OUTPUT_COMPLETE = config.BASE_DIR + "Analysis.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS  + "Chain Ladder Selections - LDFs.xlsx"
INPUT_TAIL_EXCEL    = config.SELECTIONS  + "Chain Ladder Selections - Tail.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# Excel external reference paths - using absolute paths for reliability
# (openpyxl doesn't set proper workbook relationships for relative paths to work)
import pathlib
_BASE_ABS = str(pathlib.Path(OUTPUT_COMPLETE).parent.resolve())
_ULT_WB  = _BASE_ABS + "\\selections\\[Ultimates.xlsx]"
_CL_LDF_WB = _BASE_ABS + "\\selections\\[Chain Ladder Selections - LDFs.xlsx]"
_CL_TAIL_WB = _BASE_ABS + "\\selections\\[Chain Ladder Selections - Tail.xlsx]"

def load_tail_selections(tail_excel_path):
    """
    Load tail factor selections from Chain Ladder Selections - Tail.xlsx.
    Priority: User Selection (if populated) > Rules-Based AI Selection.
    Returns {measure: (cutoff_age: int, tail_factor: float)}.  Returns {} if file absent.
    """
    path = pathlib.Path(tail_excel_path)
    if not path.exists():
        print(f"  Note: {tail_excel_path} not found -- no tail selections loaded")
        return {}

    wb = load_workbook(tail_excel_path, data_only=True)
    tail_map = {}

    for measure in wb.sheetnames:
        ws = wb[measure]
        in_tail = False
        cutoff_col = tail_col = reason_col = None
        user_entry = rb_entry = None

        for row in ws.iter_rows():
            col1 = row[0].value if row else None
            if col1 == "Tail Factor Selection":
                in_tail = True
                continue
            if not in_tail:
                continue
            if col1 == "Label":
                for cell in row:
                    if cell.value == "Cutoff Age":
                        cutoff_col = cell.column
                    elif cell.value == "Tail Factor":
                        tail_col = cell.column
                    elif cell.value == "Reasoning":
                        reason_col = cell.column
                continue
            if cutoff_col is None or tail_col is None:
                continue
            if col1 in ("User Selection", "Rules-Based AI Selection"):
                cv = row[cutoff_col - 1].value
                tv = row[tail_col - 1].value
                rv = row[reason_col - 1].value if reason_col else None
                if cv is not None and tv is not None:
                    try:
                        entry = (int(float(str(cv))), float(str(tv)), rv)
                        if col1 == "User Selection":
                            user_entry = entry
                        elif rb_entry is None:
                            rb_entry = entry
                    except (ValueError, TypeError):
                        pass

        chosen = user_entry if user_entry is not None else rb_entry
        if chosen:
            tail_map[measure] = chosen

    wb.close()
    summary = {m: (v[0], v[1]) for m, v in tail_map.items()}
    print(f"  Loaded tail selections: {summary}")
    return tail_map


def load_exposure_row_map(cl_ldf_excel_path):
    """
    Load exposure row map from Chain Ladder Selections - LDFs.xlsx.
    Returns {period: row_number} for building formula references.
    Returns {} if file absent or Exposure sheet not found.
    """
    path = pathlib.Path(cl_ldf_excel_path)
    if not path.exists():
        print(f"  Note: {cl_ldf_excel_path} not found -- no exposure row map loaded")
        return {}
    
    wb = load_workbook(cl_ldf_excel_path, data_only=True)
    if "Exposure" not in wb.sheetnames:
        wb.close()
        print("  Note: Exposure sheet not found in Chain Ladder LDFs workbook")
        return {}
    
    ws = wb["Exposure"]
    row_map = {}
    
    # Find Period column (should be in row 2)
    period_col = None
    for cell in ws[2]:
        if cell.value == "Period":
            period_col = cell.column
            break
    
    if period_col is None:
        wb.close()
        print("  Note: Period column not found in Exposure sheet")
        return {}
    
    # Read period values starting from row 3 (data starts after title and headers)
    for row_idx in range(3, ws.max_row + 1):
        period_val = ws.cell(row=row_idx, column=period_col).value
        if period_val is not None:
            row_map[str(period_val)] = row_idx
    
    wb.close()
    print(f"  Loaded exposure row map for {len(row_map)} periods")
    return row_map


# ── Generated sheet writers ───────────────────────────────────────────────────

def _get_cdf_cell_ref(ws_triangle, target_age):
    # Search row 1 or 2 for the age
    for row_idx in [1, 2]:
        for cell in ws_triangle[row_idx]:
            try:
                if cell.value is not None and int(float(cell.value)) == int(float(target_age)):
                    # Found the column! Now find CDF row (usually 'CDF' in column A)
                    for r in range(1, 100):
                        if str(ws_triangle.cell(row=r, column=1).value).strip() == "CDF":
                            return f"'{ws_triangle.title}'!{get_column_letter(cell.column)}{r}"
            except (ValueError, TypeError):
                pass
    return None

def _formula_cell(ws, r, c, formula, num_fmt):
    cell = ws.cell(r, c, formula)
    cell.number_format = num_fmt
    cell.font = DATA_FONT
    cell.border = THIN_BORDER


def _ult_ref(ult_col_map, sheet, col_header, row):
    col = ult_col_map.get((sheet, col_header))
    return f"='{_ULT_WB}{sheet}'!{col}{row}" if col else '""'


def _exp_ref(exp_row_map, period):
    """Build formula reference to exposure value in Chain Ladder Selections - LDFs.xlsx."""
    row = exp_row_map.get(str(period))
    if row:
        return f"='{_CL_LDF_WB}Exposure'!B{row}"
    return '""'


def _build_cl_ldfs_col_row_maps(wb_cl, measure_sheet_name):
    """
    Build column and row maps for referencing Chain Ladder Selections - LDFs.xlsx.
    Returns (col_map, user_row, rb_row, user_reason_row, rb_reason_row) where:
      col_map: {column_header: column_letter}
      user_row: row number of "User Selection"
      rb_row: row number of "Rules-Based AI Selection"
      user_reason_row: row number of "User Reasoning"
      rb_reason_row: row number of "Rules-Based AI Reasoning"
    """
    if measure_sheet_name not in wb_cl.sheetnames:
        return {}, None, None, None, None
    
    ws = wb_cl[measure_sheet_name]
    col_map = {}
    user_row = None
    rb_row = None
    user_reason_row = None
    rb_reason_row = None
    
    # Find column headers and selection rows
    for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
        col1 = row_cells[0].value if row_cells else None
        
        # Track header rows (usually None in col A)
        if col1 is None and row_idx <= 10:  # Headers in first few rows
            for cell in row_cells[1:]:
                if cell.value is not None:
                    col_map[str(cell.value)] = get_column_letter(cell.column)
        
        # Track selection rows
        if col1 == "User Selection":
            user_row = row_idx
        elif col1 == "Rules-Based AI Selection":
            rb_row = row_idx
        elif col1 == "User Reasoning":
            user_reason_row = row_idx
        elif col1 == "Rules-Based AI Reasoning":
            rb_reason_row = row_idx
    
    return col_map, user_row, rb_row, user_reason_row, rb_reason_row


def _copy_ws_with_ldfs_formulas(ws_src, ws_dst, measure_short_name, cl_ldfs_path):
    """
    Copy worksheet writing formulas that link to Chain Ladder Selections - LDFs.xlsx.
    Selection rows use IF formula: User Selection if not blank, else Rules-Based AI.
    Data rows reference the source workbook.
    """
    # Load source workbook to build maps
    if not pathlib.Path(cl_ldfs_path).exists():
        # Fallback to value copy if source doesn't exist
        _copy_ws_filtered(ws_src, ws_dst, None, None)
        return
    
    wb_cl = load_workbook(cl_ldfs_path, data_only=False)
    col_map, user_row, rb_row, user_reason_row, rb_reason_row = _build_cl_ldfs_col_row_maps(wb_cl, measure_short_name)
    wb_cl.close()
    
    if not col_map or user_row is None or rb_row is None:
        # Fallback if structure not found
        sel_vals = _find_selected_values(ws_src)
        sel_reason = _find_selected_reasoning(ws_src)
        _copy_ws_filtered(ws_src, ws_dst, sel_vals, sel_reason)
        return
    
    selection_done = False
    dst_row = 1
    # Excel external ref format: '[Workbook.xlsx]SheetName'!CellRef
    ext_ref = f"'{_CL_LDF_WB}{measure_short_name}'"
    
    src_rows = list(ws_src.iter_rows())
    skip_next = False
    
    for row_idx, src_cells in enumerate(src_rows):
        col1 = src_cells[0].value if src_cells else None
        
        # Skip if marked by previous iteration
        if skip_next:
            skip_next = False
            continue
        
        # Skip AI option/reasoning rows
        if col1 in _SKIP_ROW_LABELS:
            continue
        
        # Check if this is a section title row (Age-to-Age Factors, Averages, LDF Selections)
        is_section_title = (col1 in ("Age-to-Age Factors", "Averages", "LDF Selections") and 
                           all(c.value in (None, "") for c in src_cells[1:]))
        
        # If section title and next row has headers, merge them
        if is_section_title and row_idx + 1 < len(src_rows):
            next_row = src_rows[row_idx + 1]
            next_col1 = next_row[0].value if next_row else None
            has_headers = next_col1 in (None, "") and any(c.value not in (None, "") for c in next_row[1:])
            
            if has_headers:
                # Write title in col A and headers in cols 2+ on same row
                for col_idx, src_cell in enumerate(src_cells):
                    dst_cell = ws_dst.cell(dst_row, col_idx + 1)
                    if col_idx == 0:
                        dst_cell.value = col1
                        if src_cell.has_style:
                            dst_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=False)
                            dst_cell.border = copy.copy(src_cell.border)
                            dst_cell.fill = copy.copy(src_cell.fill)
                            dst_cell.alignment = copy.copy(src_cell.alignment)
                    elif col_idx < len(next_row):
                        # Take header from next row
                        next_cell = next_row[col_idx]
                        dst_cell.value = next_cell.value
                        if next_cell.has_style:
                            dst_cell.font = copy.copy(next_cell.font)
                            dst_cell.border = copy.copy(next_cell.border)
                            dst_cell.fill = copy.copy(next_cell.fill)
                            dst_cell.number_format = next_cell.number_format
                            dst_cell.alignment = copy.copy(next_cell.alignment)
                dst_row += 1
                skip_next = True
                continue
        
        # Handle selection rows
        if col1 in _SELECTION_LABELS:
            if selection_done:
                continue
            selection_done = True
            
            # Write "Selected" row with IF formulas
            for src_cell in src_cells:
                dst_cell = ws_dst.cell(dst_row, src_cell.column)
                if src_cell.column == 1:
                    dst_cell.value = "Selected"
                else:
                    # Build IF formula: if User not blank, use User, else use RB-AI
                    col_letter = get_column_letter(src_cell.column)
                    formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_row},{ext_ref}!{col_letter}{rb_row})'
                    dst_cell.value = formula
                
                if src_cell.has_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_row += 1
            
            # Write "Selected Reasoning" row with IF formula
            if user_reason_row and rb_reason_row:
                for src_cell in src_cells:
                    dst_cell = ws_dst.cell(dst_row, src_cell.column)
                    if src_cell.column == 1:
                        dst_cell.value = "Selected Reasoning"
                    else:
                        col_letter = get_column_letter(src_cell.column)
                        formula = f'=IF({ext_ref}!{col_letter}{user_row}<>"",{ext_ref}!{col_letter}{user_reason_row},{ext_ref}!{col_letter}{rb_reason_row})'
                        dst_cell.value = formula
                    
                    if src_cell.has_style:
                        dst_cell.font = copy.copy(src_cell.font)
                        dst_cell.border = copy.copy(src_cell.border)
                        dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.number_format = ""
                    dst_cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
                dst_row += 1
            continue
        
        # Copy all other rows with formulas linking to source
        for src_cell in src_cells:
            dst_cell = ws_dst.cell(dst_row, src_cell.column)
            
            # For data cells (not labels), write formula; for labels, copy value
            if src_cell.column == 1 or src_cell.value is None or isinstance(src_cell.value, str):
                dst_cell.value = src_cell.value
            else:
                # Reference the source workbook
                col_letter = get_column_letter(src_cell.column)
                row_num = src_cells[0].row
                dst_cell.value = f"={ext_ref}!{col_letter}{row_num}"
            
            if src_cell.has_style:
                # Remove bold from column A cells
                if src_cell.column == 1:
                    dst_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=False)
                else:
                    dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy.copy(src_cell.alignment)
        
        dst_row += 1


def _add_cdf_formulas_to_triangle(ws):
    """
    Add CDF row formulas below Selected Reasoning in the LDF Selections section.
    CDF formulas multiply Selected LDF by the next CDF to the right (right-to-left cumulative product).
    Rightmost column references the Selected tail factor.
    """
    selected_row = None
    selected_reasoning_row = None
    cdf_row_num = None
    last_data_col = None
    
    # Find the Selected row and determine where CDF row should be
    for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
        col1 = row_cells[0].value if row_cells else None
        
        if col1 == "Selected":
            selected_row = row_idx
            # Find the last column with data
            for cell in reversed(list(row_cells)):
                if cell.value not in (None, ""):
                    last_data_col = cell.column
                    break
        elif col1 == "Selected Reasoning":
            selected_reasoning_row = row_idx
        elif col1 == "CDF":
            cdf_row_num = row_idx
            break
    
    if not selected_row or last_data_col is None:
        return  # No Selected row found
    
    # If CDF row doesn't exist, create it after Selected Reasoning (or Selected if no reasoning)
    if cdf_row_num is None:
        cdf_row_num = (selected_reasoning_row or selected_row) + 1
        # Insert a new row by shifting cells down if needed
        # Actually, if the row exists but is empty, we can just use it
    
    # Set CDF label
    ws.cell(cdf_row_num, 1).value = "CDF"
    
    # Copy style from Selected row (but remove bold)
    ref_cell = ws.cell(selected_row, 1)
    if ref_cell.has_style:
        lbl = ws.cell(cdf_row_num, 1)
        lbl.font = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=False)
        lbl.border = copy.copy(ref_cell.border)
        lbl.fill = copy.copy(ref_cell.fill)
        lbl.alignment = copy.copy(ref_cell.alignment)
    
    # Write CDF formulas from right to left
    # Rightmost column: reference Selected tail factor
    tail_col_letter = get_column_letter(last_data_col)
    tail_cell = ws.cell(cdf_row_num, last_data_col)
    tail_cell.value = f'={tail_col_letter}{selected_row}'
    tail_cell.number_format = _DEC_FMT
    
    # Other columns: =Selected_LDF * Next_CDF
    for col in range(last_data_col - 1, 1, -1):
        col_letter = get_column_letter(col)
        next_col_letter = get_column_letter(col + 1)
        
        cell = ws.cell(cdf_row_num, col)
        cell.value = f'=IFERROR({col_letter}{selected_row}*{next_col_letter}{cdf_row_num},"")'
        cell.number_format = _DEC_FMT
        if ref_cell.has_style:
            cell.font = copy.copy(ref_cell.font)
            cell.border = copy.copy(ref_cell.border)


def _tri_col_map_from_df(triangles_df, measure, max_age=None):
    """Build {age_str: col_letter} for the triangle sheet from parquet data.
    Triangle sheet col A = period, so ages start at col B (index 2).
    Ages sorted ascending match the order written by _copy_ws_filtered.
    max_age caps the map to the cutoff age — add_tail_to_triangle_ws deletes
    data columns beyond the cutoff, so referencing those cols returns empty.
    """
    tri_m = triangles_df[triangles_df["measure"].astype(str) == measure]
    if tri_m.empty:
        return {}
    ages = sorted(tri_m["age"].dropna().unique(), key=lambda a: float(str(a)))
    if max_age is not None:
        ages = [a for a in ages if float(str(a)) <= max_age]
    return {str(int(float(str(a)))): get_column_letter(i + 2) for i, a in enumerate(ages)}


def write_method_cl(gen_wb, combined, measure, ult_col_map, tri_col_map=None):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title=f"{short_name} CL"[:31])
    headers = ["Accident Period", "Current Age", short_name, "CDF", "Ultimate", "IBNR", "Unpaid"]
    _write_headers(ws, headers)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()
    ult_sheet  = ultimates_sheet_for_measure(measure)
    actual_hdr = ultimates_col_header(measure, "actual")

    # Direct cell refs into the triangle sheet (row 1=title, row 2=age header, row 3+=data).
    # CL sheet row r maps to triangle row r+1 — both sorted ascending period_int from rows 2/3.
    tri_col_map = tri_col_map or {}
    # Tail column is the last column in the triangle (contains ultimate values for mature periods)
    tail_col_letter = get_column_letter(len(tri_col_map) + 1) if tri_col_map else None

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])
        try:
            age_key = str(int(float(str(row["current_age"]))))
        except (ValueError, TypeError):
            age_key = None
        tri_col = tri_col_map.get(age_key) if age_key else None
        
        # Always reference the triangle sheet for consistency
        if tri_col:
            # Current age has a specific column in the triangle
            _formula_cell(ws, r, 3, f"='{short_name}'!{tri_col}{r+1}", _NUM_FMT)
        elif tail_col_letter:
            # Age beyond cutoff -> use tail/ultimate column
            _formula_cell(ws, r, 3, f"='{short_name}'!{tail_col_letter}{r+1}", _NUM_FMT)
        else:
            # No triangle data available -> use static value
            _data_cell(ws.cell(r, 3), row["actual"], _NUM_FMT)
        _data_cell(ws.cell(r, 4), row.get("cdf"), _DEC_FMT)
        _formula_cell(ws, r, 5, f'=IF(AND(ISNUMBER(C{r}),ISNUMBER(D{r})),C{r}*D{r},"")', _NUM_FMT)
        _formula_cell(ws, r, 6, f"=E{r}-C{r}", _NUM_FMT)
        if proxy_exists:
            _formula_cell(ws, r, 7, f"=E{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT)
        else:
            _formula_cell(ws, r, 7, f"=E{r}-C{r}", _NUM_FMT)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell(ws.cell(t, 1), "Total")
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 5, f"=SUM(E2:E{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT)

def write_method_bf(gen_wb, combined, measure, ult_col_map):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title=f"{short_name} BF"[:31])
    
    # bf canonical structure: 
    # Accident Period, Current Age, Initial Expected, CDF, % Unreported, Unreported, Actual, Ultimate, IBNR, Unpaid
    headers = ["Accident Period", "Current Age", "Initial Expected", "CDF", "% Unreported", "Unreported", short_name, "Ultimate", "IBNR", "Unpaid"]
    _write_headers(ws, headers)
    
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")

    ws_triangle = gen_wb[short_name] if short_name in gen_wb.sheetnames else None

    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()
    ult_sheet  = ultimates_sheet_for_measure(measure)
    actual_hdr = ultimates_col_header(measure, "actual")
    ie_hdr     = ultimates_col_header(measure, "ie")
    has_ie_method = _has_method(combined, measure, "ultimate_ie")

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])
        ie_formula = f"='IE'!D{r}" if has_ie_method else _ult_ref(ult_col_map, ult_sheet, ie_hdr, r)
        _formula_cell(ws, r, 3, ie_formula, _NUM_FMT)
        _formula_cell(ws, r, 4, f"='{short_name} CL'!D{r}", _DEC_FMT)
        _formula_cell(ws, r, 5, f"=1-(1/D{r})", "0.0%")
        _formula_cell(ws, r, 6, f"=C{r}*E{r}", _NUM_FMT)
        _formula_cell(ws, r, 7, f"='{short_name} CL'!C{r}", _NUM_FMT)
        _formula_cell(ws, r, 8, f"=F{r}+G{r}", _NUM_FMT)
        _formula_cell(ws, r, 9, f"=H{r}-G{r}", _NUM_FMT)
        if proxy_exists:
            _formula_cell(ws, r, 10, f"=H{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT)
        else:
            _formula_cell(ws, r, 10, f"=H{r}-G{r}", _NUM_FMT)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell(ws.cell(t, 1), "Total")
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT)
    # Col 5: % Unreported total = weighted avg = SUM(F)/SUM(C)
    _formula_cell(ws, t, 5, f"=SUM(F2:F{t-2})/SUM(C2:C{t-2})", "0.0%")
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 8, f"=SUM(H2:H{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 9, f"=SUM(I2:I{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 10, f"=SUM(J2:J{t-2})", _NUM_FMT)


def write_method_ie(gen_wb, combined, measure, exp_row_map, ult_col_map):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title="IE")
    headers = ["Accident Period", "Current Age", "Exposure", "IE Ultimate", "Selected Loss Rate"]
    _write_headers(ws, headers)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    ult_sheet = ultimates_sheet_for_measure(measure)
    ie_hdr    = ultimates_col_header(measure, "ie")

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])

        # Link to Chain Ladder Selections - LDFs.xlsx Exposure sheet
        exp_formula = _exp_ref(exp_row_map, row["period"])
        _formula_cell(ws, r, 3, exp_formula, _NUM_FMT)

        _formula_cell(ws, r, 4, _ult_ref(ult_col_map, ult_sheet, ie_hdr, r), _NUM_FMT)
        _formula_cell(ws, r, 5, f"=D{r}/C{r}", _DEC_FMT)

        ie_ref = _ult_ref(ult_col_map, ult_sheet, ie_hdr, r)
        if ie_ref == '""':
            _data_cell(ws.cell(r, 4), row.get("ultimate_ie"), _NUM_FMT)
        else:
            _formula_cell(ws, r, 4, ie_ref, _NUM_FMT)
        _formula_cell(ws, r, 5, f"=D{r}/C{r}", _DEC_FMT)


def write_selection_grouped(gen_wb, combined, measures_group, title, ult_col_map):
    ws = gen_wb.create_sheet(title=title)
    
    # columns e.g.: Accident Period, Current Age, Incurred, Paid, Incurred CL, Paid CL, Initial Expected, Incurred BF, Paid BF, Selected Ultimate, IBNR, Unpaid
    # Column ordering convention: CL, IE, BF
    
    # Find active measures in this group
    active_m = [
        m for m in measures_group
        if m in combined["measure"].unique()
        and combined[combined["measure"] == m]["actual"].notna().any()
    ]
    if not active_m:
        return
        
    main_m = active_m[0]
    sub = combined[combined["measure"] == main_m].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    short_names = [measure_short_name(m) for m in active_m]
    bf_m  = [m for m in active_m if _has_method(combined, m, "ultimate_bf")]
    has_group_ie = any(_has_method(combined, m, "ultimate_ie") for m in active_m)

    ult_sheet  = ultimates_sheet_for_measure(active_m[0])
    ext_ref    = f"'{_ULT_WB}{ult_sheet}'"
    user_col   = ult_col_map.get((ult_sheet, "User Selection"), "")
    rb_col     = ult_col_map.get((ult_sheet, "Rules-Based AI Selection"), "")
    user_r_col = ult_col_map.get((ult_sheet, "User Reasoning"), "")
    rb_r_col   = ult_col_map.get((ult_sheet, "Rules-Based AI Reasoning"), "")
    ie_measure = next((m for m in active_m if _has_method(combined, m, "ultimate_ie")), active_m[0])
    ie_short   = measure_short_name(ie_measure)

    # Build headers in CL, IE, BF order
    headers = ["Accident Period", "Current Age"]
    for s in short_names:
        headers.append(s)
    for s in short_names:
        headers.append(f"{s} CL")
    if has_group_ie:
        headers.append("Initial Expected")
    for m in bf_m:
        headers.append(f"{measure_short_name(m)} BF")
    headers.extend(["Selected Ultimate", "IBNR", "Unpaid", "Selected Reasoning"])

    _write_headers(ws, headers)
    ws.column_dimensions[get_column_letter(len(headers))].width = 40

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])

        col_idx = 3
        # Actuals
        for s in short_names:
            _formula_cell(ws, r, col_idx, f"='{s} CL'!C{r}", _NUM_FMT)
            col_idx += 1
        # CL ultimates
        for s in short_names:
            _formula_cell(ws, r, col_idx, f"='{s} CL'!E{r}", _NUM_FMT)
            col_idx += 1
        # IE ultimate (if available)
        if has_group_ie:
            _formula_cell(ws, r, col_idx, f"='IE'!D{r}", _NUM_FMT)
            col_idx += 1
        # BF ultimates
        for m in bf_m:
            _formula_cell(ws, r, col_idx, f"='{measure_short_name(m)} BF'!H{r}", _NUM_FMT)
            col_idx += 1
        # Selected Ultimate → links Ultimates.xlsx (User Selection, falls back to RB-AI)
        sel_formula = (
            f'=IF({ext_ref}!{user_col}{r}<>"",{ext_ref}!{user_col}{r},{ext_ref}!{rb_col}{r})'
            if user_col and rb_col else '""'
        )
        _formula_cell(ws, r, col_idx, sel_formula, _NUM_FMT)
        col_idx += 1
        _formula_cell(ws, r, col_idx, f"={get_column_letter(col_idx-1)}{r}-C{r}", _NUM_FMT)
        col_idx += 1
        unpaid_actual = "D" if len(active_m) > 1 else "C"
        _formula_cell(ws, r, col_idx, f"={get_column_letter(col_idx-2)}{r}-{unpaid_actual}{r}", _NUM_FMT)
        col_idx += 1
        # Selected Reasoning → links Ultimates.xlsx (User Reasoning, falls back to RB-AI)
        reason_formula = (
            f'=IF({ext_ref}!{user_col}{r}<>"",{ext_ref}!{user_r_col}{r},{ext_ref}!{rb_r_col}{r})'
            if user_col and user_r_col and rb_r_col else '""'
        )
        _formula_cell(ws, r, col_idx, reason_formula, "@")

    # Totals row — SUM every numeric column, skip Period and Age
    n_cols = 2 + len(active_m) * 2 + len(bf_m) + (1 if has_group_ie else 0) + 3
    t = 3 + len(sub)
    _data_cell(ws.cell(t, 1), "Total")
    for c in range(3, n_cols + 1):
        col_letter = get_column_letter(c)
        _formula_cell(ws, t, c, f"=SUM({col_letter}2:{col_letter}{t-2})", _NUM_FMT)


def _build_tail_workbook_maps(tail_wb_path, measure_sheet_name):
    """
    Build row/column maps for Tail Factor Selection in Chain Ladder Selections - Tail.xlsx.
    Returns (cutoff_col, tail_col, reason_col, user_row, rb_row) where columns/rows are
    dynamically discovered from the workbook structure.
    """
    if not pathlib.Path(tail_wb_path).exists():
        return None, None, None, None, None
    
    try:
        wb = load_workbook(tail_wb_path, data_only=False)
        if measure_sheet_name not in wb.sheetnames:
            wb.close()
            return None, None, None, None, None
        
        ws = wb[measure_sheet_name]
        in_tail_section = False
        cutoff_col = tail_col = reason_col = None
        user_row = rb_row = None
        
        for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
            col1 = row_cells[0].value if row_cells else None
            
            if col1 == "Tail Factor Selection":
                in_tail_section = True
                continue
            
            if not in_tail_section:
                continue
            
            # Find column headers
            if col1 == "Label":
                for cell in row_cells:
                    if cell.value == "Cutoff Age":
                        cutoff_col = get_column_letter(cell.column)
                    elif cell.value == "Tail Factor":
                        tail_col = get_column_letter(cell.column)
                    elif cell.value == "Reasoning":
                        reason_col = get_column_letter(cell.column)
                continue
            
            # Find selection rows
            if col1 == "User Selection":
                user_row = row_idx
            elif col1 == "Rules-Based AI Selection":
                rb_row = row_idx
        
        wb.close()
        return cutoff_col, tail_col, reason_col, user_row, rb_row
    except Exception:
        return None, None, None, None, None


def add_tail_to_triangle_ws(ws, cutoff_age, tail_factor, reasoning=None,
                            df2=None, measure=None):
    """
    After _copy_ws_filtered:
      1. Delete all interval columns whose ending age > cutoff_age.
      2. Append a {cutoff_age}-Ult column header to ATA / Averages / LDF sections.
      3. Fill the ATA data rows at the tail column with the observed cumulative
         factor for each period that has development beyond the cutoff age
         (product of per-period ATAs from cutoff_age onward, sourced from df2).
      4. Write a CDF row in LDF Selections with right-to-left cumulative factors.
      5. Write the tail reasoning in the Selected Reasoning row at the tail column.
    """
    tail_label = f"{cutoff_age}-Ult"

    # ── Single pass: locate section header rows and ATA data rows ────────────
    in_ata = in_avg = in_ldf = False
    in_ata_data = False
    ata_hdr_row = avg_hdr_row = ldf_hdr_row = None
    ata_data_rows = {}          # {period_str: row_num} for ATA section
    avg_data_row_nums = []      # row nums of averages data rows (to clear at tail_col)
    ldf_option_row_nums = []    # row nums of LDF section rows that aren't Selected/Reasoning
    last_keep_col = None
    selected_row_num = None
    last_sel_row = None
    sel_vals_by_col = {}

    for row_cells in ws.iter_rows():
        row_num = row_cells[0].row
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Age-to-Age Factors":
            in_ata = True; in_avg = in_ldf = False; in_ata_data = False; continue
        if col1 == "Averages":
            in_avg = True; in_ata = in_ldf = False; in_ata_data = False; continue
        if col1 == "LDF Selections":
            in_ldf = True; in_ata = in_avg = False; in_ata_data = False; continue

        if in_ata:
            if ata_hdr_row is None:
                if col1 is None:
                    ata_hdr_row = row_num
                    in_ata_data = True
            elif in_ata_data:
                if col1 is None:
                    in_ata_data = False
                else:
                    try:
                        ata_data_rows[str(int(float(str(col1))))] = row_num
                    except (ValueError, TypeError):
                        pass
            continue

        if in_avg:
            if avg_hdr_row is None and col1 in ("Metric", None):
                avg_hdr_row = row_num
            elif avg_hdr_row is not None and col1 not in (None, ""):
                avg_data_row_nums.append(row_num)
            continue

        if in_ldf:
            if ldf_hdr_row is None:
                if col1 is None:
                    ldf_hdr_row = row_num
                    for cell in row_cells:
                        v = cell.value
                        if isinstance(v, str) and '-' in v:
                            parts = v.split('-')
                            if len(parts) == 2:
                                try:
                                    if int(parts[1]) <= cutoff_age:
                                        last_keep_col = cell.column
                                except ValueError:
                                    pass
            elif col1 == "Selected":
                selected_row_num = row_num
                last_sel_row = row_num
                sel_vals_by_col = {
                    c.column: float(c.value)
                    for c in row_cells[1:]
                    if c.value is not None
                    and isinstance(c.value, (int, float))
                    and (last_keep_col is None or c.column <= last_keep_col)
                }
            elif col1 == "Selected Reasoning":
                last_sel_row = row_num
            elif col1 is not None and col1 != "":
                ldf_option_row_nums.append(row_num)
            continue

    if last_keep_col is None:
        return

    # ── Column layout ─────────────────────────────────────────────────────────
    # Each ATA column K stores formula (K+1)/K, so the "191-203" interval at
    # last_keep_col references last_keep_col+1 as its numerator (age-203 data).
    # We must KEEP last_keep_col+1 or that formula breaks.
    # tail_col = last_keep_col+1 (the age-203 data column, relabelled to tail).
    # Delete everything from last_keep_col+2 onward.
    tail_col   = last_keep_col + 1
    del_start  = last_keep_col + 2
    del_count  = ws.max_column - last_keep_col - 1
    if del_count > 0:
        ws.delete_cols(del_start, del_count)

    ws.column_dimensions[get_column_letter(tail_col)].width = 14

    # ── Overwrite stale content at tail_col ───────────────────────────────────
    # ATA data rows: formula was (col tail_col+1)/col tail_col → now broken;
    # clear so Excel doesn't show a shifted wrong reference.
    for row_num in ata_data_rows.values():
        ws.cell(row_num, tail_col).value = None

    # Averages data rows: were averages of "203-215" ATAs — blank for tail col.
    for row_num in avg_data_row_nums:
        ws.cell(row_num, tail_col).value = None

    # LDF option rows (Vol Wtd, Simple, etc.): stale "203-215" LDFs — blank.
    for row_num in ldf_option_row_nums:
        ws.cell(row_num, tail_col).value = None

    # ── Replace section headers at tail_col with tail label ──────────────────
    for hdr_row in (ata_hdr_row, avg_hdr_row, ldf_hdr_row):
        if hdr_row is None:
            continue
        ref = ws.cell(hdr_row, last_keep_col)
        new = ws.cell(hdr_row, tail_col)
        new.value = tail_label
        if ref.has_style:
            new.font      = copy.copy(ref.font)
            new.border    = copy.copy(ref.border)
            new.fill      = copy.copy(ref.fill)
            new.number_format = ref.number_format
            new.alignment = copy.copy(ref.alignment)

    if selected_row_num is None:
        return

    # Set Selected row at tail_col to formula referencing Tail workbook
    # IF(User Selection not blank, User Selection, Rules-Based AI Selection)
    c = ws.cell(selected_row_num, tail_col)
    if measure:
        # Dynamically discover Tail workbook structure
        tail_path = config.SELECTIONS + "/Chain Ladder Selections - Tail.xlsx"
        t_cutoff_col, t_tail_col, t_reason_col, t_user_row, t_rb_row = _build_tail_workbook_maps(tail_path, measure)
        
        if t_tail_col and t_user_row and t_rb_row:
            tail_ref = f"'{_CL_TAIL_WB}{measure}'"
            c.value = f'=IF({tail_ref}!{t_tail_col}{t_user_row}<>"",{tail_ref}!{t_tail_col}{t_user_row},{tail_ref}!{t_tail_col}{t_rb_row})'
        else:
            # Fallback to static value if structure not found
            c.value = tail_factor
    else:
        c.value = tail_factor
    c.number_format = _DEC_FMT

    # ── Compute cumulative CDFs right-to-left, seeded by tail_factor ─────────
    cdf_by_col = {tail_col: tail_factor}
    for col in range(last_keep_col, 1, -1):
        ldf = sel_vals_by_col.get(col)
        if ldf is not None and ldf > 0:
            cdf_by_col[col] = ldf * cdf_by_col.get(col + 1, tail_factor)

    # ── Write CDF row after Selected Reasoning ────────────────────────────────
    cdf_row = (last_sel_row or selected_row_num) + 1
    ref_cell = ws.cell(selected_row_num, 1)
    lbl = ws.cell(cdf_row, 1)
    lbl.value = "CDF"
    if ref_cell.has_style:
        lbl.font      = Font(name=ref_cell.font.name, size=ref_cell.font.size, bold=False)
        lbl.border    = copy.copy(ref_cell.border)
        lbl.fill      = copy.copy(ref_cell.fill)
        lbl.alignment = copy.copy(ref_cell.alignment)

    for col in range(2, tail_col + 1):
        if cdf_by_col.get(col) is None:
            continue
        c = ws.cell(cdf_row, col)
        if col == tail_col:
            # Reference the Selected row's tail formula
            c.value = f'={get_column_letter(col)}{selected_row_num}'
        else:
            sel_ref  = f"{get_column_letter(col)}{selected_row_num}"
            next_ref = f"{get_column_letter(col + 1)}{cdf_row}"
            c.value  = f'=IFERROR({sel_ref}*{next_ref},"")'
        c.number_format = _DEC_FMT
        c.font          = DATA_FONT
        c.border        = THIN_BORDER

    # ── Write tail reasoning in Selected Reasoning row at tail column ─────────
    if last_sel_row and last_sel_row > selected_row_num:
        rc = ws.cell(last_sel_row, tail_col)
        if measure:
            # Dynamically discover Tail workbook structure
            tail_path = config.SELECTIONS + "/Chain Ladder Selections - Tail.xlsx"
            t_cutoff_col, t_tail_col, t_reason_col, t_user_row, t_rb_row = _build_tail_workbook_maps(tail_path, measure)
            
            if t_tail_col and t_reason_col and t_user_row and t_rb_row:
                tail_ref = f"'{_CL_TAIL_WB}{measure}'"
                # Use tail column to check if User Selection is populated, then pull from reasoning column
                rc.value = f'=IF({tail_ref}!{t_tail_col}{t_user_row}<>"",{tail_ref}!{t_reason_col}{t_user_row},{tail_ref}!{t_reason_col}{t_rb_row})'
            else:
                # Fallback to static value if structure not found
                rc.value = reasoning
        else:
            rc.value = reasoning
        ref_r = ws.cell(last_sel_row, last_keep_col)
        if ref_r.has_style:
            rc.font   = copy.copy(ref_r.font)
            rc.border = copy.copy(ref_r.border)
            rc.fill   = copy.copy(ref_r.fill)
        rc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")


def create_triangle_sheets(gen_wb, measures):
    # Copy the triangle sheets from the CL selections workbook with formulas
    tail_sel = load_tail_selections(INPUT_TAIL_EXCEL)
    df2_enh = (pd.read_parquet(INPUT_CL_ENHANCED)
               if pathlib.Path(INPUT_CL_ENHANCED).exists()
               else pd.DataFrame())
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        wb_cl_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
        for measure in measures:
            if measure in wb_cl_vals.sheetnames:
                short_name = measure_short_name(measure)[:31]
                ws = gen_wb.create_sheet(title=short_name)
                # Use formula-based copy to link to Chain Ladder Selections - LDFs.xlsx
                _copy_ws_with_ldfs_formulas(wb_cl_form[measure], ws, measure, INPUT_CL_EXCEL)
                # Add CDF formulas below Selected row
                _add_cdf_formulas_to_triangle(ws)
                if measure in tail_sel:
                    cutoff_age, tf, reasoning = tail_sel[measure]
                    add_tail_to_triangle_ws(
                        ws, cutoff_age, tf, reasoning,
                        df2=df2_enh, measure=measure,
                    )



def write_exposure_sheet(gen_wb, triangles_path):
    """
    Write Exposure sheet from triangles parquet.
    Single age per period -> two-column table (Period | Exposure).
    Multiple ages        -> full triangle (Period | age1 | age2 | ...).
    """
    if not pathlib.Path(triangles_path).exists():
        return
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return

    exp["period_int"] = exp["period"].apply(lambda x: _period_int(str(x)))
    exp["age_num"]    = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    exp = exp.sort_values(["period_int", "age_num"])

    ages = sorted(exp["age_num"].dropna().unique())
    ws   = gen_wb.create_sheet(title="Exposure")

    if len(ages) <= 1:
        headers  = ["Period", "Exposure"]
        data_row = _write_title_and_headers(ws, "Exposure", headers, col_width=18)
        for r, (_, row) in enumerate(exp.iterrows(), start=data_row):
            _data_cell(ws.cell(r, 1), row["period_int"])
            _data_cell(ws.cell(r, 2), _safe(row["value"]), _NUM_FMT)
    else:
        pivot = (
            exp.pivot_table(
                index="period_int", columns="age_num",
                values="value", aggfunc="first", observed=True,
            )
            .sort_index()
        )
        pivot.columns = [int(c) for c in pivot.columns]
        headers  = ["Period"] + [str(c) for c in pivot.columns]
        data_row = _write_title_and_headers(ws, "Exposure", headers, col_width=14)
        for r, (idx, row) in enumerate(pivot.iterrows(), start=data_row):
            _data_cell(ws.cell(r, 1), idx)
            for c, val in enumerate(row, start=2):
                _data_cell(ws.cell(r, c), _safe(val), _NUM_FMT)


def write_diagnostics_sheet(ws, combined, exp_map):
    """
    Write the 'Diagnostics' sheet: Ultimate Severity, Loss Rate, Frequency.
    Loss Rate and Frequency columns are omitted when no exposure data is present.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    if inc.empty:
        return

    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    data_row = _write_title_and_headers(ws, "Post-Method Series Diagnostics", headers)

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else str(x)),
    )
    for r, p in enumerate(periods, start=data_row):
        ult_loss   = inc.loc[p, "selected_ultimate"] if p in inc.index else np.nan
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        def _div(a, b):
            if pd.notna(a) and pd.notna(b) and b != 0:
                return a / b
            return None

        _data_cell(ws.cell(r, 1), _period_int(p))
        _data_cell(ws.cell(r, 2), _div(ult_loss, ult_counts), _DEC_FMT)
        if has_exposure:
            _data_cell(ws.cell(r, 3), _div(ult_loss, exp),    _DEC_FMT)
            _data_cell(ws.cell(r, 4), _div(ult_counts, exp),  _DEC_FMT)


def build_post_method_triangle_data(triangles_df, combined):
    """
    Build post-method triangle DataFrames as a list of (sheet_name, index_name, df).
    Each df has period ints as index and age ints as columns.
    """
    if triangles_df.empty:
        return []

    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    def _to_int_series(s):
        try:
            return s.apply(lambda x: int(float(str(x))))
        except (ValueError, TypeError):
            return s.astype(str)

    def x_to_ult(measure, label):
        sub = triangles_df[triangles_df["measure"] == measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        return label, result

    sheets = []
    for measure, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        result = x_to_ult(measure, label)
        if result is not None:
            sheets.append(result)

    def avg_triangle(sub_measure, sheet_name, diff_fn):
        sub = triangles_df[triangles_df["measure"] == sub_measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), sub_measure), np.nan)
            if pd.notna(sel):
                result.loc[period] = diff_fn(sel, result.loc[period])
        return sheet_name, result

    inc_sub  = triangles_df[triangles_df["measure"] == "Incurred Loss"]
    paid_sub = triangles_df[triangles_df["measure"] == "Paid Loss"]

    if not inc_sub.empty:
        r = avg_triangle("Incurred Loss", "Average IBNR", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    if not paid_sub.empty:
        r = avg_triangle("Paid Loss", "Average Unpaid", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    return sheets


def write_triangle_sheet(ws, sheet_name, df):
    """
    Write a triangle DataFrame to a worksheet.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    headers = ["Period"] + [str(c) for c in df.columns]
    data_row = _write_title_and_headers(ws, sheet_name, headers, col_width=14)

    for r, (idx, row) in enumerate(df.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), idx)
        for c, val in enumerate(row, start=2):
            _data_cell(ws.cell(r, c), _safe(val), _DEC_FMT)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def assemble_workbook(output_path, generated_wb):
    """
    Build master workbook from the pre-built generated sheets.
    All content is already computed values — no formula copying needed.
    """
    master = Workbook()
    master.remove(master.active)
    sheet_list = []

    for sname in generated_wb.sheetnames:
        ws_dst = master.create_sheet(title=sname)
        _copy_ws(generated_wb[sname], ws_dst)
        sheet_list.append((sname, *_sheet_desc(sname)))

    notes_ws = master.create_sheet(title="Notes", index=0)
    write_notes_sheet(notes_ws, sheet_list)

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    sel_lookup = load_selections(INPUT_SELECTIONS_EXCEL)

    ult_wb = load_workbook(INPUT_SELECTIONS_EXCEL, data_only=True)
    ult_col_map = {}
    for _sname in ult_wb.sheetnames:
        for _cell in ult_wb[_sname][1]:
            if _cell.value:
                ult_col_map[(_sname, str(_cell.value).strip())] = get_column_letter(_cell.column)
    ult_wb.close()

    combined, available_methods = load_combined(INPUT_ULTIMATES, sel_lookup)
    reason_lookup = load_selection_reasoning(INPUT_SELECTIONS_EXCEL)
    combined["selected_reasoning"] = combined.apply(
        lambda row: reason_lookup.get(
            (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]),
            ""
        ) or "",
        axis=1
    )
    exp_map = get_exposure(INPUT_TRIANGLES)
    has_exposure = bool(exp_map)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]
    
    # Load exposure row map for linking to Chain Ladder LDFs workbook
    exp_row_map = load_exposure_row_map(INPUT_CL_EXCEL)
    
    # We also need triangles_df for diagnostics
    triangles_df = pd.read_parquet(INPUT_TRIANGLES)
    # Load tail selections early to know which ages survive add_tail_to_triangle_ws.
    # _tri_col_map_from_df caps at cutoff so CL actual refs don't point to deleted columns.
    tail_cutoff = {}
    if pathlib.Path(INPUT_TAIL_EXCEL).exists():
        _ts = load_tail_selections(INPUT_TAIL_EXCEL)
        tail_cutoff = {m: info[0] for m, info in _ts.items()}

    print(f"Measures: {measures}")
    print(f"Methods: {available_methods}")
    
    gen_wb = Workbook()
    gen_wb.remove(gen_wb.active)
    
    loss_m  = [m for m in measures if "Loss"  in m]
    count_m = [m for m in measures if "Count" in m]
    other_m = [m for m in measures if m not in loss_m and m not in count_m]
    has_ie  = any(col == "ultimate_ie" for col, _ in available_methods)

    print("Building Loss sheets...")
    if loss_m:
        write_selection_grouped(gen_wb, combined, ["Incurred Loss", "Paid Loss"], "Loss Selection", ult_col_map)
        for m in loss_m:
            write_method_cl(gen_wb, combined, m, ult_col_map, _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)))
        for m in loss_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m, ult_col_map)
        # Create single IE sheet for Incurred Loss only
        if "Incurred Loss" in loss_m and _has_method(combined, "Incurred Loss", "ultimate_ie"):
            write_method_ie(gen_wb, combined, "Incurred Loss", exp_row_map, ult_col_map)

    print("Building Counts sheets...")
    if count_m:
        write_selection_grouped(gen_wb, combined, ["Reported Count", "Closed Count"], "Count Selection", ult_col_map)
        for m in count_m:
            write_method_cl(gen_wb, combined, m, ult_col_map, _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)))
        for m in count_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m, ult_col_map)

    if other_m:
        print("Building other method sheets...")
        for m in other_m:
            write_method_cl(gen_wb, combined, m, ult_col_map, _tri_col_map_from_df(triangles_df, m, max_age=tail_cutoff.get(m)))
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m, ult_col_map)
        
    print("Copying CL LDF sheets (Triangle + Averages)...")
    create_triangle_sheets(gen_wb, measures)
    if has_exposure:
        write_exposure_sheet(gen_wb, INPUT_TRIANGLES)
    
    # Also copy remaining Diagnostics and CV & Slopes sheets
    print("Copying remaining Diag and CV & Slopes sheets...")
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
        for sname in wb_cl_form.sheetnames:
            if sname in ("Diagnostics", "CV & Slopes"):
                ws = gen_wb.create_sheet(title=sname[:31])
                _copy_ws_filtered(wb_cl_form[sname], ws, {}, {})
        wb_cl_form.close()

    print("Building post-method triangle sheets...")
    for sheet_name, df in build_post_method_triangle_data(triangles_df, combined):
        ws_t = gen_wb.create_sheet(title=sheet_name[:31])
        write_triangle_sheet(ws_t, sheet_name, df)

    print("Building diagnostics sheet...")
    ws_diag = gen_wb.create_sheet(title="Summary Diagnostics")
    write_diagnostics_sheet(ws_diag, combined, exp_map)

    print("\nAssembling Analysis.xlsx...")
    assemble_workbook(OUTPUT_COMPLETE, gen_wb)
    print("Done. Run 6b-create-values-only.py to produce the plain-numbers copy.")

if __name__ == "__main__":
    main()
