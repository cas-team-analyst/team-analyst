# Reads projected ultimates, actuary selections, and triangle data to produce the
# Complete Analysis workbooks.
#
# Outputs:
#   Complete Analysis.xlsx            Source Excel files assembled with formulas intact.
#                                     Open in Excel to evaluate any cross-sheet references.
#   Complete Analysis - Values Only.xlsx  Same content as plain computed numbers.
#                                         Safe to read with openpyxl/pandas without requiring
#                                         Excel to re-evaluate formulas first (used by script 7+).
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6-create-complete-analysis.py

import copy  # used by _copy_ws (local full-copy helper)
import os
import pathlib
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import (
    HEADER_FILL, SUBHEADER_FILL,
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)
from modules.xl_selections import (
    SKIP_ROW_LABELS as _SKIP_ROW_LABELS,
    SELECTION_LABELS as _SELECTION_LABELS,
    find_selected_values as _find_selected_values,
    find_selected_reasoning as _find_selected_reasoning,
    copy_ws_filtered as _copy_ws_filtered,
)

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
OUTPUT_PATH            = config.OUTPUT

OUTPUT_COMPLETE = OUTPUT_PATH + "Complete Analysis.xlsx"
OUTPUT_VALUES   = OUTPUT_PATH + "Complete Analysis - Values Only.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS  + "Chain Ladder Selections - LDFs.xlsx"
INPUT_TAIL_EXCEL    = config.SELECTIONS  + "Chain Ladder Selections - Tail.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"

# Unpaid = selected ultimate - latest actual of the proxy measure for that period.
UNPAID_PROXY = {
    "Incurred Loss":  "Paid Loss",
    "Paid Loss":      "Paid Loss",
    "Reported Count": "Closed Count",
    "Closed Count":   "Closed Count",
}

# Preferred display order for method columns — discovered dynamically from parquet.
_METHOD_COLS = [
    ("ultimate_cl", "Chain Ladder"),
    ("ultimate_ie", "Initial Expected"),
    ("ultimate_bf", "BF"),
]

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# ── Selection loading ─────────────────────────────────────────────────────────

def load_selections(excel_path):
    """
    Load final selections from Ultimates.xlsx.
    Priority per (measure, period): User Selection > Rules-Based AI Selection.
    Open-Ended AI is intentionally excluded — it is not a trusted final selection source.
    Columns are located by header name — adapts to layout changes automatically.
    Returns dict {(measure, period): float}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        print(f"  Note: {excel_path} not found -- no selections loaded")
        return {}

    wb = load_workbook(excel_path, data_only=True)
    sel_lookup = {}
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Period":
                period_col = cell.column
            elif v == "User Selection":
                user_col = cell.column
            elif v == "Rules-Based AI Selection":
                rb_col = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in sel_lookup:
                continue
            for col_idx in [user_col, rb_col]:
                if col_idx is None:
                    continue
                v = ws.cell(row=row, column=col_idx).value
                if isinstance(v, (int, float)) and not (isinstance(v, float) and v != v):
                    sel_lookup[key] = float(v)
                    total += 1
                    break

    wb.close()
    print(f"  Loaded {total} selections from {excel_path}")
    return sel_lookup


def load_selection_reasoning(excel_path):
    """
    Load reasoning text for each (measure, period) from Ultimates.xlsx.
    Priority matches load_selections: User Selection reasoning first, then RB-AI.
    Returns {(measure, period): str}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        return {}

    wb = load_workbook(excel_path, data_only=True)
    reason_lookup = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = user_r_col = rb_r_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Period":                   period_col = cell.column
            elif v == "User Selection":         user_col   = cell.column
            elif v == "Rules-Based AI Selection": rb_col   = cell.column
            elif v == "User Reasoning":         user_r_col = cell.column
            elif v == "Rules-Based AI Reasoning": rb_r_col = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in reason_lookup:
                continue

            user_v = ws.cell(row=row, column=user_col).value if user_col else None
            rb_v   = ws.cell(row=row, column=rb_col).value   if rb_col   else None
            is_num = lambda v: isinstance(v, (int, float)) and not (isinstance(v, float) and v != v)

            if is_num(user_v) and user_r_col:
                reason_lookup[key] = ws.cell(row=row, column=user_r_col).value or ""
            elif is_num(rb_v) and rb_r_col:
                reason_lookup[key] = ws.cell(row=row, column=rb_r_col).value or ""

    wb.close()
    return reason_lookup


def load_tail_selections(tail_excel_path):
    """
    Load tail factor selections from Chain Ladder Selections - Tail.xlsx.
    Priority: User Selection (if populated) > Rules-Based AI Selection.
    Returns {measure: (cutoff_age: int, tail_factor: float)}.  Returns {} if file absent.
    """
    path = pathlib.Path(tail_excel_path)
    if not path.exists():
        print(f"  Note: {tail_excel_path} not found -- no tail selections loaded")
        return {}

    wb = load_workbook(tail_excel_path, data_only=True)
    tail_map = {}

    for measure in wb.sheetnames:
        ws = wb[measure]
        in_tail = False
        cutoff_col = tail_col = reason_col = None
        user_entry = rb_entry = None

        for row in ws.iter_rows():
            col1 = row[0].value if row else None
            if col1 == "Tail Factor Selection":
                in_tail = True
                continue
            if not in_tail:
                continue
            if col1 == "Label":
                for cell in row:
                    if cell.value == "Cutoff Age":
                        cutoff_col = cell.column
                    elif cell.value == "Tail Factor":
                        tail_col = cell.column
                    elif cell.value == "Reasoning":
                        reason_col = cell.column
                continue
            if cutoff_col is None or tail_col is None:
                continue
            if col1 in ("User Selection", "Rules-Based AI Selection"):
                cv = row[cutoff_col - 1].value
                tv = row[tail_col - 1].value
                rv = row[reason_col - 1].value if reason_col else None
                if cv is not None and tv is not None:
                    try:
                        entry = (int(float(str(cv))), float(str(tv)), rv)
                        if col1 == "User Selection":
                            user_entry = entry
                        elif rb_entry is None:
                            rb_entry = entry
                    except (ValueError, TypeError):
                        pass

        chosen = user_entry if user_entry is not None else rb_entry
        if chosen:
            tail_map[measure] = chosen

    wb.close()
    summary = {m: (v[0], v[1]) for m, v in tail_map.items()}
    print(f"  Loaded tail selections: {summary}")
    return tail_map


# ── Data loading ──────────────────────────────────────────────────────────────

def load_combined(ultimates_path, sel_lookup):
    """
    Load projected ultimates, apply final selections, compute IBNR and Unpaid.
    Methods (CL/IE/BF) are discovered dynamically from non-empty parquet columns.
    Raises ValueError for any non-Exposure (measure, period) missing from sel_lookup.
    Returns (df, available_methods) where available_methods = [(col, label), ...].
    """
    df = pd.read_parquet(ultimates_path)
    df["period"]  = df["period"].astype(str)
    df["measure"] = df["measure"].astype(str)

    available_methods = [
        (col, label) for col, label in _METHOD_COLS
        if col in df.columns and df[col].notna().any()
    ]

    # Validate: every non-Exposure row must have a selection.
    missing = [
        (row["measure"], row["period"])
        for _, row in df.iterrows()
        if row["measure"] != "Exposure" and (row["measure"], row["period"]) not in sel_lookup
    ]
    if missing:
        lines = "\n  ".join(f"{m} / {p}" for m, p in missing)
        raise ValueError(
            f"{len(missing)} (measure, period) pair(s) have no User Selection or "
            f"Rules-Based AI Selection in Ultimates.xlsx:\n  {lines}\n"
            "Populate Rules-Based AI Selection (run 5b) or add a User Selection."
        )

    actual_lookup = df.set_index(["period", "measure"])["actual"].to_dict()

    def _pick_unpaid(row):
        proxy = UNPAID_PROXY.get(row["measure"])
        if proxy is None:
            return np.nan
        proxy_actual = actual_lookup.get((row["period"], proxy), np.nan)
        if pd.isna(proxy_actual) or pd.isna(row["selected_ultimate"]):
            return np.nan
        return row["selected_ultimate"] - proxy_actual

    df["selected_ultimate"] = df.apply(
        lambda row: sel_lookup.get((row["measure"], row["period"]), np.nan), axis=1
    )
    df["selected_ibnr"]   = df["selected_ultimate"] - df["actual"]
    df["selected_unpaid"] = df.apply(_pick_unpaid, axis=1)

    return df, available_methods


def get_exposure(triangles_path):
    """Latest Exposure value per period from triangles. Returns {} if absent."""
    if not pathlib.Path(triangles_path).exists():
        return {}
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return {}
    exp["period"]  = exp["period"].astype(str)
    exp["age_num"] = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    latest = exp.sort_values("age_num").groupby("period").last()
    return latest["value"].to_dict()


def get_triangles(triangles_path):
    """Load triangle data. Returns empty DataFrame when file is absent."""
    if not pathlib.Path(triangles_path).exists():
        return pd.DataFrame()
    tri = pd.read_parquet(triangles_path)
    tri["period"]  = tri["period"].astype(str)
    tri["measure"] = tri["measure"].astype(str)
    return tri


# ── Cell / layout helpers ─────────────────────────────────────────────────────

def _safe(v):
    """Convert NaN/NA to None so openpyxl writes a blank cell."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _to_py(v):
    """
    Coerce a value to a plain Python type so openpyxl writes a number, not text.
    - numpy scalars → int / float
    - numeric strings ('120', '1.05') → int / float
    - everything else unchanged
    """
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.floating):
        return float(v)
    if isinstance(v, str):
        try:
            f = float(v)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            pass
    return v


def _period_int(v):
    """Display a period as int when it is a whole number, else return as-is."""
    try:
        f = float(v)
        if f == int(f):
            return int(f)
    except (ValueError, TypeError):
        pass
    return v


def _data_cell(cell, value, num_fmt=None):
    """Write a styled data cell with border and alignment."""
    value = _to_py(_safe(value))
    cell.value     = value
    cell.font      = DATA_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="right" if isinstance(value, (int, float)) else "left",
        vertical="center",
    )
    if num_fmt and value is not None and isinstance(value, (int, float)):
        cell.number_format = num_fmt


def _write_title_and_headers(ws, title, headers, col_width=18):
    """
    Write a title row (row 1) and a styled header row (row 2).
    Returns 3 — the first data row.
    This format matches script 7's read_with_title() convention.
    """
    title_cell = ws.cell(1, 1, title)
    style_header(title_cell, "header")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 20

    for c, text in enumerate(headers, 1):
        cell = ws.cell(2, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A3"
    return 3


def _write_headers(ws, headers, col_width=18):
    """
    Write a single styled header row (row 1), no title row.
    Returns 2 — the first data row.
    Matches script 7's read_no_title() convention ("Sel - " sheets).
    """
    for c, text in enumerate(headers, 1):
        cell = ws.cell(1, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A2"
    return 2


# ── Generated sheet writers ───────────────────────────────────────────────────

def _has_method(combined, measure, col):
    """True if col has at least one non-NaN value for this measure."""
    return col in combined.columns and combined[combined["measure"] == measure][col].notna().any()


def measure_short_name(measure):
    """Strip ' Loss' or ' Count' for sheet names."""
    return measure.replace(" Loss", "").replace(" Count", "")

def _get_cdf_cell_ref(ws_triangle, target_age):
    # Search row 1 or 2 for the age
    for row_idx in [1, 2]:
        for cell in ws_triangle[row_idx]:
            try:
                if cell.value is not None and int(float(cell.value)) == int(float(target_age)):
                    # Found the column! Now find CDF row (usually 'CDF' in column A)
                    for r in range(1, 100):
                        if str(ws_triangle.cell(row=r, column=1).value).strip() == "CDF":
                            return f"'{ws_triangle.title}'!{get_column_letter(cell.column)}{r}"
            except (ValueError, TypeError):
                pass
    return None

def _formula_cell(ws, r, c, formula, num_fmt):
    cell = ws.cell(r, c, formula)
    cell.number_format = num_fmt
    cell.font = DATA_FONT
    cell.border = THIN_BORDER


def write_method_cl(gen_wb, combined, measure):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title=f"{short_name} CL"[:31])
    headers = ["Accident Period", "Current Age", short_name, "CDF", "Ultimate", "IBNR", "Unpaid"]
    _write_headers(ws, headers)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])
        _data_cell(ws.cell(r, 3), row["actual"], _NUM_FMT)
        _data_cell(ws.cell(r, 4), row.get("cdf"), _DEC_FMT)
        _formula_cell(ws, r, 5, f"=C{r}*D{r}", _NUM_FMT)
        _formula_cell(ws, r, 6, f"=E{r}-C{r}", _NUM_FMT)
        if proxy_exists:
            _formula_cell(ws, r, 7, f"=E{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT)
        else:
            _formula_cell(ws, r, 7, f"=E{r}-C{r}", _NUM_FMT)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell(ws.cell(t, 1), "Total")
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 5, f"=SUM(E2:E{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT)

def write_method_bf(gen_wb, combined, measure):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title=f"{short_name} BF"[:31])
    
    # bf canonical structure: 
    # Accident Period, Current Age, Initial Expected, CDF, % Unreported, Unreported, Actual, Ultimate, IBNR, Unpaid
    headers = ["Accident Period", "Current Age", "Initial Expected", "CDF", "% Unreported", "Unreported", short_name, "Ultimate", "IBNR", "Unpaid"]
    _write_headers(ws, headers)
    
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    ws_triangle = gen_wb[short_name] if short_name in gen_wb.sheetnames else None
    
    proxy = UNPAID_PROXY.get(measure)
    proxy_exists = proxy and proxy != measure and proxy in combined["measure"].unique()

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])
        _data_cell(ws.cell(r, 3), row.get("ultimate_ie", np.nan), _NUM_FMT)
        _data_cell(ws.cell(r, 4), row.get("cdf"), _DEC_FMT)
        _formula_cell(ws, r, 5, f"=1-(1/D{r})", "0.0%")
        _formula_cell(ws, r, 6, f"=C{r}*E{r}", _NUM_FMT)
        _data_cell(ws.cell(r, 7), row["actual"], _NUM_FMT)
        _formula_cell(ws, r, 8, f"=F{r}+G{r}", _NUM_FMT)
        _formula_cell(ws, r, 9, f"=H{r}-G{r}", _NUM_FMT)
        if proxy_exists:
            _formula_cell(ws, r, 10, f"=H{r}-'{measure_short_name(proxy)} CL'!C{r}", _NUM_FMT)
        else:
            _formula_cell(ws, r, 10, f"=H{r}-G{r}", _NUM_FMT)

    t = 3 + len(sub)  # blank row at t-1, totals at t
    _data_cell(ws.cell(t, 1), "Total")
    _formula_cell(ws, t, 3, f"=SUM(C2:C{t-2})", _NUM_FMT)
    # Col 5: % Unreported total = weighted avg = SUM(F)/SUM(C)
    _formula_cell(ws, t, 5, f"=SUM(F2:F{t-2})/SUM(C2:C{t-2})", "0.0%")
    _formula_cell(ws, t, 6, f"=SUM(F2:F{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 7, f"=SUM(G2:G{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 8, f"=SUM(H2:H{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 9, f"=SUM(I2:I{t-2})", _NUM_FMT)
    _formula_cell(ws, t, 10, f"=SUM(J2:J{t-2})", _NUM_FMT)


def write_method_ie(gen_wb, combined, measure, exp_map):
    short_name = measure_short_name(measure)
    ws = gen_wb.create_sheet(title=f"{short_name} IE"[:31])
    headers = ["Accident Period", "Current Age", "Exposure", "Selected Loss Rate", "IE Ultimate"]
    _write_headers(ws, headers)

    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])

        exp = exp_map.get(row["period"], np.nan)
        _data_cell(ws.cell(r, 3), _safe(exp), _NUM_FMT)

        ult_ie = row.get("ultimate_ie", np.nan)
        if pd.notna(ult_ie) and pd.notna(exp) and exp != 0:
            elr = float(ult_ie) / float(exp)
        else:
            elr = np.nan
        _data_cell(ws.cell(r, 4), _safe(elr), _DEC_FMT)
        _data_cell(ws.cell(r, 5), _safe(ult_ie), _NUM_FMT)


def write_selection_grouped(gen_wb, combined, measures_group, title):
    ws = gen_wb.create_sheet(title=title)
    
    # columns e.g.: Accident Period, Current Age, Incurred, Paid, Incurred CL, Paid CL, Incurred BF, Paid BF, Initial Expected, Selected Ultimate, IBNR, Unpaid
    
    # Find active measures in this group
    active_m = [m for m in measures_group if m in combined["measure"].unique()]
    if not active_m:
        return
        
    main_m = active_m[0]
    sub = combined[combined["measure"] == main_m].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    
    short_names = [measure_short_name(m) for m in active_m]
    bf_m  = [m for m in active_m if _has_method(combined, m, "ultimate_bf")]
    has_group_ie = any(_has_method(combined, m, "ultimate_ie") for m in active_m)

    headers = ["Accident Period", "Current Age"]
    for s in short_names:
        headers.append(s)
    for s in short_names:
        headers.append(f"{s} CL")
    for m in bf_m:
        headers.append(f"{measure_short_name(m)} BF")
    if has_group_ie:
        headers.append("Initial Expected")
    headers.extend(["Selected Ultimate", "IBNR", "Unpaid"])

    _write_headers(ws, headers)

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        _data_cell(ws.cell(r, 1), row["period_int"])
        _data_cell(ws.cell(r, 2), row["current_age"])

        col_idx = 3
        for s in short_names:
            _formula_cell(ws, r, col_idx, f"='{s} CL'!C{r}", _NUM_FMT)
            col_idx += 1
        for s in short_names:
            _formula_cell(ws, r, col_idx, f"='{s} CL'!E{r}", _NUM_FMT)
            col_idx += 1
        for m in bf_m:
            _formula_cell(ws, r, col_idx, f"='{measure_short_name(m)} BF'!H{r}", _NUM_FMT)
            col_idx += 1
        if has_group_ie:
            _data_cell(ws.cell(r, col_idx), row.get("ultimate_ie", np.nan), _NUM_FMT)
            col_idx += 1
        _data_cell(ws.cell(r, col_idx), row["selected_ultimate"], _NUM_FMT)
        col_idx += 1
        _formula_cell(ws, r, col_idx, f"={get_column_letter(col_idx-1)}{r}-C{r}", _NUM_FMT)
        col_idx += 1
        unpaid_actual = "D" if len(active_m) > 1 else "C"
        _formula_cell(ws, r, col_idx, f"={get_column_letter(col_idx-2)}{r}-{unpaid_actual}{r}", _NUM_FMT)

    # Totals row — SUM every numeric column, skip Period and Age
    n_cols = 2 + len(active_m) * 2 + len(bf_m) + (1 if has_group_ie else 0) + 3
    t = 3 + len(sub)
    _data_cell(ws.cell(t, 1), "Total")
    for c in range(3, n_cols + 1):
        col_letter = get_column_letter(c)
        _formula_cell(ws, t, c, f"=SUM({col_letter}2:{col_letter}{t-2})", _NUM_FMT)


def add_tail_to_triangle_ws(ws, cutoff_age, tail_factor, reasoning=None,
                            df2=None, measure=None):
    """
    After _copy_ws_filtered:
      1. Delete all interval columns whose ending age > cutoff_age.
      2. Append a {cutoff_age}-Ult column header to ATA / Averages / LDF sections.
      3. Fill the ATA data rows at the tail column with the observed cumulative
         factor for each period that has development beyond the cutoff age
         (product of per-period ATAs from cutoff_age onward, sourced from df2).
      4. Write a CDF row in LDF Selections with right-to-left cumulative factors.
      5. Write the tail reasoning in the Selected Reasoning row at the tail column.
    """
    tail_label = f"{cutoff_age}-Ult"

    # ── Single pass: locate section header rows and ATA data rows ────────────
    in_ata = in_avg = in_ldf = False
    in_ata_data = False
    ata_hdr_row = avg_hdr_row = ldf_hdr_row = None
    ata_data_rows = {}          # {period_str: row_num} for ATA section
    avg_data_row_nums = []      # row nums of averages data rows (to clear at tail_col)
    ldf_option_row_nums = []    # row nums of LDF section rows that aren't Selected/Reasoning
    last_keep_col = None
    selected_row_num = None
    last_sel_row = None
    sel_vals_by_col = {}

    for row_cells in ws.iter_rows():
        row_num = row_cells[0].row
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Age-to-Age Factors":
            in_ata = True; in_avg = in_ldf = False; in_ata_data = False; continue
        if col1 == "Averages":
            in_avg = True; in_ata = in_ldf = False; in_ata_data = False; continue
        if col1 == "LDF Selections":
            in_ldf = True; in_ata = in_avg = False; in_ata_data = False; continue

        if in_ata:
            if ata_hdr_row is None:
                if col1 is None:
                    ata_hdr_row = row_num
                    in_ata_data = True
            elif in_ata_data:
                if col1 is None:
                    in_ata_data = False
                else:
                    try:
                        ata_data_rows[str(int(float(str(col1))))] = row_num
                    except (ValueError, TypeError):
                        pass
            continue

        if in_avg:
            if avg_hdr_row is None and col1 in ("Metric", None):
                avg_hdr_row = row_num
            elif avg_hdr_row is not None and col1 not in (None, ""):
                avg_data_row_nums.append(row_num)
            continue

        if in_ldf:
            if ldf_hdr_row is None:
                if col1 is None:
                    ldf_hdr_row = row_num
                    for cell in row_cells:
                        v = cell.value
                        if isinstance(v, str) and '-' in v:
                            parts = v.split('-')
                            if len(parts) == 2:
                                try:
                                    if int(parts[1]) <= cutoff_age:
                                        last_keep_col = cell.column
                                except ValueError:
                                    pass
            elif col1 == "Selected":
                selected_row_num = row_num
                last_sel_row = row_num
                sel_vals_by_col = {
                    c.column: float(c.value)
                    for c in row_cells[1:]
                    if c.value is not None
                    and isinstance(c.value, (int, float))
                    and (last_keep_col is None or c.column <= last_keep_col)
                }
            elif col1 == "Selected Reasoning":
                last_sel_row = row_num
            elif col1 is not None and col1 != "":
                ldf_option_row_nums.append(row_num)
            continue

    if last_keep_col is None:
        return

    # ── Column layout ─────────────────────────────────────────────────────────
    # Each ATA column K stores formula (K+1)/K, so the "191-203" interval at
    # last_keep_col references last_keep_col+1 as its numerator (age-203 data).
    # We must KEEP last_keep_col+1 or that formula breaks.
    # tail_col = last_keep_col+1 (the age-203 data column, relabelled to tail).
    # Delete everything from last_keep_col+2 onward.
    tail_col   = last_keep_col + 1
    del_start  = last_keep_col + 2
    del_count  = ws.max_column - last_keep_col - 1
    if del_count > 0:
        ws.delete_cols(del_start, del_count)

    ws.column_dimensions[get_column_letter(tail_col)].width = 14

    # ── Overwrite stale content at tail_col ───────────────────────────────────
    # ATA data rows: formula was (col tail_col+1)/col tail_col → now broken;
    # clear so Excel doesn't show a shifted wrong reference.
    for row_num in ata_data_rows.values():
        ws.cell(row_num, tail_col).value = None

    # Averages data rows: were averages of "203-215" ATAs — blank for tail col.
    for row_num in avg_data_row_nums:
        ws.cell(row_num, tail_col).value = None

    # LDF option rows (Vol Wtd, Simple, etc.): stale "203-215" LDFs — blank.
    for row_num in ldf_option_row_nums:
        ws.cell(row_num, tail_col).value = None

    # ── Replace section headers at tail_col with tail label ──────────────────
    for hdr_row in (ata_hdr_row, avg_hdr_row, ldf_hdr_row):
        if hdr_row is None:
            continue
        ref = ws.cell(hdr_row, last_keep_col)
        new = ws.cell(hdr_row, tail_col)
        new.value = tail_label
        if ref.has_style:
            new.font      = copy.copy(ref.font)
            new.border    = copy.copy(ref.border)
            new.fill      = copy.copy(ref.fill)
            new.number_format = ref.number_format
            new.alignment = copy.copy(ref.alignment)

    if selected_row_num is None:
        return

    # Set Selected row at tail_col to the tail_factor (replacing stale interval LDF).
    c = ws.cell(selected_row_num, tail_col)
    c.value = tail_factor
    c.number_format = _DEC_FMT

    # ── Compute cumulative CDFs right-to-left, seeded by tail_factor ─────────
    cdf_by_col = {tail_col: tail_factor}
    for col in range(last_keep_col, 1, -1):
        ldf = sel_vals_by_col.get(col)
        if ldf is not None and ldf > 0:
            cdf_by_col[col] = ldf * cdf_by_col.get(col + 1, tail_factor)

    # ── Write CDF row after Selected Reasoning ────────────────────────────────
    cdf_row = (last_sel_row or selected_row_num) + 1
    ref_cell = ws.cell(selected_row_num, 1)
    lbl = ws.cell(cdf_row, 1)
    lbl.value = "CDF"
    if ref_cell.has_style:
        lbl.font      = copy.copy(ref_cell.font)
        lbl.border    = copy.copy(ref_cell.border)
        lbl.fill      = copy.copy(ref_cell.fill)
        lbl.alignment = copy.copy(ref_cell.alignment)

    for col in range(2, tail_col + 1):
        if cdf_by_col.get(col) is None:
            continue
        c = ws.cell(cdf_row, col)
        if col == tail_col:
            c.value = tail_factor          # tail: literal seed value
        else:
            sel_ref  = f"{get_column_letter(col)}{selected_row_num}"
            next_ref = f"{get_column_letter(col + 1)}{cdf_row}"
            c.value  = f'=IFERROR({sel_ref}*{next_ref},"")'
        c.number_format = _DEC_FMT
        c.font          = DATA_FONT
        c.border        = THIN_BORDER

    # ── Write tail reasoning in Selected Reasoning row at tail column ─────────
    if reasoning and last_sel_row and last_sel_row > selected_row_num:
        rc = ws.cell(last_sel_row, tail_col)
        rc.value = reasoning
        ref_r = ws.cell(last_sel_row, last_keep_col)
        if ref_r.has_style:
            rc.font   = copy.copy(ref_r.font)
            rc.border = copy.copy(ref_r.border)
            rc.fill   = copy.copy(ref_r.fill)
        rc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")


def create_triangle_sheets(gen_wb, measures):
    # Copy the triangle sheets from the CL selections workbook
    tail_sel = load_tail_selections(INPUT_TAIL_EXCEL)
    df2_enh = (pd.read_parquet(INPUT_CL_ENHANCED)
               if pathlib.Path(INPUT_CL_ENHANCED).exists()
               else pd.DataFrame())
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        wb_cl_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
        for measure in measures:
            if measure in wb_cl_vals.sheetnames:
                sel_vals   = _find_selected_values(wb_cl_vals[measure])
                sel_reason = _find_selected_reasoning(wb_cl_vals[measure])
                short_name = measure_short_name(measure)[:31]
                ws = gen_wb.create_sheet(title=short_name)
                _copy_ws_filtered(wb_cl_form[measure], ws, sel_vals, sel_reason)
                if measure in tail_sel:
                    cutoff_age, tf, reasoning = tail_sel[measure]
                    add_tail_to_triangle_ws(
                        ws, cutoff_age, tf, reasoning,
                        df2=df2_enh, measure=measure,
                    )



def write_exposure_sheet(gen_wb, triangles_path):
    """
    Write Exposure sheet from triangles parquet.
    Single age per period -> two-column table (Period | Exposure).
    Multiple ages        -> full triangle (Period | age1 | age2 | ...).
    """
    if not pathlib.Path(triangles_path).exists():
        return
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return

    exp["period_int"] = exp["period"].apply(lambda x: _period_int(str(x)))
    exp["age_num"]    = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    exp = exp.sort_values(["period_int", "age_num"])

    ages = sorted(exp["age_num"].dropna().unique())
    ws   = gen_wb.create_sheet(title="Exposure")

    if len(ages) <= 1:
        headers  = ["Period", "Exposure"]
        data_row = _write_title_and_headers(ws, "Exposure", headers, col_width=18)
        for r, (_, row) in enumerate(exp.iterrows(), start=data_row):
            _data_cell(ws.cell(r, 1), row["period_int"])
            _data_cell(ws.cell(r, 2), _safe(row["value"]), _NUM_FMT)
    else:
        pivot = (
            exp.pivot_table(
                index="period_int", columns="age_num",
                values="value", aggfunc="first", observed=True,
            )
            .sort_index()
        )
        pivot.columns = [int(c) for c in pivot.columns]
        headers  = ["Period"] + [str(c) for c in pivot.columns]
        data_row = _write_title_and_headers(ws, "Exposure", headers, col_width=14)
        for r, (idx, row) in enumerate(pivot.iterrows(), start=data_row):
            _data_cell(ws.cell(r, 1), idx)
            for c, val in enumerate(row, start=2):
                _data_cell(ws.cell(r, c), _safe(val), _NUM_FMT)


def write_diagnostics_sheet(ws, combined, exp_map):
    """
    Write the 'Diagnostics' sheet: Ultimate Severity, Loss Rate, Frequency.
    Loss Rate and Frequency columns are omitted when no exposure data is present.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    if inc.empty:
        return

    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    data_row = _write_title_and_headers(ws, "Post-Method Series Diagnostics", headers)

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else str(x)),
    )
    for r, p in enumerate(periods, start=data_row):
        ult_loss   = inc.loc[p, "selected_ultimate"] if p in inc.index else np.nan
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        def _div(a, b):
            if pd.notna(a) and pd.notna(b) and b != 0:
                return a / b
            return None

        _data_cell(ws.cell(r, 1), _period_int(p))
        _data_cell(ws.cell(r, 2), _div(ult_loss, ult_counts), _DEC_FMT)
        if has_exposure:
            _data_cell(ws.cell(r, 3), _div(ult_loss, exp),    _DEC_FMT)
            _data_cell(ws.cell(r, 4), _div(ult_counts, exp),  _DEC_FMT)


def build_post_method_triangle_data(triangles_df, combined):
    """
    Build post-method triangle DataFrames as a list of (sheet_name, index_name, df).
    Each df has period ints as index and age ints as columns.
    """
    if triangles_df.empty:
        return []

    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    def _to_int_series(s):
        try:
            return s.apply(lambda x: int(float(str(x))))
        except (ValueError, TypeError):
            return s.astype(str)

    def x_to_ult(measure, label):
        sub = triangles_df[triangles_df["measure"] == measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        return label, result

    sheets = []
    for measure, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        result = x_to_ult(measure, label)
        if result is not None:
            sheets.append(result)

    def avg_triangle(sub_measure, sheet_name, diff_fn):
        sub = triangles_df[triangles_df["measure"] == sub_measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), sub_measure), np.nan)
            if pd.notna(sel):
                result.loc[period] = diff_fn(sel, result.loc[period])
        return sheet_name, result

    inc_sub  = triangles_df[triangles_df["measure"] == "Incurred Loss"]
    paid_sub = triangles_df[triangles_df["measure"] == "Paid Loss"]

    if not inc_sub.empty:
        r = avg_triangle("Incurred Loss", "Average IBNR", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    if not paid_sub.empty:
        r = avg_triangle("Paid Loss", "Average Unpaid", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    return sheets


def write_triangle_sheet(ws, sheet_name, df):
    """
    Write a triangle DataFrame to a worksheet.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    headers = ["Period"] + [str(c) for c in df.columns]
    data_row = _write_title_and_headers(ws, sheet_name, headers, col_width=14)

    for r, (idx, row) in enumerate(df.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), idx)
        for c, val in enumerate(row, start=2):
            _data_cell(ws.cell(r, c), _safe(val), _DEC_FMT)


# ── Filtered copy of source selection workbooks ───────────────────────────────
# _find_selected_values and _copy_ws_filtered imported from modules.xl_selections.


def _fill_method_cl_values(ws, measure, combined, actual_lookup):
    """Replace CL method sheet formula cells with computed values."""
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    proxy = UNPAID_PROXY.get(measure)

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        ult = row.get("ultimate_cl", np.nan)
        ult = None if pd.isna(ult) else float(ult)
        actual = row["actual"]
        actual = None if pd.isna(actual) else float(actual)

        ws.cell(r, 5).value = ult
        ws.cell(r, 6).value = (ult - actual) if ult is not None and actual is not None else None
        if proxy and proxy != measure:
            proxy_actual = actual_lookup.get((row["period"], proxy))
            ws.cell(r, 7).value = (ult - proxy_actual) if ult is not None and proxy_actual is not None else None
        else:
            ws.cell(r, 7).value = (ult - actual) if ult is not None and actual is not None else None


def _fill_method_bf_values(ws, measure, combined, actual_lookup):
    """Replace BF method sheet formula cells with computed values."""
    sub = combined[combined["measure"] == measure].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")
    proxy = UNPAID_PROXY.get(measure)

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        cdf = row.get("cdf", np.nan)
        cdf = None if pd.isna(cdf) else float(cdf)
        actual = row["actual"]
        actual = None if pd.isna(actual) else float(actual)
        ult_bf = row.get("ultimate_bf", np.nan)
        ult_bf = None if pd.isna(ult_bf) else float(ult_bf)
        ult_ie = row.get("ultimate_ie", np.nan)
        ult_ie = None if pd.isna(ult_ie) else float(ult_ie)

        pct_unr = (1.0 - 1.0 / cdf) if cdf is not None and cdf != 0 else None
        unreported = (ult_bf - actual) if ult_bf is not None and actual is not None else None

        ws.cell(r, 5).value = pct_unr
        ws.cell(r, 6).value = unreported
        ws.cell(r, 8).value = ult_bf
        ws.cell(r, 9).value = (ult_bf - actual) if ult_bf is not None and actual is not None else None
        if proxy and proxy != measure:
            proxy_actual = actual_lookup.get((row["period"], proxy))
            ws.cell(r, 10).value = (ult_bf - proxy_actual) if ult_bf is not None and proxy_actual is not None else None
        else:
            ws.cell(r, 10).value = (ult_bf - actual) if ult_bf is not None and actual is not None else None


def _fill_selection_values(ws, measures_group, combined, actual_lookup):
    """Replace Selection sheet formula cells with computed values."""
    active_m = [m for m in measures_group if m in combined["measure"].unique()]
    if not active_m:
        return

    main_m = active_m[0]
    sub = combined[combined["measure"] == main_m].copy()
    sub["period_int"] = sub["period"].apply(_period_int)
    sub = sub.sort_values("period_int")

    ult_cl = {m: combined[combined["measure"] == m].set_index("period")["ultimate_cl"].to_dict() for m in active_m}
    ult_bf = {m: combined[combined["measure"] == m].set_index("period")["ultimate_bf"].to_dict() for m in active_m}
    actuals = {m: combined[combined["measure"] == m].set_index("period")["actual"].to_dict() for m in active_m}

    n = len(active_m)
    bf_m = [m for m in active_m if _has_method(combined, m, "ultimate_bf")]
    has_group_ie = any(_has_method(combined, m, "ultimate_ie") for m in active_m)

    for r, (_, row) in enumerate(sub.iterrows(), start=2):
        period = row["period"]
        col = 3
        for m in active_m:
            ws.cell(r, col).value = actuals[m].get(period)
            col += 1
        for m in active_m:
            v = ult_cl[m].get(period)
            ws.cell(r, col).value = None if v is None or (isinstance(v, float) and v != v) else v
            col += 1
        for m in bf_m:
            v = ult_bf[m].get(period)
            ws.cell(r, col).value = None if v is None or (isinstance(v, float) and v != v) else v
            col += 1
        if has_group_ie:
            col += 1  # Initial Expected already hardcoded
        # Selected Ultimate already hardcoded
        sel_ult = row["selected_ultimate"]
        sel_ult = None if pd.isna(sel_ult) else float(sel_ult)
        first_act = actuals[active_m[0]].get(period)
        ws.cell(r, col).value = (sel_ult - first_act) if sel_ult is not None and first_act is not None else None
        col += 1
        proxy_act = actuals[active_m[1]].get(period) if n > 1 else first_act
        ws.cell(r, col).value = (sel_ult - proxy_act) if sel_ult is not None and proxy_act is not None else None


def _fill_cdf_row_values(ws):
    """
    Resolve CDF formula strings in a triangle sheet's LDF Selections section to
    Python-computed values.  Used for the Values Only output.
    Right-to-left product: Selected[col] * CDF[col+1], seeded by the tail literal.
    """
    in_ldf = False
    selected_vals = {}
    cdf_row_num = None
    tail_literal_col = None

    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None
        if col1 == "LDF Selections":
            in_ldf = True; continue
        if not in_ldf:
            continue
        if col1 == "Selected":
            selected_vals = {
                c.column: float(c.value)
                for c in row_cells[1:]
                if isinstance(c.value, (int, float))
            }
        elif col1 == "CDF":
            cdf_row_num = row_cells[0].row
            for c in reversed(row_cells[1:]):
                if isinstance(c.value, (int, float)):
                    tail_literal_col = c.column
                    break
            break

    if cdf_row_num is None or not selected_vals or tail_literal_col is None:
        return

    prev_cdf = ws.cell(cdf_row_num, tail_literal_col).value
    for col in range(tail_literal_col - 1, 1, -1):
        cell = ws.cell(cdf_row_num, col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            ldf = selected_vals.get(col)
            if ldf is not None and prev_cdf is not None:
                prev_cdf = ldf * prev_cdf
                cell.value = round(prev_cdf, 6)
            else:
                cell.value = None
                prev_cdf = None
        elif isinstance(cell.value, (int, float)):
            prev_cdf = cell.value


def _strip_formulas(ws):
    """Replace formula strings with None so downstream readers see blank, not a formula string."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None


def _fill_cl_main_values(ws, measure, df2, df4):
    """
    Replace formula strings in a CL main-measure sheet with Python-computed values.
    Only called for main measure sheets (e.g. Incurred Loss) -- Diag-* and CV-&-Slopes
    sheets have no formula cells.

    ATA section: each period/interval cell replaced from df2['ldf'].
    Averages section: each display-name/interval cell replaced from df4.

    Sheet structure in gen_wb after _copy_ws_filtered:
      Loss triangle title -> header -> data rows -> blank
      "Age-to-Age Factors" title -> "" header -> ATA data rows (formulas) -> blank
      "Averages" title -> "Metric" header -> avg data rows (formulas) -> blank
      "LDF Selections" title -> header -> "Selected" row
    """
    df_m = df2[df2['measure'].astype(str) == measure].copy()

    ata_lookup = {
        (str(r['period']), str(r['interval'])): r['ldf']
        for _, r in df_m[df_m['ldf'].notna()].iterrows()
    }

    avg_lookup = {}
    if df4 is not None and not df4.empty:
        df4_m = df4[df4['measure'].astype(str) == measure].copy()
        avg_data_cols = [c for c in df4_m.columns
                         if c not in ('measure', 'interval')
                         and not c.startswith('cv_')
                         and not c.startswith('slope_')]
        for _, r in df4_m.iterrows():
            intv = str(r['interval'])
            for col in avg_data_cols:
                display = col.replace('avg_exclude_high_low', 'exclude_high_low')
                avg_lookup[(display, intv)] = r[col]

    section = None
    col_headers = {}

    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Age-to-Age Factors":
            section = "ata_pre_header"
            continue
        if col1 == "Averages":
            section = "avg_pre_header"
            continue
        if col1 == "LDF Selections":
            break

        if section == "ata_pre_header":
            col_headers = {c.column: str(c.value) for c in row_cells[1:] if c.value not in (None, "")}
            section = "ata"
            continue

        if section == "avg_pre_header":
            col_headers = {c.column: str(c.value) for c in row_cells[1:] if c.value not in (None, "")}
            section = "avg"
            continue

        if section == "ata":
            if col1 is None or col1 == "":
                section = None
                continue
            period = str(col1)
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_headers.get(cell.column)
                    cell.value = ata_lookup.get((period, intv)) if intv else None

        elif section == "avg":
            if col1 is None or col1 == "":
                section = None
                continue
            display = str(col1)
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_headers.get(cell.column)
                    cell.value = avg_lookup.get((display, intv)) if intv else None


def _fill_tail_values(ws, measure, df2):
    """
    Replace formula strings in a Tail sheet with Python-computed values.
    Only the "Average" and "CV" rows contain formulas; they summarise the observed
    ATA factors in the column above.  Header row col1 = "Accident Year" identifies
    the age columns so we can map them to df2 intervals.
    """
    df_m = df2[df2['measure'].astype(str) == measure].copy()

    # Map "to" age string -> interval string  e.g. "23" -> "11-23"
    to_age_map = {}
    for intv in df_m['interval'].dropna().unique():
        parts = str(intv).split('-')
        if len(parts) == 2:
            to_age_map[parts[1]] = str(intv)

    def _mean(intv):
        vals = df_m[df_m['interval'].astype(str) == intv]['ldf'].dropna()
        return vals.mean() if not vals.empty else None

    def _cv(intv):
        vals = df_m[df_m['interval'].astype(str) == intv]['ldf'].dropna()
        if vals.empty:
            return None
        m = vals.mean()
        return (vals.std() / m) if m and m != 0 else None

    col_to_intv = {}
    header_found = False

    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Accident Year":
            col_to_intv = {c.column: to_age_map.get(str(c.value))
                           for c in row_cells[1:] if c.value is not None}
            header_found = True
            continue

        if not header_found:
            continue

        if col1 == "Average":
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_to_intv.get(cell.column)
                    cell.value = _mean(intv) if intv else None

        elif col1 == "CV":
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_to_intv.get(cell.column)
                    cell.value = _cv(intv) if intv else None


# ── Notes sheet ───────────────────────────────────────────────────────────────
# Each entry: (purpose, key columns string).  Matched by exact name, then prefix, then suffix.

_SHEET_DESCS_EXACT = {
    "Loss Selection": (
        "Actuary-selected ultimates for loss measures. Compare CL/BF/IE methods and record final selection.",
        "Period · Age · Incurred · Paid · Incurred CL · Paid CL · Incurred BF · Paid BF · Initial Expected · Selected Ultimate · IBNR · Unpaid",
    ),
    "Counts Selection": (
        "Actuary-selected ultimates for count measures. Compare CL/BF/IE methods and record final selection.",
        "Period · Age · Reported · Closed · Reported CL · Closed CL · Reported BF · Closed BF · Initial Expected · Selected Ultimate · IBNR · Unpaid",
    ),
    "Diagnostics": (
        "Post-selection reasonableness checks. Severity = Ultimate Loss / Ultimate Count. Loss Rate and Frequency require Exposure.",
        "Period · Ultimate Severity · Ultimate Loss Rate · Ultimate Frequency",
    ),
    "Incurred-to-Ult": (
        "Incurred loss as a percent of selected ultimate by age. Values approach 1.0 as periods mature.",
        "Period rows · Development age columns",
    ),
    "Paid-to-Ult": (
        "Paid loss as a percent of selected ultimate by age. Values approach 1.0 as periods mature.",
        "Period rows · Development age columns",
    ),
    "Reported-to-Ult": (
        "Reported count as a percent of selected ultimate by age. Values approach 1.0 as periods mature.",
        "Period rows · Development age columns",
    ),
    "Closed-to-Ult": (
        "Closed count as a percent of selected ultimate by age. Values approach 1.0 as periods mature.",
        "Period rows · Development age columns",
    ),
    "Average IBNR": (
        "Selected Ultimate minus Incurred at each development age. Shows average reserve need remaining by age.",
        "Period rows · Development age columns",
    ),
    "Average Unpaid": (
        "Selected Ultimate minus Paid at each development age. Shows average unpaid reserve remaining by age.",
        "Period rows · Development age columns",
    ),
}

_SHEET_DESCS_SUFFIX = {
    " CL": (
        "Chain Ladder method results. Ultimate = Actual × CDF. IBNR = Ultimate − Actual. "
        "Unpaid = Ultimate − latest paid (proxy measure).",
        "Period · Age · Actual · CDF · Ultimate · IBNR · Unpaid",
    ),
    " BF": (
        "Bornhuetter-Ferguson method. Blends Initial Expected with emerged experience. "
        "% Unreported = 1 − 1/CDF. Ultimate = (IE × % Unreported) + Actual.",
        "Period · Age · Initial Expected · CDF · % Unreported · Unreported · Actual · Ultimate · IBNR · Unpaid",
    ),
    " IE": (
        "Initial Expected method. Ultimate = Exposure × Selected Loss Rate (ELR). "
        "ELR back-calculated from IE ultimate and exposure.",
        "Period · Age · Exposure · Selected Loss Rate · IE Ultimate",
    ),
    " - CV & Slopes": (
        "Coefficient of variation and regression slope statistics for age-to-age factors by development interval.",
        "Interval columns · CV row · Slope row",
    ),
}

_SHEET_DESCS_PREFIX = {
    "Diag - ": (
        "Diagnostic scatter plot data for development pattern review.",
        "Period · Development age · Age-to-age factor",
    ),
}

_TRIANGLE_SHEET_NAMES = {"Incurred", "Paid", "Reported", "Closed", "Exposure"}


def _sheet_desc(name):
    if name in _SHEET_DESCS_EXACT:
        return _SHEET_DESCS_EXACT[name]
    for prefix, desc in _SHEET_DESCS_PREFIX.items():
        if name.startswith(prefix):
            return desc
    for suffix, desc in _SHEET_DESCS_SUFFIX.items():
        if name.endswith(suffix):
            return desc
    if name in _TRIANGLE_SHEET_NAMES:
        return (
            f"{name} development triangle with age-to-age factors, weighted/simple averages, and selected LDFs.",
            "Period rows · Age columns (triangle values) · ATA factor rows · Average rows · LDF Selection row",
        )
    return ("Analysis results", "")


def write_notes_sheet(ws, sheet_list):
    """Write Notes sheet with metadata header and table of contents."""
    r = 1

    title_cell = ws.cell(r, 1, "Reserve Analysis - Complete Analysis")
    style_header(title_cell, "header")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.row_dimensions[r].height = 24
    r += 1

    meta = ws.cell(r, 1, "Workbook Information")
    style_header(meta, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for label, value in [
        ("Created:",     datetime.now().strftime("%B %d, %Y %I:%M %p")),
        ("Description:", "Complete actuarial reserve analysis combining selections, ultimates, and diagnostics"),
    ]:
        lbl = ws.cell(r, 1, label)
        lbl.font = LABEL_FONT; lbl.border = THIN_BORDER
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        val = ws.cell(r, 2, value)
        val.font = DATA_FONT; val.border = THIN_BORDER
        val.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        r += 1

    r += 1
    toc_hdr = ws.cell(r, 1, "Table of Contents")
    style_header(toc_hdr, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for col_num, text, width in [
        (1, "Name", 28),
        (2, "Description", 80),
    ]:
        cell = ws.cell(r, col_num, text)
        cell.font = SUBHEADER_FONT; cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = width
    r += 1

    for name, purpose, _cols in sheet_list:
        nc = ws.cell(r, 1, name)
        nc.font = DATA_FONT; nc.border = THIN_BORDER
        nc.alignment = Alignment(horizontal="left", vertical="center")
        pc = ws.cell(r, 2, purpose)
        pc.font = DATA_FONT; pc.border = THIN_BORDER
        pc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    ws.freeze_panes = "A8"


# ── Workbook assembly ─────────────────────────────────────────────────────────

def _copy_ws(ws_src, ws_dst):
    """Copy all cells and styles from source to destination worksheet."""
    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst[cell.coordinate]
            dst.value = cell.value
            if cell.has_style:
                dst.font       = copy.copy(cell.font)
                dst.border     = copy.copy(cell.border)
                dst.fill       = copy.copy(cell.fill)
                dst.number_format = cell.number_format
                dst.protection = copy.copy(cell.protection)
                dst.alignment  = copy.copy(cell.alignment)

    for col in ws_src.column_dimensions:
        ws_dst.column_dimensions[col].width = ws_src.column_dimensions[col].width
    for row_num in ws_src.row_dimensions:
        ws_dst.row_dimensions[row_num].height = ws_src.row_dimensions[row_num].height
    for rng in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(rng))
    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes


def assemble_workbook(output_path, generated_wb):
    """
    Build master workbook from the pre-built generated sheets.
    All content is already computed values — no formula copying needed.
    """
    master = Workbook()
    master.remove(master.active)
    sheet_list = []

    for sname in generated_wb.sheetnames:
        ws_dst = master.create_sheet(title=sname)
        _copy_ws(generated_wb[sname], ws_dst)
        sheet_list.append((sname, *_sheet_desc(sname)))

    notes_ws = master.create_sheet(title="Notes", index=0)
    write_notes_sheet(notes_ws, sheet_list)

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)
    sel_lookup = load_selections(INPUT_SELECTIONS_EXCEL)
    combined, available_methods = load_combined(INPUT_ULTIMATES, sel_lookup)
    exp_map = get_exposure(INPUT_TRIANGLES)
    has_exposure = bool(exp_map)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]
    
    # We also need triangles_df for diagnostics
    triangles_df = pd.read_parquet(INPUT_TRIANGLES)
    
    print(f"Measures: {measures}")
    print(f"Methods: {available_methods}")
    
    gen_wb = Workbook()
    gen_wb.remove(gen_wb.active)
    
    loss_m  = [m for m in measures if "Loss"  in m]
    count_m = [m for m in measures if "Count" in m]
    other_m = [m for m in measures if m not in loss_m and m not in count_m]
    has_ie  = any(col == "ultimate_ie" for col, _ in available_methods)

    print("Building Loss sheets...")
    if loss_m:
        write_selection_grouped(gen_wb, combined, ["Incurred Loss", "Paid Loss"], "Loss Selection")
        for m in loss_m:
            write_method_cl(gen_wb, combined, m)
        for m in loss_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m)
        for m in loss_m:
            if _has_method(combined, m, "ultimate_ie"):
                write_method_ie(gen_wb, combined, m, exp_map)

    print("Building Counts sheets...")
    if count_m:
        write_selection_grouped(gen_wb, combined, ["Reported Count", "Closed Count"], "Counts Selection")
        for m in count_m:
            write_method_cl(gen_wb, combined, m)
        for m in count_m:
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m)
        for m in count_m:
            if _has_method(combined, m, "ultimate_ie"):
                write_method_ie(gen_wb, combined, m, exp_map)

    if other_m:
        print("Building other method sheets...")
        for m in other_m:
            write_method_cl(gen_wb, combined, m)
            if _has_method(combined, m, "ultimate_bf"):
                write_method_bf(gen_wb, combined, m)
            if _has_method(combined, m, "ultimate_ie"):
                write_method_ie(gen_wb, combined, m, exp_map)
        
    print("Copying CL LDF sheets (Triangle + Averages)...")
    create_triangle_sheets(gen_wb, measures)
    if has_exposure:
        write_exposure_sheet(gen_wb, INPUT_TRIANGLES)
    
    # Also copy remaining Diag and CV & Slopes sheets
    print("Copying remaining Diag and CV & Slopes sheets...")
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)
        for sname in wb_cl_form.sheetnames:
            if sname.startswith("Diag - ") or sname.endswith(" - CV & Slopes"):
                ws = gen_wb.create_sheet(title=sname[:31])
                _copy_ws_filtered(wb_cl_form[sname], ws, {}, {})
        wb_cl_form.close()

    print("Building post-method triangle sheets...")
    for sheet_name, df in build_post_method_triangle_data(triangles_df, combined):
        ws_t = gen_wb.create_sheet(title=sheet_name[:31])
        write_triangle_sheet(ws_t, sheet_name, df)

    print("Building diagnostics sheet...")
    ws_diag = gen_wb.create_sheet(title="Diagnostics")
    write_diagnostics_sheet(ws_diag, combined, exp_map)

    print("\nAssembling Complete Analysis.xlsx...")
    assemble_workbook(OUTPUT_COMPLETE, gen_wb)

    print("Computing formula values for Values Only...")
    df2_enh = pd.read_parquet(INPUT_CL_ENHANCED) if pathlib.Path(INPUT_CL_ENHANCED).exists() else pd.DataFrame()
    df4_avg = pd.read_parquet(INPUT_LDF_AVERAGES) if pathlib.Path(INPUT_LDF_AVERAGES).exists() else None
    measures_short_set = {measure_short_name(m) for m in measures}
    actual_lookup_full = combined.set_index(["period", "measure"])["actual"].to_dict()

    for sname in gen_wb.sheetnames:
        ws = gen_wb[sname]
        if sname in measures_short_set and not df2_enh.empty:
            full_m = next((m for m in measures if measure_short_name(m) == sname), None)
            if full_m:
                _fill_cl_main_values(ws, full_m, df2_enh, df4_avg)
                _fill_cdf_row_values(ws)
            else:
                _strip_formulas(ws)
        elif sname.endswith(" CL"):
            full_m = next((m for m in measures if measure_short_name(m) == sname[:-3]), None)
            if full_m:
                _fill_method_cl_values(ws, full_m, combined, actual_lookup_full)
            else:
                _strip_formulas(ws)
        elif sname.endswith(" BF"):
            full_m = next((m for m in measures if measure_short_name(m) == sname[:-3]), None)
            if full_m:
                _fill_method_bf_values(ws, full_m, combined, actual_lookup_full)
            else:
                _strip_formulas(ws)
        elif sname.endswith(" IE"):
            pass  # IE sheet has no formula cells
        elif sname == "Loss Selection":
            _fill_selection_values(ws, ["Incurred Loss", "Paid Loss"], combined, actual_lookup_full)
        elif sname == "Counts Selection":
            _fill_selection_values(ws, ["Reported Count", "Closed Count"], combined, actual_lookup_full)
        else:
            _strip_formulas(ws)

    print("Assembling Complete Analysis - Values Only.xlsx...")
    assemble_workbook(OUTPUT_VALUES, gen_wb)

if __name__ == "__main__":
    main()
