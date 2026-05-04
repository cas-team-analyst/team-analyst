# Creates a formatted Excel workbook for tail curve review and selection. Each measure gets its
# own sheet with observed data, scenario comparisons, and selection areas for actuarial judgment.

"""
goal: Create Chain Ladder Selections - Tail.xlsx for actuarial tail curve review and selection.
      Agents select only the curve/method; the tail factor is derived from the selected method
      and the fitted curve parameters in tail-scenarios.parquet.

Sheet layout per measure:
  - Section A: Selected LDFs from Chain Ladder Selections - LDFs.xlsx
  - Section B: Curve Comparison Table (method × diagnostics)
  - Section C: Tail Curve Selection Area (prior, rules-based, open-ended, user — method only)

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 2d-tail-create-excel.py
"""

import pandas as pd
import numpy as np
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook

from modules import config
from modules.xl_styles import create_xlsxwriter_formats, COLORS
from modules.markdown_utils import df_to_markdown
from modules.xl_writers import col_letter
from modules.average_names import pretty_average_name

# Paths
TAIL_SCENARIOS_PATH = config.PROCESSED_DATA + "tail-scenarios.parquet"
ENHANCED_PATH = config.PROCESSED_DATA + "2_enhanced.parquet"
DIAGNOSTICS_PATH = config.PROCESSED_DATA + "3_diagnostics.parquet"
LDF_AVERAGES_PATH = config.PROCESSED_DATA + "4_ldf_averages.parquet"
SELECTIONS_OUTPUT_PATH = config.SELECTIONS
OUTPUT_FILE_NAME = "Chain Ladder Selections - Tail.xlsx"
CL_LDF_EXCEL = config.SELECTIONS + "Chain Ladder Selections - LDFs.xlsx"


def write_section_header(ws, row, col_span, title, fmt, level="section"):
    """Write a section header spanning multiple columns."""
    format_key = {'header': 'header', 'section': 'section', 'subheader': 'subheader', 'label_row': 'label_row'}.get(level, 'section')
    header_fmt = fmt.get(format_key, fmt['section'])
    
    ws.merge_range(row, 0, row, col_span - 1, title, header_fmt)
    return row + 1


def find_selected_ldfs_in_cl_excel(cl_excel_path, measure):
    """
    Read Chain Ladder Excel to find selected LDFs and reasoning for a measure.
    Finds ALL THREE selection+reasoning rows for cascaded value priority logic.
    
    Returns:
        {
            'intervals': {interval: col_idx},
            'selection_rows': {'user': row, 'rules_based': row, 'open_ended': row},
            'reasoning_rows': {'user': row, 'rules_based': row, 'open_ended': row},
            'cached_values': {interval: value},  # Using priority: User > Rules-Based > Open-Ended
            'cached_reasoning': {interval: {'user': text, 'rules_based': text, 'open_ended': text}}
        }
        or {} if file not found or no selections
    """
    path = Path(cl_excel_path)
    if not path.exists():
        print(f"  WARNING: {cl_excel_path} not found")
        return {}
    
    try:
        wb = load_workbook(cl_excel_path, data_only=True)
        if measure not in wb.sheetnames:
            print(f"  WARNING: Sheet '{measure}' not found in {cl_excel_path}")
            return {}
        
        ws = wb[measure]
        
        # Find ALL THREE selection rows
        selection_rows = {}
        selection_labels = {
            "User Selection": "user",
            "Rules-Based AI Selection": "rules_based",
            "Open-Ended AI Selection": "open_ended"
        }
        
        for label, key in selection_labels.items():
            for row_idx in range(1, ws.max_row + 1):
                cell_val = ws.cell(row_idx, 1).value
                if cell_val and str(cell_val).strip() == label:
                    selection_rows[key] = row_idx
                    break
        
        if not selection_rows:
            print(f"  WARNING: No selection rows found in {measure} sheet")
            return {}
        
        # Find reasoning rows (selection row + 1)
        reasoning_rows = {}
        reasoning_labels = {
            "User Reasoning": "user",
            "Rules-Based AI Reasoning": "rules_based",
            "Open-Ended AI Reasoning": "open_ended"
        }
        
        for label, key in reasoning_labels.items():
            for row_idx in range(1, ws.max_row + 1):
                cell_val = ws.cell(row_idx, 1).value
                if cell_val and str(cell_val).strip() == label:
                    reasoning_rows[key] = row_idx
                    break
        
        # Find interval header row by searching backwards from first found selection row
        first_selection_row = min(selection_rows.values())
        
        import re
        interval_re = re.compile(r'^\d+-\d+$')
        interval_row = None
        interval_map = {}  # {interval: col_idx}
        
        # Search backwards from first selection row
        for row_idx in range(first_selection_row - 1, max(1, first_selection_row - 10), -1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            interval_count = sum(1 for v in row_vals[1:] if v and interval_re.match(str(v).strip()))
            if interval_count >= 3:
                interval_row = row_idx
                for col_idx, val in enumerate(row_vals):
                    if val and interval_re.match(str(val).strip()):
                        interval_key = str(val).strip()
                        interval_map[interval_key] = col_idx + 1  # +1 for 1-based Excel
                break
        
        if not interval_map:
            print(f"  WARNING: No interval headers found near selection rows in {measure} sheet")
            return {}
        
        # Calculate cached values using priority logic: User > Rules-Based > Open-Ended
        cached_values = {}
        cached_reasoning = {}
        
        for interval, col_idx in interval_map.items():
            value = None
            # Check User first
            if 'user' in selection_rows:
                value = ws.cell(selection_rows['user'], col_idx).value
            # Fall back to Rules-Based if User is None
            if value is None and 'rules_based' in selection_rows:
                value = ws.cell(selection_rows['rules_based'], col_idx).value
            # Fall back to Open-Ended if both above are None
            if value is None and 'open_ended' in selection_rows:
                value = ws.cell(selection_rows['open_ended'], col_idx).value
            
            cached_values[interval] = value
            
            # Read reasoning for all three
            reasoning_dict = {}
            for key in ['user', 'rules_based', 'open_ended']:
                if key in reasoning_rows:
                    reason_val = ws.cell(reasoning_rows[key], col_idx).value
                    reasoning_dict[key] = reason_val if reason_val else None
                else:
                    reasoning_dict[key] = None
            
            cached_reasoning[interval] = reasoning_dict
        
        return {
            'intervals': interval_map,
            'selection_rows': selection_rows,
            'reasoning_rows': reasoning_rows,
            'cached_values': cached_values,
            'cached_reasoning': cached_reasoning
        }
    
    except Exception as e:
        print(f"  ERROR reading Chain Ladder Excel: {e}")
        return {}


_AVG_PERIOD_ORDER = ['all', '3yr', '5yr', '10yr']
_AVG_TYPE_ORDER = ['simple', 'weighted', 'exclude_high_low']

def _sort_avg_cols(cols):
    def key(col):
        normalized = col.replace('avg_', '')
        for p_idx, p in enumerate(_AVG_PERIOD_ORDER):
            for t_idx, t in enumerate(_AVG_TYPE_ORDER):
                if normalized == f"{t}_{p}":
                    return (p_idx, t_idx)
        return (99, 99)
    return sorted(cols, key=key)

def _avg_display_name(col):
    """Map average column names to prettier display names."""
    return pretty_average_name(col)

def write_ata_triangle_section(ws, start_row, measure, df_enhanced, fmt):
    """
    Write Age-to-Age Factors triangle - empirical ATAs by period x interval.
    Returns: next_row (int)
    """
    df_m = df_enhanced[(df_enhanced['measure'] == measure) & df_enhanced['ldf'].notna()].copy()
    if df_m.empty:
        ws.write(start_row, 0, "Age-to-Age Factors", fmt['subheader'])
        ws.write(start_row + 1, 0, "(No empirical ATAs found)", fmt['label'])
        return start_row + 3
    
    periods = df_m['period'].dropna().unique().tolist()
    if hasattr(df_m['period'], 'cat'):
        periods = df_m['period'].cat.categories.tolist()
    else:
        periods = sorted(periods, key=lambda x: (isinstance(x, str), x))
    
    intervals = df_m['interval'].dropna().unique().tolist()
    if hasattr(df_m['interval'], 'cat'):
        intervals = df_m['interval'].cat.categories.tolist()
    else:
        def interval_sort_key(interval):
            try:
                return int(str(interval).split('-')[0])
            except (ValueError, IndexError, AttributeError):
                return 999
        intervals = sorted(intervals, key=interval_sort_key)
    
    # Build LDF lookup dict
    ldf_dict = {}
    for _, row_data in df_m.iterrows():
        key = (str(row_data['period']), str(row_data['interval']))
        ldf_dict[key] = row_data['ldf']
    
    # Header row
    row = start_row
    ws.write(row, 0, "Age-to-Age Factors", fmt['subheader_left'])
    for c_idx, interval in enumerate(intervals):
        ws.write(row, c_idx + 1, interval, fmt['subheader_right'])
    row += 1
    
    # Data rows
    for period in periods:
        ws.write(row, 0, period, fmt['label'])
        for c_idx, interval in enumerate(intervals):
            ldf_val = ldf_dict.get((str(period), str(interval)))
            if ldf_val is not None and pd.notna(ldf_val):
                ws.write(row, c_idx + 1, ldf_val, fmt['data_ldf'])
        row += 1
    
    return row + 1  # Skip blank row

def write_averages_section(ws, start_row, measure, df_ldf_averages, intervals, fmt):
    """
    Write averages/CVs/slopes section from df4_ldf_averages.
    Returns: next_row (int)
    """
    df_avg = df_ldf_averages[df_ldf_averages['measure'] == measure].copy()
    if df_avg.empty:
        ws.write(start_row, 0, "Averages", fmt['subheader'])
        ws.write(start_row + 1, 0, "(No averages found)", fmt['label'])
        return start_row + 3
    
    # Include all metrics: averages, min/max, CVs, slopes
    all_cols = [col for col in df_avg.columns if col not in ['measure', 'interval']]
    
    # Order: averages first, then min/max, then CVs, then slopes
    # Exclude all-year averages (simple_all, weighted_all, exclude_high_low_all) but keep min_all, max_all
    avg_cols = [c for c in all_cols if not c.startswith('cv_') and not c.startswith('slope_') 
                and c not in ['min_all', 'max_all'] and not c.endswith('_all')]
    minmax_cols = [c for c in all_cols if c in ['min_all', 'max_all']]
    cv_cols = [c for c in all_cols if c.startswith('cv_')]
    slope_cols = [c for c in all_cols if c.startswith('slope_')]
    
    avg_cols = _sort_avg_cols(avg_cols)
    cv_cols = sorted(cv_cols)  # cv_3yr, cv_5yr, cv_10yr
    slope_cols = sorted(slope_cols)  # slope_3yr, slope_5yr, slope_10yr
    
    ordered_cols = avg_cols + minmax_cols + cv_cols + slope_cols
    display_cols = [_avg_display_name(c) for c in ordered_cols]
    
    row = start_row
    ws.write(row, 0, "Averages", fmt['subheader_left'])
    for c_idx, interval in enumerate(intervals):
        ws.write(row, c_idx + 1, interval, fmt['subheader_right'])
    row += 1
    
    # Write all metric values directly from df_avg
    for col, display in zip(ordered_cols, display_cols):
        ws.write(row, 0, display, fmt['label'])
        
        for c_idx, interval in enumerate(intervals):
            val_row = df_avg[df_avg['interval'] == interval]
            if not val_row.empty and col in val_row.columns:
                val = val_row[col].iloc[0]
                if pd.notna(val):
                    ws.write(row, c_idx + 1, val, fmt['data_ldf'])
        row += 1
    
    return row + 1  # Skip blank row

def write_selected_ldfs_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt):
    """
    Write selected LDFs section with values and reasoning from Chain Ladder Excel.
    Intervals as columns, one row for selections, one row for reasoning.
    Uses cascaded priority: User Selection > Rules-Based AI > Open-Ended AI.
    
    Returns:
        next_row (int)
    """
    # Find selected LDFs in Chain Ladder Excel
    ldf_data = find_selected_ldfs_in_cl_excel(cl_excel_path, measure)
    
    if not ldf_data:
        # No selections found - write empty section
        ws.write(start_row, 0, "Selected LDFs", fmt['label_row'])
        ws.write(start_row + 1, 0, "(No selections found in Chain Ladder Excel)", fmt['label'])
        return start_row + 3
    
    intervals = ldf_data['intervals']
    cached_values = ldf_data['cached_values']
    cached_reasoning = ldf_data.get('cached_reasoning', {})
    
    # Sort intervals by numeric start value
    def interval_sort_key(interval):
        try:
            return int(interval.split('-')[0])
        except (ValueError, IndexError):
            return 999
    
    sorted_intervals = sorted(intervals.keys(), key=interval_sort_key)
    
    # Write label row
    ws.write(start_row, 0, "Selected LDFs", fmt['label_row'])
    for c_idx, interval in enumerate(sorted_intervals):
        ws.write(start_row, c_idx + 1, interval, fmt['label_row'])
    
    # Write selection value row
    row = start_row + 1
    ws.write(row, 0, "Selection", fmt['label'])
    
    for c_idx, interval in enumerate(sorted_intervals):
        cached_value = cached_values.get(interval)
        if cached_value is not None and pd.notna(cached_value):
            ws.write(row, c_idx + 1, cached_value, fmt['data_ldf'])
    row += 1
    
    # Write reasoning row
    ws.write(row, 0, "Reasoning", fmt['label'])
    for c_idx, interval in enumerate(sorted_intervals):
        reasoning_dict = cached_reasoning.get(interval, {})
        reasoning = reasoning_dict.get('user') or reasoning_dict.get('rules_based') or reasoning_dict.get('open_ended') or ""
        if reasoning:
            ws.write(row, c_idx + 1, reasoning, fmt['label'])
    
    return row + 2  # Skip blank row


def write_observed_factors_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt):
    """
    Section B: Selected LDFs from Chain Ladder Excel (hard-coded values).
    Renamed but kept old function name for compatibility.
    """
    return write_selected_ldfs_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt)


def write_scenario_comparison_section(ws, start_row, measure, df_scenarios, fmt):
    """Section C: Curve Comparison Table with all diagnostics (no color formatting)."""
    df_m = df_scenarios[df_scenarios['measure'] == measure].copy()
    if df_m.empty:
        return start_row
    
    col_headers = [
        'Starting Age', 'Monotone', 'CV', 'Slope Breaks', 'Method', 'Params',
        'Tail Factor', 'R2', 'LOO Std Dev', 'Gap to Last', 'Gap Flag',
        '% of CDF', 'Materiality OK', '+10% Reserve Delta', '+20% Reserve Delta'
    ]
    
    row = write_section_header(ws, start_row, len(col_headers), "Curve Comparison", fmt, "label_row")
    
    # Header row
    for c_idx, header in enumerate(col_headers):
        ws.write(row, c_idx, header, fmt['subheader'])
    row += 1
    
    # Sort scenarios by starting_age then method
    df_m = df_m.sort_values(['starting_age', 'method'])
    
    # Helper to safely convert boolean columns (handles pandas NA)
    def safe_bool(val):
        if pd.isna(val):
            return False
        return bool(val)
    
    # Data rows - simple formatting, no colors
    for _, scenario in df_m.iterrows():
        values = [
            scenario['starting_age'],
            'Y' if safe_bool(scenario.get('is_monotone_from_here', False)) else 'N',
            scenario.get('cv_at_starting_age'),
            scenario.get('slope_sign_changes', 0),
            scenario['method'],
            str(scenario.get('method_params', '')) if pd.notna(scenario.get('method_params')) else '',
            scenario['tail_factor'],
            scenario.get('r_squared'),
            scenario.get('loo_std_dev'),
            scenario.get('gap_to_last_observed'),
            'Y' if safe_bool(scenario.get('gap_flag', False)) else 'N',
            scenario.get('pct_of_cdf'),
            'Y' if safe_bool(scenario.get('materiality_ok', False)) else 'N',
            scenario.get('sensitivity_plus10_reserve_delta'),
            scenario.get('sensitivity_plus20_reserve_delta'),
        ]
        
        for c_idx, val in enumerate(values):
            # Only write cells with valid data - match 2a pattern
            if c_idx == 0:  # Starting Age
                ws.write(row, c_idx, val, fmt['data'])
            elif c_idx == 2:  # CV
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx in [4, 5]:  # Method, Params (text)
                if val:  # Only write non-empty strings
                    ws.write(row, c_idx, val, fmt['label'])
            elif c_idx == 6:  # Tail Factor
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx in [7, 8, 9]:  # R2, LOO Std Dev, Gap to Last
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx == 11:  # % of CDF
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_pct'])
            elif c_idx in [13, 14]:  # Reserve deltas
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_num'])
            else:  # Monotone (1), Slope Breaks (3), Gap Flag (10), Materiality OK (12)
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data'])
        
        row += 1
    
    return row + 1


def write_selection_section(ws, start_row, measure, fmt, prior_selections=None):
    """Section C: Tail Curve Selection Area (prior, rules-based AI, open-ended AI, user selection)."""
    row = write_section_header(ws, start_row, 4, "Tail Curve Selection", fmt, "label_row")
    
    headers = ['Label', 'Method', 'Reasoning', 'Additional Notes']
    for c_idx, header in enumerate(headers):
        ws.write(row, c_idx, header, fmt['subheader'])
    row += 1
    
    # Prior selection row
    if prior_selections is not None:
        prior_m = prior_selections[prior_selections['measure'] == measure]
        if not prior_m.empty:
            prior_row_data = prior_m.iloc[0]
            values = [
                'Prior Selection',
                prior_row_data.get('method', ''),
                prior_row_data.get('reasoning', ''),
                ''
            ]
            
            for c_idx, val in enumerate(values):
                if c_idx == 0:  # Label column
                    ws.write(row, c_idx, val, fmt['prior'])
                elif c_idx in [1, 2, 3]:  # Text columns
                    ws.write(row, c_idx, val, fmt['prior_text'])
                else:
                    ws.write(row, c_idx, val, fmt['prior'])
            
            row += 1
            row += 1  # Blank row
    
    # Rules-Based AI Selection row
    ws.write(row, 0, "Rules-Based AI Selection", fmt['selection'])
    for c_idx in range(1, 4):
        ws.write(row, c_idx, '', fmt['selection_text'])
    row += 1
    
    # Open-Ended AI Selection row
    ws.write(row, 0, "Open-Ended AI Selection", fmt['ai'])
    for c_idx in range(1, 4):
        ws.write(row, c_idx, '', fmt['ai_text'])
    row += 1
    
    row += 1  # Blank row
    
    # User Selection row
    ws.write(row, 0, "User Selection", fmt['user'])
    for c_idx in range(1, 4):
        ws.write(row, c_idx, '', fmt['user_text'])
    row += 1
    
    return row + 1



def build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, df_ldf_averages, fmt, 
                        cl_excel_path, output_file_path, prior_selections=None):
    """Build complete sheet for one measure."""
    row = 0  # xlsxwriter uses 0-based indexing, so row 0 = Excel row 1
    
    # Get intervals from enhanced data
    df_m = df_enhanced[(df_enhanced['measure'] == measure) & df_enhanced['ldf'].notna()].copy()
    if not df_m.empty:
        intervals = df_m['interval'].dropna().unique().tolist()
        if hasattr(df_m['interval'], 'cat'):
            intervals = df_m['interval'].cat.categories.tolist()
        else:
            def interval_sort_key(interval):
                try:
                    return int(str(interval).split('-')[0])
                except (ValueError, IndexError, AttributeError):
                    return 999
            intervals = sorted(intervals, key=interval_sort_key)
        intervals = [str(i) for i in intervals]
    else:
        intervals = []
    
    # Section 1: Age-to-Age Factors triangle
    row = write_ata_triangle_section(ws, row, measure, df_enhanced, fmt)
    
    # Section 2: Averages
    row = write_averages_section(ws, row, measure, df_ldf_averages, intervals, fmt)
    
    # Section 3: Selected LDFs from Chain Ladder Excel (cutoff point, with reasoning)
    row = write_observed_factors_section(ws, row, measure, cl_excel_path, output_file_path, fmt)
    
    # Section 4: Curve Comparison
    row = write_scenario_comparison_section(ws, row, measure, df_scenarios, fmt)
    
    # Section 5: Selection Area
    row = write_selection_section(ws, row, measure, fmt, prior_selections)
    
    # Column widths
    ws.set_column(0, 0, 30)  # A: Interval/Label
    ws.set_column(1, 1, 18)  # B: Method/LDF value
    ws.set_column(2, 2, 18)  # C: Data
    ws.set_column(3, 3, 30)  # D: Additional Notes



def export_md_data(measures, df_scenarios, df_enhanced, df_diagnostics, df_ldf_averages, exp_md, prior_selections=None):
    """
    Export markdown context files for tail curve selection.
    
    ALIGNMENT PRINCIPLE: Markdown files should mirror Excel workbook structure and content.
    - Sections in MD should correspond to sections in Excel
    - Tables should use same column ordering and display names
    - This allows subagents to use MD as canonical context that matches user's Excel view
    """
    for measure in measures:
        safe_name = measure.lower().replace(' ', '_')
        md_path = Path(SELECTIONS_OUTPUT_PATH) / f"tail-context-{safe_name}.md"
        
        # Get selected LDFs from Chain Ladder Excel instead of observed factors
        ldf_data = find_selected_ldfs_in_cl_excel(CL_LDF_EXCEL, measure)
        if ldf_data and 'intervals' in ldf_data and 'cached_values' in ldf_data:
            # Sort intervals by numeric start value
            def interval_sort_key(interval):
                try:
                    return int(interval.split('-')[0])
                except (ValueError, IndexError):
                    return 999
            
            sorted_intervals = sorted(ldf_data['intervals'].keys(), key=interval_sort_key)
            cached_values = ldf_data['cached_values']
            cached_reasoning = ldf_data.get('cached_reasoning', {})
            
            # Create markdown table with intervals as columns, selections + reasoning rows
            header = "| | " + " | ".join(sorted_intervals) + " |"
            separator = "|---|" + "|".join(["---"] * len(sorted_intervals)) + "|"
            
            # Selection row
            values = "| Selection | " + " | ".join(
                [f"{cached_values[interval]:.4f}" if cached_values[interval] is not None else "" 
                 for interval in sorted_intervals]
            ) + " |"
            
            # Reasoning row - truncate to first 80 chars per cell
            reasoning_vals = []
            for interval in sorted_intervals:
                reasoning_dict = cached_reasoning.get(interval, {})
                reasoning = reasoning_dict.get('user') or reasoning_dict.get('rules_based') or reasoning_dict.get('open_ended') or ""
                truncated = reasoning[:80] + "..." if len(reasoning) > 80 else reasoning
                reasoning_vals.append(truncated)
            
            reasoning_row = "| Reasoning | " + " | ".join(reasoning_vals) + " |"
            
            selected_md = header + "\n" + separator + "\n" + values + "\n" + reasoning_row + "\n"
        else:
            selected_md = "(No selected LDFs found in Chain Ladder Excel)\n"
        
        # --- Averages section (mirrors Excel averages section) ---
        df_avg = df_ldf_averages[df_ldf_averages['measure'] == measure].copy()
        if not df_avg.empty:
            # Get intervals from averages data
            avg_intervals = df_avg['interval'].dropna().unique().tolist()
            def interval_sort_key(interval):
                try:
                    return int(str(interval).split('-')[0])
                except (ValueError, IndexError, AttributeError):
                    return 999
            avg_intervals = sorted(avg_intervals, key=interval_sort_key)
            
            # Build ordered columns matching Excel: avg_cols + minmax + cv + slope
            # Exclude all-year averages (simple_all, weighted_all, exclude_high_low_all) but keep min_all, max_all
            all_cols = [col for col in df_avg.columns if col not in ['measure', 'interval']]
            avg_cols = [c for c in all_cols if not c.startswith('cv_') and not c.startswith('slope_') 
                        and c not in ['min_all', 'max_all'] and not c.endswith('_all')]
            minmax_cols = [c for c in all_cols if c in ['min_all', 'max_all']]
            cv_cols = sorted([c for c in all_cols if c.startswith('cv_')])  # cv_3yr, cv_5yr, cv_10yr
            slope_cols = sorted([c for c in all_cols if c.startswith('slope_')])  # slope_3yr, slope_5yr, slope_10yr
            
            avg_cols = _sort_avg_cols(avg_cols)
            ordered_cols = avg_cols + minmax_cols + cv_cols + slope_cols
            
            # Create markdown table with intervals as columns
            header = "| Metric | " + " | ".join([str(i) for i in avg_intervals]) + " |"
            separator = "|---|" + "|".join(["---"] * len(avg_intervals)) + "|"
            
            rows = [header, separator]
            for col in ordered_cols:
                display_name = pretty_average_name(col)
                row_vals = [display_name]
                for interval in avg_intervals:
                    val_row = df_avg[df_avg['interval'] == interval]
                    if not val_row.empty and col in val_row.columns:
                        val = val_row[col].iloc[0]
                        if pd.notna(val):
                            row_vals.append(f"{val:.4f}")
                        else:
                            row_vals.append("")
                    else:
                        row_vals.append("")
                rows.append("| " + " | ".join(row_vals) + " |")
            
            averages_md = "\n".join(rows) + "\n"
        else:
            averages_md = "(No averages found)\n"
        
        # --- Curves section (all tail scenario diagnostics) ---
        scen_sub = df_scenarios[df_scenarios['measure'] == measure].drop(columns=['measure', 'starting_age', 'cutoff_age'], errors='ignore')
        scen_md = df_to_markdown(scen_sub, index=False)
        
        # Add empirical ATAs as triangle table (period × interval)
        ata_sub = df_enhanced[(df_enhanced['measure'] == measure) & df_enhanced['ldf'].notna()].copy()
        if not ata_sub.empty:
            # Get unique periods and intervals
            periods = ata_sub['period'].dropna().unique().tolist()
            if hasattr(ata_sub['period'], 'cat'):
                periods = ata_sub['period'].cat.categories.tolist()
            else:
                periods = sorted(periods, key=lambda x: (isinstance(x, str), x))
            
            intervals = ata_sub['interval'].dropna().unique().tolist()
            if hasattr(ata_sub['interval'], 'cat'):
                intervals = ata_sub['interval'].cat.categories.tolist()
            else:
                def interval_sort_key(interval):
                    try:
                        return int(str(interval).split('-')[0])
                    except (ValueError, IndexError, AttributeError):
                        return 999
                intervals = sorted(intervals, key=interval_sort_key)
            
            # Build lookup dict: (period, interval) -> ldf
            ldf_dict = {}
            for _, row_data in ata_sub.iterrows():
                key = (str(row_data['period']), str(row_data['interval']))
                ldf_dict[key] = row_data['ldf']
            
            # Create markdown triangle table
            header = "| Period | " + " | ".join([str(i) for i in intervals]) + " |"
            separator = "|---|" + "|".join(["---"] * len(intervals)) + "|"
            
            rows = [header, separator]
            for period in periods:
                row_vals = [str(period)]
                for interval in intervals:
                    ldf_val = ldf_dict.get((str(period), str(interval)))
                    if ldf_val is not None and pd.notna(ldf_val):
                        row_vals.append(f"{ldf_val:.4f}")
                    else:
                        row_vals.append("")
                rows.append("| " + " | ".join(row_vals) + " |")
            
            ata_md = "\n".join(rows) + "\n"
        else:
            ata_md = "No empirical ATAs found\n"
            
        # Prior selections section
        prior_md = ""
        if prior_selections is not None:
            prior_m = prior_selections[prior_selections['measure'] == measure]
            if not prior_m.empty:
                prior_row = prior_m.iloc[0]
                prior_md = "## Prior Selection\n\n"
                prior_md += "| Field | Value |\n"
                prior_md += "|---|---|\n"
                prior_md += f"| Cutoff Age | {prior_row.get('cutoff_age', 'N/A')} |\n"
                prior_md += f"| Tail Factor | {prior_row.get('tail_factor', 'N/A')} |\n"
                prior_md += f"| Method | {prior_row.get('method', 'N/A')} |\n"
                prior_md += f"| Reasoning | {prior_row.get('reasoning', 'N/A')} |\n\n"
        
        if not prior_md:
            prior_md = "## Prior Selection\n\nNo prior tail selections found for this analysis.\n\n"
        
        md_content = f"# Tail Context: {measure}\n\n"
        md_content += "## Table of Contents\n"
        if prior_md:
            md_content += "- [Prior Selection](#prior-selection)\n"
        md_content += "- [Exposure](#exposure)\n"
        md_content += "- [Averages](#averages)\n"
        md_content += "- [Selected LDFs](#selected-ldfs)\n"
        md_content += "- [Empirical Age-to-Age Factors](#empirical-age-to-age-factors)\n"
        md_content += "- [Curves](#curves)\n\n"
        if prior_md:
            md_content += prior_md
        md_content += "## Exposure\n" + exp_md + "\n"
        md_content += "## Averages\n"
        md_content += "Averages, min/max, CVs, and slopes for each interval (mirrors Excel averages section).\n\n"
        md_content += averages_md + "\n"
        md_content += "## Selected LDFs\n"
        md_content += selected_md + "\n"
        md_content += "## Empirical Age-to-Age Factors\n" + ata_md + "\n"
        md_content += "## Curves\n"
        md_content += "All tail curve scenarios with diagnostics.\n\n"
        md_content += scen_md + "\n"
        
        with open(md_path, 'w') as f:
            f.write(md_content)
        print(f"  Exported MD: {md_path}")

def main():
    """Create Tail Factor Selections Excel file."""
    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    if Path(output_file).exists():
        raise FileExistsError(
            f"Output file already exists: {output_file}\n"
            "Delete or rename the file before re-running to avoid overwriting manual edits."
        )
    
    if not Path(TAIL_SCENARIOS_PATH).exists():
        raise FileNotFoundError(
            f"Tail scenarios file not found: {TAIL_SCENARIOS_PATH}\n"
            "Run 2c-tail-methods-diagnostics.py first to generate tail scenarios."
        )
    
    print("Loading data...")
    df_scenarios = pd.read_parquet(TAIL_SCENARIOS_PATH)
    df_enhanced = pd.read_parquet(ENHANCED_PATH)
    df_diagnostics = pd.read_parquet(DIAGNOSTICS_PATH)
    df_ldf_averages = pd.read_parquet(LDF_AVERAGES_PATH)
    
    print(f"  {len(df_scenarios)} tail scenarios")
    print(f"  {len(df_enhanced)} enhanced rows")
    print(f"  {len(df_diagnostics)} diagnostic rows")
    print(f"  {len(df_ldf_averages)} LDF averages rows")
    
    # Load prior selections if available
    prior_selections_path = Path(config.SELECTIONS) / "tail-factor-prior.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"  {len(df_prior)} prior tail selections loaded")
    else:
        print("  No prior tail selections found (optional)")
    
    # Create workbook - match 2a script exactly
    wb = xlsxwriter.Workbook(output_file)
    fmt = create_xlsxwriter_formats(wb)
    fmt['wb'] = wb  # Store workbook reference
    
    raw_measures = sorted(df_scenarios['measure'].unique())
    
    exp_sub = df_enhanced[(df_enhanced['measure'] == 'Exposure') & df_enhanced['value'].notna()]
    if not exp_sub.empty:
        # Format Exposure as simple 2-column table (period, value)
        # Exposure doesn't develop over age, so we take the last value per period
        exp_simple = exp_sub.groupby('period', observed=True).agg({'value': 'last'}).reset_index()
        exp_simple.columns = ['Period', 'Exposure']
        exp_simple['Exposure'] = exp_simple['Exposure'].round(0)
        exp_md = df_to_markdown(exp_simple, index=False)
    else:
        exp_md = "No Exposure data\n"
        
    measures = [m for m in raw_measures if m != 'Exposure']
    
    print(f"\nCreating sheets for measures: {measures}")
    
    for measure in measures:
        # xlsxwriter sheet names limited to 31 chars
        ws = wb.add_worksheet(measure[:31])
        build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, df_ldf_averages, fmt, 
                          CL_LDF_EXCEL, output_file, df_prior)
        print(f"  Built sheet: {measure[:31]}")
    
    wb.close()
    export_md_data(measures, df_scenarios, df_enhanced, df_diagnostics, df_ldf_averages, exp_md, df_prior)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    main()
