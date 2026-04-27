# Creates a formatted Excel workbook that displays all the claims data, development factors,
# diagnostics, and calculated averages in an organized layout. Actuaries review this workbook and
# manually enter their selected factors based on the information presented.

"""
goal: Create Chain Ladder Selections.xlsx for actuarial LDF review and selection. This file should be manually edited by the actuary to make selections.

Sheet layout:
  - One main sheet per measure: loss triangle, LDF triangle, averages (no CV/slopes), selections
  - "Exposure": exposure triangle (if exposure data is available)
  - One CV & slopes sheet per measure: "{Measure} - CV & Slopes"
  - One sheet per diagnostic triangle: "Diag - {Name}" (shared across measures)

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 2a-chainladder-create-excel.py
"""

import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pathlib import Path

from modules import config
from modules.xl_styles import (
    SUBHEADER_FILL, SELECTION_FILL, PRIOR_FILL, AI_FILL, USER_FILL,
    SUBHEADER_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)

# Paths from modules/config.py — override here if needed:
OUTPUT_PATH = config.PROCESSED_DATA
SELECTIONS_OUTPUT_PATH = config.SELECTIONS
METHOD_ID = "chainladder"
OUTPUT_FILE_NAME = "Chain Ladder Selections - LDFs.xlsx"

DIAG_SHEET_LABELS = {
    'reported_claims': 'Rptd Claims',
    'incurred_severity': 'Inc Severity',
    'paid_severity': 'Paid Severity',
    'paid_to_incurred': 'Paid to Incurred',
    'open_counts': 'Open Counts',
    'average_case_reserve': 'Avg Case Rsrv',
    'claim_closure_rate': 'Clm Closure Rt',
    'incremental_incurred_severity': 'Incr Inc Svty',
    'incremental_paid_severity': 'Incr Paid Svty',
    'incremental_closure_rate': 'Incr Closure Rt',
}

DIAG_NUMBER_FORMATS = {
    'reported_claims': '#,##0',
    'incurred_severity': '#,##0',
    'paid_severity': '#,##0',
    'paid_to_incurred': '0.00%',
    'open_counts': '#,##0',
    'average_case_reserve': '#,##0',
    'claim_closure_rate': '0.00%',
    'incremental_incurred_severity': '#,##0',
    'incremental_paid_severity': '#,##0',
    'incremental_closure_rate': '0.00%',
}

def diag_sheet_name(diag_col):
    label = DIAG_SHEET_LABELS.get(diag_col, diag_col.replace('_', ' ').title()[:14])
    return f"Diag - {label}"[:31]

def cv_slopes_sheet_name(measure):
    return f"{measure} - CV & Slopes"[:31]

def write_triangle(ws, start_row, title, row_labels, col_labels, data_dict, number_format="#,##0", is_formula=False):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    data_start_row = start_row
    for r_idx, row_label in enumerate(row_labels):
        row_cell = ws.cell(row=start_row, column=1, value=row_label)
        row_cell.font = LABEL_FONT
        row_cell.alignment = Alignment(horizontal="left")
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((str(row_label), str(col)))
            if val is not None:
                cell = ws.cell(row=start_row, column=c_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                cell.number_format = number_format
        start_row += 1

    return start_row + 1, data_start_row, start_row - 1

def write_metrics_table(ws, start_row, title, metric_rows, col_labels, data_dict, is_formula=False):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="Metric").border = THIN_BORDER
    ws.cell(row=start_row, column=1).font = SUBHEADER_FONT
    ws.cell(row=start_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    data_start_row = start_row
    for r_idx, metric in enumerate(metric_rows):
        row_cell = ws.cell(row=start_row, column=1, value=metric)
        row_cell.font = LABEL_FONT
        row_cell.border = THIN_BORDER
        for c_idx, col in enumerate(col_labels, start=2):
            val = data_dict.get((metric, str(col)))
            if val is not None:
                cell = ws.cell(row=start_row, column=c_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                cell.number_format = "0.0000"
        start_row += 1

    return start_row + 1, data_start_row, start_row - 1

def write_selections_section(ws, start_row, col_labels, prior_selections=None, measure=None):
    title_cell = ws.cell(row=start_row, column=1, value="LDF Selections")
    style_header(title_cell, "section")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(col_labels) + 1)
    start_row += 1

    ws.cell(row=start_row, column=1, value="").border = THIN_BORDER
    for c_idx, col in enumerate(col_labels, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        style_header(cell, "subheader")
    start_row += 1

    if prior_selections is not None and measure is not None:
        measure_prior = prior_selections[prior_selections['measure'] == measure]
        if not measure_prior.empty:
            prior_dict = {}
            for _, row in measure_prior.iterrows():
                prior_dict[str(row['interval'])] = {
                    'selection': row['selection'],
                    'reasoning': row['reasoning']
                }

            cell = ws.cell(row=start_row, column=1, value="Prior Selection")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['selection']
                    c.number_format = "0.0000"
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="right")
            start_row += 1

            cell = ws.cell(row=start_row, column=1, value="Prior Reasoning")
            style_header(cell, "prior")
            for c_idx, col in enumerate(col_labels, start=2):
                c = ws.cell(row=start_row, column=c_idx)
                if str(col) in prior_dict:
                    c.value = prior_dict[str(col)]['reasoning']
                c.fill = PRIOR_FILL
                c.border = THIN_BORDER
                c.font = DATA_FONT
                c.alignment = Alignment(horizontal="left", wrap_text=True)
            start_row += 1

            start_row += 1

    for label in ["Rules-Based AI Selection", "Rules-Based AI Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "selection")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = SELECTION_FILL
            c.border = THIN_BORDER
            if label == "Rules-Based AI Selection":
                c.number_format = "0.0000"
            if label == "Rules-Based AI Reasoning":
                c.alignment = Alignment(horizontal="left", wrap_text=True)
                c.font = DATA_FONT
        start_row += 1

    start_row += 1

    for label in ["Open-Ended AI Selection", "Open-Ended AI Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "ai")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = AI_FILL
            c.border = THIN_BORDER
            if label == "Open-Ended AI Selection":
                c.number_format = "0.0000"
            if label == "Open-Ended AI Reasoning":
                c.alignment = Alignment(horizontal="left", wrap_text=True)
                c.font = DATA_FONT
        start_row += 1

    start_row += 1

    sel_row = None
    for label in ["User Selection", "User Reasoning"]:
        cell = ws.cell(row=start_row, column=1, value=label)
        style_header(cell, "user")
        for c_idx in range(2, len(col_labels) + 2):
            c = ws.cell(row=start_row, column=c_idx, value="")
            c.fill = USER_FILL
            c.border = THIN_BORDER
            if label == "User Selection":
                c.number_format = "0.0000"
                if sel_row is None:
                    sel_row = start_row
            if label == "User Reasoning":
                c.alignment = Alignment(horizontal="left", wrap_text=True)
                c.font = DATA_FONT
        start_row += 1

    return start_row, sel_row

_AVG_PERIOD_ORDER = ['all', '3yr', '5yr', '10yr']
_AVG_TYPE_ORDER = ['simple', 'weighted', 'exclude_high_low']

def _sort_avg_cols(cols):
    def key(col):
        normalized = col.replace('avg_', '')
        for p_idx, p in enumerate(_AVG_PERIOD_ORDER):
            for t_idx, t in enumerate(_AVG_TYPE_ORDER):
                if normalized == f"{t}_{p}":
                    return (p_idx, t_idx)
        return (99, 99)
    return sorted(cols, key=key)

def _avg_display_name(col):
    return col.replace('avg_exclude_high_low', 'exclude_high_low')

def atoa_formula(r, c, tri_data_start, ata_data_start):
    tri_row = tri_data_start + (r - ata_data_start)
    num_col = get_column_letter(c + 1)
    den_col = get_column_letter(c)
    return f"=IFERROR(IF({num_col}{tri_row}=\"\",\"\",{num_col}{tri_row}/{den_col}{tri_row}),\"\")"

def simple_avg_formula(col, row_range, ata_range):
    a, b = row_range
    ata_a, ata_b = ata_range
    return f"=IFERROR(AVERAGE({col}{a}:{col}{b}),AVERAGE({col}{ata_a}:{col}{ata_b}))"

def weighted_avg_formula(col, ata_range, tri_range, all_ata_range, all_tri_range):
    ata_a, ata_b = ata_range
    tri_a, tri_b = tri_range
    ata_ref = f"{col}{ata_a}:{col}{ata_b}"
    tri_ref = f"{col}{tri_a}:{col}{tri_b}"
    all_ata = f"{col}{all_ata_range[0]}:{col}{all_ata_range[1]}"
    all_tri = f"{col}{all_tri_range[0]}:{col}{all_tri_range[1]}"
    primary = f"SUMPRODUCT({ata_ref},{tri_ref})/SUMPRODUCT(({ata_ref}<>\"\")*{tri_ref})"
    fallback = f"SUMPRODUCT({all_ata},{all_tri})/SUMPRODUCT(({all_ata}<>\"\")*{all_tri})"
    return f"=IFERROR({primary},{fallback})"

def excl_hl_formula(col, row_range, ata_range):
    a, b = row_range
    ata_a, ata_b = ata_range
    rng = f"{col}{a}:{col}{b}"
    all_rng = f"{col}{ata_a}:{col}{ata_b}"
    primary = f"TRIMMEAN({rng},2/COUNT({rng}))"
    fallback = f"AVERAGE({all_rng})"
    return f"=IFERROR({primary},{fallback})"

def build_main_sheet(ws, measure, df2, df4, df_prior=None):
    df_m = df2[df2['measure'] == measure].copy()
    periods = df_m['period'].cat.categories.tolist()
    ages = [str(a) for a in df_m['age'].cat.categories.tolist()]
    intervals = [str(i) for i in df_m['interval'].dropna().cat.categories.tolist()]

    # 1. Loss Triangle
    loss_dict = {}
    for _, row in df_m.iterrows():
        loss_dict[(str(row['period']), str(row['age']))] = row['value']
    row_ptr, tri_start, tri_end = write_triangle(ws, 1, measure, periods, ages, loss_dict, "#,##0")
    
    # 2. Age-to-Age Factors (formulas)
    ldf_dict = {}
    # We populate ldf_dict with dummy string to reserve cells, will overwrite with formula
    for p in periods:
        for i in intervals:
            ldf_dict[(str(p), str(i))] = ""
            
    row_ptr, ata_start, ata_end = write_triangle(ws, row_ptr, "Age-to-Age Factors", periods, intervals, ldf_dict, "0.0000", is_formula=True)
    
    # Overwrite LDFs with formula
    for r in range(ata_start, ata_end + 1):
        for c in range(2, len(intervals) + 2):
            ws.cell(row=r, column=c).value = atoa_formula(r, c, tri_start, ata_start)

    # 3. Averages (formulas)
    df_avg = df4[df4['measure'] == measure].copy()
    raw_avg_cols = [col for col in df_avg.columns if col not in ['measure', 'interval'] and not col.startswith('cv_') and not col.startswith('slope_')]
    avg_cols = _sort_avg_cols(raw_avg_cols)
    display_cols = [_avg_display_name(c) for c in avg_cols]
    intervals_with_tail = intervals + ["Tail"]

    avg_dict = {}
    for d in display_cols:
        for i in intervals_with_tail:
            avg_dict[(d, str(i))] = ""
            
    row_ptr, avg_start, avg_end = write_metrics_table(ws, row_ptr, "Averages", display_cols, intervals_with_tail, avg_dict, is_formula=True)
    
    # Map periods to ranges
    all_ata_rng = (ata_start, ata_end)
    all_tri_rng = (tri_start, tri_end)
    n_periods = len(periods)
    
    def get_ranges(n):
        if n >= n_periods:
            return all_ata_rng, all_tri_rng
        return (ata_end - n + 1, ata_end), (tri_end - n + 1, tri_end)

    # Overwrite Averages with formulas
    for r_idx, display in enumerate(display_cols):
        r = avg_start + r_idx
        for c in range(2, len(intervals) + 2):
            col = get_column_letter(c)
            if "all" in display:
                ata_rng, tri_rng = all_ata_rng, all_tri_rng
            elif "3yr" in display:
                ata_rng, tri_rng = get_ranges(3)
            elif "5yr" in display:
                ata_rng, tri_rng = get_ranges(5)
            elif "10yr" in display:
                ata_rng, tri_rng = get_ranges(10)
            else:
                ata_rng, tri_rng = all_ata_rng, all_tri_rng
                
            if "simple" in display:
                ws.cell(row=r, column=c).value = simple_avg_formula(col, ata_rng, all_ata_rng)
            elif "weighted" in display:
                ws.cell(row=r, column=c).value = weighted_avg_formula(col, ata_rng, tri_rng, all_ata_rng, all_tri_rng)
            elif "exclude_high_low" in display:
                ws.cell(row=r, column=c).value = excl_hl_formula(col, ata_rng, all_ata_rng)

    # 4. Selections
    row_ptr, sel_row = write_selections_section(ws, row_ptr, intervals, prior_selections=df_prior, measure=measure)
    
    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(ages) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12

    return {
        "tri_start": tri_start, "tri_end": tri_end,
        "sel_row": sel_row
    }

def build_exposure_sheet(ws, df2):
    """Build exposure sheet displaying exposure values per period.
    
    Exposure is typically provided in simple 2-column format (period, value).
    If provided as a triangle, displays it accordingly.
    """
    exp_sub = df2[(df2['measure'] == 'Exposure') & df2['value'].notna()].copy()
    if exp_sub.empty:
        return
    
    periods = exp_sub['period'].cat.categories.tolist()
    
    # Check if exposure has meaningful age values or is in simple format
    has_ages = exp_sub['age'].notna().any()
    
    if has_ages:
        # Triangle format - display as triangle
        ages = [str(a) for a in exp_sub['age'].cat.categories.tolist() if pd.notna(a)]
        exp_dict = {}
        for _, row in exp_sub.iterrows():
            if pd.notna(row['age']):
                exp_dict[(str(row['period']), str(row['age']))] = row['value']
        write_triangle(ws, 1, "Exposure", periods, ages, exp_dict, number_format="#,##0")
    else:
        # Simple format - display as single column
        ws['A1'] = "Exposure by Period"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A3'] = "Period"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = "Exposure"
        ws['B3'].font = Font(bold=True)
        
        for idx, period in enumerate(periods, start=4):
            period_val = exp_sub[exp_sub['period'] == period]['value'].iloc[0] if len(exp_sub[exp_sub['period'] == period]) > 0 else None
            ws[f'A{idx}'] = period
            if period_val is not None:
                ws[f'B{idx}'] = period_val
                ws[f'B{idx}'].number_format = "#,##0"
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15

def build_diagnostic_sheet(ws, diag_col, df2, df3):
    first_measure = df2['measure'].cat.categories[0]
    df_m = df2[df2['measure'] == first_measure].copy()
    periods = df_m['period'].cat.categories.tolist()
    ages = [str(a) for a in df_m['age'].cat.categories.tolist()]

    number_format = DIAG_NUMBER_FORMATS.get(diag_col, "0.0000")

    diag_dict = {}
    for _, row in df3.iterrows():
        diag_dict[(str(row['period']), str(row['age']))] = row[diag_col]

    title = diag_col.replace('_', ' ').title()
    write_triangle(ws, 1, title, periods, ages, diag_dict, number_format)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(ages) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12

def build_cv_slopes_sheet(ws, measure, df2, df4):
    df_m = df2[df2['measure'] == measure].copy()
    intervals = [str(i) for i in df_m['interval'].dropna().cat.categories.tolist()]
    intervals_with_tail = intervals + ["Tail"]

    df_avg = df4[df4['measure'] == measure].copy()
    cv_slope_cols = [col for col in df_avg.columns
                     if col not in ['measure', 'interval']
                     and (col.startswith('cv_') or col.startswith('slope_'))]

    if not cv_slope_cols:
        return

    cv_dict = {}
    for _, row in df_avg.iterrows():
        for metric in cv_slope_cols:
            cv_dict[(metric, str(row['interval']))] = row[metric]

    write_metrics_table(ws, 1, f"{measure} - CV & Slopes", cv_slope_cols, intervals_with_tail, cv_dict)

    ws.column_dimensions['A'].width = 22
    for c_idx in range(2, len(intervals_with_tail) + 2):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12


def df_to_markdown(df, index=False):
    if df.empty:
        return "No data\n"
    if index:
        df = df.reset_index()
    str_df = df.astype(str)
    headers = list(str_df.columns)
    header_str = "| " + " | ".join(headers) + " |"
    sep_str = "|" + "|".join(["---"] * len(headers)) + "|"
    rows = []
    for _, row in str_df.iterrows():
        rows.append("| " + " | ".join(row.values) + " |")
    return "\n".join([header_str, sep_str] + rows) + "\n"
def export_md_data(measures, df2, df3, df4, exp_md):
    # Subagents should use these markdown files as canonical context.
    # The workbook contains formulas that may not be recalculated in headless environments,
    # so cached Excel values can be stale or missing.
    for measure in measures:
        safe_name = measure.lower().replace(' ', '_')
        md_path = Path(SELECTIONS_OUTPUT_PATH) / f"chainladder-context-{safe_name}.md"
        
        tri_sub = df2[(df2['measure'] == measure) & df2['value'].notna()]
        if not tri_sub.empty:
            tri_piv = tri_sub.pivot(index='period', columns='age', values='value')
            tri_md = df_to_markdown(tri_piv, index=True)
        else:
            tri_md = "No data\n"
            
        diag_md = df_to_markdown(df3, index=False)
        avg_md = df_to_markdown(df4[df4['measure'] == measure], index=False)
        
        md_content = f"# Chain Ladder Context: {measure}\n\n"
        md_content += "## Table of Contents\n"
        md_content += "- [Exposure](#exposure)\n"
        md_content += "- [Triangle](#triangle)\n"
        md_content += "- [Diagnostics](#diagnostics)\n"
        md_content += "- [Averages](#averages)\n\n"
        md_content += "## Exposure\n" + exp_md + "\n"
        md_content += "## Triangle\n" + tri_md + "\n"
        md_content += "## Diagnostics\n" + diag_md + "\n"
        md_content += "## Averages\n" + avg_md + "\n"
        
        with open(md_path, 'w') as f:
            f.write(md_content)
        print(f"Exported MD: {md_path}")

def main():
    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    if Path(output_file).exists():
        raise FileExistsError(
            f"Output file already exists: {output_file}\n"
            "Delete or rename the file before re-running to avoid overwriting manual edits."
        )

    df2 = pd.read_parquet(OUTPUT_PATH + "2_enhanced.parquet")
    df3 = pd.read_parquet(OUTPUT_PATH + "3_diagnostics.parquet")
    df4 = pd.read_parquet(OUTPUT_PATH + "4_ldf_averages.parquet")

    print(f"Loaded data: {len(df2)} enhanced rows, {len(df3)} diagnostic rows, {len(df4)} average rows")

    prior_selections_path = Path(OUTPUT_PATH) / "../prior-selections.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"Loaded {len(df_prior)} prior selections")
    else:
        print("No prior selections found (optional)")

    for col in df3.columns:
        if col not in ['period', 'age']:
            if ('rate' in col.lower() or 'ratio' in col.lower() or 'to_' in col.lower()):
                if df3[col].dropna().abs().max() <= 2.0:
                    df3[col] = df3[col].astype(float)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    raw_measures = df2['measure'].cat.categories.tolist()
    
    exp_sub = df2[(df2['measure'] == 'Exposure') & df2['value'].notna()]
    if not exp_sub.empty:
        exp_piv = exp_sub.pivot(index='period', columns='age', values='value')
        exp_md = df_to_markdown(exp_piv, index=True)
    else:
        exp_md = "No Exposure data\n"
        
    measures = [m for m in raw_measures if m != 'Exposure']
    
    diagnostic_cols = [col for col in df3.columns if col not in ['period', 'age']]
    
    layout_infos = {}

    for measure in measures:
        ws = wb.create_sheet(title=measure[:31])
        layout_infos[measure] = build_main_sheet(ws, measure, df2, df4, df_prior=df_prior)
        print(f"Built main sheet: {measure[:31]}")

    # Add exposure sheet if exposure data is available
    if not exp_sub.empty:
        ws = wb.create_sheet(title="Exposure")
        build_exposure_sheet(ws, df2)
        print("Built Exposure sheet")

    for diag_col in diagnostic_cols:
        sheet_title = diag_sheet_name(diag_col)
        ws = wb.create_sheet(title=sheet_title)
        build_diagnostic_sheet(ws, diag_col, df2, df3)
        print(f"Built diagnostic sheet: {sheet_title}")

    for measure in measures:
        sheet_title = cv_slopes_sheet_name(measure)
        ws = wb.create_sheet(title=sheet_title)
        build_cv_slopes_sheet(ws, measure, df2, df4)
        print(f"Built CV & Slopes sheet: {sheet_title}")

    wb.save(output_file)
    print(f"\nSaved: {output_file}")
    
    export_md_data(measures, df2, df3, df4, exp_md)

if __name__ == "__main__":
    main()
