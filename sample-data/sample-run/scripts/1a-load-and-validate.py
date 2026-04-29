"""
Data preparation script for triangle analysis.

Reads raw triangle data, optional prior selections, and optional expected loss rates,
validates them, and saves to standardized format.

Run from scripts/ directory: python 1a-load-and-validate.py
"""

import pandas as pd
from pathlib import Path
from typing import Optional

from modules import config
from modules.validators import (
    validate_combined_data,
    validate_prior_selections, 
    validate_expected_loss_rates
)

# Paths from modules/config.py — override here if needed:
DATA_FILE_PATH = config.RAW_DATA
OUTPUT_PATH = config.PROCESSED_DATA
EXPECTED_LOSS_RATES_FILE = None  # Optional: path to expected loss rates file


# =============================================================================
# TODO: IMPLEMENT DATA READING FUNCTIONS BELOW
# =============================================================================

def read_triangle_data(
    file_path: str,
    sheet_name: Optional[str] = None,
    **kwargs
) -> pd.DataFrame:
    """
    Reads a loss/count triangle from Triangle Examples 1.xlsx.

    Two header formats are handled:
      - Paid/Incurred (Paid 1, Inc 1): Row 0 = "Age of Evaluation" label row,
        Row 1 = ["Accident Year", 11, 23, ...], Row 2+ = data
      - Count (Ct 1): Row 0 = ["Accident Year", 11, 23, ...], Row 1+ = data

    kwargs must supply:
        measure    (str)  : "Paid Loss", "Incurred Loss", "Reported Count", etc.
        unit_type  (str)  : "Dollars" or "Count"
        has_label_row (bool): True for Paid/Incurred format, False for Count
    """
    import openpyxl
    measure   = kwargs['measure']
    unit_type = kwargs['unit_type']
    has_label_row = kwargs.get('has_label_row', True)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    if has_label_row:
        # Row 0 = "Age of Evaluation" label row — skip
        # Row 1 = ["Accident Year", 11, 23, 35, ...]
        age_row = all_rows[1]
        data_rows = all_rows[2:]
    else:
        # Row 0 = ["Accident Year", 11, 23, 35, ...]
        age_row = all_rows[0]
        data_rows = all_rows[1:]

    # Extract age labels (skip column 0 which is "Accident Year")
    ages = [int(a) for a in age_row[1:] if a is not None]

    records = []
    for row in data_rows:
        if row[0] is None:
            continue
        period = str(int(row[0]))
        values = list(row[1:len(ages)+1])
        for age_val, cell_val in zip(ages, values):
            if cell_val is not None:
                records.append({
                    'period': period,
                    'age': str(age_val),
                    'value': float(cell_val),
                    'measure': measure,
                    'unit_type': unit_type,
                    'source': f'{sheet_name} ({file_path})',
                    'details': '',
                })

    return pd.DataFrame(records)


def read_and_process_triangles():
    """
    *** IMPLEMENT THIS FUNCTION TO PROCESS YOUR TRIANGLES ***
    
    This function should:
    1. Call read_triangle_data() for each triangle type you have
    2. Combine them into a single DataFrame
    3. Ensure proper categorical ordering (CRITICAL: age must be ordered categorical)
    4. Pass validation
    5. Save to parquet and CSV
    
    Example Implementation:
    -----------------------
    # Read each triangle type
    incurred = read_triangle_data(
        file_path=DATA_FILE_PATH + "my_data.xlsx",
        sheet_name="Incurred",
        measure="Incurred Loss",
        unit_type="Dollars"
    )
    
    paid = read_triangle_data(...)
    
    # For Exposure: create with age=None, but define age categories from other triangles
    exposure_data = []
    for period in incurred['period'].cat.categories:
        exposure_data.append({
            'period': period,
            'age': None,  # Exposure doesn't develop
            'value': get_exposure_for_period(period),
            'measure': 'Exposure',
            'unit_type': 'Count',
            'source': 'your_source',
            'details': ''
        })
    exposure = pd.DataFrame(exposure_data)
    
    # Combine
    all_data = pd.concat([incurred, paid, exposure, ...], ignore_index=True)
    
    # CRITICAL: Re-apply categorical ordering after concat
    # The age column MUST be an ordered categorical even though Exposure rows have None values
    all_data['period'] = pd.Categorical(
        all_data['period'], 
        categories=incurred['period'].cat.categories, 
        ordered=True
    )
    all_data['age'] = pd.Categorical(
        all_data['age'],  # This will have None values for Exposure rows
        categories=incurred['age'].cat.categories,  # Use age categories from a non-Exposure triangle
        ordered=True
    )
    all_data['measure'] = all_data['measure'].astype('category')
    all_data['unit_type'] = all_data['unit_type'].astype('category')
    all_data['source'] = all_data['source'].astype('category')
    
    # Validate using the combined validator that handles both triangles and exposure
    validate_combined_data(all_data)
    
    # Save
    all_data.to_parquet(OUTPUT_PATH + "1_triangles.parquet", index=False)
    all_data.to_csv(OUTPUT_PATH + "1_triangles.csv", index=False)
    """
    import openpyxl

    xl_file = DATA_FILE_PATH + "Triangle Examples 1.xlsx"

    # --- Paid Loss ---
    paid = read_triangle_data(
        xl_file, sheet_name="Paid 1",
        measure="Paid Loss", unit_type="Dollars", has_label_row=True
    )

    # --- Incurred Loss ---
    incurred = read_triangle_data(
        xl_file, sheet_name="Inc 1",
        measure="Incurred Loss", unit_type="Dollars", has_label_row=True
    )

    # --- Reported Count (Ct 1 has no "Age of Evaluation" label row) ---
    count = read_triangle_data(
        xl_file, sheet_name="Ct 1",
        measure="Reported Count", unit_type="Count", has_label_row=False
    )

    # --- Exposure (payroll) from 2-column sheet ---
    wb = openpyxl.load_workbook(xl_file, data_only=True)
    exp_ws = wb['Exposure']
    exp_records = []
    for row in exp_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        try:
            val = float(row[1])
        except (TypeError, ValueError):
            continue
        exp_records.append({
            'period': str(int(row[0])),
            'age': None,
            'value': val,
            'measure': 'Exposure',
            'unit_type': 'Count',
            'source': f'Exposure (Triangle Examples 1.xlsx)',
            'details': 'Payroll',
        })
    exposure = pd.DataFrame(exp_records)

    # Derive ordered categories from paid triangle (most complete triangle)
    period_cats = sorted(paid['period'].unique(), key=lambda x: int(x))
    age_cats    = sorted(paid['age'].unique(), key=lambda x: int(x))

    # Combine all
    all_data = pd.concat([paid, incurred, count, exposure], ignore_index=True)

    # Apply ordered categoricals
    all_data['period'] = pd.Categorical(all_data['period'], categories=period_cats, ordered=True)
    all_data['age']    = pd.Categorical(all_data['age'],    categories=age_cats,    ordered=True)
    all_data['measure']   = all_data['measure'].astype('category')
    all_data['unit_type'] = all_data['unit_type'].astype('category')
    all_data['source']    = all_data['source'].astype('category')

    # Validate
    validate_combined_data(all_data)

    # Save
    all_data.to_parquet(OUTPUT_PATH + "1_triangles.parquet", index=False)
    all_data.to_csv(OUTPUT_PATH + "1_triangles.csv", index=False)
    print(f"  Saved {len(all_data)} rows to processed-data/")


def read_and_process_prior_selections(triangle_data: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Read and validate prior selections if they exist."""
    prior_file = OUTPUT_PATH + "../prior-selections.csv"
    
    if not Path(prior_file).exists():
        print("  No prior selections file found (optional)")
        return None
    
    print(f"  Found prior selections file: {prior_file}")
    df_prior = pd.read_csv(prior_file)
    
    # Ensure required columns
    required = ['measure', 'interval', 'selection']
    if not all(col in df_prior.columns for col in required):
        raise ValueError(f"Prior selections must have columns: {', '.join(required)}")
    
    if 'reasoning' not in df_prior.columns:
        df_prior['reasoning'] = ''
    
    if df_prior.empty:
        print("  Prior selections file is empty")
        return None
    
    validate_prior_selections(df_prior, triangle_data)
    
    # Ensure consistent types
    df_prior = df_prior[['measure', 'interval', 'selection', 'reasoning']].copy()
    df_prior['measure'] = df_prior['measure'].astype(str)
    df_prior['interval'] = df_prior['interval'].astype(str)
    df_prior['selection'] = df_prior['selection'].astype(float)
    df_prior['reasoning'] = df_prior['reasoning'].fillna('').astype(str)
    
    print(f"  Validated {len(df_prior)} prior selections")
    return df_prior



def read_and_process_expected_loss_rates(triangle_data: pd.DataFrame, file_path: Optional[str] = None) -> Optional[pd.DataFrame]:
    """Read and validate expected loss rates if they exist."""
    if file_path is None:
        if EXPECTED_LOSS_RATES_FILE is None:
            print("  Expected loss rates not configured")
            return None
        file_path = DATA_FILE_PATH + EXPECTED_LOSS_RATES_FILE
    
    if not Path(file_path).exists():
        print("  No expected loss rates file found (optional)")
        return None
    
    print(f"  Found expected loss rates file: {file_path}")
    
    # Read file
    if file_path.lower().endswith('.csv'):
        df = pd.read_csv(file_path)
    elif file_path.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(file_path, engine='openpyxl', engine_kwargs={'data_only': True})
    else:
        raise ValueError(f"Unsupported file format: {file_path}")
    
    # Clean data
    df = df.dropna(subset=['period', 'expected_loss_rate', 'expected_freq'], how='all')
    df = df.dropna(subset=['period'])
    df['period'] = df['period'].astype(str).str.strip()
    df = df.dropna(subset=['expected_loss_rate', 'expected_freq'], how='all')
    return df

if __name__ == "__main__":
    """Run the data preparation process."""
    print("="*70)
    print("TRIANGLE DATA PREPARATION SCRIPT")
    print("="*70)
    
    try:
        # Process triangle data
        print("\n[1/3] Processing triangle data...")
        read_and_process_triangles()
        print(f"✓ Triangle data complete: {OUTPUT_PATH}1_triangles.parquet/.csv")
        
        # Load triangles and validate
        print("  Validating combined data...")
        df_triangles = pd.read_parquet(OUTPUT_PATH + "1_triangles.parquet")
        validate_combined_data(df_triangles)
        print(f"  ✓ Validation passed: {len(df_triangles)} rows")
        
        # Process prior selections (optional)
        print("\n[2/3] Processing prior selections (if available)...")
        df_prior = read_and_process_prior_selections(df_triangles)
        if df_prior is not None:
            validate_prior_selections(df_prior, df_triangles)
            output_file = OUTPUT_PATH + "../prior-selections.csv"
            df_prior.to_csv(output_file, index=False)
            print(f"✓ Prior selections validated and saved to: {output_file}")
        
        # Process expected loss rates (optional)
        print("\n[3/3] Processing expected loss rates (if available)...")
        df_expected = read_and_process_expected_loss_rates(df_triangles)
        if df_expected is not None:
            validate_expected_loss_rates(df_expected, df_triangles)
            print(f"\u2713 Expected loss rates validated and saved to: {OUTPUT_PATH}1_expected_loss_rates.parquet/.csv")
        
        print("\n" + "="*70)
        print("\u2713 DATA PREPARATION COMPLETE!")
        print("="*70)
        
    except NotImplementedError as e:
        print("\n" + str(e))
        print("\nNEXT STEPS:")
        print("1. Implement read_triangle_data() to read your raw data source")
        print("2. Implement read_and_process_triangles() to combine triangles")
        print("3. Ensure output passes validate_combined_data() checks")
        print("4. Run this script again")
        print("\nSee function docstrings for detailed requirements.")
        print("="*70)
        exit(1)
    except Exception as e:
        print(f"\n\u2717 ERROR: {type(e).__name__}: {e}")
        import traceback; traceback.print_exc()
        print("="*70)
        exit(1)
