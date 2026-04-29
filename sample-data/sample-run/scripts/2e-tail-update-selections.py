# Refreshes only the tail factor selections section of the Excel workbook based on saved 
# selections from JSON files. Allows you to update the displayed selections without rebuilding 
# the entire workbook, which is useful when you need to revert changes or apply selections from 
# another source.

"""
goal: Update ONLY the tail factor selections section of Chain Ladder Selections - Tail.xlsx 
      from tail-factor JSON files. Re-run this script any time selections change without 
      needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 2e-tail-update-selections.py
"""

import json
import os
import glob
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from modules import config
from modules.xl_styles import SELECTION_FILL, AI_FILL, USER_FILL, LABEL_FONT, DATA_FONT, THIN_BORDER

# Paths from modules/config.py — override here if needed:
METHOD_ID = "tail"
SELECTIONS_PATTERN    = config.SELECTIONS + f"{METHOD_ID}-ai-rules-based-*.json"
AI_SELECTIONS_PATTERN = config.SELECTIONS + f"{METHOD_ID}-ai-open-ended-*.json"
EXCEL_FILE            = config.SELECTIONS + "Chain Ladder Selections - Tail.xlsx"


def find_selection_rows(ws):
    """Find the row numbers for Rules-Based AI Selection, Open-Ended AI Selection, and User Selection rows."""
    rules_row = None
    ai_row = None
    user_row = None
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Rules-Based AI Selection":
                rules_row = cell.row
            elif cell.value == "Open-Ended AI Selection":
                ai_row = cell.row
            elif cell.value == "User Selection":
                user_row = cell.row
    
    return rules_row, ai_row, user_row


def has_existing_final_selection(ws):
    """Return True if the User Selection row already has values."""
    _, _, user_row = find_selection_rows(ws)
    if user_row is None:
        return False
    for cell in ws[user_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def has_existing_rules_selection(ws):
    """Return True if the Rules-Based AI Selection row already has values."""
    rules_row, _, _ = find_selection_rows(ws)
    if rules_row is None:
        return False
    for cell in ws[rules_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def has_existing_ai_selection(ws):
    """Return True if the Open-Ended AI Selection row already has values."""
    _, ai_row, _ = find_selection_rows(ws)
    if ai_row is None:
        return False
    for cell in ws[ai_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def update_rules_based_selection(ws, selection):
    """Update the Rules-Based AI Selection row with data from JSON."""
    rules_row, _, _ = find_selection_rows(ws)
    if rules_row is None:
        print(f"  WARNING: Could not find 'Rules-Based AI Selection' row in sheet '{ws.title}'")
        return False
    
    # Column mapping (based on Section D in 2d-tail-create-excel.py)
    # ['Label', 'Cutoff Age', 'Tail Factor', 'Method', 'Reasoning', 'Additional Notes']
    col_map = {
        'cutoff_age': 2,
        'tail_factor': 3,
        'method': 4,
        'reasoning': 5,
        'additional_notes': 6
    }
    
    # Write values
    for field, col_idx in col_map.items():
        value = selection.get(field, '')
        cell = ws.cell(row=rules_row, column=col_idx)
        cell.value = value
        cell.fill = SELECTION_FILL
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Number formatting
        if field == 'cutoff_age' and value:
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="right")
        elif field == 'tail_factor' and value:
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="right")
    
    return True


def update_ai_selection(ws, selection):
    """Update the Open-Ended AI Selection row with data from JSON."""
    _, ai_row, _ = find_selection_rows(ws)
    if ai_row is None:
        print(f"  WARNING: Could not find 'Open-Ended AI Selection' row in sheet '{ws.title}'")
        return False
    
    # Column mapping
    col_map = {
        'cutoff_age': 2,
        'tail_factor': 3,
        'method': 4,
        'reasoning': 5,
        'additional_notes': 6
    }
    
    # Write values
    for field, col_idx in col_map.items():
        value = selection.get(field, '')
        cell = ws.cell(row=ai_row, column=col_idx)
        cell.value = value
        cell.fill = AI_FILL
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Number formatting
        if field == 'cutoff_age' and value:
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="right")
        elif field == 'tail_factor' and value:
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="right")
    
    return True


def load_per_measure_json_files(pattern, selection_type):
    """Load and combine all per-measure JSON files matching the pattern.
    
    Extracts measure name from filename (e.g., 'tail-ai-rules-based-paid_loss.json' -> 'Paid Loss')
    and adds it to each selection object.
    """
    combined = []
    files = glob.glob(pattern)
    
    if not files:
        return []
    
    for filepath in sorted(files):
        try:
            # Extract measure from filename: tail-ai-rules-based-paid_loss.json -> paid_loss
            filename = os.path.basename(filepath)
            # Remove extension
            name_no_ext = os.path.splitext(filename)[0]
            # Extract measure part (after last hyphen)
            parts = name_no_ext.split('-')
            measure_slug = parts[-1] if parts else ""
            # Convert slug to display name: paid_loss -> Paid Loss
            measure = measure_slug.replace('_', ' ').title()
            
            with open(filepath, 'r') as f:
                data = json.load(f)
                # If it's a single object, wrap it in a list
                if isinstance(data, dict):
                    data = [data]
                # Add measure field to each selection
                for sel in data:
                    sel['measure'] = measure
                combined.extend(data)
        except Exception as e:
            print(f"  WARNING: Failed to load {filepath}: {e}")
            continue
    
    print(f"Loaded {len(combined)} {selection_type} tail selections from {len(files)} file(s)")
    return combined


def main():
    """Update Tail Factor Selections Excel file with selections from per-measure JSON files."""
    # Load selections from per-measure JSON files
    selections = load_per_measure_json_files(SELECTIONS_PATTERN, "rules-based")
    
    if not selections:
        print(f"No selections found matching pattern: {SELECTIONS_PATTERN}")
        print("Skipping update - no selections to apply.")
        return

    # Load AI selections if available
    ai_selections = load_per_measure_json_files(AI_SELECTIONS_PATTERN, "open-ended AI")
    if not ai_selections:
        print(f"No open-ended AI selections found matching pattern: {AI_SELECTIONS_PATTERN} (optional)")

    # Check if Excel file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel file not found: {EXCEL_FILE}")
        print("Run 2d-tail-create-excel.py first to create the Excel file.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Abort if any sheet has final selections to avoid overwriting actuary work
    sheets_with_final = [s for s in wb.sheetnames if has_existing_final_selection(wb[s])]
    if sheets_with_final:
        raise ValueError(
            f"Existing user selections found in sheet(s): {sheets_with_final}\n"
            "Clear user selections manually before re-running to avoid overwriting actuary edits."
        )

    # Abort if rules-based rows already have values
    sheets_with_rules = [s for s in wb.sheetnames if has_existing_rules_selection(wb[s])]
    if sheets_with_rules:
        raise ValueError(
            f"Existing rules-based selections found in sheet(s): {sheets_with_rules}\n"
            "Clear rules-based selections manually before re-running."
        )

    # Abort if AI rows already have values
    sheets_with_ai = [s for s in wb.sheetnames if has_existing_ai_selection(wb[s])]
    if sheets_with_ai:
        raise ValueError(
            f"Existing AI selections found in sheet(s): {sheets_with_ai}\n"
            "Clear AI selections manually before re-running."
        )

    # Build measure -> selection mapping (tail selections have one per measure, not per interval)
    by_measure = {}
    for sel in selections:
        m = sel["measure"]
        by_measure[m] = sel

    by_measure_ai = {}
    for sel in ai_selections:
        m = sel["measure"]
        by_measure_ai[m] = sel

    updated_count = 0
    ai_updated_count = 0

    for sheet_name in wb.sheetnames:
        # Match sheet name to measure (sheet name may be truncated to 31 chars)
        matched = None
        for measure in by_measure:
            if sheet_name == measure[:31]:
                matched = measure
                break

        if matched:
            # Update rules-based selection
            if update_rules_based_selection(wb[sheet_name], by_measure[matched]):
                print(f"  Updated rules-based selection in '{sheet_name}'")
                updated_count += 1

            # Also write AI selection for this measure if available
            if matched in by_measure_ai:
                if update_ai_selection(wb[sheet_name], by_measure_ai[matched]):
                    print(f"  Updated AI selection in '{sheet_name}'")
                    ai_updated_count += 1
        else:
            print(f"No selections found for sheet '{sheet_name}' - skipping")

    wb.save(EXCEL_FILE)
    print(f"\nUpdated {updated_count} rules-based selections and {ai_updated_count} AI selections")
    print(f"Saved: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
