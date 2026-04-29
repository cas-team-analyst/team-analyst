from openpyxl import load_workbook

wb = load_workbook('../Analysis.xlsx')
ws = wb['Count Selection']

print('=== Count Selection Headers ===')
for i in range(1, 15):
    if ws.cell(1, i).value:
        print(f'Col {i}: {ws.cell(1, i).value}')

print('\n=== Row 2 Values ===')
sel_col = None
ibnr_col = None
unpaid_col = None

for i in range(1, 15):
    hdr = ws.cell(1, i).value
    if hdr == 'Selected Ultimate':
        sel_col = i
    elif hdr == 'IBNR':
        ibnr_col = i
    elif hdr == 'Unpaid':
        unpaid_col = i

if sel_col:
    print(f'Selected Ultimate: {ws.cell(2, sel_col).value}')
print(f'Reported: {ws.cell(2, 3).value}')
print(f'Closed: {ws.cell(2, 4).value}')

if ibnr_col:
    print(f'IBNR: {ws.cell(2, ibnr_col).value}')
    print(f'IBNR format: {ws.cell(2, ibnr_col).number_format}')

if unpaid_col:
    print(f'Unpaid column exists at col {unpaid_col}')
    print(f'Unpaid value: {ws.cell(2, unpaid_col).value}')
else:
    print('Unpaid column does NOT exist')

wb.close()
