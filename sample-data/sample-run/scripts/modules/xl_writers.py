from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules.xl_styles import DATA_FONT, THIN_BORDER, style_header
from modules.xl_utils import _safe, _to_py


def _data_cell(cell, value, num_fmt=None):
    """Write a styled data cell with border and alignment."""
    value = _to_py(_safe(value))
    cell.value     = value
    cell.font      = DATA_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="right" if isinstance(value, (int, float)) else "left",
        vertical="center",
    )
    if num_fmt and value is not None and isinstance(value, (int, float)):
        cell.number_format = num_fmt


def _write_title_and_headers(ws, title, headers, col_width=18):
    """
    Write a title row (row 1) and a styled header row (row 2).
    Returns 3 — the first data row.
    This format matches script 7's read_with_title() convention.
    """
    title_cell = ws.cell(1, 1, title)
    style_header(title_cell, "header")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 20

    for c, text in enumerate(headers, 1):
        cell = ws.cell(2, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A3"
    return 3


def _write_headers(ws, headers, col_width=18):
    """
    Write a single styled header row (row 1), no title row.
    Returns 2 — the first data row.
    Matches script 7's read_no_title() convention ("Sel - " sheets).
    """
    for c, text in enumerate(headers, 1):
        cell = ws.cell(1, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A2"
    return 2
