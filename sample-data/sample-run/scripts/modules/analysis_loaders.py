import pathlib

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from modules.xl_values import UNPAID_PROXY

# Preferred display order for method columns — discovered dynamically from parquet.
_METHOD_COLS = [
    ("ultimate_cl", "Chain Ladder"),
    ("ultimate_ie", "Initial Expected"),
    ("ultimate_bf", "BF"),
]

# Map individual measure names to Ultimates.xlsx selection category sheets.
# projected-ultimates.parquet has per-measure data (Incurred Loss, Paid Loss, etc.)
# but Ultimates.xlsx has category sheets (Losses, Counts) where one ultimate is
# selected per category. This mapping converts measure → category for lookups.
MEASURE_TO_CATEGORY = {
    "Incurred Loss":   "Losses",
    "Paid Loss":       "Losses",
    "Reported Count":  "Counts",
    "Closed Count":    "Counts",
}


def load_selections(excel_path):
    """
    Load final selections from Ultimates.xlsx.
    Priority per (measure, period): User Selection > Rules-Based AI Selection.
    Open-Ended AI is intentionally excluded — it is not a trusted final selection source.
    Columns are located by header name — adapts to layout changes automatically.
    Returns dict {(measure, period): float}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        print(f"  Note: {excel_path} not found -- no selections loaded")
        return {}

    wb = load_workbook(excel_path, data_only=True)
    sel_lookup = {}
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Accident Period":
                period_col = cell.column
            elif v == "User Selection":
                user_col = cell.column
            elif v == "Rules-Based AI Selection":
                rb_col = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in sel_lookup:
                continue
            for col_idx in [user_col, rb_col]:
                if col_idx is None:
                    continue
                v = ws.cell(row=row, column=col_idx).value
                if isinstance(v, (int, float)) and not (isinstance(v, float) and v != v):
                    sel_lookup[key] = float(v)
                    total += 1
                    break

    wb.close()
    print(f"  Loaded {total} selections from {excel_path}")
    return sel_lookup


def load_selection_reasoning(excel_path):
    """
    Load reasoning text for each (measure, period) from Ultimates.xlsx.
    Priority matches load_selections: User Selection reasoning first, then RB-AI.
    Returns {(measure, period): str}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        return {}

    wb = load_workbook(excel_path, data_only=True)
    reason_lookup = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = user_r_col = rb_r_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Accident Period":            period_col = cell.column
            elif v == "User Selection":           user_col   = cell.column
            elif v == "Rules-Based AI Selection": rb_col     = cell.column
            elif v == "User Reasoning":           user_r_col = cell.column
            elif v == "Rules-Based AI Reasoning": rb_r_col   = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in reason_lookup:
                continue

            user_v = ws.cell(row=row, column=user_col).value if user_col else None
            rb_v   = ws.cell(row=row, column=rb_col).value   if rb_col   else None
            is_num = lambda v: isinstance(v, (int, float)) and not (isinstance(v, float) and v != v)

            if is_num(user_v) and user_r_col:
                reason_lookup[key] = ws.cell(row=row, column=user_r_col).value or ""
            elif is_num(rb_v) and rb_r_col:
                reason_lookup[key] = ws.cell(row=row, column=rb_r_col).value or ""

    wb.close()
    return reason_lookup


def load_combined(ultimates_path, sel_lookup):
    """
    Load projected ultimates, apply final selections, compute IBNR and Unpaid.
    Methods (CL/IE/BF) are discovered dynamically from non-empty parquet columns.
    Raises ValueError for any non-Exposure (measure, period) missing from sel_lookup.
    Returns (df, available_methods) where available_methods = [(col, label), ...].
    """
    df = pd.read_parquet(ultimates_path)
    df["period"]  = df["period"].astype(str)
    df["measure"] = df["measure"].astype(str)

    available_methods = [
        (col, label) for col, label in _METHOD_COLS
        if col in df.columns and df[col].notna().any()
    ]

    missing = [
        (row["measure"], row["period"])
        for _, row in df.iterrows()
        if row["measure"] != "Exposure" and
           (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]) not in sel_lookup
    ]
    if missing:
        lines = "\n  ".join(f"{m} / {p}" for m, p in missing)
        raise ValueError(
            f"{len(missing)} (measure, period) pair(s) have no User Selection or "
            f"Rules-Based AI Selection in Ultimates.xlsx:\n  {lines}\n"
            "Populate Rules-Based AI Selection (run 5b) or add a User Selection."
        )

    actual_lookup = df.set_index(["period", "measure"])["actual"].to_dict()

    def _pick_unpaid(row):
        proxy = UNPAID_PROXY.get(row["measure"])
        if proxy is None:
            return np.nan
        proxy_actual = actual_lookup.get((row["period"], proxy), np.nan)
        if pd.isna(proxy_actual) or pd.isna(row["selected_ultimate"]):
            return np.nan
        return row["selected_ultimate"] - proxy_actual

    df["selected_ultimate"] = df.apply(
        lambda row: sel_lookup.get(
            (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]),
            np.nan
        ),
        axis=1
    )
    df["selected_ibnr"]   = df["selected_ultimate"] - df["actual"]
    df["selected_unpaid"] = df.apply(_pick_unpaid, axis=1)

    return df, available_methods


def get_exposure(triangles_path):
    """Latest Exposure value per period from triangles. Returns {} if absent."""
    if not pathlib.Path(triangles_path).exists():
        return {}
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return {}
    exp["period"]  = exp["period"].astype(str)
    exp["age_num"] = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    latest = exp.sort_values("age_num").groupby("period").last()
    return latest["value"].to_dict()


def get_triangles(triangles_path):
    """Load triangle data. Returns empty DataFrame when file is absent."""
    if not pathlib.Path(triangles_path).exists():
        return pd.DataFrame()
    tri = pd.read_parquet(triangles_path)
    tri["period"]  = tri["period"].astype(str)
    tri["measure"] = tri["measure"].astype(str)
    return tri
