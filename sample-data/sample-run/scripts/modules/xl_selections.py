# Shared helpers for copying selection workbook sheets into output workbooks
# with AI option/reasoning rows filtered out and selection rows collapsed to one "Selected" row.
#
# Used by:
#   6-analysis-create-excel.py  -- CL LDF and Tail sheets
#   2a-chainladder-create-excel.py -- chain ladder update scripts (future)

import copy

from openpyxl.styles import Alignment

# Rows in the selection section that are never shown in the output.
SKIP_ROW_LABELS = frozenset({
    "Prior Selection", "Prior Reasoning", "Prior Delta",
    "Rules-Based AI Reasoning",
    "Open-Ended AI Selection", "Open-Ended AI Reasoning",
    "User Reasoning",
})

# Rows that each contain one selection option -- collapsed into a single "Selected" row.
SELECTION_LABELS = frozenset({"Rules-Based AI Selection", "User Selection"})


def find_selected_values(ws):
    """
    Return the priority-correct selected row values (cols 2+) from a worksheet.
    Priority: User Selection (if populated) > Rules-Based AI Selection.
    Returns list of values or None if neither row has data.
    """
    user_vals = rb_vals = None
    for src_cells in ws.iter_rows():
        col1 = src_cells[0].value if src_cells else None
        if col1 in ("User Selection", "Rules-Based AI Selection"):
            vals = [c.value for c in src_cells[1:]]
            has_data = any(v not in (None, "") for v in vals)
            if not has_data:
                continue
            if col1 == "User Selection":
                user_vals = vals
            elif col1 == "Rules-Based AI Selection" and rb_vals is None:
                rb_vals = vals
    return user_vals if user_vals is not None else rb_vals


def find_selected_reasoning(ws):
    """
    Return the reasoning text values (cols 2+) for the priority selection.
    Priority: User Selection (if populated) -> User Reasoning; else RB-AI Reasoning.
    Returns list of values or None if no reasoning found.
    """
    user_sel_has_data = False
    rb_sel_has_data   = False
    user_reason = None
    rb_reason   = None

    for src_cells in ws.iter_rows():
        col1 = src_cells[0].value if src_cells else None
        if col1 in ("User Selection", "Rules-Based AI Selection"):
            vals = [c.value for c in src_cells[1:]]
            if any(v not in (None, "") for v in vals):
                if col1 == "User Selection":
                    user_sel_has_data = True
                else:
                    rb_sel_has_data = True
        elif col1 == "User Reasoning":
            user_reason = [c.value for c in src_cells[1:]]
        elif col1 == "Rules-Based AI Reasoning":
            rb_reason = [c.value for c in src_cells[1:]]

    if user_sel_has_data:
        return user_reason
    elif rb_sel_has_data:
        return rb_reason
    return None


def copy_ws_filtered(ws_src, ws_dst, selected_row_values=None, selected_reasoning=None):
    """
    Copy a worksheet filtering out AI option/reasoning rows.
    All SELECTION_LABELS rows are collapsed into one "Selected" row.
    When selected_reasoning is provided, a "Selected Reasoning" row is emitted
    immediately after the "Selected" row.

    selected_row_values: priority-correct list of values for cols 2+ of the Selected row.
                         None = write the first selection row encountered as-is (relabeled).
    selected_reasoning:  list of text values for cols 2+ of the Selected Reasoning row.
                         None = no reasoning row written.
    """
    selection_done = False
    dst_row = 1

    for src_cells in ws_src.iter_rows():
        col1 = src_cells[0].value if src_cells else None

        if col1 in SKIP_ROW_LABELS:
            continue

        if col1 in SELECTION_LABELS:
            if selection_done:
                continue
            selection_done = True

            # Write "Selected" row
            for src_cell in src_cells:
                dst_cell = ws_dst.cell(dst_row, src_cell.column)
                if src_cell.column == 1:
                    dst_cell.value = "Selected"
                elif selected_row_values is not None:
                    idx = src_cell.column - 2
                    dst_cell.value = selected_row_values[idx] if idx < len(selected_row_values) else None
                else:
                    dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font          = copy.copy(src_cell.font)
                    dst_cell.border        = copy.copy(src_cell.border)
                    dst_cell.fill          = copy.copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment     = copy.copy(src_cell.alignment)
            dst_row += 1

            # Write "Selected Reasoning" row immediately after, reusing fill from selection row
            if selected_reasoning is not None:
                for src_cell in src_cells:
                    dst_cell = ws_dst.cell(dst_row, src_cell.column)
                    if src_cell.column == 1:
                        dst_cell.value = "Selected Reasoning"
                    else:
                        idx = src_cell.column - 2
                        dst_cell.value = (selected_reasoning[idx]
                                          if idx < len(selected_reasoning) else None)
                    if src_cell.has_style:
                        dst_cell.font   = copy.copy(src_cell.font)
                        dst_cell.border = copy.copy(src_cell.border)
                        dst_cell.fill   = copy.copy(src_cell.fill)
                    dst_cell.number_format = ""
                    dst_cell.alignment = Alignment(wrap_text=True, horizontal="left",
                                                   vertical="top")
                dst_row += 1

            continue

        for src_cell in src_cells:
            dst_cell = ws_dst.cell(dst_row, src_cell.column)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font          = copy.copy(src_cell.font)
                dst_cell.border        = copy.copy(src_cell.border)
                dst_cell.fill          = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection    = copy.copy(src_cell.protection)
                dst_cell.alignment     = copy.copy(src_cell.alignment)
        dst_row += 1

    for col in ws_src.column_dimensions:
        ws_dst.column_dimensions[col].width = ws_src.column_dimensions[col].width
    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes
