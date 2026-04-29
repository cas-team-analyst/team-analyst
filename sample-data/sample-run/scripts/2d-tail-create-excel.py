# Creates a formatted Excel workbook for tail factor review and selection. Each measure gets its
# own sheet with observed data, scenario comparisons, and selection areas for actuarial judgment.

"""
goal: Create Chain Ladder Selections - Tail.xlsx for actuarial tail factor review and selection.
      This file should be manually edited by the actuary to make tail factor selections.

Sheet layout per measure:
  - Section A: Triangle Type Banner (measure name + key stats)
  - Section B: Observed Factors Table (AY x age triangle + weighted avg + CV)
  - Section C: Scenario Comparison Table (method x starting_age scenarios with diagnostics)
  - Section D: Selection Area (prior, rule-based, AI, final selection rows)

run-note: When copied to a project, run from the scripts/ directory:
    cd scripts/
    python 2d-tail-create-excel.py
"""

import pandas as pd
import numpy as np
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook

from modules import config
from modules.xl_styles import create_xlsxwriter_formats, COLORS
from modules.markdown_utils import df_to_markdown
from modules.xl_writers import col_letter

# Paths
TAIL_SCENARIOS_PATH = config.PROCESSED_DATA + "tail-scenarios.parquet"
ENHANCED_PATH = config.PROCESSED_DATA + "2_enhanced.parquet"
DIAGNOSTICS_PATH = config.PROCESSED_DATA + "3_diagnostics.parquet"
SELECTIONS_OUTPUT_PATH = config.SELECTIONS
OUTPUT_FILE_NAME = "Chain Ladder Selections - Tail.xlsx"
CL_LDF_EXCEL = config.SELECTIONS + "Chain Ladder Selections - LDFs.xlsx"


def write_section_header(ws, row, col_span, title, fmt, level="section"):
    """Write a section header spanning multiple columns."""
    format_key = {'header': 'header', 'section': 'section', 'subheader': 'subheader'}.get(level, 'section')
    header_fmt = fmt.get(format_key, fmt['section'])
    
    ws.merge_range(row, 0, row, col_span - 1, title, header_fmt)
    return row + 1


CL_LDF_EXCEL = config.SELECTIONS + "Chain Ladder Selections - LDFs.xlsx"


def find_selected_ldfs_in_cl_excel(cl_excel_path, measure):
    """
    Read Chain Ladder Excel to find selected LDFs for a measure.
    Finds ALL THREE selection rows for cascaded value priority logic.
    
    Returns:
        {
            'intervals': {interval: col_idx},
            'selection_rows': {'user': row, 'rules_based': row, 'open_ended': row},
            'cached_values': {interval: value}  # Using priority: User > Rules-Based > Open-Ended
        }
        or {} if file not found or no selections
    """
    path = Path(cl_excel_path)
    if not path.exists():
        print(f"  WARNING: {cl_excel_path} not found")
        return {}
    
    try:
        wb = load_workbook(cl_excel_path, data_only=True)
        if measure not in wb.sheetnames:
            print(f"  WARNING: Sheet '{measure}' not found in {cl_excel_path}")
            return {}
        
        ws = wb[measure]
        
        # Find ALL THREE selection rows
        selection_rows = {}
        selection_labels = {
            "User Selection": "user",
            "Rules-Based AI Selection": "rules_based",
            "Open-Ended AI Selection": "open_ended"
        }
        
        for label, key in selection_labels.items():
            for row_idx in range(1, ws.max_row + 1):
                cell_val = ws.cell(row_idx, 1).value
                if cell_val and str(cell_val).strip() == label:
                    selection_rows[key] = row_idx
                    break
        
        if not selection_rows:
            print(f"  WARNING: No selection rows found in {measure} sheet")
            return {}
        
        # Find interval header row by searching backwards from first found selection row
        first_selection_row = min(selection_rows.values())
        
        import re
        interval_re = re.compile(r'^\d+-\d+$')
        interval_row = None
        interval_map = {}  # {interval: col_idx}
        
        # Search backwards from first selection row
        for row_idx in range(first_selection_row - 1, max(1, first_selection_row - 10), -1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            interval_count = sum(1 for v in row_vals[1:] if v and interval_re.match(str(v).strip()))
            if interval_count >= 3:
                interval_row = row_idx
                for col_idx, val in enumerate(row_vals):
                    if val and interval_re.match(str(val).strip()):
                        interval_key = str(val).strip()
                        interval_map[interval_key] = col_idx + 1  # +1 for 1-based Excel
                break
        
        if not interval_map:
            print(f"  WARNING: No interval headers found near selection rows in {measure} sheet")
            return {}
        
        # Calculate cached values using priority logic: User > Rules-Based > Open-Ended
        cached_values = {}
        for interval, col_idx in interval_map.items():
            value = None
            # Check User first
            if 'user' in selection_rows:
                value = ws.cell(selection_rows['user'], col_idx).value
            # Fall back to Rules-Based if User is None
            if value is None and 'rules_based' in selection_rows:
                value = ws.cell(selection_rows['rules_based'], col_idx).value
            # Fall back to Open-Ended if both above are None
            if value is None and 'open_ended' in selection_rows:
                value = ws.cell(selection_rows['open_ended'], col_idx).value
            
            cached_values[interval] = value
        
        return {
            'intervals': interval_map,
            'selection_rows': selection_rows,
            'cached_values': cached_values
        }
    
    except Exception as e:
        print(f"  ERROR reading Chain Ladder Excel: {e}")
        return {}


def write_selected_ldfs_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt):
    """
    Write selected LDFs section with values from Chain Ladder Excel.
    Uses cascaded priority: User Selection > Rules-Based AI > Open-Ended AI.
    
    Returns:
        next_row (int)
    """
    # Find selected LDFs in Chain Ladder Excel
    ldf_data = find_selected_ldfs_in_cl_excel(cl_excel_path, measure)
    
    if not ldf_data:
        # No selections found - write empty section
        ws.write(start_row, 0, "Selected LDFs", fmt['subheader'])
        ws.write(start_row + 1, 0, "(No selections found in Chain Ladder Excel)", fmt['label'])
        return start_row + 3
    
    intervals = ldf_data['intervals']
    selection_rows = ldf_data['selection_rows']
    cached_values = ldf_data['cached_values']
    
    # Sort intervals by numeric start value
    def interval_sort_key(interval):
        try:
            return int(interval.split('-')[0])
        except (ValueError, IndexError):
            return 999
    
    sorted_intervals = sorted(intervals.keys(), key=interval_sort_key)
    
    # Write header row
    ws.write(start_row, 0, "Selected LDFs", fmt['subheader'])
    for c_idx, interval in enumerate(sorted_intervals):
        ws.write(start_row, c_idx + 1, interval, fmt['subheader'])
    
    # Write value row with cascading IF logic
    row = start_row + 1
    ws.write(row, 0, "(User > Rules-Based > Open-Ended)", fmt['label'])
    
    for c_idx, interval in enumerate(sorted_intervals):
        # Write value directly from cached data
        cached_value = cached_values.get(interval)
        if cached_value is not None and pd.notna(cached_value):
            ws.write(row, c_idx + 1, cached_value, fmt['data_ldf'])
    
    return row + 2  # Skip blank row


def write_observed_factors_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt):
    """
    Section B: Selected LDFs from Chain Ladder Excel (hard-coded values).
    Renamed but kept old function name for compatibility.
    """
    return write_selected_ldfs_section(ws, start_row, measure, cl_excel_path, output_file_path, fmt)


def write_scenario_comparison_section(ws, start_row, measure, df_scenarios, fmt):
    """Section C: Scenario Comparison Table with all diagnostics (no color formatting)."""
    df_m = df_scenarios[df_scenarios['measure'] == measure].copy()
    if df_m.empty:
        return start_row
    
    col_headers = [
        'Starting Age', 'Monotone', 'CV', 'Slope Breaks', 'Method', 'Params',
        'Tail Factor', 'R2', 'LOO Std Dev', 'Gap to Last', 'Gap Flag',
        '% of CDF', 'Materiality OK', '+10% Reserve Delta', '+20% Reserve Delta'
    ]
    
    row = write_section_header(ws, start_row, len(col_headers), "Scenario Comparison", fmt, "section")
    
    # Header row
    for c_idx, header in enumerate(col_headers):
        ws.write(row, c_idx, header, fmt['subheader'])
    row += 1
    
    # Sort scenarios by starting_age then method
    df_m = df_m.sort_values(['starting_age', 'method'])
    
    # Helper to safely convert boolean columns (handles pandas NA)
    def safe_bool(val):
        if pd.isna(val):
            return False
        return bool(val)
    
    # Data rows - simple formatting, no colors
    for _, scenario in df_m.iterrows():
        values = [
            scenario['starting_age'],
            'Y' if safe_bool(scenario.get('is_monotone_from_here', False)) else 'N',
            scenario.get('cv_at_starting_age'),
            scenario.get('slope_sign_changes', 0),
            scenario['method'],
            str(scenario.get('method_params', '')) if pd.notna(scenario.get('method_params')) else '',
            scenario['tail_factor'],
            scenario.get('r_squared'),
            scenario.get('loo_std_dev'),
            scenario.get('gap_to_last_observed'),
            'Y' if safe_bool(scenario.get('gap_flag', False)) else 'N',
            scenario.get('pct_of_cdf'),
            'Y' if safe_bool(scenario.get('materiality_ok', False)) else 'N',
            scenario.get('sensitivity_plus10_reserve_delta'),
            scenario.get('sensitivity_plus20_reserve_delta'),
        ]
        
        for c_idx, val in enumerate(values):
            # Only write cells with valid data - match 2a pattern
            if c_idx == 0:  # Starting Age
                ws.write(row, c_idx, val, fmt['data'])
            elif c_idx == 2:  # CV
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx in [4, 5]:  # Method, Params (text)
                if val:  # Only write non-empty strings
                    ws.write(row, c_idx, val, fmt['label'])
            elif c_idx == 6:  # Tail Factor
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx in [7, 8, 9]:  # R2, LOO Std Dev, Gap to Last
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_ldf'])
            elif c_idx == 11:  # % of CDF
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_pct'])
            elif c_idx in [13, 14]:  # Reserve deltas
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data_num'])
            else:  # Monotone (1), Slope Breaks (3), Gap Flag (10), Materiality OK (12)
                if val is not None and pd.notna(val):
                    ws.write(row, c_idx, val, fmt['data'])
        
        row += 1
    
    return row + 1


def write_selection_section(ws, start_row, measure, fmt, prior_selections=None):
    """Section D: Selection Area (prior, rules-based AI, open-ended AI, user selection)."""
    row = write_section_header(ws, start_row, 6, "Tail Factor Selection", fmt, "section")
    
    headers = ['Label', 'Cutoff Age', 'Tail Factor', 'Method', 'Reasoning', 'Additional Notes']
    for c_idx, header in enumerate(headers):
        ws.write(row, c_idx, header, fmt['subheader'])
    row += 1
    
    # Prior selection row
    if prior_selections is not None:
        prior_m = prior_selections[prior_selections['measure'] == measure]
        if not prior_m.empty:
            prior_row_data = prior_m.iloc[0]
            values = [
                'Prior Selection',
                prior_row_data.get('cutoff_age', ''),
                prior_row_data.get('tail_factor', ''),
                prior_row_data.get('method', ''),
                prior_row_data.get('reasoning', ''),
                ''
            ]
            
            for c_idx, val in enumerate(values):
                if c_idx == 0:  # Label column
                    ws.write(row, c_idx, val, fmt['prior'])
                elif c_idx == 1 and val:  # Cutoff Age
                    ws.write(row, c_idx, val, fmt['prior_num'])
                elif c_idx == 2 and val:  # Tail Factor
                    ws.write(row, c_idx, val, fmt['prior_data'])
                elif c_idx in [3, 4, 5]:  # Text columns
                    ws.write(row, c_idx, val, fmt['prior_text'])
                else:
                    ws.write(row, c_idx, val, fmt['prior'])
            
            row += 1
            
            # Prior Delta row (blank - user can calculate manually if needed)
            ws.write(row, 0, "Prior Delta", fmt['prior'])
            for c_idx in range(1, 6):
                ws.write(row, c_idx, '', fmt['prior'])
            
            # Add note in Reasoning column
            ws.write(row, 4, "(current - prior)", fmt['prior_note'])
            row += 1
            
            row += 1  # Blank row
    
    # Rules-Based AI Selection row
    ws.write(row, 0, "Rules-Based AI Selection", fmt['selection'])
    for c_idx in range(1, 6):
        ws.write(row, c_idx, '', fmt['selection_text'])
    row += 1
    
    # Open-Ended AI Selection row
    ws.write(row, 0, "Open-Ended AI Selection", fmt['ai'])
    for c_idx in range(1, 6):
        ws.write(row, c_idx, '', fmt['ai_text'])
    row += 1
    
    row += 1  # Blank row
    
    # User Selection row
    ws.write(row, 0, "User Selection", fmt['user'])
    for c_idx in range(1, 6):
        ws.write(row, c_idx, '', fmt['user_text'])
    row += 1
    
    return row + 1



def build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, fmt, 
                        cl_excel_path, output_file_path, prior_selections=None):
    """Build complete sheet for one measure."""
    row = 0  # xlsxwriter uses 0-based indexing, so row 0 = Excel row 1
    
    # Section A: Selected LDFs from Chain Ladder Excel (hard-coded values)
    row = write_observed_factors_section(ws, row, measure, cl_excel_path, output_file_path, fmt)
    
    # Section B: Scenario Comparison
    row = write_scenario_comparison_section(ws, row, measure, df_scenarios, fmt)
    
    # Section C: Selection Area
    row = write_selection_section(ws, row, measure, fmt, prior_selections)
    
    # Column widths (xlsxwriter uses 0-based indexing for columns)
    ws.set_column(0, 0, 30)  # A: Label
    ws.set_column(1, 1, 12)  # B: Cutoff Age
    ws.set_column(2, 2, 12)  # C: Tail Factor
    ws.set_column(3, 3, 15)  # D: Method
    ws.set_column(4, 4, 25)  # E: Reasoning
    ws.set_column(5, 5, 30)  # F: Additional Notes



def export_md_data(measures, df_scenarios, df_enhanced, df_diagnostics, exp_md):
    # Subagents should use these markdown files as canonical context.
    # Workbook contains hard-coded values from source data.
    for measure in measures:
        safe_name = measure.lower().replace(' ', '_')
        md_path = Path(SELECTIONS_OUTPUT_PATH) / f"tail-context-{safe_name}.md"
        
        scen_sub = df_scenarios[df_scenarios['measure'] == measure].drop(columns=['measure'], errors='ignore')
        scen_md = df_to_markdown(scen_sub, index=False)
        
        # Get selected LDFs from Chain Ladder Excel instead of observed factors
        ldf_data = find_selected_ldfs_in_cl_excel(CL_LDF_EXCEL, measure)
        if ldf_data and 'intervals' in ldf_data and 'cached_values' in ldf_data:
            # Sort intervals by numeric start value
            def interval_sort_key(interval):
                try:
                    return int(interval.split('-')[0])
                except (ValueError, IndexError):
                    return 999
            
            sorted_intervals = sorted(ldf_data['intervals'].keys(), key=interval_sort_key)
            cached_values = ldf_data['cached_values']
            
            # Create markdown table with intervals as columns
            header = "| Interval | " + " | ".join(sorted_intervals) + " |"
            separator = "|---|" + "|".join(["---"] * len(sorted_intervals)) + "|"
            values = "| Selected LDF | " + " | ".join(
                [f"{cached_values[interval]:.4f}" if cached_values[interval] is not None else "" 
                 for interval in sorted_intervals]
            ) + " |"
            selected_md = header + "\n" + separator + "\n" + values + "\n"
        else:
            selected_md = "(No selected LDFs found in Chain Ladder Excel)\n"
            
        md_content = f"# Tail Context: {measure}\n\n"
        md_content += "## Table of Contents\n"
        md_content += "- [Exposure](#exposure)\n"
        md_content += "- [Scenarios](#scenarios)\n"
        md_content += "- [Selected LDFs](#selected-ldfs)\n\n"
        md_content += "## Exposure\n" + exp_md + "\n"
        md_content += "## Scenarios\n" + scen_md + "\n"
        md_content += "## Selected LDFs\n"
        md_content += "Priority: User Selection > Rules-Based AI > Open-Ended AI\n\n"
        md_content += selected_md
        
        with open(md_path, 'w') as f:
            f.write(md_content)
        print(f"  Exported MD: {md_path}")

def main():
    """Create Tail Factor Selections Excel file."""
    output_file = SELECTIONS_OUTPUT_PATH + OUTPUT_FILE_NAME
    if Path(output_file).exists():
        raise FileExistsError(
            f"Output file already exists: {output_file}\n"
            "Delete or rename the file before re-running to avoid overwriting manual edits."
        )
    
    if not Path(TAIL_SCENARIOS_PATH).exists():
        raise FileNotFoundError(
            f"Tail scenarios file not found: {TAIL_SCENARIOS_PATH}\n"
            "Run 2c-tail-methods-diagnostics.py first to generate tail scenarios."
        )
    
    print("Loading data...")
    df_scenarios = pd.read_parquet(TAIL_SCENARIOS_PATH)
    df_enhanced = pd.read_parquet(ENHANCED_PATH)
    df_diagnostics = pd.read_parquet(DIAGNOSTICS_PATH)
    
    print(f"  {len(df_scenarios)} tail scenarios")
    print(f"  {len(df_enhanced)} enhanced rows")
    print(f"  {len(df_diagnostics)} diagnostic rows")
    
    # Load prior selections if available
    prior_selections_path = Path(config.SELECTIONS) / "tail-factor-prior.csv"
    df_prior = None
    if prior_selections_path.exists():
        df_prior = pd.read_csv(prior_selections_path)
        print(f"  {len(df_prior)} prior tail selections loaded")
    else:
        print("  No prior tail selections found (optional)")
    
    # Create workbook - match 2a script exactly
    wb = xlsxwriter.Workbook(output_file)
    fmt = create_xlsxwriter_formats(wb)
    fmt['wb'] = wb  # Store workbook reference
    
    raw_measures = sorted(df_scenarios['measure'].unique())
    
    exp_sub = df_enhanced[(df_enhanced['measure'] == 'Exposure') & df_enhanced['value'].notna()]
    if not exp_sub.empty:
        # Format Exposure as simple 2-column table (period, value)
        # Exposure doesn't develop over age, so we take the last value per period
        exp_simple = exp_sub.groupby('period', observed=True).agg({'value': 'last'}).reset_index()
        exp_simple.columns = ['Period', 'Exposure']
        exp_simple['Exposure'] = exp_simple['Exposure'].round(0)
        exp_md = df_to_markdown(exp_simple, index=False)
    else:
        exp_md = "No Exposure data\n"
        
    measures = [m for m in raw_measures if m != 'Exposure']
    
    print(f"\nCreating sheets for measures: {measures}")
    
    for measure in measures:
        # xlsxwriter sheet names limited to 31 chars
        ws = wb.add_worksheet(measure[:31])
        build_measure_sheet(ws, measure, df_scenarios, df_enhanced, df_diagnostics, fmt, 
                          CL_LDF_EXCEL, output_file, df_prior)
        print(f"  Built sheet: {measure[:31]}")
    
    wb.close()
    export_md_data(measures, df_scenarios, df_enhanced, df_diagnostics, exp_md)
    print(f"\nSaved: {output_file}")


if __name__ == "__main__":
    main()
