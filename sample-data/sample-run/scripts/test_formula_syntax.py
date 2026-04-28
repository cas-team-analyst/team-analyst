from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Test'

# Try different external reference syntaxes
ws['A1'] = "='selections\\[Chain Ladder Selections - LDFs.xlsx]Incurred Loss'!B3"
ws['A2'] = "='selections/Chain Ladder Selections - LDFs.xlsx'!B3"
ws['A3'] = "='..\\selections\\Chain Ladder Selections - LDFs.xlsx'!B3"
ws['A4'] = "='../selections/Chain Ladder Selections - LDFs.xlsx'!B3"

wb.save('../test-formulas.xlsx')
print('Test workbook saved with 4 different formula syntaxes')
print('A1 (current syntax):', ws['A1'].value)
print('A2 (no brackets, backslash):', ws['A2'].value)
print('A3 (relative .., backslash):', ws['A3'].value)
print('A4 (relative .., forward slash):', ws['A4'].value)
wb.close()
print('\nNow open test-formulas.xlsx in Excel to see which works')
