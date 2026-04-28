"""Quick verification of Ultimates.xlsx structure"""
import openpyxl

wb = openpyxl.load_workbook("../selections/Ultimates.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{sheet_name} sheet:")
    print(f"  Dimensions: {ws.dimensions}")
    
    # Check headers (row 1)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    print(f"  Column count: {len(headers)}")
    print(f"  First 5 headers: {headers[:5]}")
    print(f"  Last 5 headers: {headers[-5:]}")
    
    # Check data rows
    row_count = ws.max_row - 1  # Exclude header
    print(f"  Data rows: {row_count}")
    
    # Check if second row has data
    if ws.max_row >= 2:
        period = ws.cell(row=2, column=1).value
        age = ws.cell(row=2, column=2).value
        print(f"  First period: {period}, Age: {age}")

print("\n✓ Ultimates.xlsx verification complete!")
