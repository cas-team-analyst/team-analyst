# Verify that formula cells in Complete Analysis.xlsx evaluate to the same values
# as those written in Complete Analysis - Values Only.xlsx.
#
# Uses win32com to open Complete Analysis in Excel (resolving cross-workbook links
# and running a full recalc), then compares key columns against the Values Only file.
#
# Checks per CL sheet  : col C (Actual), col D (CDF), col E (Ultimate)
# Checks per IE sheet  : col D (IE Ultimate)
# Checks per BF sheet  : col C (Initial Expected), col D (CDF), col G (Actual)
# Checks per Selection : col with Selected Ultimate
#
# run-note: Run from the scripts/ directory with Excel closed:
#     cd scripts/
#     python verify-formulas.py

import os
import pathlib
import sys

import time

import win32com.client
from openpyxl import load_workbook

from modules import config
from modules.xl_utils import measure_short_name

COMPLETE   = pathlib.Path(config.BASE_DIR + "Complete Analysis.xlsx").resolve()
VALUES_ONLY = pathlib.Path(config.BASE_DIR + "Complete Analysis - Values Only.xlsx").resolve()

TOL_ABS = 1.0       # absolute tolerance — $1 for dollar amounts
TOL_REL = 1e-5     # relative tolerance — 0.001% for large values

# ── Helpers ──────────────────────────────────────────────────────────────────

def _is_blank(v):
    return v is None or v == "" or v == '""'


def _close(a, b):
    if _is_blank(a) and _is_blank(b):
        return True
    if _is_blank(a) or _is_blank(b):
        return False
    try:
        fa, fb = float(a), float(b)
        return abs(fa - fb) <= TOL_ABS or abs(fa - fb) <= TOL_REL * max(abs(fa), abs(fb), 1)
    except (TypeError, ValueError):
        return str(a) == str(b)


def _xl_sheet_to_dict(xl_ws, max_col):
    """Read Excel COM worksheet into {(row, col): value} for data rows (row >= 2)."""
    out = {}
    last = xl_ws.UsedRange.Rows.Count + 1
    for r in range(2, last + 1):
        col1 = xl_ws.Cells(r, 1).Value
        if col1 in (None, "Total", ""):
            break
        for c in range(1, max_col + 1):
            out[(r, c)] = xl_ws.Cells(r, c).Value
    return out


def _opx_sheet_to_dict(opx_ws, max_col):
    """Read openpyxl worksheet into {(row, col): value} for data rows (row >= 2)."""
    out = {}
    for r in range(2, opx_ws.max_row + 1):
        col1 = opx_ws.cell(r, 1).value
        if col1 in (None, "Total", ""):
            break
        for c in range(1, max_col + 1):
            out[(r, c)] = opx_ws.cell(r, c).value
    return out

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not COMPLETE.exists():
        sys.exit(f"Not found: {COMPLETE}")
    if not VALUES_ONLY.exists():
        sys.exit(f"Not found: {VALUES_ONLY}")

    print(f"Opening {COMPLETE.name} in Excel...")
    xl = win32com.client.DispatchEx("Excel.Application")
    # Wait for app to finish initializing before setting properties.
    for _ in range(10):
        try:
            xl.Visible = False
            xl.DisplayAlerts = False
            break
        except Exception:
            time.sleep(0.5)
    else:
        xl.Visible = False
        xl.DisplayAlerts = False
    try:
        wb_xl = xl.Workbooks.Open(str(COMPLETE), UpdateLinks=True)
        print("  Recalculating (resolves cross-workbook links)...")
        xl.CalculateFullRebuild()

        print(f"Loading {VALUES_ONLY.name} with openpyxl...")
        wb_opx = load_workbook(str(VALUES_ONLY), data_only=False)

        errors = []
        checked = 0

        xl_names = [s.Name for s in wb_xl.Sheets]
        opx_names = wb_opx.sheetnames

        def check_sheet(sname, cols_to_check):
            nonlocal checked
            if sname not in xl_names or sname not in opx_names:
                return
            xl_ws  = wb_xl.Sheets(sname)
            opx_ws = wb_opx[sname]
            max_c  = max(cols_to_check)
            xl_data  = _xl_sheet_to_dict(xl_ws, max_c)
            opx_data = _opx_sheet_to_dict(opx_ws, max_c)
            rows = sorted({r for (r, _) in xl_data} | {r for (r, _) in opx_data})
            for r in rows:
                for c in cols_to_check:
                    v_xl  = xl_data.get((r, c))
                    v_opx = opx_data.get((r, c))
                    if not _close(v_xl, v_opx):
                        errors.append(
                            f"  [{sname}] row {r} col {c}: Excel={v_xl!r}  ValuesOnly={v_opx!r}"
                        )
                    checked += 1

        # Detect sheet names present in workbook
        cl_sheets  = [s for s in opx_names if s.endswith(" CL")]
        bf_sheets  = [s for s in opx_names if s.endswith(" BF")]
        ie_sheets  = [s for s in opx_names if s.endswith(" IE")]
        sel_sheets = [s for s in opx_names if s.endswith("Selection")]

        print("Checking CL sheets (cols C=Actual, D=CDF, E=Ultimate)...")
        for s in cl_sheets:
            check_sheet(s, [3, 4, 5])

        print("Checking IE sheets (col D=IE Ultimate)...")
        for s in ie_sheets:
            check_sheet(s, [4])

        print("Checking BF sheets (cols C=InitExp, D=CDF, G=Actual)...")
        for s in bf_sheets:
            check_sheet(s, [3, 4, 7])

        print("Checking Selection sheets (cols C-onward, spot up to col 10)...")
        for s in sel_sheets:
            check_sheet(s, list(range(3, 11)))

    finally:
        try:
            wb_xl.Close(SaveChanges=False)
        except Exception:
            pass
        xl.Quit()

    print(f"\n{'='*60}")
    print(f"Checked {checked} cells across formula sheets.")
    if errors:
        print(f"FAIL — {len(errors)} mismatches:")
        for e in errors:
            print(e)
        sys.exit(1)
    else:
        print("PASS — all formula values match Values Only.")

if __name__ == "__main__":
    main()
