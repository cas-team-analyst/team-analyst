# Reads projected ultimates, actuary selections, and triangle data to produce the
# Complete Analysis workbooks.
#
# Outputs:
#   Complete Analysis.xlsx            Source Excel files assembled with formulas intact.
#                                     Open in Excel to evaluate any cross-sheet references.
#   Complete Analysis - Values Only.xlsx  Same content as plain computed numbers.
#                                         Safe to read with openpyxl/pandas without requiring
#                                         Excel to re-evaluate formulas first (used by script 7+).
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6-create-complete-analysis.py

import copy  # used by _copy_ws (local full-copy helper)
import os
import pathlib
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules import config
from modules.xl_styles import (
    HEADER_FILL, SUBHEADER_FILL,
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, DATA_FONT,
    THIN_BORDER, style_header,
)
from modules.xl_selections import (
    SKIP_ROW_LABELS as _SKIP_ROW_LABELS,
    SELECTION_LABELS as _SELECTION_LABELS,
    find_selected_values as _find_selected_values,
    find_selected_reasoning as _find_selected_reasoning,
    copy_ws_filtered as _copy_ws_filtered,
)

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
OUTPUT_PATH            = config.OUTPUT

OUTPUT_COMPLETE = OUTPUT_PATH + "Complete Analysis.xlsx"
OUTPUT_VALUES   = OUTPUT_PATH + "Complete Analysis - Values Only.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS  + "Chain Ladder Selections - LDFs.xlsx"
INPUT_TAIL_EXCEL    = config.SELECTIONS  + "Chain Ladder Selections - Tail.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"

# Unpaid = selected ultimate - latest actual of the proxy measure for that period.
UNPAID_PROXY = {
    "Incurred Loss":  "Paid Loss",
    "Paid Loss":      "Paid Loss",
    "Reported Count": "Closed Count",
    "Closed Count":   "Closed Count",
}

# Preferred display order for method columns — discovered dynamically from parquet.
_METHOD_COLS = [
    ("ultimate_cl", "Chain Ladder"),
    ("ultimate_ie", "Initial Expected"),
    ("ultimate_bf", "BF"),
]

_NUM_FMT = "#,##0"
_DEC_FMT = "#,##0.000"

# ── Selection loading ─────────────────────────────────────────────────────────

def load_selections(excel_path):
    """
    Load final selections from Ultimates.xlsx.
    Priority per (measure, period): User Selection > Rules-Based AI Selection.
    Open-Ended AI is intentionally excluded — it is not a trusted final selection source.
    Columns are located by header name — adapts to layout changes automatically.
    Returns dict {(measure, period): float}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        print(f"  Note: {excel_path} not found -- no selections loaded")
        return {}

    wb = load_workbook(excel_path, data_only=True)
    sel_lookup = {}
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Period":
                period_col = cell.column
            elif v == "User Selection":
                user_col = cell.column
            elif v == "Rules-Based AI Selection":
                rb_col = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in sel_lookup:
                continue
            for col_idx in [user_col, rb_col]:
                if col_idx is None:
                    continue
                v = ws.cell(row=row, column=col_idx).value
                if isinstance(v, (int, float)) and not (isinstance(v, float) and v != v):
                    sel_lookup[key] = float(v)
                    total += 1
                    break

    wb.close()
    print(f"  Loaded {total} selections from {excel_path}")
    return sel_lookup


def load_selection_reasoning(excel_path):
    """
    Load reasoning text for each (measure, period) from Ultimates.xlsx.
    Priority matches load_selections: User Selection reasoning first, then RB-AI.
    Returns {(measure, period): str}.  Returns {} when file is absent.
    """
    path = pathlib.Path(excel_path)
    if not path.exists():
        return {}

    wb = load_workbook(excel_path, data_only=True)
    reason_lookup = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_col = user_col = rb_col = user_r_col = rb_r_col = None
        for cell in ws[1]:
            v = cell.value
            if v == "Period":                   period_col = cell.column
            elif v == "User Selection":         user_col   = cell.column
            elif v == "Rules-Based AI Selection": rb_col   = cell.column
            elif v == "User Reasoning":         user_r_col = cell.column
            elif v == "Rules-Based AI Reasoning": rb_r_col = cell.column

        if period_col is None:
            continue

        measure = sheet_name
        for row in range(2, ws.max_row + 1):
            period_val = ws.cell(row=row, column=period_col).value
            if period_val is None:
                continue
            period = str(period_val).strip()
            key = (measure, period)
            if key in reason_lookup:
                continue

            user_v = ws.cell(row=row, column=user_col).value if user_col else None
            rb_v   = ws.cell(row=row, column=rb_col).value   if rb_col   else None
            is_num = lambda v: isinstance(v, (int, float)) and not (isinstance(v, float) and v != v)

            if is_num(user_v) and user_r_col:
                reason_lookup[key] = ws.cell(row=row, column=user_r_col).value or ""
            elif is_num(rb_v) and rb_r_col:
                reason_lookup[key] = ws.cell(row=row, column=rb_r_col).value or ""

    wb.close()
    return reason_lookup


# ── Data loading ──────────────────────────────────────────────────────────────

def load_combined(ultimates_path, sel_lookup):
    """
    Load projected ultimates, apply final selections, compute IBNR and Unpaid.
    Methods (CL/IE/BF) are discovered dynamically from non-empty parquet columns.
    Raises ValueError for any non-Exposure (measure, period) missing from sel_lookup.
    Returns (df, available_methods) where available_methods = [(col, label), ...].
    """
    df = pd.read_parquet(ultimates_path)
    df["period"]  = df["period"].astype(str)
    df["measure"] = df["measure"].astype(str)

    available_methods = [
        (col, label) for col, label in _METHOD_COLS
        if col in df.columns and df[col].notna().any()
    ]

    # Validate: every non-Exposure row must have a selection.
    missing = [
        (row["measure"], row["period"])
        for _, row in df.iterrows()
        if row["measure"] != "Exposure" and (row["measure"], row["period"]) not in sel_lookup
    ]
    if missing:
        lines = "\n  ".join(f"{m} / {p}" for m, p in missing)
        raise ValueError(
            f"{len(missing)} (measure, period) pair(s) have no User Selection or "
            f"Rules-Based AI Selection in Ultimates.xlsx:\n  {lines}\n"
            "Populate Rules-Based AI Selection (run 5b) or add a User Selection."
        )

    actual_lookup = df.set_index(["period", "measure"])["actual"].to_dict()

    def _pick_unpaid(row):
        proxy = UNPAID_PROXY.get(row["measure"])
        if proxy is None:
            return np.nan
        proxy_actual = actual_lookup.get((row["period"], proxy), np.nan)
        if pd.isna(proxy_actual) or pd.isna(row["selected_ultimate"]):
            return np.nan
        return row["selected_ultimate"] - proxy_actual

    df["selected_ultimate"] = df.apply(
        lambda row: sel_lookup.get((row["measure"], row["period"]), np.nan), axis=1
    )
    df["selected_ibnr"]   = df["selected_ultimate"] - df["actual"]
    df["selected_unpaid"] = df.apply(_pick_unpaid, axis=1)

    return df, available_methods


def get_exposure(triangles_path):
    """Latest Exposure value per period from triangles. Returns {} if absent."""
    if not pathlib.Path(triangles_path).exists():
        return {}
    tri = pd.read_parquet(triangles_path)
    exp = tri[tri["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return {}
    exp["period"]  = exp["period"].astype(str)
    exp["age_num"] = pd.to_numeric(exp["age"].astype(str), errors="coerce")
    latest = exp.sort_values("age_num").groupby("period").last()
    return latest["value"].to_dict()


def get_triangles(triangles_path):
    """Load triangle data. Returns empty DataFrame when file is absent."""
    if not pathlib.Path(triangles_path).exists():
        return pd.DataFrame()
    tri = pd.read_parquet(triangles_path)
    tri["period"]  = tri["period"].astype(str)
    tri["measure"] = tri["measure"].astype(str)
    return tri


# ── Cell / layout helpers ─────────────────────────────────────────────────────

def _safe(v):
    """Convert NaN to None so openpyxl writes a blank cell."""
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    return v


def _period_int(v):
    """Display a period as int when it is a whole number, else return as-is."""
    try:
        f = float(v)
        if f == int(f):
            return int(f)
    except (ValueError, TypeError):
        pass
    return v


def _data_cell(cell, value, num_fmt=None):
    """Write a styled data cell with border and alignment."""
    value = _safe(value)
    cell.value     = value
    cell.font      = DATA_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="right" if isinstance(value, (int, float)) else "left",
        vertical="center",
    )
    if num_fmt and value is not None and isinstance(value, (int, float)):
        cell.number_format = num_fmt


def _write_title_and_headers(ws, title, headers, col_width=18):
    """
    Write a title row (row 1) and a styled header row (row 2).
    Returns 3 — the first data row.
    This format matches script 7's read_with_title() convention.
    """
    title_cell = ws.cell(1, 1, title)
    style_header(title_cell, "header")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 20

    for c, text in enumerate(headers, 1):
        cell = ws.cell(2, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A3"
    return 3


def _write_headers(ws, headers, col_width=18):
    """
    Write a single styled header row (row 1), no title row.
    Returns 2 — the first data row.
    Matches script 7's read_no_title() convention ("Sel - " sheets).
    """
    for c, text in enumerate(headers, 1):
        cell = ws.cell(1, c, text)
        style_header(cell, "subheader")
        ws.column_dimensions[get_column_letter(c)].width = col_width

    ws.freeze_panes = "A2"
    return 2


# ── Generated sheet writers ───────────────────────────────────────────────────

def write_selected_ultimates_sheet(ws, measure, combined, available_methods,
                                    reason_lookup=None):
    """
    Write one 'Sel - {measure}' sheet.
    Columns: Accident Period, Current Age, Actual, [method cols],
             Selected Ultimate, IBNR, Unpaid[, Selected Reasoning].
    Reasoning column appended when reason_lookup is provided.
    No title row -- matches script 7's read_no_title() convention for 'Sel - ' sheets.
    """
    df_m = combined[combined["measure"] == measure].copy()
    if df_m.empty:
        return

    active_methods = [
        (col, label) for col, label in available_methods
        if col in df_m.columns and df_m[col].notna().any()
    ]

    has_reasoning = bool(reason_lookup)
    headers = ["Accident Period", "Current Age", "Actual"]
    headers += [label for _, label in active_methods]
    headers += ["Selected Ultimate", "IBNR", "Unpaid"]
    if has_reasoning:
        headers += ["Selected Reasoning"]

    data_row = _write_headers(ws, headers)
    reason_col = len(headers) if has_reasoning else None
    if reason_col:
        ws.column_dimensions[get_column_letter(reason_col)].width = 50

    for r, (_, row) in enumerate(df_m.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), _period_int(row["period"]))
        _data_cell(ws.cell(r, 2), _safe(row.get("current_age")))
        _data_cell(ws.cell(r, 3), _safe(row["actual"]), _NUM_FMT)

        c = 4
        for col, _ in active_methods:
            _data_cell(ws.cell(r, c), _safe(row.get(col)), _NUM_FMT)
            c += 1

        _data_cell(ws.cell(r, c),     _safe(row["selected_ultimate"]), _NUM_FMT)
        _data_cell(ws.cell(r, c + 1), _safe(row["selected_ibnr"]),     _NUM_FMT)
        _data_cell(ws.cell(r, c + 2), _safe(row["selected_unpaid"]),   _NUM_FMT)

        if has_reasoning:
            period = str(row["period"])
            reason = (reason_lookup or {}).get((measure, period), "")
            cell = ws.cell(r, reason_col, reason)
            cell.font      = DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")


def write_diagnostics_sheet(ws, combined, exp_map):
    """
    Write the 'Diagnostics' sheet: Ultimate Severity, Loss Rate, Frequency.
    Loss Rate and Frequency columns are omitted when no exposure data is present.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    if inc.empty:
        return

    has_exposure = bool(exp_map)
    headers = ["Accident Period", "Ultimate Severity"]
    if has_exposure:
        headers += ["Ultimate Loss Rate", "Ultimate Frequency"]

    data_row = _write_title_and_headers(ws, "Post-Method Series Diagnostics", headers)

    periods = sorted(
        inc.index,
        key=lambda x: (int(x) if str(x).isdigit() else str(x)),
    )
    for r, p in enumerate(periods, start=data_row):
        ult_loss   = inc.loc[p, "selected_ultimate"] if p in inc.index else np.nan
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        def _div(a, b):
            if pd.notna(a) and pd.notna(b) and b != 0:
                return a / b
            return None

        _data_cell(ws.cell(r, 1), _period_int(p))
        _data_cell(ws.cell(r, 2), _div(ult_loss, ult_counts), _DEC_FMT)
        if has_exposure:
            _data_cell(ws.cell(r, 3), _div(ult_loss, exp),    _DEC_FMT)
            _data_cell(ws.cell(r, 4), _div(ult_counts, exp),  _DEC_FMT)


def build_post_method_triangle_data(triangles_df, combined):
    """
    Build post-method triangle DataFrames as a list of (sheet_name, index_name, df).
    Each df has period ints as index and age ints as columns.
    """
    if triangles_df.empty:
        return []

    sel_lookup = combined.set_index(["period", "measure"])["selected_ultimate"].to_dict()

    def _to_int_series(s):
        try:
            return s.apply(lambda x: int(float(str(x))))
        except (ValueError, TypeError):
            return s.astype(str)

    def x_to_ult(measure, label):
        sub = triangles_df[triangles_df["measure"] == measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        return label, result

    sheets = []
    for measure, label in [
        ("Incurred Loss",  "Incurred-to-Ult"),
        ("Paid Loss",      "Paid-to-Ult"),
        ("Reported Count", "Reported-to-Ult"),
        ("Closed Count",   "Closed-to-Ult"),
    ]:
        result = x_to_ult(measure, label)
        if result is not None:
            sheets.append(result)

    def avg_triangle(sub_measure, sheet_name, diff_fn):
        sub = triangles_df[triangles_df["measure"] == sub_measure].copy()
        if sub.empty:
            return None
        sub["period_int"] = _to_int_series(sub["period"])
        sub["age_int"]    = _to_int_series(sub["age"])
        pivot = sub.pivot_table(
            index="period_int", columns="age_int", values="value", aggfunc="first", observed=True
        )
        pivot.columns.name = None
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), sub_measure), np.nan)
            if pd.notna(sel):
                result.loc[period] = diff_fn(sel, result.loc[period])
        return sheet_name, result

    inc_sub  = triangles_df[triangles_df["measure"] == "Incurred Loss"]
    paid_sub = triangles_df[triangles_df["measure"] == "Paid Loss"]

    if not inc_sub.empty:
        r = avg_triangle("Incurred Loss", "Average IBNR", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    if not paid_sub.empty:
        r = avg_triangle("Paid Loss", "Average Unpaid", lambda sel, row: sel - row)
        if r is not None:
            sheets.append(r)

    return sheets


def write_triangle_sheet(ws, sheet_name, df):
    """
    Write a triangle DataFrame to a worksheet.
    Title row + header row format matches script 7's read_with_title() convention.
    """
    headers = ["Period"] + [str(c) for c in df.columns]
    data_row = _write_title_and_headers(ws, sheet_name, headers, col_width=14)

    for r, (idx, row) in enumerate(df.iterrows(), start=data_row):
        _data_cell(ws.cell(r, 1), idx)
        for c, val in enumerate(row, start=2):
            _data_cell(ws.cell(r, c), _safe(val), _DEC_FMT)


# ── Filtered copy of source selection workbooks ───────────────────────────────
# _find_selected_values and _copy_ws_filtered imported from modules.xl_selections.


def _strip_formulas(ws):
    """Replace formula strings with None so downstream readers see blank, not a formula string."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None


def _fill_cl_main_values(ws, measure, df2, df4):
    """
    Replace formula strings in a CL main-measure sheet with Python-computed values.
    Only called for main measure sheets (e.g. Incurred Loss) -- Diag-* and CV-&-Slopes
    sheets have no formula cells.

    ATA section: each period/interval cell replaced from df2['ldf'].
    Averages section: each display-name/interval cell replaced from df4.

    Sheet structure in gen_wb after _copy_ws_filtered:
      Loss triangle title -> header -> data rows -> blank
      "Age-to-Age Factors" title -> "" header -> ATA data rows (formulas) -> blank
      "Averages" title -> "Metric" header -> avg data rows (formulas) -> blank
      "LDF Selections" title -> header -> "Selected" row
    """
    df_m = df2[df2['measure'].astype(str) == measure].copy()

    ata_lookup = {
        (str(r['period']), str(r['interval'])): r['ldf']
        for _, r in df_m[df_m['ldf'].notna()].iterrows()
    }

    avg_lookup = {}
    if df4 is not None and not df4.empty:
        df4_m = df4[df4['measure'].astype(str) == measure].copy()
        avg_data_cols = [c for c in df4_m.columns
                         if c not in ('measure', 'interval')
                         and not c.startswith('cv_')
                         and not c.startswith('slope_')]
        for _, r in df4_m.iterrows():
            intv = str(r['interval'])
            for col in avg_data_cols:
                display = col.replace('avg_exclude_high_low', 'exclude_high_low')
                avg_lookup[(display, intv)] = r[col]

    section = None
    col_headers = {}

    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Age-to-Age Factors":
            section = "ata_pre_header"
            continue
        if col1 == "Averages":
            section = "avg_pre_header"
            continue
        if col1 == "LDF Selections":
            break

        if section == "ata_pre_header":
            col_headers = {c.column: str(c.value) for c in row_cells[1:] if c.value not in (None, "")}
            section = "ata"
            continue

        if section == "avg_pre_header":
            col_headers = {c.column: str(c.value) for c in row_cells[1:] if c.value not in (None, "")}
            section = "avg"
            continue

        if section == "ata":
            if col1 is None or col1 == "":
                section = None
                continue
            period = str(col1)
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_headers.get(cell.column)
                    cell.value = ata_lookup.get((period, intv)) if intv else None

        elif section == "avg":
            if col1 is None or col1 == "":
                section = None
                continue
            display = str(col1)
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_headers.get(cell.column)
                    cell.value = avg_lookup.get((display, intv)) if intv else None


def _fill_tail_values(ws, measure, df2):
    """
    Replace formula strings in a Tail sheet with Python-computed values.
    Only the "Average" and "CV" rows contain formulas; they summarise the observed
    ATA factors in the column above.  Header row col1 = "Accident Year" identifies
    the age columns so we can map them to df2 intervals.
    """
    df_m = df2[df2['measure'].astype(str) == measure].copy()

    # Map "to" age string -> interval string  e.g. "23" -> "11-23"
    to_age_map = {}
    for intv in df_m['interval'].dropna().unique():
        parts = str(intv).split('-')
        if len(parts) == 2:
            to_age_map[parts[1]] = str(intv)

    def _mean(intv):
        vals = df_m[df_m['interval'].astype(str) == intv]['ldf'].dropna()
        return vals.mean() if not vals.empty else None

    def _cv(intv):
        vals = df_m[df_m['interval'].astype(str) == intv]['ldf'].dropna()
        if vals.empty:
            return None
        m = vals.mean()
        return (vals.std() / m) if m and m != 0 else None

    col_to_intv = {}
    header_found = False

    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None

        if col1 == "Accident Year":
            col_to_intv = {c.column: to_age_map.get(str(c.value))
                           for c in row_cells[1:] if c.value is not None}
            header_found = True
            continue

        if not header_found:
            continue

        if col1 == "Average":
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_to_intv.get(cell.column)
                    cell.value = _mean(intv) if intv else None

        elif col1 == "CV":
            for cell in row_cells[1:]:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    intv = col_to_intv.get(cell.column)
                    cell.value = _cv(intv) if intv else None


# ── Notes sheet ───────────────────────────────────────────────────────────────

_SHEET_DESCS = {
    "CL - ":   "Chain Ladder LDF triangle, averages, and selected LDFs",
    "Tail - ": "Tail factor analysis and selected tail",
    "Sel - ":  "Final selected ultimates, IBNR, and Unpaid",
    "Diagnostics": "Post-method diagnostics: severity, loss rate, frequency",
    "Incurred-to-Ult":  "Incurred-to-Ultimate development ratios",
    "Paid-to-Ult":      "Paid-to-Ultimate development ratios",
    "Reported-to-Ult":  "Reported-to-Ultimate development ratios",
    "Closed-to-Ult":    "Closed-to-Ultimate development ratios",
    "Average IBNR":  "Average IBNR by development age",
    "Average Unpaid":"Average Unpaid by development age",
}


def _sheet_desc(name):
    for key, desc in _SHEET_DESCS.items():
        if name == key or name.startswith(key):
            return desc
    return "Analysis results"


def write_notes_sheet(ws, sheet_list):
    """Write Notes sheet with metadata header and table of contents."""
    r = 1

    title_cell = ws.cell(r, 1, "Reserve Analysis - Complete Analysis")
    style_header(title_cell, "header")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 24
    r += 1

    meta = ws.cell(r, 1, "Workbook Information")
    style_header(meta, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    for label, value in [
        ("Created:",     datetime.now().strftime("%B %d, %Y %I:%M %p")),
        ("Description:", "Complete actuarial reserve analysis combining selections, ultimates, and diagnostics"),
    ]:
        lbl = ws.cell(r, 1, label)
        lbl.font = LABEL_FONT; lbl.border = THIN_BORDER
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        val = ws.cell(r, 2, value)
        val.font = DATA_FONT; val.border = THIN_BORDER
        val.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1

    r += 1
    toc_hdr = ws.cell(r, 1, "Table of Contents")
    style_header(toc_hdr, "section")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    for col_num, text, width in [(1, "Sheet Name", 35), (2, "Description", 60)]:
        cell = ws.cell(r, col_num, text)
        cell.font = SUBHEADER_FONT; cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = width
    ws.column_dimensions["C"].width = 15
    r += 1

    for name, desc in sheet_list:
        nc = ws.cell(r, 1, name)
        nc.font = DATA_FONT; nc.border = THIN_BORDER
        nc.alignment = Alignment(horizontal="left", vertical="center")
        dc = ws.cell(r, 2, desc)
        dc.font = DATA_FONT; dc.border = THIN_BORDER
        dc.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    ws.freeze_panes = "A8"


# ── Workbook assembly ─────────────────────────────────────────────────────────

def _copy_ws(ws_src, ws_dst):
    """Copy all cells and styles from source to destination worksheet."""
    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst[cell.coordinate]
            dst.value = cell.value
            if cell.has_style:
                dst.font       = copy.copy(cell.font)
                dst.border     = copy.copy(cell.border)
                dst.fill       = copy.copy(cell.fill)
                dst.number_format = cell.number_format
                dst.protection = copy.copy(cell.protection)
                dst.alignment  = copy.copy(cell.alignment)

    for col in ws_src.column_dimensions:
        ws_dst.column_dimensions[col].width = ws_src.column_dimensions[col].width
    for row_num in ws_src.row_dimensions:
        ws_dst.row_dimensions[row_num].height = ws_src.row_dimensions[row_num].height
    for rng in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(rng))
    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes


def assemble_workbook(output_path, generated_wb):
    """
    Build master workbook from the pre-built generated sheets.
    All content is already computed values — no formula copying needed.
    """
    master = Workbook()
    master.remove(master.active)
    sheet_list = []

    for sname in generated_wb.sheetnames:
        ws_dst = master.create_sheet(title=sname)
        _copy_ws(generated_wb[sname], ws_dst)
        sheet_list.append((sname, _sheet_desc(sname)))

    notes_ws = master.create_sheet(title="Notes", index=0)
    write_notes_sheet(notes_ws, sheet_list)

    os.makedirs(pathlib.Path(output_path).parent, exist_ok=True)
    master.save(output_path)
    print(f"  Saved -> {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_PATH, exist_ok=True)

    print("Loading selections...")
    sel_lookup    = load_selections(INPUT_SELECTIONS_EXCEL)
    reason_lookup = load_selection_reasoning(INPUT_SELECTIONS_EXCEL)

    print("Loading ultimates...")
    combined, available_methods = load_combined(INPUT_ULTIMATES, sel_lookup)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]
    print(f"  Measures: {measures}")
    print(f"  Methods:  {[label for _, label in available_methods]}")

    print("Loading triangles...")
    exp_map      = get_exposure(INPUT_TRIANGLES)
    triangles_df = get_triangles(INPUT_TRIANGLES)
    print(f"  Exposure periods: {len(exp_map)}  Triangle rows: {len(triangles_df)}")

    # Build generated workbook -- all content computed from parquet and selection files.
    # No Excel formulas written; both output files use this same content.
    gen_wb = Workbook()
    gen_wb.remove(gen_wb.active)

    print("Copying CL LDF sheets (full triangle + selected row)...")
    if pathlib.Path(INPUT_CL_EXCEL).exists():
        # data_only=False preserves formula strings (averages, LDF calcs).
        # openpyxl never caches formula results on save, so data_only=True returns None.
        # The "Selected" row is always plain numeric values written by the update scripts.
        wb_cl_vals = load_workbook(INPUT_CL_EXCEL, data_only=True)   # for selected-row extraction
        wb_cl_form = load_workbook(INPUT_CL_EXCEL, data_only=False)  # for formula copy
        for sname in wb_cl_vals.sheetnames:
            sel_vals   = _find_selected_values(wb_cl_vals[sname])
            sel_reason = _find_selected_reasoning(wb_cl_vals[sname])
            ws = gen_wb.create_sheet(title=f"CL - {sname}"[:31])
            _copy_ws_filtered(wb_cl_form[sname], ws, sel_vals, sel_reason)
            print(f"  CL - {sname}")
        wb_cl_vals.close()
        wb_cl_form.close()
    else:
        print(f"  Note: {INPUT_CL_EXCEL} not found -- skipped")

    print("Copying Tail sheets (full analysis + selected row)...")
    if pathlib.Path(INPUT_TAIL_EXCEL).exists():
        wb_tail_vals = load_workbook(INPUT_TAIL_EXCEL, data_only=True)
        wb_tail_form = load_workbook(INPUT_TAIL_EXCEL, data_only=False)
        for sname in wb_tail_vals.sheetnames:
            sel_vals   = _find_selected_values(wb_tail_vals[sname])
            sel_reason = _find_selected_reasoning(wb_tail_vals[sname])
            ws = gen_wb.create_sheet(title=f"Tail - {sname}"[:31])
            _copy_ws_filtered(wb_tail_form[sname], ws, sel_vals, sel_reason)
            print(f"  Tail - {sname}")
        wb_tail_vals.close()
        wb_tail_form.close()
    else:
        print(f"  Note: {INPUT_TAIL_EXCEL} not found -- skipped")

    print("Building selected ultimates sheets...")
    for measure in measures:
        ws = gen_wb.create_sheet(title=f"Sel - {measure}"[:31])
        write_selected_ultimates_sheet(ws, measure, combined, available_methods,
                                        reason_lookup=reason_lookup)
        print(f"  Sel - {measure}")

    print("Building diagnostics sheet...")
    ws_diag = gen_wb.create_sheet(title="Diagnostics")
    write_diagnostics_sheet(ws_diag, combined, exp_map)

    print("Building post-method triangle sheets...")
    for sheet_name, df in build_post_method_triangle_data(triangles_df, combined):
        ws_t = gen_wb.create_sheet(title=sheet_name[:31])
        write_triangle_sheet(ws_t, sheet_name, df)
        print(f"  {sheet_name}")

    print("\nAssembling Complete Analysis.xlsx...")
    assemble_workbook(OUTPUT_COMPLETE, gen_wb)

    # Values Only: replace formula strings in CL/Tail sheets with Python-computed values
    # so downstream readers (script 7, pandas) see real numbers, not "=..." strings.
    # Main measure sheets (Incurred Loss etc.) have ATA and average formula cells -- we
    # fill those from df2/df4 parquet.  Diag-* and CV-&-Slopes sheets have no formulas.
    print("Computing CL/Tail formula values for Values Only...")
    df2_enh = pd.read_parquet(INPUT_CL_ENHANCED) if pathlib.Path(INPUT_CL_ENHANCED).exists() else pd.DataFrame()
    df4_avg = pd.read_parquet(INPUT_LDF_AVERAGES) if pathlib.Path(INPUT_LDF_AVERAGES).exists() else None
    measures_set = set(measures)
    for sname in gen_wb.sheetnames:
        if sname.startswith("CL - "):
            measure_name = sname[5:]
            if measure_name in measures_set and not df2_enh.empty:
                _fill_cl_main_values(gen_wb[sname], measure_name, df2_enh, df4_avg)
            else:
                _strip_formulas(gen_wb[sname])
        elif sname.startswith("Tail - "):
            measure_name = sname[7:]
            if measure_name in measures_set and not df2_enh.empty:
                _fill_tail_values(gen_wb[sname], measure_name, df2_enh)
            else:
                _strip_formulas(gen_wb[sname])

    print("Assembling Complete Analysis - Values Only.xlsx...")
    assemble_workbook(OUTPUT_VALUES, gen_wb)

    print("\n=== IBNR Summary ===")
    for m in measures:
        sub = combined[combined["measure"] == m]
        if sub.empty or sub["selected_ultimate"].isna().all():
            continue
        parts = [f"Actual={sub['actual'].sum():,.0f}"]
        for col, label in available_methods:
            if col in sub.columns and sub[col].notna().any():
                parts.append(f"{label}={sub[col].sum():,.0f}")
        parts += [
            f"Selected={sub['selected_ultimate'].sum():,.0f}",
            f"IBNR={sub['selected_ibnr'].sum():,.0f}",
        ]
        print(f"  {m}: " + "  ".join(parts))


if __name__ == "__main__":
    print("=== Step 6: Creating Complete Analysis workbooks ===")
    main()
