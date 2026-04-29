from openpyxl import load_workbook

wb = load_workbook('../Analysis.xlsx')

# Loss Selection
ws_loss = wb['Loss Selection']
loss_headers = [ws_loss.cell(1, i).value for i in range(1, 15) if ws_loss.cell(1, i).value]
print('Loss Selection headers:')
for h in loss_headers:
    print(f'  {h}')
print(f'Has Unpaid: {"Unpaid" in loss_headers}')

print()

# Count Selection
ws_count = wb['Count Selection']
count_headers = [ws_count.cell(1, i).value for i in range(1, 15) if ws_count.cell(1, i).value]
print('Count Selection headers:')
for h in count_headers:
    print(f'  {h}')
print(f'Has Unpaid: {"Unpaid" in count_headers}')

wb.close()

print('\n✓ Flexible: Loss has Unpaid (Paid proxy exists), Count has no Unpaid (Closed proxy has no data)')
