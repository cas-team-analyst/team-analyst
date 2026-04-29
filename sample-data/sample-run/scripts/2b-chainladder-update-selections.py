# Refreshes only the selections section of the Excel workbook based on saved selections from a
# JSON file. Allows you to update the displayed selections without rebuilding the entire workbook,
# which is useful when you need to revert changes or apply selections from another source.

"""
goal: Update ONLY the selections section of Chain Ladder Selections.xlsx from chainladder.json.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 2b-chainladder-update-selections.py
"""

import json
import os
import glob
import openpyxl
from openpyxl.styles import Alignment, Font

from modules import config
from modules.xl_styles import SELECTION_FILL, AI_FILL, DATA_FONT, THIN_BORDER

# Paths from modules/config.py — override here if needed:
METHOD_ID = "chainladder"
SELECTIONS_PATTERN    = config.SELECTIONS + f"{METHOD_ID}-ai-rules-based-*.json"
AI_SELECTIONS_PATTERN = config.SELECTIONS + f"{METHOD_ID}-ai-open-ended-*.json"
EXCEL_FILE            = config.SELECTIONS + "Chain Ladder Selections - LDFs.xlsx"


def find_selections_section(ws):
    """Find the row where interval headers and selection rows are located."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Rules-Based AI Selection":
                selection_row = cell.row
                reasoning_row = selection_row + 1
                # Header row is the row just before the Rules-Based AI Selection row
                # (or before Prior Selection if it exists)
                header_row = selection_row - 1
                # Check if there are Prior Selection rows above
                check_label = ws.cell(row=header_row, column=1).value
                while check_label and check_label.startswith("Prior"):
                    header_row -= 1
                    check_label = ws.cell(row=header_row, column=1).value
                return header_row, selection_row, reasoning_row
    return None, None, None


def get_interval_columns(ws, header_row):
    """Build a dict mapping interval string -> column index from the header row."""
    interval_to_col = {}
    for cell in ws[header_row]:
        if cell.value:
            interval_to_col[str(cell.value).strip()] = cell.column
    return interval_to_col


def find_ai_section(ws):
    """Find the row numbers for the Open-Ended AI Selection and Open-Ended AI Reasoning rows."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Open-Ended AI Selection":
                ai_selection_row = cell.row
                ai_reasoning_row = ai_selection_row + 1
                return ai_selection_row, ai_reasoning_row
    return None, None


def find_user_section(ws):
    """Find the row numbers for the User Selection and User Reasoning rows."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "User Selection":
                user_selection_row = cell.row
                user_reasoning_row = user_selection_row + 1
                return user_selection_row, user_reasoning_row
    return None, None


def has_existing_selections(ws):
    """Return True if the Rules-Based AI Selection row already has values."""
    _, selection_row, _ = find_selections_section(ws)
    if selection_row is None:
        return False
    for cell in ws[selection_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def has_existing_ai_selections(ws):
    """Return True if the Open-Ended AI Selection row already has values."""
    ai_row, _ = find_ai_section(ws)
    if ai_row is None:
        return False
    for cell in ws[ai_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def has_existing_user_selections(ws):
    """Return True if the User Selection row already has values."""
    user_row, _ = find_user_section(ws)
    if user_row is None:
        return False
    for cell in ws[user_row]:
        if cell.column > 1 and cell.value not in (None, ""):
            return True
    return False


def update_sheet(ws, measure_selections):
    """Update the selections section of a single sheet."""
    header_row, selection_row, reasoning_row = find_selections_section(ws)
    if header_row is None:
        print(f"  WARNING: Could not find selections section in sheet '{ws.title}'")
        return

    interval_to_col = get_interval_columns(ws, header_row)

    for sel in measure_selections:
        interval = sel["interval"]
        if interval not in interval_to_col:
            print(f"  WARNING: Interval '{interval}' not found in sheet '{ws.title}'")
            continue

        col = interval_to_col[interval]

        # Cutoff markers have reasoning but no selection value — write reasoning only
        if "selection" not in sel:
            reason_cell = ws.cell(row=reasoning_row, column=col)
            reason_cell.value = "[CUTOFF] " + sel.get("reasoning", "")
            reason_cell.fill = SELECTION_FILL
            reason_cell.font = Font(size=8, italic=True)
            reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            reason_cell.border = THIN_BORDER
            continue

        # Write selection value (this is rules-based now)
        sel_cell = ws.cell(row=selection_row, column=col)
        sel_cell.value = sel["selection"]
        sel_cell.fill = SELECTION_FILL
        sel_cell.font = DATA_FONT
        sel_cell.alignment = Alignment(horizontal="right")
        sel_cell.border = THIN_BORDER
        sel_cell.number_format = "0.0000"

        # Write reasoning
        reason_cell = ws.cell(row=reasoning_row, column=col)
        reason_cell.value = sel.get("reasoning", "")
        reason_cell.fill = SELECTION_FILL
        reason_cell.font = Font(size=8, italic=True)
        reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        reason_cell.border = THIN_BORDER

    # Set row height for reasoning row
    ws.row_dimensions[reasoning_row].height = 60

    print(f"  Updated {len(measure_selections)} rules-based selections in '{ws.title}'")


def update_ai_sheet(ws, measure_selections, interval_to_col):
    """Write AI selections and reasoning into the AI Selection / AI Reasoning rows."""
    ai_selection_row, ai_reasoning_row = find_ai_section(ws)
    if ai_selection_row is None:
        print(f"  WARNING: Could not find 'AI Selection' row in sheet '{ws.title}'")
        return

    for sel in measure_selections:
        interval = sel["interval"]
        if interval not in interval_to_col:
            print(f"  WARNING: Interval '{interval}' not found in sheet '{ws.title}'")
            continue

        col = interval_to_col[interval]

        # Cutoff markers have reasoning but no selection value — write reasoning only
        if "selection" not in sel:
            reason_cell = ws.cell(row=ai_reasoning_row, column=col)
            reason_cell.value = "[CUTOFF] " + sel.get("reasoning", "")
            reason_cell.fill = AI_FILL
            reason_cell.font = Font(size=8, italic=True)
            reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            reason_cell.border = THIN_BORDER
            continue

        sel_cell = ws.cell(row=ai_selection_row, column=col)
        sel_cell.value = sel["selection"]
        sel_cell.fill = AI_FILL
        sel_cell.font = DATA_FONT
        sel_cell.alignment = Alignment(horizontal="right")
        sel_cell.border = THIN_BORDER
        sel_cell.number_format = "0.0000"

        reason_cell = ws.cell(row=ai_reasoning_row, column=col)
        reason_cell.value = sel.get("reasoning", "")
        reason_cell.fill = AI_FILL
        reason_cell.font = Font(size=8, italic=True)
        reason_cell.alignment = Alignment(horizontal="left", wrap_text=True)
        reason_cell.border = THIN_BORDER

    ws.row_dimensions[ai_reasoning_row].height = 60
    print(f"  Updated {len(measure_selections)} open-ended AI selections in '{ws.title}'")


def load_per_measure_json_files(pattern, selection_type):
    """Load and combine all per-measure JSON files matching the pattern.
    
    Extracts measure name from filename (e.g., 'chainladder-ai-rules-based-paid_loss.json' -> 'Paid Loss')
    and adds it to each selection object.
    """
    combined = []
    files = glob.glob(pattern)
    
    if not files:
        return []
    
    for filepath in sorted(files):
        try:
            # Extract measure from filename: chainladder-ai-rules-based-paid_loss.json -> paid_loss
            filename = os.path.basename(filepath)
            # Remove extension
            name_no_ext = os.path.splitext(filename)[0]
            # Extract measure part (after last hyphen)
            parts = name_no_ext.split('-')
            measure_slug = parts[-1] if parts else ""
            # Convert slug to display name: paid_loss -> Paid Loss
            measure = measure_slug.replace('_', ' ').title()
            
            with open(filepath, 'r') as f:
                data = json.load(f)
                # If it's a single object, wrap it in a list
                if isinstance(data, dict):
                    data = [data]
                # Add measure field to each selection
                for sel in data:
                    sel['measure'] = measure
                combined.extend(data)
        except Exception as e:
            print(f"  WARNING: Failed to load {filepath}: {e}")
            continue
    
    print(f"Loaded {len(combined)} {selection_type} selections from {len(files)} file(s)")
    return combined


def main():
    """Update Chain Ladder Selections Excel file with selections from per-measure JSON files."""
    # Load selections from per-measure JSON files
    selections = load_per_measure_json_files(SELECTIONS_PATTERN, "rules-based")
    
    if not selections:
        print(f"No selections found matching pattern: {SELECTIONS_PATTERN}")
        print("Skipping update - no selections to apply.")
        return

    # Load AI selections if available
    ai_selections = load_per_measure_json_files(AI_SELECTIONS_PATTERN, "open-ended AI")
    if not ai_selections:
        print(f"No open-ended AI selections found matching pattern: {AI_SELECTIONS_PATTERN} (optional)")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Abort if any sheet has user selections to avoid overwriting manual work
    sheets_with_user = [s for s in wb.sheetnames if has_existing_user_selections(wb[s])]
    if sheets_with_user:
        raise ValueError(
            f"Existing user selections found in sheet(s): {sheets_with_user}\n"
            "Clear user selections manually before re-running to avoid overwriting manual edits."
        )

    # Abort if any sheet already has manual selections to avoid overwriting actuary work
    sheets_with_selections = [s for s in wb.sheetnames if has_existing_selections(wb[s])]
    if sheets_with_selections:
        raise ValueError(
            f"Existing selections found in sheet(s): {sheets_with_selections}\n"
            "Clear selections manually before re-running to avoid overwriting actuary edits."
        )

    # Abort if AI rows already have values
    sheets_with_ai = [s for s in wb.sheetnames if has_existing_ai_selections(wb[s])]
    if sheets_with_ai:
        raise ValueError(
            f"Existing AI selections found in sheet(s): {sheets_with_ai}\n"
            "Clear AI selections manually before re-running."
        )

    # Group selections by measure
    by_measure = {}
    for sel in selections:
        m = sel["measure"]
        by_measure.setdefault(m, []).append(sel)

    by_measure_ai = {}
    for sel in ai_selections:
        m = sel["measure"]
        by_measure_ai.setdefault(m, []).append(sel)

    for sheet_name in wb.sheetnames:
        # Match sheet name to measure (sheet name may be truncated to 31 chars)
        matched = None
        for measure in by_measure:
            if measure[:31] == sheet_name[:31] or sheet_name[:31] in measure[:31]:
                matched = measure
                break
        if matched:
            ws = wb[sheet_name]
            # update_sheet finds header/selection rows internally
            update_sheet(ws, by_measure[matched])
            # update_ai_sheet needs interval_to_col — derive from find_selections_section
            header_row, _, _ = find_selections_section(ws)
            if header_row and matched in by_measure_ai:
                interval_to_col = get_interval_columns(ws, header_row)
                update_ai_sheet(ws, by_measure_ai[matched], interval_to_col)
        else:
            print(f"No selections found for sheet '{sheet_name}' - skipping")

    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
