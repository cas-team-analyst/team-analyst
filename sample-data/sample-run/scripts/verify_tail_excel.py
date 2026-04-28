"""Quick verification that tail excel file has formulas with cached values."""
import openpyxl

# Open the generated file - data_only=False to see formulas
wb = openpyxl.load_workbook('../selections/Chain Ladder Selections - Tail.xlsx', data_only=False)
ws = wb['Incurred Loss']

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print()

# Find the "Average" row by scanning column A
avg_row = None
cv_row = None
for row in range(1, min(30, ws.max_row + 1)):
    cell_val = ws.cell(row, 1).value
    if cell_val == "Average":
        avg_row = row
    elif cell_val == "CV":
        cv_row = row

print(f"Found 'Average' at row: {avg_row}")
print(f"Found 'CV' at row: {cv_row}")
print()

if avg_row:
    # Check the Average row (should have formula with cached value)
    avg_cell = ws.cell(avg_row, 2)  # Column B
    print(f"Average cell B{avg_row}:")
    print(f"  Value: {avg_cell.value}")
    print(f"  Type: {type(avg_cell.value).__name__}")
    
    # Check if it's a formula
    if isinstance(avg_cell.value, str) and avg_cell.value.startswith('='):
        print(f"  ✓ Formula present: {avg_cell.value[:50]}...")
    else:
        print(f"  ✗ No formula found")
    
    print()

if cv_row:
    # Check the CV row
    cv_cell = ws.cell(cv_row, 2)  # Column B
    print(f"CV cell B{cv_row}:")
    print(f"  Value: {cv_cell.value}")
    print(f"  Type: {type(cv_cell.value).__name__}")
    
    if isinstance(cv_cell.value, str) and cv_cell.value.startswith('='):
        print(f"  ✓ Formula present: {cv_cell.value[:60]}...")
    else:
        print(f"  ✗ No formula found")
    
    print()

print("\nChecking all rows of column A for 'Average' and 'CV':")
avg_row = None
cv_row = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val:
        if val == 'Average':
            avg_row = row
            print(f"  Row {row}: {val}")
        elif val == 'CV':
            cv_row = row
            print(f"  Row {row}: {val}")
        elif 'selection' in str(val).lower():
            print(f"  Row {row}: {val}")

if avg_row:
    print(f"\nChecking Average row {avg_row}:")
    for col in range(1, min(6, ws.max_column + 1)):
        cell = ws.cell(avg_row, col)
        col_letter_str = chr(64 + col)  # A=65
        print(f"  {col_letter_str}{avg_row}: {cell.value} (type: {type(cell.value).__name__})")

if cv_row:
    print(f"\nChecking CV row {cv_row}:")
    for col in range(1, min(6, ws.max_column + 1)):
        cell = ws.cell(cv_row, col)
        col_letter_str = chr(64 + col)  # A=65
        val_str = str(cell.value)[:60] if cell.value else 'None'
        print(f"  {col_letter_str}{cv_row}: {val_str} (type: {type(cell.value).__name__})")


