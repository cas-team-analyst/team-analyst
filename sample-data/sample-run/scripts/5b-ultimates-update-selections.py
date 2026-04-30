# Refreshes only the selections section of the Ultimates Excel workbook based on saved selections 
# from the ultimates-ai-rules-based-loss.json, ultimates-ai-rules-based-count.json, 
# ultimates-ai-open-ended-loss.json, and ultimates-ai-open-ended-count.json files. 
# Allows you to update the displayed selections without rebuilding the entire workbook, 
# which is useful when you need to revert changes or apply selections from another source.

"""
goal: Update ONLY the selections section of Ultimates.xlsx from ultimates-ai-rules-based-*.json 
      and ultimates-ai-open-ended-*.json files.
      Re-run this script any time selections change without needing to rebuild the full Excel.

run-note: When copied to a project, run from the scripts/ directory. Close the Excel file before running.
    cd scripts/
    python 5b-ultimates-update-selections.py
"""

import json
import pathlib
from openpyxl import load_workbook

from modules import config
from modules.xl_utils import build_column_map

# Paths from modules/config.py — override here if needed:
RULES_BASED_LOSS_FILE = config.SELECTIONS + "ultimates-ai-rules-based-loss.json"
RULES_BASED_COUNT_FILE = config.SELECTIONS + "ultimates-ai-rules-based-count.json"
OPEN_ENDED_LOSS_FILE = config.SELECTIONS + "ultimates-ai-open-ended-loss.json"
OPEN_ENDED_COUNT_FILE = config.SELECTIONS + "ultimates-ai-open-ended-count.json"
EXCEL_FILE = config.SELECTIONS + "Ultimates.xlsx"


def update_sheet_selections(ws, periods_data, selection_type="rules-based"):
    """
    Update the selection and reasoning columns in a single sheet.
    
    Args:
        ws: openpyxl worksheet object
        periods_data: Dict mapping period -> {'selection': value, 'reasoning': text}
        selection_type: "rules-based" or "open-ended" - determines which columns to update
    
    Returns:
        Number of selections updated
    """
    updates_made = 0
    
    # Build column map for dynamic column lookups
    col_map = build_column_map(ws, header_row=1)
    
    # Determine which columns to update based on selection type
    if selection_type == "rules-based":
        sel_header = "Rules-Based AI Selection"
        reason_header = "Rules-Based AI Reasoning"
    else:  # open-ended
        sel_header = "Open-Ended AI Selection"
        reason_header = "Open-Ended AI Reasoning"
    
    # Get column indices from the map
    sel_col = col_map.get(sel_header)
    reason_col = col_map.get(reason_header)
    
    if sel_col is None or reason_col is None:
        print(f"  WARNING: Could not find columns '{sel_header}' or '{reason_header}' in sheet {ws.title}")
        return 0
    
    # Iterate through rows starting at 2 (row 1 is headers)
    row = 2
    while True:
        period_cell = ws.cell(row=row, column=1)
        if not period_cell.value:
            break
            
        period = str(period_cell.value).strip()
        if period in periods_data:
            ws.cell(row=row, column=sel_col).value = float(periods_data[period]['selection'])
            ws.cell(row=row, column=reason_col).value = str(periods_data[period]['reasoning'])
            updates_made += 1
        row += 1
        
    return updates_made


def load_category_json_file(filepath, category_name):
    """Load a single category JSON file (loss or count)."""
    if not pathlib.Path(filepath).exists():
        return []
    
    try:
        with open(filepath, 'r') as f:
            data = json.load(f)
            # If it's a single object, wrap it in a list
            if isinstance(data, dict):
                data = [data]
            return data
    except Exception as e:
        print(f"  WARNING: Failed to load {filepath}: {e}")
        return []


def main():
    """Update Ultimates Excel file with selections from category JSON files."""
    # Load rules-based selections for Loss and Count
    rules_based_loss = load_category_json_file(RULES_BASED_LOSS_FILE, "Loss")
    rules_based_count = load_category_json_file(RULES_BASED_COUNT_FILE, "Count")
    
    # Load open-ended selections for Loss and Count
    open_ended_loss = load_category_json_file(OPEN_ENDED_LOSS_FILE, "Loss")
    open_ended_count = load_category_json_file(OPEN_ENDED_COUNT_FILE, "Count")
    
    # Check if we have any selections to apply
    has_selections = any([rules_based_loss, rules_based_count, open_ended_loss, open_ended_count])
    
    if not has_selections:
        print("No selections found in any file")
        print("Skipping update - selections arrays are empty.")
        return
        
    print(f"Loaded {len(rules_based_loss)} rules-based Loss selections")
    print(f"Loaded {len(rules_based_count)} rules-based Count selections")
    print(f"Loaded {len(open_ended_loss)} open-ended Loss selections")
    print(f"Loaded {len(open_ended_count)} open-ended Count selections")
    
    # Organize selections by period for each category
    def organize_by_period(selections):
        by_period = {}
        for entry in selections:
            period = str(entry['period'])
            by_period[period] = {
                'selection': entry.get('selection', entry.get('selected_ultimate')),
                'reasoning': entry['reasoning']
            }
        return by_period
    
    rules_based_loss_by_period = organize_by_period(rules_based_loss)
    rules_based_count_by_period = organize_by_period(rules_based_count)
    open_ended_loss_by_period = organize_by_period(open_ended_loss)
    open_ended_count_by_period = organize_by_period(open_ended_count)
    
    print(f"Opening workbook {EXCEL_FILE}")
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"Excel file not found: {EXCEL_FILE}")
        return
        
    total_updates_rb = 0
    total_updates_oe = 0
    
    # Update Losses sheet
    if 'Losses' in wb.sheetnames:
        ws = wb['Losses']
        if rules_based_loss_by_period:
            updates = update_sheet_selections(ws, rules_based_loss_by_period, "rules-based")
            total_updates_rb += updates
            print(f"  Updated {updates} rules-based selections in 'Losses'")
        if open_ended_loss_by_period:
            updates = update_sheet_selections(ws, open_ended_loss_by_period, "open-ended")
            total_updates_oe += updates
            print(f"  Updated {updates} open-ended selections in 'Losses'")
    else:
        print("  WARNING: Sheet 'Losses' not found in workbook")
    
    # Update Counts sheet
    if 'Counts' in wb.sheetnames:
        ws = wb['Counts']
        if rules_based_count_by_period:
            updates = update_sheet_selections(ws, rules_based_count_by_period, "rules-based")
            total_updates_rb += updates
            print(f"  Updated {updates} rules-based selections in 'Counts'")
        if open_ended_count_by_period:
            updates = update_sheet_selections(ws, open_ended_count_by_period, "open-ended")
            total_updates_oe += updates
            print(f"  Updated {updates} open-ended selections in 'Counts'")
    else:
        print("  WARNING: Sheet 'Counts' not found in workbook")
                
    wb.save(EXCEL_FILE)
    print(f"\nSaved: {EXCEL_FILE}")
    print(f"Total rules-based updates: {total_updates_rb}")
    print(f"Total open-ended updates: {total_updates_oe}")


if __name__ == "__main__":
    main()
