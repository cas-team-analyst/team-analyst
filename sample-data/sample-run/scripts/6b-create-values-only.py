# Reads Analysis.xlsx (produced by 6a) and fills formula cells with
# values computed from selection Excel workbooks to produce Analysis - Values Only.xlsx.
# The Values Only file is safe to read with openpyxl/pandas without Excel.
#
# This script reads LDF selections from Chain Ladder Selections - LDFs.xlsx
# and Chain Ladder Selections - Tail.xlsx (respecting User > Rules-Based AI priority),
# then computes triangle and method values using those selections.
#
# Prereq: 6a-create-complete-analysis.py must have run first.
#
# run-note: Run from the scripts/ directory:
#     cd scripts/
#     python 6b-create-values-only.py

import os
import pathlib

import pandas as pd
from openpyxl import load_workbook

from modules import config
from modules.xl_utils import measure_short_name
from modules.xl_selections import find_selected_values
from modules.xl_values import (
    _has_method,
    _fill_method_cl_values,
    _fill_method_bf_values,
    _fill_method_ie_values,
    _fill_selection_values,
    _strip_formulas,
)
from modules.analysis_loaders import (
    MEASURE_TO_CATEGORY,
    load_selections,
    load_selection_reasoning,
    load_combined,
    get_exposure,
)


def load_selected_ldfs_from_excel(cl_excel_path):
    """
    Load selected LDF values from Chain Ladder Selections - LDFs.xlsx.
    Returns dict: {measure: {interval: ldf_value}}
    Priority: User Selection > Rules-Based AI Selection
    """
    if not pathlib.Path(cl_excel_path).exists():
        return {}
    
    wb = load_workbook(cl_excel_path, data_only=True)
    ldf_map = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find LDF Selections section
        in_ldf_section = False
        interval_headers = []
        selected_vals = None
        
        for row_cells in ws.iter_rows():
            col1 = row_cells[0].value if row_cells else None
            
            if col1 == "LDF Selections":
                in_ldf_section = True
                continue
            
            if in_ldf_section:
                # Next row after "LDF Selections" should be interval headers
                if not interval_headers:
                    interval_headers = [c.value for c in row_cells[1:]]
                    continue
                
                # Use find_selected_values to get priority selection
                if col1 in ("User Selection", "Rules-Based AI Selection"):
                    # We've reached the selection rows - use xl_selections logic
                    selected_vals = find_selected_values(ws)
                    break
        
        if selected_vals and interval_headers:
            # Build map: interval -> LDF value
            interval_map = {}
            for idx, (interval, val) in enumerate(zip(interval_headers, selected_vals)):
                if interval and val is not None:
                    interval_map[str(interval)] = val
            
            ldf_map[sheet_name] = interval_map
    
    wb.close()
    return ldf_map


def fill_triangle_with_excel_ldfs(ws, measure, ldf_map, df2):
    """
    Fill triangle sheet using LDF selections from Excel workbook.
    Falls back to df2 parquet data for ATA values (raw triangle data).
    """
    measure_ldfs = ldf_map.get(measure, {})
    
    # Get triangle data from parquet for ATA section
    df_m = df2[df2['measure'].astype(str) == measure].copy() if not df2.empty else pd.DataFrame()
    
    ata_lookup = {
        (str(r['period']), str(r['interval'])): r['ldf']
        for _, r in df_m[df_m['ldf'].notna()].iterrows()
    } if not df_m.empty else {}
    
    section = None
    col_headers = {}
    
    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None
        
        if col1 == "Age-to-Age Factors":
            section = "ata_pre_header"
            continue
        if col1 == "Averages":
            section = "avg"  # Skip averages - use Excel selections
            continue
        if col1 == "LDF Selections":
            section = "ldf"
            continue
        
        if section == "ata_pre_header":
            col_headers = {c.column: str(c.value) for c in row_cells[1:] if c.value not in (None, "")}
            section = "ata"
            continue
        
        if section == "ata":
            if col1 is None or col1 == "":
                section = None
                continue
            period = str(col1)
            for cell in row_cells[1:]:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    interval = col_headers.get(cell.column, "")
                    val = ata_lookup.get((period, interval))
                    if val is not None:
                        cell.value = val
        
        if section == "ldf":
            # Fill Selected row with values from Excel
            if col1 == "Selected":
                for cell in row_cells[1:]:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        interval = col_headers.get(cell.column, "")
                        if interval in measure_ldfs:
                            cell.value = measure_ldfs[interval]


def fill_cdf_row(ws):
    """Compute CDF row from Selected LDF row (right-to-left cumulative product)"""
    selected_row = None
    cdf_row = None
    last_col = 0
    
    for row_cells in ws.iter_rows():
        col1 = row_cells[0].value if row_cells else None
        if col1 == "Selected":
            selected_row = row_cells
            last_col = max(c.column for c in row_cells if c.value not in (None, ""))
        elif col1 == "CDF" and selected_row:
            cdf_row = row_cells
            break
    
    if not selected_row or not cdf_row:
        return
    
    # Compute CDFs right-to-left
    cdf_vals = {}
    for col in range(last_col, 1, -1):
        ldf = selected_row[col - 1].value
        if ldf and isinstance(ldf, (int, float)):
            next_cdf = cdf_vals.get(col + 1, 1.0)
            cdf_vals[col] = ldf * next_cdf
    
    # Write CDF values
    for col, cdf_val in cdf_vals.items():
        if col - 1 < len(cdf_row):
            cdf_row[col - 1].value = cdf_val


def fill_cv_slopes_values(ws, df4):
    """
    Fill CV & Slopes sheet with values from 4_ldf_averages.parquet.
    Sheet has sections for each measure with metrics: cv_3yr, cv_5yr, cv_10yr, slope_3yr, slope_5yr, slope_10yr.
    """
    if df4.empty:
        return
    
    # Scan sheet to find sections and intervals
    row_idx = 1
    max_row = ws.max_row if ws.max_row else 100
    
    while row_idx <= max_row:
        cell_val = ws.cell(row_idx, 1).value
        
        # Check if this is a measure section header (ALL CAPS like INCURRED LOSS)
        if cell_val and isinstance(cell_val, str) and cell_val.isupper() and len(cell_val) > 5:
            # This is a section header
            measure_upper = cell_val
            # Convert back to title case for matching parquet data
            measure = measure_upper.title()
            
            # Next row should be "Metric" header with intervals
            if row_idx + 1 <= max_row and ws.cell(row_idx + 1, 1).value == "Metric":
                # Get intervals from header row
                intervals = []
                col_idx = 2
                while col_idx <= 20:  # Reasonable max columns
                    interval_val = ws.cell(row_idx + 1, col_idx).value
                    if interval_val:
                        intervals.append((col_idx, str(interval_val)))
                    else:
                        break
                    col_idx += 1
                
                # Get data for this measure
                df_measure = df4[df4['measure'] == measure].copy()
                
                # Process metric rows
                metric_row_idx = row_idx + 2
                metrics = ['cv_3yr', 'cv_5yr', 'cv_10yr', 'slope_3yr', 'slope_5yr', 'slope_10yr']
                for metric in metrics:
                    if metric_row_idx > max_row:
                        break
                    
                    metric_label = ws.cell(metric_row_idx, 1).value
                    if metric_label == metric:
                        # Fill values for each interval
                        for col_idx, interval in intervals:
                            # Find value in parquet
                            interval_key = "Tail" if interval == "Tail" else interval
                            val_row = df_measure[df_measure['interval'] == interval_key]
                            if not val_row.empty and metric in val_row.columns:
                                val = val_row[metric].iloc[0]
                                if pd.notna(val):
                                    ws.cell(metric_row_idx, col_idx).value = val
                        
                        metric_row_idx += 1
                
                row_idx = metric_row_idx
            else:
                row_idx += 1
        else:
            row_idx += 1


def fill_diagnostics_values(ws, df3):
    """
    Fill Diagnostics sheet with values from 3_diagnostics.parquet.
    Sheet has sections for each metric with periods as rows and ages as columns.
    Metrics: incurred_severity, paid_to_incurred, incremental_incurred_severity
    """
    if df3.empty:
        return
    
    # Scan sheet to find sections and ages
    row_idx = 1
    max_row = ws.max_row if ws.max_row else 200
    
    while row_idx <= max_row:
        cell_val = ws.cell(row_idx, 1).value
        
        # Check if this is a metric section header (ALL CAPS)
        if cell_val and isinstance(cell_val, str) and cell_val.isupper() and len(cell_val) > 5:
            # This is a section header - convert to parquet column name
            metric_upper = cell_val
            metric_lower = metric_upper.lower().replace(' ', '_')
            
            # Next row should be "Period" header with ages
            if row_idx + 1 <= max_row and ws.cell(row_idx + 1, 1).value == "Period":
                # Get ages from header row - keep as strings to match parquet
                ages = []
                col_idx = 2
                while col_idx <= 20:  # Reasonable max columns
                    age_val = ws.cell(row_idx + 1, col_idx).value
                    if age_val and age_val != "Tail":
                        # Keep as string (parquet has string category dtype)
                        age_str = str(age_val)
                        ages.append((col_idx, age_str))
                    elif age_val == "Tail":
                        break
                    else:
                        break
                    col_idx += 1
                
                # Process period rows
                period_row_idx = row_idx + 2
                while period_row_idx <= max_row:
                    period_val = ws.cell(period_row_idx, 1).value
                    if not period_val:
                        break
                    
                    # Convert period to string to match parquet (category dtype with string values)
                    period_str = str(period_val)
                    # Check if it looks like a valid period (all digits)
                    if not period_str.isdigit():
                        # Not a valid period, end of section
                        break
                    
                    # Fill values for each age
                    for col_idx, age in ages:
                        # Find value in parquet
                        val_row = df3[(df3['period'] == period_str) & (df3['age'] == age)]
                        if not val_row.empty and metric_lower in val_row.columns:
                            val = val_row[metric_lower].iloc[0]
                            if pd.notna(val):
                                ws.cell(period_row_idx, col_idx).value = val
                    
                    period_row_idx += 1
                
                row_idx = period_row_idx
            else:
                row_idx += 1
        else:
            row_idx += 1

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_ULTIMATES        = config.ULTIMATES + "projected-ultimates.parquet"
INPUT_TRIANGLES        = config.PROCESSED_DATA + "1_triangles.parquet"
INPUT_SELECTIONS_EXCEL = config.SELECTIONS + "Ultimates.xlsx"
INPUT_COMPLETE         = config.BASE_DIR + "Analysis.xlsx"
OUTPUT_VALUES          = config.BASE_DIR + "Analysis - Values Only.xlsx"

INPUT_CL_EXCEL      = config.SELECTIONS + "Chain Ladder Selections - LDFs.xlsx"
INPUT_CL_ENHANCED   = config.PROCESSED_DATA + "2_enhanced.parquet"
INPUT_DIAGNOSTICS   = config.PROCESSED_DATA + "3_diagnostics.parquet"
INPUT_LDF_AVERAGES  = config.PROCESSED_DATA + "4_ldf_averages.parquet"

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not pathlib.Path(INPUT_COMPLETE).exists():
        raise FileNotFoundError(
            f"{INPUT_COMPLETE} not found. Run 6a-create-complete-analysis.py first."
        )

    # Load ultimate selections from Ultimates.xlsx (for method selections)
    sel_lookup = load_selections(INPUT_SELECTIONS_EXCEL)
    combined, _ = load_combined(INPUT_ULTIMATES, sel_lookup)
    reason_lookup = load_selection_reasoning(INPUT_SELECTIONS_EXCEL)
    combined["selected_reasoning"] = combined.apply(
        lambda row: reason_lookup.get(
            (MEASURE_TO_CATEGORY.get(row["measure"], row["measure"]), row["period"]),
            ""
        ) or "",
        axis=1
    )
    exp_map  = get_exposure(INPUT_TRIANGLES)
    measures = [m for m in combined["measure"].unique() if m != "Exposure"]

    print(f"Measures: {measures}")
    print("Loading Analysis.xlsx...")
    wb = load_workbook(INPUT_COMPLETE, data_only=False)

    # Load LDF selections from Chain Ladder Selections - LDFs.xlsx (User > Rules-Based AI)
    print("Reading LDF selections from Chain Ladder Selections - LDFs.xlsx...")
    ldf_map = load_selected_ldfs_from_excel(INPUT_CL_EXCEL)
    
    # Load triangle data for ATA section
    df2_enh = pd.read_parquet(INPUT_CL_ENHANCED) if pathlib.Path(INPUT_CL_ENHANCED).exists() else pd.DataFrame()
    # Load diagnostics for Diagnostics sheet
    df3 = pd.read_parquet(INPUT_DIAGNOSTICS) if pathlib.Path(INPUT_DIAGNOSTICS).exists() else pd.DataFrame()
    # Load LDF averages for CV & Slopes sheet
    df4 = pd.read_parquet(INPUT_LDF_AVERAGES) if pathlib.Path(INPUT_LDF_AVERAGES).exists() else pd.DataFrame()
    measures_short_set = {measure_short_name(m) for m in measures}
    actual_lookup_full = combined.set_index(["period", "measure"])["actual"].to_dict()

    print("Computing formula values for Values Only...")
    for sname in wb.sheetnames:
        ws = wb[sname]
        # Triangle sheets: fill with selected LDF values from Chain Ladder Selections - LDFs.xlsx
        if sname in measures_short_set:
            full_m = next((m for m in measures if measure_short_name(m) == sname), None)
            if full_m:
                fill_triangle_with_excel_ldfs(ws, full_m, ldf_map, df2_enh)
                fill_cdf_row(ws)
            _strip_formulas(ws)
        # Method sheets: use combined ultimates
        elif sname.endswith(" CL"):
            full_m = next((m for m in measures if measure_short_name(m) == sname[:-3]), None)
            if full_m:
                _fill_method_cl_values(ws, full_m, combined, actual_lookup_full)
            _strip_formulas(ws)
        elif sname.endswith(" BF"):
            full_m = next((m for m in measures if measure_short_name(m) == sname[:-3]), None)
            if full_m:
                _fill_method_bf_values(ws, full_m, combined, actual_lookup_full)
            _strip_formulas(ws)
        elif sname == "IE":
            # IE sheet is always for Incurred Loss
            full_m = "Incurred Loss"
            if _has_method(combined, full_m, "ultimate_ie"):
                _fill_method_ie_values(ws, full_m, combined, exp_map)
            _strip_formulas(ws)
        elif sname == "Loss Selection":
            _fill_selection_values(ws, ["Incurred Loss", "Paid Loss"], combined, actual_lookup_full)
            _strip_formulas(ws)
        elif sname == "Count Selection":
            _fill_selection_values(ws, ["Reported Count", "Closed Count"], combined, actual_lookup_full)
            _strip_formulas(ws)
        elif sname == "Diagnostics":
            fill_diagnostics_values(ws, df3)
            _strip_formulas(ws)
        elif sname == "CV & Slopes":
            fill_cv_slopes_values(ws, df4)
            _strip_formulas(ws)
        else:
            _strip_formulas(ws)

    os.makedirs(pathlib.Path(OUTPUT_VALUES).parent, exist_ok=True)
    wb.save(OUTPUT_VALUES)
    print(f"  Saved -> {OUTPUT_VALUES}")

    # Verify formulas if on Windows with Excel available
    if os.name == "nt":
        try:
            import win32com.client
            from modules.verify_formulas import run_verify
            run_verify()
        except ImportError:
            print("  (Skipping formula verification - pywin32 not installed)")
    else:
        print("  (Skipping formula verification - not on Windows)")

if __name__ == "__main__":
    main()
