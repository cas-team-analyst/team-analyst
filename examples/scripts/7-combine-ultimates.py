"""
Script 7: Combine ultimates, compute post-method diagnostics, write full analysis.

Merges CL, BF, and IE ultimates. Computes post-method diagnostics.
Writes the canonical output files and a master full-analysis workbook.

Usage (run from the project root):
    python scripts/7-combine-ultimates.py

Inputs:
    output/prep/triangles.pkl
    output/prep/diagonal.pkl
    output/chain-ladder/cl_ultimates.csv
    output/initial-expected/ie_ultimates.csv
    output/bornhuetter-ferguson/bf_ultimates.csv

Outputs:
    output/selected-ultimates/
        selected-ultimates.xlsx     (Loss + Counts sheets)
    output/post-method/
        post-method-series-diagnostics.xlsx     (ultimate severity, loss rate, frequency)
        post-method-triangle-diagnostics.xlsx   (X-to-ultimate triangles + avg IBNR/Unpaid)
    output/
        full-analysis.xlsx          (all sheets from all output files combined)
"""

import pandas as pd
import numpy as np
import os


def _try_int(series):
    """Convert series to int if all values are numeric, else return as-is."""
    try:
        return series.astype(str).apply(lambda x: int(float(x)))
    except (ValueError, TypeError):
        return series.astype(str)


INPUT_TRIANGLES = "output/prep/triangles.pkl"
INPUT_DIAGONAL  = "output/prep/diagonal.pkl"
INPUT_CL        = "output/chain-ladder/cl_ultimates.csv"
INPUT_IE        = "output/initial-expected/ie_ultimates.csv"
INPUT_BF        = "output/bornhuetter-ferguson/bf_ultimates.csv"
OUTPUT_SEL      = "output/selected-ultimates"
OUTPUT_POST     = "output/post-method"
OUTPUT_ROOT     = "output"

UNPAID_PROXY = {
    "Incurred Loss":  "Paid Loss",
    "Paid Loss":      "Paid Loss",
    "Reported Count": "Closed Count",
    "Closed Count":   "Closed Count",
}


def load_all(diag, cl, ie, bf):
    """Merge CL, IE, BF into one wide table per period x measure."""
    for df in (diag, cl, ie, bf):
        for col in ("period", "measure"):
            if col in df.columns:
                df[col] = df[col].astype(str)

    diag_lookup = diag.set_index(["period", "measure"])["value"].to_dict()

    # Build unified frame indexed by (period, measure)
    cl_map  = cl.set_index(["period","measure"])[["actual","cl_ultimate","cl_ibnr"]].to_dict("index")
    ie_map  = ie.set_index(["period","measure"])["expected_ultimate"].to_dict()
    bf_map  = bf.set_index(["period","measure"])[["bf_ultimate","bf_ibnr"]].to_dict("index")

    periods  = sorted(diag["period"].unique(), key=lambda x: (int(x) if x.isdigit() else x))
    measures = [m for m in ["Incurred Loss","Paid Loss","Reported Count","Closed Count"]
                if m in diag["measure"].values]

    rows = []
    for period in periods:
        for measure in measures:
            actual     = diag_lookup.get((period, measure), np.nan)
            cl_row     = cl_map.get((period, measure), {})
            ie_ult     = ie_map.get((period, measure), np.nan)
            bf_row     = bf_map.get((period, measure), {})

            cl_ult  = cl_row.get("cl_ultimate",  np.nan)
            cl_ibnr = cl_row.get("cl_ibnr",      np.nan)
            ie_ult_v= ie_ult
            bf_ult  = bf_row.get("bf_ultimate",  np.nan)
            bf_ibnr = bf_row.get("bf_ibnr",      np.nan)

            # Default selected = BF (good blend of credibility)
            sel_ult = bf_ult if pd.notna(bf_ult) else (cl_ult if pd.notna(cl_ult) else ie_ult_v)

            proxy  = UNPAID_PROXY.get(measure)
            proxy_actual = diag_lookup.get((period, proxy), actual)
            sel_ibnr     = (sel_ult - actual)      if pd.notna(sel_ult) else np.nan
            sel_unpaid   = (sel_ult - proxy_actual) if pd.notna(sel_ult) else np.nan

            rows.append(dict(
                period          = period,
                measure         = measure,
                actual          = actual,
                cl_ultimate     = cl_ult,
                ie_ultimate     = ie_ult_v,
                bf_ultimate     = bf_ult,
                selected_ultimate = sel_ult,
                selected_ibnr   = sel_ibnr,
                selected_unpaid = sel_unpaid,
            ))

    return pd.DataFrame(rows)


def get_current_ages(diag):
    diag = diag.copy()
    diag["period"]  = diag["period"].astype(str)
    diag["measure"] = diag["measure"].astype(str)
    diag["age_int"] = _try_int(diag["age"])
    latest = diag.sort_values("age_int").groupby(["period","measure"]).last()
    return latest["age_int"].to_dict()


def get_exposure(diag):
    exp = diag[diag["measure"].astype(str) == "Exposure"].copy()
    if exp.empty:
        return {}
    exp["period"]  = exp["period"].astype(str)
    exp["age_int"] = _try_int(exp["age"])
    latest = exp.sort_values("age_int").groupby("period").last()
    return latest["value"].to_dict()


def write_selected_ultimates(combined, ages_map, exp_map, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # ── Loss sheet ──────────────────────────────────────────────────────
        inc  = combined[combined["measure"] == "Incurred Loss"].set_index("period")
        paid = combined[combined["measure"] == "Paid Loss"].set_index("period")

        if not inc.empty:
            out = pd.DataFrame({
                "Accident Period":  _try_int(inc.index.to_series()),
                "Current Age":      [ages_map.get((p,"Incurred Loss"), np.nan) for p in inc.index],
                "Incurred":         inc["actual"],
                "Paid":             [combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss"), "actual"].values[0]
                                     if len(combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss"), "actual"]) else np.nan
                                     for p in inc.index],
                "Incurred CL":      inc["cl_ultimate"],
                "Paid CL":          [combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss"), "cl_ultimate"].values[0]
                                     if len(combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss")]) else np.nan
                                     for p in inc.index],
                "Incurred BF":      inc["bf_ultimate"],
                "Paid BF":          [combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss"), "bf_ultimate"].values[0]
                                     if len(combined.loc[(combined["period"]==p) &
                                     (combined["measure"]=="Paid Loss")]) else np.nan
                                     for p in inc.index],
                "Initial Expected": inc["ie_ultimate"],
                "Selected Ultimate":inc["selected_ultimate"],
                "IBNR":             inc["selected_ibnr"],
                "Unpaid":           inc["selected_unpaid"],
            })
            out.to_excel(writer, sheet_name="Loss", index=False)

        # ── Counts sheet ─────────────────────────────────────────────────────
        rep = combined[combined["measure"] == "Reported Count"].set_index("period")
        cls = combined[combined["measure"] == "Closed Count"].set_index("period")

        if not rep.empty:
            def get_val(period, measure, col):
                sub = combined[(combined["period"]==period) & (combined["measure"]==measure)]
                return sub[col].values[0] if not sub.empty else np.nan

            out_c = pd.DataFrame({
                "Accident Period":   _try_int(rep.index.to_series()),
                "Current Age":       [ages_map.get((p,"Reported Count"), np.nan) for p in rep.index],
                "Reported":          rep["actual"],
                "Closed":            [get_val(p,"Closed Count","actual") for p in rep.index],
                "Reported CL":       rep["cl_ultimate"],
                "Closed CL":         [get_val(p,"Closed Count","cl_ultimate") for p in rep.index],
                "Reported BF":       rep["bf_ultimate"],
                "Closed BF":         [get_val(p,"Closed Count","bf_ultimate") for p in rep.index],
                "Initial Expected":  rep["ie_ultimate"],
                "Selected Ultimate": rep["selected_ultimate"],
                "IBNR":              rep["selected_ibnr"],
                "Unpaid":            rep["selected_unpaid"],
            })
            out_c.to_excel(writer, sheet_name="Counts", index=False)

    print(f"  Saved -> {path}")


def write_post_method_series(combined, exp_map, path):
    """Post-method series diagnostics: Ultimate Severity, Loss Rate, Frequency."""
    # Use selected ultimates for Incurred and Reported
    inc = combined[combined["measure"] == "Incurred Loss"].set_index("period")
    rep = combined[combined["measure"] == "Reported Count"].set_index("period")

    periods = sorted(inc.index, key=lambda x: (int(x) if str(x).isdigit() else str(x)))
    rows = []
    for p in periods:
        if p not in inc.index:
            continue
        ult_loss   = inc.loc[p, "selected_ultimate"]
        ult_counts = rep.loc[p, "selected_ultimate"] if p in rep.index else np.nan
        exp        = exp_map.get(p, np.nan)

        sev  = ult_loss / ult_counts  if pd.notna(ult_loss) and pd.notna(ult_counts) and ult_counts != 0 else np.nan
        lr   = ult_loss / exp         if pd.notna(ult_loss) and pd.notna(exp) and exp != 0 else np.nan
        freq = ult_counts / exp       if pd.notna(ult_counts) and pd.notna(exp) and exp != 0 else np.nan

        rows.append(dict(
            **{"Accident Period": _try_int(pd.Series([p])).iloc[0],
               "Ultimate Severity":   round(sev,  6) if pd.notna(sev)  else np.nan,
               "Ultimate Loss Rate":  round(lr,   6) if pd.notna(lr)   else np.nan,
               "Ultimate Frequency":  round(freq, 6) if pd.notna(freq) else np.nan}
        ))

    diag_out = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        diag_out.to_excel(writer, sheet_name="Diagnostics", index=False)
    print(f"  Saved -> {path}")


def write_post_method_triangles(triangles_df, combined, path):
    """Post-method triangle diagnostics: X-to-Ultimate and Average IBNR/Unpaid."""
    # Selected ultimates lookup
    sel_lookup = combined.set_index(["period","measure"])["selected_ultimate"].to_dict()

    # Prep the main triangle data
    df = triangles_df.copy()
    df["period_int"] = _try_int(df["period"])
    df["age_int"]    = _try_int(df["age"])

    def x_to_ult(measure, label):
        sub = df[df["measure"] == measure]
        if sub.empty:
            return None
        pivot = sub.pivot(index="period_int", columns="age_int", values="value")
        pivot.columns.name = None
        # Divide each row by its selected ultimate
        result = pivot.copy().astype(float)
        for period in result.index:
            sel = sel_lookup.get((str(period), measure), np.nan)
            if pd.notna(sel) and sel != 0:
                result.loc[period] = result.loc[period] / sel
        result.index.name   = label
        result.columns.name = None
        return result

    # Average IBNR = selected IBNR per period / selected count of ultimate
    # = (Ultimate - Incurred) / Reported
    inc = combined[combined["measure"]=="Incurred Loss"].set_index("period")
    rep = combined[combined["measure"]=="Reported Count"].set_index("period")
    exp_df = df[df["measure"]=="Exposure"]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for meas, label in [("Incurred Loss","Incurred-to-Ultimate"),
                             ("Paid Loss","Paid-to-Ultimate"),
                             ("Reported Count","Reported-to-Ultimate"),
                             ("Closed Count","Closed-to-Ultimate")]:
            t = x_to_ult(meas, label)
            if t is not None:
                t.to_excel(writer, sheet_name=label)

        # Average IBNR = (Ultimate - Incurred diagonal) / Reported at each age
        rep_triangle = df[df["measure"]=="Reported Count"].pivot(
            index="period_int", columns="age_int", values="value")
        rep_triangle.columns.name = None

        if not rep_triangle.empty and not inc.empty:
            avg_ibnr = rep_triangle.copy().astype(float)
            for period in avg_ibnr.index:
                sel_ult = sel_lookup.get((str(period), "Incurred Loss"), np.nan)
                if pd.notna(sel_ult):
                    avg_ibnr.loc[period] = (sel_ult - avg_ibnr.loc[period]) / 1
            avg_ibnr.index.name   = "Average IBNR"
            avg_ibnr.columns.name = None
            avg_ibnr.to_excel(writer, sheet_name="Average IBNR")

        # Average Unpaid = (Ultimate - Paid diagonal) / Reported at each age
        paid_triangle = df[df["measure"]=="Paid Loss"].pivot(
            index="period_int", columns="age_int", values="value")
        paid_triangle.columns.name = None

        if not paid_triangle.empty and not inc.empty:
            avg_unpaid = paid_triangle.copy().astype(float)
            for period in avg_unpaid.index:
                sel_ult = sel_lookup.get((str(period), "Incurred Loss"), np.nan)
                if pd.notna(sel_ult):
                    avg_unpaid.loc[period] = sel_ult - avg_unpaid.loc[period]
            avg_unpaid.index.name   = "Average Unpaid"
            avg_unpaid.columns.name = None
            avg_unpaid.to_excel(writer, sheet_name="Average Unpaid")

    print(f"  Saved -> {path}")


def write_full_analysis(path):
    """Combine all output Excel files into a single master workbook."""
    import glob
    from openpyxl import load_workbook

    source_files = [
        ("data-processing/long.xlsx",                         None),
        ("data-processing/triangles.xlsx",                    None),
        ("data-processing/pre-method-triangle-diagnostics.xlsx", None),
        ("chain-ladder/chain-ladder.xlsx",                    "CL - "),
        ("initial-expected/initial-expected.xlsx",            "IE - "),
        ("bornhuetter-ferguson/bornhuetter-ferguson.xlsx",    "BF - "),
        ("selected-ultimates/selected-ultimates.xlsx",        "Sel - "),
        ("post-method/post-method-series-diagnostics.xlsx",   "Post - "),
        ("post-method/post-method-triangle-diagnostics.xlsx", "Post - "),
    ]

    from openpyxl import Workbook
    master = Workbook()
    master.remove(master.active)  # remove default sheet

    for rel_path, prefix in source_files:
        full_path = f"output/{rel_path}"
        if not os.path.exists(full_path):
            print(f"  Skipping (not found): {full_path}")
            continue
        wb = load_workbook(full_path)
        for sname in wb.sheetnames:
            ws_src = wb[sname]
            new_name = f"{prefix}{sname}" if prefix else sname
            new_name = new_name[:31]  # Excel sheet name limit
            ws_dst = master.create_sheet(title=new_name)
            for row in ws_src.iter_rows():
                for cell in row:
                    ws_dst[cell.coordinate].value = cell.value

    master.save(path)
    print(f"  Saved -> {path}")


def main():
    os.makedirs(OUTPUT_SEL,  exist_ok=True)
    os.makedirs(OUTPUT_POST, exist_ok=True)

    triangles_df = pd.read_pickle(INPUT_TRIANGLES)
    diag         = pd.read_pickle(INPUT_DIAGONAL)
    diag["period"]  = diag["period"].astype(str)
    diag["measure"] = diag["measure"].astype(str)
    diag["age"]     = diag["age"].astype(str)

    cl   = pd.read_csv(INPUT_CL)
    ie   = pd.read_csv(INPUT_IE)
    bf   = pd.read_csv(INPUT_BF)

    combined = load_all(diag, cl, ie, bf)
    ages_map = get_current_ages(diag)
    exp_map  = get_exposure(diag)

    # ── Selected Ultimates ─────────────────────────────────────────────────
    write_selected_ultimates(
        combined, ages_map, exp_map,
        f"{OUTPUT_SEL}/selected-ultimates.xlsx"
    )

    # ── Post-Method Diagnostics ────────────────────────────────────────────
    write_post_method_series(
        combined, exp_map,
        f"{OUTPUT_POST}/post-method-series-diagnostics.xlsx"
    )
    write_post_method_triangles(
        triangles_df, combined,
        f"{OUTPUT_POST}/post-method-triangle-diagnostics.xlsx"
    )

    # ── Full Analysis Workbook ─────────────────────────────────────────────
    write_full_analysis(f"{OUTPUT_ROOT}/full-analysis.xlsx")

    # ── Print IBNR summary ─────────────────────────────────────────────────
    print("\n=== IBNR Summary ===")
    pd.set_option("display.float_format", lambda x: f"{x:,.0f}")
    for m in ["Incurred Loss", "Paid Loss", "Reported Count", "Closed Count"]:
        sub = combined[combined["measure"] == m]
        if sub.empty or sub["selected_ultimate"].isna().all():
            continue
        print(f"  {m}: "
              f"Actual={sub['actual'].sum():,.0f}  "
              f"CL={sub['cl_ultimate'].sum():,.0f}  "
              f"IE={sub['ie_ultimate'].sum():,.0f}  "
              f"BF={sub['bf_ultimate'].sum():,.0f}  "
              f"Selected={sub['selected_ultimate'].sum():,.0f}  "
              f"IBNR={sub['selected_ibnr'].sum():,.0f}")


if __name__ == "__main__":
    print("=== Step 7: Combining ultimates and writing final outputs ===")
    main()
